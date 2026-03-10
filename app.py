"""
Sistema de Refeições — Interface Web (Flask)
============================================
Corre com:  python app.py
Acede em:   http://localhost:8080
"""

import os
import re
import sys
import secrets
import time
from datetime import date, datetime, timedelta
from functools import wraps

from markupsafe import Markup, escape
from werkzeug.security import (
    generate_password_hash as _gen_pw_hash,
    check_password_hash,
)
from werkzeug.middleware.proxy_fix import ProxyFix
from flask import (
    Flask,
    Response,
    render_template_string,
    request,
    redirect,
    url_for,
    session,
    flash,
    g,
    abort,
    send_file,
)

sys.path.insert(0, os.path.dirname(__file__))
import sistema_refeicoes_v8_4 as sr  # noqa: E402
import config as cfg  # noqa: E402


def generate_password_hash(password: str) -> str:
    """Wrapper que usa pbkdf2:sha256 como fallback quando scrypt não está disponível."""
    try:
        return _gen_pw_hash(password)
    except (ValueError, AttributeError):
        return _gen_pw_hash(password, method="pbkdf2:sha256")


# Garantir colunas extra e FTS no arranque
def _ensure_extra_schema():
    """Garante colunas extra e FTS5 nos utilizadores."""
    try:
        with sr.db() as conn:
            cols = [
                r["name"]
                for r in conn.execute("PRAGMA table_info(utilizadores)").fetchall()
            ]
            if "email" not in cols:
                conn.execute("ALTER TABLE utilizadores ADD COLUMN email TEXT")
            if "telemovel" not in cols:
                conn.execute("ALTER TABLE utilizadores ADD COLUMN telemovel TEXT")
            if "is_active" not in cols:
                conn.execute(
                    "ALTER TABLE utilizadores ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1"
                )
            if "turma_id" not in cols:
                conn.execute(
                    "ALTER TABLE utilizadores ADD COLUMN turma_id INTEGER REFERENCES turmas(id)"
                )

            # Colunas extra na tabela licencas (entradas/saídas)
            lic_cols = [
                r["name"]
                for r in conn.execute("PRAGMA table_info(licencas)").fetchall()
            ]
            if "hora_saida" not in lic_cols:
                conn.execute("ALTER TABLE licencas ADD COLUMN hora_saida TEXT")
            if "hora_entrada" not in lic_cols:
                conn.execute("ALTER TABLE licencas ADD COLUMN hora_entrada TEXT")

            # Verificar e reparar FTS5 se necessário (sem writable_schema)
            try:
                conn.execute("SELECT COUNT(*) FROM utilizadores_fts").fetchone()
            except Exception:
                print("[AVISO] FTS corrompida — a recriar...", flush=True)
                for trg in (
                    "utilizadores_ai_fts",
                    "utilizadores_ad_fts",
                    "utilizadores_au_fts",
                ):
                    try:
                        conn.execute(f"DROP TRIGGER IF EXISTS {trg}")
                    except Exception:
                        pass
                try:
                    conn.execute("DROP TABLE IF EXISTS utilizadores_fts")
                except Exception as e2:
                    print(f"[AVISO] DROP utilizadores_fts: {e2}", flush=True)

                conn.execute("""CREATE VIRTUAL TABLE IF NOT EXISTS utilizadores_fts
USING fts5(Nome_completo, content='utilizadores', content_rowid='id')""")
                conn.execute(
                    "INSERT OR IGNORE INTO utilizadores_fts(rowid, Nome_completo) SELECT id, Nome_completo FROM utilizadores"
                )
                conn.execute("""CREATE TRIGGER IF NOT EXISTS utilizadores_ai_fts
AFTER INSERT ON utilizadores BEGIN
  INSERT INTO utilizadores_fts(rowid, Nome_completo) VALUES (NEW.id, NEW.Nome_completo);
END""")
                conn.execute("""CREATE TRIGGER IF NOT EXISTS utilizadores_ad_fts
AFTER DELETE ON utilizadores BEGIN
  INSERT INTO utilizadores_fts(utilizadores_fts, rowid) VALUES('delete', OLD.id);
END""")
                conn.execute("""CREATE TRIGGER IF NOT EXISTS utilizadores_au_fts
AFTER UPDATE OF Nome_completo ON utilizadores BEGIN
  INSERT INTO utilizadores_fts(utilizadores_fts, rowid) VALUES('delete', OLD.id);
  INSERT INTO utilizadores_fts(rowid, Nome_completo) VALUES (NEW.id, NEW.Nome_completo);
END""")
                print("[INFO] FTS recriada com sucesso.", flush=True)

            # === Migrações pontuais (tabela de controlo) ===
            conn.execute(
                "CREATE TABLE IF NOT EXISTS _migracoes (nome TEXT PRIMARY KEY, aplicada_em TEXT)"
            )
            done = {
                r["nome"]
                for r in conn.execute("SELECT nome FROM _migracoes").fetchall()
            }

            # 1) Corrigir NI da aluna Reis 4º ano: 382 → 482
            if "reis_ni_382_482" not in done:
                reis = conn.execute(
                    "SELECT id FROM utilizadores WHERE NI='382' AND ano='4'"
                ).fetchone()
                if reis:
                    conn.execute(
                        "UPDATE utilizadores SET NI='482' WHERE id=?",
                        (reis["id"],),
                    )
                    print(
                        "[MIGRAÇÃO] NI da aluna Reis corrigido: 382→482",
                        flush=True,
                    )
                conn.execute(
                    "INSERT INTO _migracoes VALUES('reis_ni_382_482', datetime('now','localtime'))"
                )

            # 2) Corrigir NII da aluna Rafaela Fernandes: 20223 → 21223
            if "rafaela_nii_20223_21223" not in done:
                try:
                    cur = conn.execute(
                        "UPDATE utilizadores SET NII='21223' WHERE NII='20223'"
                    )
                    if cur.rowcount:
                        print(
                            f"[MIGRAÇÃO] NII Rafaela Fernandes corrigido: 20223→21223 (linhas={cur.rowcount})",
                            flush=True,
                        )
                except Exception as exc:
                    print(
                        f"[AVISO] Migração Rafaela NII 20223→21223 falhou: {exc}",
                        flush=True,
                    )
                conn.execute(
                    "INSERT INTO _migracoes VALUES('rafaela_nii_20223_21223', datetime('now','localtime'))"
                )

            # 3) Reset credenciais alunos: password→hash(NII), must_change=1
            #    Login = NII (ex: 24123), pw inicial = NII
            if "reset_creds_nii_v2" not in done:
                alunos_reset = conn.execute(
                    "SELECT id, NII FROM utilizadores WHERE perfil='aluno'"
                ).fetchall()
                for al in alunos_reset:
                    al = dict(al)
                    nii = al["NII"]
                    if not nii:
                        continue
                    pw_hash = generate_password_hash(nii)
                    conn.execute(
                        "UPDATE utilizadores SET Palavra_chave=?, must_change_password=1 WHERE id=?",
                        (pw_hash, al["id"]),
                    )
                conn.execute(
                    "INSERT INTO _migracoes VALUES('reset_creds_nii_v2', datetime('now','localtime'))"
                )
                print(
                    "[MIGRAÇÃO] Credenciais alunos resetadas: pw=hash(NII), must_change=1",
                    flush=True,
                )

            _bootstrap_dev_system_accounts(conn)
            conn.commit()
    except Exception as e:
        print(f"[ERRO] _ensure_extra_schema: {e}", flush=True)


# ═══════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
app.config.from_object(cfg.Config)
cfg.configure_logging(app)

# ── Teardown: fechar conexão SQLite no fim de cada request ────────────────
app.teardown_appcontext(sr.close_request_db)

# Expor constantes de config usadas localmente
_is_production = cfg.is_production
CRON_API_TOKEN = cfg.CRON_API_TOKEN
DIAS_ANTECEDENCIA = cfg.DIAS_ANTECEDENCIA

_APP_BOOTSTRAPPED = False


def _init_app_once() -> None:
    """Inicialização segura para app.py importado via gunicorn ou execução direta."""
    global _APP_BOOTSTRAPPED
    if _APP_BOOTSTRAPPED:
        return
    sr.ensure_schema()
    _ensure_extra_schema()
    try:
        sr.ensure_daily_backup()
    except Exception as exc:
        app.logger.warning("Backup no bootstrap falhou: %s", exc)
    _APP_BOOTSTRAPPED = True


@app.before_request
def _bootstrap_before_request():
    """Garante que o schema da BD é criado antes do primeiro pedido (Gunicorn).
    Auto-remove-se após a primeira execução para não correr em cada request."""
    _init_app_once()
    # Remover este handler — já não é necessário
    app.before_request_funcs.setdefault(None, [])
    try:
        app.before_request_funcs[None].remove(_bootstrap_before_request)
    except ValueError:
        pass


def _verify_cron_token() -> bool:
    """Verifica o token Bearer no header Authorization para endpoints de cron."""
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return False
    token = auth[len("Bearer ") :]
    if not CRON_API_TOKEN:
        # Sem token configurado: bloquear em produção, avisar fora
        if _is_production:
            return False
        app.logger.warning(
            "CRON_API_TOKEN não definido — endpoint de cron desprotegido!"
        )
        return True  # permite apenas fora de produção sem token
    return secrets.compare_digest(token, CRON_API_TOKEN)


def _client_ip() -> str:
    """IP do cliente (com ProxyFix activo atrás de proxy)."""
    try:
        if request.access_route:
            return str(request.access_route[0])[:64]
    except Exception:
        pass
    return str(request.remote_addr or "")[:64]


def _bootstrap_dev_system_accounts(conn=None) -> None:
    """Sincroniza PERFIS_ADMIN/PERFIS_TESTE para a BD em desenvolvimento (passwords hashed)."""
    if _is_production:
        return
    perfis = {**getattr(sr, "PERFIS_ADMIN", {}), **getattr(sr, "PERFIS_TESTE", {})}
    if not perfis:
        return
    owns_conn = conn is None
    try:
        if owns_conn:
            conn = sr.db()
        cols = {
            r["name"]
            for r in conn.execute("PRAGMA table_info(utilizadores)").fetchall()
        }
        if not cols:
            return
        for nii, p in perfis.items():
            row = conn.execute(
                "SELECT id, Palavra_chave FROM utilizadores WHERE NII=?", (nii,)
            ).fetchone()
            pw_hash = generate_password_hash(p.get("senha", ""))
            nome = p.get("nome", nii)
            perfil = p.get("perfil", "aluno")
            ano = str(p.get("ano", "") or "")
            if row is None:
                conn.execute(
                    """INSERT INTO utilizadores
                    (NII,NI,Nome_completo,Palavra_chave,ano,perfil,must_change_password,password_updated_at,is_active)
                    VALUES (?,?,?,?,?,?,0,datetime('now','localtime'),1)""",
                    (nii, nii, nome, pw_hash, ano, perfil),
                )
            else:
                stored = row["Palavra_chave"] or ""
                # Sempre sincroniza perfil, nome e ano — garante que perfis de sistema
                # nao ficam "presos" como aluno se o campo estiver errado na BD.
                conn.execute(
                    "UPDATE utilizadores SET perfil=?, Nome_completo=?, ano=?, must_change_password=CASE WHEN ? != 'aluno' THEN 0 ELSE must_change_password END WHERE id=?",
                    (perfil, nome, ano, perfil, row["id"]),
                )
                # Migra password de plain-text para hash se ainda nao foi migrada
                if stored == p.get("senha", ""):
                    conn.execute(
                        "UPDATE utilizadores SET Palavra_chave=?, password_updated_at=datetime('now','localtime') WHERE id=?",
                        (pw_hash, row["id"]),
                    )
        if owns_conn:
            conn.commit()
    except Exception as exc:
        app.logger.warning(f"_bootstrap_dev_system_accounts falhou: {exc}")
    finally:
        if owns_conn and conn is not None:
            try:
                conn.close()
            except Exception:
                pass


_init_app_once()

# ═══════════════════════════════════════════════════════════════════════════
# HELPERS DE NEGÓCIO
# ═══════════════════════════════════════════════════════════════════════════


def _refeicao_set(uid, dt, pa, lanche, alm, jan, sai, alterado_por="sistema"):
    r = {
        "pequeno_almoco": pa,
        "lanche": lanche,
        "almoco": alm or None,
        "jantar_tipo": jan or None,
        "jantar_sai_unidade": sai,
    }
    return sr.refeicao_save(uid, dt, r, alterado_por=alterado_por)


def _audit(actor: str, action: str, detail: str = "") -> None:
    """Regista uma entrada de auditoria na tabela admin_audit_log."""
    try:
        with sr.db() as conn:
            conn.execute(
                "INSERT INTO admin_audit_log(actor,action,detail) VALUES(?,?,?)",
                (actor, action, detail),
            )
            conn.commit()
    except Exception as exc:
        app.logger.warning(f"_audit falhou [{action}]: {exc}")


def _check_password(stored_hash: str, password: str) -> bool:
    """Verifica password — suporta hashes werkzeug e passwords em claro (migração transparente)."""
    if not stored_hash:
        return False
    if stored_hash.startswith(("pbkdf2:", "scrypt:", "argon2:")):
        return check_password_hash(stored_hash, password)
    # Password em claro (legado) — comparação simples + migração automática ao login
    return password == stored_hash


def _migrate_password_hash(uid: int, plain_password: str) -> None:
    """Migra password em claro para hash werkzeug na BD (chamado após login bem-sucedido)."""
    try:
        new_hash = generate_password_hash(plain_password)
        with sr.db() as conn:
            conn.execute(
                "UPDATE utilizadores SET Palavra_chave=? WHERE id=?", (new_hash, uid)
            )
            conn.commit()
    except Exception as exc:
        app.logger.warning(f"_migrate_password_hash uid={uid}: {exc}")


def _alterar_password(nii, old, new):
    uid = sr.user_id_by_nii(nii)
    if not uid:
        return False, "Conta de sistema — não é possível alterar a password."
    with sr.db() as conn:
        row = conn.execute(
            "SELECT Palavra_chave FROM utilizadores WHERE id=?", (uid,)
        ).fetchone()
    if not row:
        return False, "Utilizador não encontrado."
    ph = row["Palavra_chave"] or ""
    if not _check_password(ph, old):
        return False, "Password atual incorreta."
    pw_ok, pw_msg = _validate_password(new)
    if not pw_ok:
        return False, pw_msg
    new_hash = generate_password_hash(new)
    with sr.db() as conn:
        conn.execute(
            """UPDATE utilizadores SET Palavra_chave=?, must_change_password=0,
                        password_updated_at=datetime('now','localtime') WHERE id=?""",
            (new_hash, uid),
        )
        conn.commit()
    return True, ""


def _validate_password(pw: str) -> tuple:
    """Valida requisitos de password: mínimo 8 caracteres, letras e números."""
    if len(pw) < 8:
        return False, "A password deve ter pelo menos 8 caracteres."
    if pw.isdigit() or pw.isalpha():
        return False, "A password deve conter letras e números."
    return True, ""


def _criar_utilizador(nii, ni, nome, ano, perfil, pw):
    try:
        if not all([nii, ni, nome, ano, perfil, pw]):
            return False, "Todos os campos são obrigatórios."
        nii = _val_nii(nii)
        if not nii:
            return False, "NII inválido (alfanumérico, máx. 20 caracteres)."
        if _val_ni(ni) is None:
            return False, "NI inválido (alfanumérico, máx. 20 caracteres)."
        ni = _val_ni(ni)
        nome = _val_nome(nome)
        if not nome:
            return False, "Nome inválido ou vazio."
        ano_int = _val_ano(ano)
        if ano_int is None:
            return False, "Ano inválido (deve ser entre 0 e 8)."
        perfil = _val_perfil(perfil)
        if not perfil:
            return False, "Perfil inválido."
        pw = str(pw).strip()[:256]
        if len(pw) < 6:
            return False, "Password deve ter pelo menos 6 caracteres."
        pw_hash = generate_password_hash(pw)
        with sr.db() as conn:
            conn.execute(
                """INSERT INTO utilizadores
              (NII,NI,Nome_completo,Palavra_chave,ano,perfil,must_change_password,password_updated_at)
              VALUES (?,?,?,?,?,?,1,datetime('now','localtime'))""",
                (nii, ni, nome, pw_hash, ano_int, perfil),
            )
            conn.commit()
        _audit(
            "sistema", "criar_utilizador", f"NII={nii} perfil={perfil} ano={ano_int}"
        )
        return True, ""
    except Exception as e:
        app.logger.error(f"_criar_utilizador({nii}): {e}")
        return False, str(e)


def _reset_pw(nii, nova_pw=None):
    import string

    if not nova_pw:
        alphabet = string.ascii_letters + string.digits
        nova_pw = "".join(secrets.choice(alphabet) for _ in range(10))
    nova_hash = generate_password_hash(nova_pw)
    with sr.db() as conn:
        cur = conn.execute(
            """UPDATE utilizadores SET Palavra_chave=?, must_change_password=1,
                              password_updated_at=datetime('now','localtime') WHERE NII=?""",
            (nova_hash, nii),
        )
        conn.commit()
    if cur.rowcount:
        _audit("sistema", "reset_password", f"NII={nii}")
        return True, nova_pw  # devolve a password em claro para mostrar ao admin
    return False, "NII não encontrado."


def _unblock_user(nii):
    with sr.db() as conn:
        conn.execute("UPDATE utilizadores SET locked_until=NULL WHERE NII=?", (nii,))
        conn.commit()


def _eliminar_utilizador(nii):
    with sr.db() as conn:
        cur = conn.execute("DELETE FROM utilizadores WHERE NII=?", (nii,))
        conn.commit()
    return cur.rowcount > 0


def _registar_ausencia(uid, de, ate, motivo, criado_por):
    try:
        datetime.strptime(de, "%Y-%m-%d")
        datetime.strptime(ate, "%Y-%m-%d")
    except ValueError:
        return False, "Data inválida."
    if de > ate:
        return False, "A data de início não pode ser posterior à data de fim."
    with sr.db() as conn:
        conn.execute(
            """INSERT INTO ausencias (utilizador_id,ausente_de,ausente_ate,motivo,criado_por)
                        VALUES (?,?,?,?,?)""",
            (uid, de, ate, motivo or None, criado_por),
        )
        conn.commit()
    return True, ""


def _remover_ausencia(aid):
    with sr.db() as conn:
        conn.execute("DELETE FROM ausencias WHERE id=?", (aid,))
        conn.commit()


def _get_ocupacao_dia(dt):
    return sr.get_ocupacao_capacidade(dt)


ANOS_LABELS = {
    1: "1º Ano",
    2: "2º Ano",
    3: "3º Ano",
    4: "4º Ano",
    5: "5º Ano",
    6: "6º Ano",
    7: "CFBO",
    8: "CFCO",
}
ANOS_OPCOES = [
    (1, "1º Ano"),
    (2, "2º Ano"),
    (3, "3º Ano"),
    (4, "4º Ano"),
    (5, "5º Ano"),
    (6, "6º Ano"),
    (7, "CFBO — Curso de Formação Básica de Oficiais"),
    (8, "CFCO — Curso de Formação Complementar de Oficiais"),
]


def _ano_label(ano):
    return ANOS_LABELS.get(int(ano) if ano else 0, f"{ano}º Ano")


def _get_anos_disponiveis():
    with sr.db() as conn:
        rows = conn.execute(
            "SELECT DISTINCT CAST(ano AS INTEGER) AS ano FROM utilizadores"
            " WHERE ano IS NOT NULL AND ano != '' AND CAST(ano AS INTEGER) > 0"
            " ORDER BY CAST(ano AS INTEGER)"
        ).fetchall()
    return [r["ano"] for r in rows]


def _editar_ausencia(aid, uid, de, ate, motivo):
    try:
        datetime.strptime(de, "%Y-%m-%d")
        datetime.strptime(ate, "%Y-%m-%d")
    except ValueError:
        return False, "Data inválida."
    if de > ate:
        return False, "A data de início não pode ser posterior à data de fim."
    with sr.db() as conn:
        conn.execute(
            """UPDATE ausencias SET ausente_de=?,ausente_ate=?,motivo=?
                        WHERE id=? AND utilizador_id=?""",
            (de, ate, motivo or None, aid, uid),
        )
        conn.commit()
    return True, ""


def _tem_ausencia_ativa(uid, d=None):
    """Verifica se utilizador tem ausência ativa na data (ou hoje)."""
    d_str = (d or date.today()).isoformat()
    with sr.db() as conn:
        row = conn.execute(
            """SELECT 1 FROM ausencias WHERE utilizador_id=?
                              AND ausente_de<=? AND ausente_ate>=?""",
            (uid, d_str, d_str),
        ).fetchone()
    return bool(row)


def _auto_marcar_refeicoes_detido(uid, d_de, d_ate, alterado_por="sistema"):
    """Auto-marca todas as refeições para dias de detenção se não estiverem marcadas."""
    try:
        d = d_de
        while d <= d_ate:
            with sr.db() as conn:
                existe = conn.execute(
                    "SELECT almoco FROM refeicoes WHERE utilizador_id=? AND data=?",
                    (uid, d.isoformat()),
                ).fetchone()
            if not existe or not existe["almoco"]:
                _refeicao_set(
                    uid,
                    d,
                    pa=1,
                    lanche=1,
                    alm="Normal",
                    jan="Normal",
                    sai=0,
                    alterado_por=alterado_por,
                )
            d += timedelta(days=1)
    except Exception as exc:
        app.logger.warning(f"_auto_marcar_refeicoes_detido uid={uid}: {exc}")


def _tem_detencao_ativa(uid, d=None):
    """Verifica se utilizador tem detenção ativa na data (ou hoje)."""
    try:
        d_str = (d or date.today()).isoformat()
        with sr.db() as conn:
            row = conn.execute(
                """SELECT 1 FROM detencoes WHERE utilizador_id=?
                              AND detido_de<=? AND detido_ate>=? LIMIT 1""",
                (uid, d_str, d_str),
            ).fetchone()
        return bool(row)
    except Exception:
        return False


def _regras_licenca(ano: int, ni: str) -> dict:
    """Devolve regras de licença para um aluno com base no ano e NI.

    Regras:
    - NI começa com '7' → pode sair todos os dias (exceção especial)
    - 4º ano e acima → pode sair todos os dias
    - 3º ano → sex/sab/dom + 3 dias úteis (seg-qui) por semana
    - 2º ano → sex/sab/dom + 2 dias úteis (seg-qui) por semana
    - 1º ano → sex/sab/dom + apenas quarta-feira
    """
    if str(ni).startswith("7"):
        return {
            "max_dias_uteis": 4,
            "dias_permitidos": [0, 1, 2, 3, 4, 5, 6],
            "excepcao_ni7": True,
        }
    if ano >= 4:
        return {
            "max_dias_uteis": 4,
            "dias_permitidos": [0, 1, 2, 3, 4, 5, 6],
            "excepcao_ni7": False,
        }
    if ano == 3:
        return {
            "max_dias_uteis": 3,
            "dias_permitidos": [0, 1, 2, 3, 4, 5, 6],
            "excepcao_ni7": False,
        }
    if ano == 2:
        return {
            "max_dias_uteis": 2,
            "dias_permitidos": [0, 1, 2, 3, 4, 5, 6],
            "excepcao_ni7": False,
        }
    # 1º ano — só quarta (2) + fim de semana (4=sex, 5=sab, 6=dom)
    return {"max_dias_uteis": 1, "dias_permitidos": [2, 4, 5, 6], "excepcao_ni7": False}


def _licencas_semana_usadas(uid: int, d: date) -> int:
    """Conta licenças de dias úteis (seg-qui) já usadas na semana ISO de 'd'."""
    # Calcular seg e qui da semana
    seg = d - timedelta(days=d.weekday())  # segunda
    qui = seg + timedelta(days=3)  # quinta
    with sr.db() as conn:
        row = conn.execute(
            """SELECT COUNT(*) c FROM licencas
            WHERE utilizador_id=? AND data>=? AND data<=?""",
            (uid, seg.isoformat(), qui.isoformat()),
        ).fetchone()
    return row["c"] or 0


def _pode_marcar_licenca(uid: int, d: date, ano: int, ni: str) -> tuple:
    """Verifica se o aluno pode marcar licença para o dia 'd'.

    Retorna (pode: bool, motivo: str).
    """
    regras = _regras_licenca(ano, ni)

    # Detido não pode sair
    if _tem_detencao_ativa(uid, d):
        return False, "Estás detido neste dia — não podes marcar licença."

    dia_semana = d.weekday()  # 0=seg ... 6=dom

    # Fim de semana (sex=4, sab=5, dom=6) — todos podem
    if dia_semana >= 4:
        return True, ""

    # Dia útil (seg-qui) — verificar se o dia é permitido
    if dia_semana not in regras["dias_permitidos"]:
        nomes = {0: "segunda", 1: "terça", 2: "quarta", 3: "quinta"}
        return False, f"O teu ano não tem licença à {nomes.get(dia_semana, '')}."

    # Verificar limite semanal de dias úteis
    usadas = _licencas_semana_usadas(uid, d)
    # Verificar se este dia já está contado (para não contar duas vezes ao editar)
    with sr.db() as conn:
        ja_tem = conn.execute(
            "SELECT 1 FROM licencas WHERE utilizador_id=? AND data=?",
            (uid, d.isoformat()),
        ).fetchone()
    if not ja_tem and usadas >= regras["max_dias_uteis"]:
        return (
            False,
            f"Já esgotaste as tuas {regras['max_dias_uteis']} saídas desta semana (seg-qui). "
            "Precisas de aprovação do Comandante de Companhia ou Oficial de Dia.",
        )

    return True, ""


def _dia_editavel_aluno(d):
    """Editável pelo aluno: futuro, dentro de DIAS_ANTECEDENCIA, prazo ok. Fins de semana permitidos."""
    hoje = date.today()
    if d < hoje:
        return False, "Data no passado."
    if (d - hoje).days > DIAS_ANTECEDENCIA:
        return (
            False,
            f"Só é possível marcar com {DIAS_ANTECEDENCIA} dias de antecedência.",
        )
    return sr.refeicao_editavel(d)


# ═══════════════════════════════════════════════════════════════════════════
# LICENÇA DE FIM DE SEMANA — marcar/cancelar FDS com um clique
# ═══════════════════════════════════════════════════════════════════════════


def _marcar_licenca_fds(uid: int, sexta: date, alterado_por: str) -> tuple:
    """
    Marca 'licença fim de semana' para um aluno:
    - Sexta: licença antes_jantar (retira jantar, marca sai_unidade)
    - Sábado e Domingo: apaga todas as refeições
    Retorna (sucesso: bool, mensagem: str)
    """
    try:
        # ── Sexta-feira: licença antes_jantar ──────────────────────────
        with sr.db() as conn:
            conn.execute(
                """INSERT INTO licencas(utilizador_id, data, tipo)
                   VALUES(?,?,?)
                   ON CONFLICT(utilizador_id, data) DO UPDATE SET tipo=excluded.tipo""",
                (uid, sexta.isoformat(), "antes_jantar"),
            )
            conn.commit()

        # Refeições da sexta: guardar sem jantar, com sai_unidade
        r_sexta = sr.refeicao_get(uid, sexta)
        r_sexta["jantar_tipo"] = None
        r_sexta["jantar_sai_unidade"] = 1
        sr.refeicao_save(uid, sexta, r_sexta, alterado_por=alterado_por)

        # ── Sábado e Domingo: apagar todas as refeições ────────────────
        for delta in (1, 2):  # sábado=+1, domingo=+2
            d = sexta + timedelta(days=delta)
            r_vazio = {
                "pequeno_almoco": 0,
                "lanche": 0,
                "almoco": None,
                "jantar_tipo": None,
                "jantar_sai_unidade": 0,
            }
            sr.refeicao_save(uid, d, r_vazio, alterado_por=alterado_por)

        return True, ""
    except Exception as exc:
        app.logger.warning(f"_marcar_licenca_fds uid={uid}: {exc}")
        return False, str(exc)


def _cancelar_licenca_fds(uid: int, sexta: date, alterado_por: str) -> tuple:
    """
    Cancela 'licença fim de semana':
    - Remove a licença da sexta
    - Repõe jantar normal na sexta, retira sai_unidade
    """
    try:
        # Remover licença da sexta
        with sr.db() as conn:
            conn.execute(
                "DELETE FROM licencas WHERE utilizador_id=? AND data=?",
                (uid, sexta.isoformat()),
            )
            conn.commit()

        # Repor sexta: jantar Normal, sem sai_unidade
        r_sexta = sr.refeicao_get(uid, sexta)
        r_sexta["jantar_sai_unidade"] = 0
        if not r_sexta.get("jantar_tipo"):
            r_sexta["jantar_tipo"] = "Normal"
        sr.refeicao_save(uid, sexta, r_sexta, alterado_por=alterado_por)

        return True, ""
    except Exception as exc:
        app.logger.warning(f"_cancelar_licenca_fds uid={uid}: {exc}")
        return False, str(exc)


# ═══════════════════════════════════════════════════════════════════════════
# TEMPLATE BASE
# ═══════════════════════════════════════════════════════════════════════════

BASE = """
<!doctype html>
<html lang="pt">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Escola Naval — Refeições</title>
  <link rel="icon" type="image/svg+xml" href="data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'%3E%3Ctext y='.9em' font-size='90'%3E⚓%3C/text%3E%3C/svg%3E">
  <style>
    :root{
      --bg:#f0f4f8;--card:#fff;--primary:#0a2d4e;--primary-light:#1a5276;
      --gold:#c9a227;--danger:#c0392b;--ok:#1e8449;--warn:#d68910;--muted:#6c757d;
      --text:#1a2533;--border:#dde3ea;--shadow:0 2px 10px rgba(0,0,0,.08);
      --radius:12px;
    }
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif;
         background:var(--bg);color:var(--text);line-height:1.5}
    a{text-decoration:none;color:inherit}

    nav{background:var(--primary);color:#fff;padding:.85rem 1.5rem;
        display:flex;align-items:center;justify-content:space-between;
        box-shadow:0 2px 8px rgba(0,0,0,.2);position:sticky;top:0;z-index:100}
    .nav-brand{font-weight:800;font-size:1.05rem}
    .nav-right{display:flex;align-items:center;gap:.75rem}
    .nav-user{opacity:.8;font-size:.82rem}
    .nav-link{color:#fff;font-weight:600;font-size:.85rem;
              background:rgba(255,255,255,.15);padding:.38rem .75rem;
              border-radius:8px;transition:.15s}
    .nav-link:hover{background:rgba(255,255,255,.25)}

    .container{max-width:1120px;margin:0 auto;padding:1.5rem 1.2rem}
    .page-header{display:flex;align-items:center;gap:.75rem;margin-bottom:1.4rem;flex-wrap:wrap}
    .page-title{font-size:1.4rem;font-weight:800;color:var(--primary);flex:1}
    .back-btn{display:inline-flex;align-items:center;gap:.3rem;padding:.42rem .85rem;
              border-radius:9px;font-weight:600;font-size:.85rem;border:1.5px solid var(--border);
              background:#fff;color:var(--text);cursor:pointer;transition:.15s;white-space:nowrap}
    .back-btn:hover{border-color:var(--primary-light);color:var(--primary);
                    box-shadow:0 2px 8px rgba(0,0,0,.08)}

    .card{background:var(--card);border-radius:var(--radius);
          box-shadow:var(--shadow);padding:1.4rem;margin-bottom:1.3rem}
    .card-title{font-size:.93rem;font-weight:700;color:var(--primary);
                margin-bottom:.9rem;padding-bottom:.5rem;
                border-bottom:2px solid var(--border)}

    .grid{display:grid;gap:1rem}
    .grid-2{grid-template-columns:repeat(auto-fit,minmax(280px,1fr))}
    .grid-3{grid-template-columns:repeat(auto-fit,minmax(210px,1fr))}
    .grid-4{grid-template-columns:repeat(auto-fit,minmax(170px,1fr))}
    .grid-6{grid-template-columns:repeat(auto-fit,minmax(110px,1fr))}

    .alert{padding:.82rem 1.05rem;border-radius:10px;margin:.7rem 0;font-size:.87rem;border:1px solid}
    .alert-ok{background:#eafaf1;border-color:#a9dfbf;color:#1e8449}
    .alert-error{background:#fdecea;border-color:#f1948a;color:#922b21}
    .alert-warn{background:#fef9e7;border-color:#f9e79f;color:#9a7d0a}

    .btn{display:inline-flex;align-items:center;gap:.32rem;padding:.55rem 1rem;
         border-radius:9px;font-weight:600;font-size:.87rem;border:none;
         cursor:pointer;transition:.15s;white-space:nowrap}
    .btn:hover{filter:brightness(1.08)}
    .btn-primary{background:var(--primary);color:#fff}
    .btn-ok{background:var(--ok);color:#fff}
    .btn-danger{background:var(--danger);color:#fff}
    .btn-warn{background:var(--warn);color:#fff}
    .btn-ghost{background:#fff;border:1.5px solid var(--border);color:var(--text)}
    .btn-ghost:hover{border-color:var(--primary-light);color:var(--primary)}
    .btn-sm{padding:.3rem .62rem;font-size:.78rem;border-radius:7px}
    .btn-gold{background:var(--gold);color:#fff}

    .form-group{margin-bottom:.85rem}
    label{font-weight:600;font-size:.83rem;display:block;margin-bottom:.28rem;color:#4a5568}
    input,select,textarea{width:100%;padding:.56rem .72rem;border:1.5px solid var(--border);
      border-radius:9px;font-size:.87rem;color:var(--text);background:#fff;transition:.15s}
    input:focus,select:focus{outline:none;border-color:var(--primary-light);
      box-shadow:0 0 0 3px rgba(36,113,163,.12)}
    input[type="checkbox"]{width:1rem;height:1rem;cursor:pointer;accent-color:var(--primary)}

    .table-wrap{overflow-x:auto;border-radius:10px;border:1px solid var(--border)}
    table{width:100%;border-collapse:collapse;font-size:.84rem}
    th{background:var(--primary);color:#fff;padding:.58rem .9rem;text-align:left;font-weight:600}
    td{padding:.5rem .9rem;border-bottom:1px solid var(--border)}
    tr:last-child td{border-bottom:none}
    tr:hover td{background:#f5f8fc}

    .badge{display:inline-block;padding:.18rem .52rem;border-radius:999px;font-size:.72rem;font-weight:700}
    .badge-ok{background:#d5f5e3;color:#1e8449}
    .badge-no{background:#fdecea;color:#922b21}
    .badge-warn{background:#fef9e7;color:#9a7d0a}
    .badge-info{background:#ebf5fb;color:#1a4a6e}
    .badge-gold{background:#fef3cd;color:#856404}
    .badge-muted{background:#ecf0f1;color:#6c757d}

    .stat-box{padding:1.05rem;border:1.5px solid var(--border);border-radius:12px;
              text-align:center;background:#fff;transition:.15s}
    .stat-box:hover{border-color:var(--primary-light);box-shadow:0 4px 14px rgba(0,0,0,.09)}
    .stat-num{font-size:1.85rem;font-weight:800;color:var(--primary);line-height:1.1}
    .stat-lbl{font-size:.77rem;color:var(--muted);margin-top:.22rem}

    .week-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(135px,1fr));gap:.65rem}
    .week-card{border:1.5px solid var(--border);border-radius:12px;padding:.82rem;
               background:#fff;transition:.15s}
    .week-card:hover{border-color:var(--primary-light);box-shadow:0 4px 12px rgba(0,0,0,.09)}
    .week-card.day-off{background:#f9fafb}
    .week-card.weekend-active{border-color:var(--gold);background:#fffdf5}
    .week-dow{font-weight:800;font-size:.88rem;color:var(--primary)}
    .week-dow.weekend{color:var(--gold)}
    .week-date{font-size:.73rem;color:var(--muted);margin-bottom:.42rem}
    .week-meals{display:flex;flex-wrap:wrap;gap:.24rem;margin-bottom:.38rem}

    .ausente-banner{background:#fff9e6;border:1.5px solid var(--gold);border-radius:10px;
                    padding:.65rem 1rem;font-size:.84rem;color:#856404;margin-bottom:.8rem}

    .occ-bar{height:7px;border-radius:999px;overflow:hidden;background:#e9ecef;margin:.2rem 0}
    .occ-bar>span{display:block;height:100%;border-radius:999px}
    .occ-label{font-size:.73rem;color:var(--muted)}

    .meal-chip{font-size:.67rem;font-weight:700;padding:.15rem .42rem;border-radius:999px;white-space:nowrap}
    .chip-ok{background:#d5f5e3;color:#1e8449}
    .chip-no{background:#fdecea;color:#922b21}
    .chip-type{background:#ebf5fb;color:#1a4a6e}

    .prazo-warn{font-size:.7rem;color:#9a7d0a;font-weight:600}
    .prazo-lock{font-size:.7rem;color:#922b21;font-weight:600}

    .login-wrap{display:flex;justify-content:center;align-items:center;
                min-height:100vh;
                background:linear-gradient(160deg,#0a2d4e 0%,#1a5276 60%,#0a2d4e 100%)}
    .login-box{background:#fff;border-radius:20px;padding:2.6rem 2.4rem;width:100%;
               max-width:420px;box-shadow:0 24px 70px rgba(0,0,0,.32)}
    .login-header{display:flex;align-items:center;gap:1rem;margin-bottom:1.6rem;justify-content:center}
    .login-title{font-size:1.45rem;font-weight:900;color:var(--primary);line-height:1.1}
    .login-subtitle{font-size:.8rem;color:var(--muted);margin-top:.15rem;letter-spacing:.4px;text-transform:uppercase}

    .year-tabs{display:flex;gap:.4rem;flex-wrap:wrap;margin-bottom:1.1rem}
    .year-tab{padding:.42rem .95rem;border-radius:9px;font-weight:700;font-size:.84rem;
              border:1.5px solid var(--border);background:#fff;color:var(--text);transition:.15s}
    .year-tab:hover{border-color:var(--primary-light);color:var(--primary)}
    .year-tab.active{background:var(--primary);color:#fff;border-color:var(--primary)}

    .action-card{border:1.5px solid var(--border);border-radius:12px;padding:.95rem;
                 background:#fff;text-align:center;transition:.15s;cursor:pointer;
                 display:flex;flex-direction:column;align-items:center;gap:.35rem}
    .action-card:hover{border-color:var(--primary-light);box-shadow:0 4px 14px rgba(0,0,0,.1);
                       transform:translateY(-1px)}
    .action-card .icon{font-size:1.55rem}
    .action-card .label{font-weight:700;font-size:.84rem}
    .action-card .desc{font-size:.74rem;color:var(--muted)}

    .gap-btn{display:flex;gap:.5rem;flex-wrap:wrap;align-items:center}
    .text-muted{color:var(--muted)}
    .small{font-size:.79rem}
    .right{text-align:right}
    .center{text-align:center}
    .flex{display:flex;gap:.5rem;align-items:center;flex-wrap:wrap}
    .flex-between{display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:.5rem}
    hr{border:none;border-top:1.5px solid var(--border);margin:.85rem 0}

    @media(max-width:640px){
      .grid-2,.grid-3,.grid-4{grid-template-columns:1fr}
      .week-grid{grid-template-columns:1fr 1fr}
    }
  </style>
</head>
<body>
{% autoescape true %}
{% if session.user %}
<nav>
  <span class="nav-brand" style="display:flex;align-items:center;gap:.6rem"><img src="/static/logo_escola_naval.jpg" style="height:36px;width:auto;border-radius:4px" alt="EN"> <span style="color:#fff">Escola Naval</span></span>
  <div class="nav-right">
    <span class="nav-user">{{ session.user.nome }} · <strong>{{ session.user.perfil }}</strong></span>
    {% if session.user.perfil == 'aluno' %}
    <a class="nav-link" href="{{ url_for('aluno_perfil') }}">👤 Perfil</a>
    {% endif %}
    <form method="post" action="{{ url_for('logout') }}" style="display:inline;margin:0">
      <input type="hidden" name="csrf_token" value="{{ session.get('_csrf_token','') }}">
      <button type="submit" class="nav-link" style="background:none;border:none;cursor:pointer;padding:0;font:inherit;color:inherit">Sair</button>
    </form>
  </div>
</nav>
{% endif %}
{% with msgs = get_flashed_messages(with_categories=true) %}
  {% if msgs %}
    <div class="container" style="margin-top:.8rem;margin-bottom:0">
    {% for cat,msg in msgs %}
      <div class="alert alert-{{ cat }}">{{ msg }}</div>
    {% endfor %}
    </div>
  {% endif %}
{% endwith %}
{{ content|safe }}
{% endautoescape %}
</body></html>
"""


def render(content, status=200):
    html = render_template_string(BASE, content=content)
    return Response(html, status=status, mimetype="text/html")


def esc(v):
    return str(escape(str(v))) if v is not None else ""


def csrf_input():
    t = session.get("_csrf_token") or secrets.token_urlsafe(32)
    session["_csrf_token"] = t
    return Markup(f'<input type="hidden" name="csrf_token" value="{t}">')


def _parse_date(s, default=None):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return default or date.today()


def _parse_date_strict(s):
    try:
        return datetime.strptime((s or "").strip(), "%Y-%m-%d").date()
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════════════════
# VALIDADORES DE INPUT — funções reutilizáveis para sanitização server-side
# ═══════════════════════════════════════════════════════════════════════════

_PERFIS_VALIDOS = {"admin", "cmd", "cozinha", "oficialdia", "aluno"}
_TIPOS_CALENDARIO = {"normal", "fim_semana", "feriado", "exercicio", "outro"}
_REFEICAO_OPCOES = {"Normal", "Vegetariano", "Dieta", ""}
_RE_EMAIL = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_RE_PHONE = re.compile(r"^[\d\s\+\-\(\)]{7,20}$")
_RE_ALNUM = re.compile(r"^[A-Za-z0-9_]+$")
_MAX_NOME = 200
_MAX_TEXT = 500
_MAX_DATE_RANGE = 366


def _val_email(v):
    """Valida email. Devolve string limpa ou None se vazio, False se inválido."""
    v = (v or "").strip()[:254]
    if not v:
        return None
    return v if _RE_EMAIL.match(v) else False


def _val_phone(v):
    """Valida telemóvel. Devolve string limpa ou None se vazio, False se inválido."""
    v = (v or "").strip()[:20]
    if not v:
        return None
    return v if _RE_PHONE.match(v) else False


def _val_nii(v):
    """Valida NII (alfanumérico, 1-20 chars). Devolve string ou None se inválido."""
    v = (v or "").strip()[:20]
    return v if v and _RE_ALNUM.match(v) else None


def _val_ni(v):
    """Valida NI (alfanumérico, até 20 chars). Pode ser vazio."""
    v = (v or "").strip()[:20]
    if not v:
        return ""
    return v if _RE_ALNUM.match(v) else None


def _val_nome(v, max_len=_MAX_NOME):
    """Valida nome (não-vazio, limitado). Devolve string ou None se vazio."""
    v = (v or "").strip()[:max_len]
    return v if v else None


def _val_ano(v):
    """Valida ano escolar (0-8). 0 = concluído. Devolve int ou None se inválido."""
    try:
        a = int(v)
        return a if 0 <= a <= 8 else None
    except (TypeError, ValueError):
        return None


def _val_perfil(v):
    """Valida perfil contra whitelist. Devolve string ou None se inválido."""
    v = (v or "").strip().lower()
    return v if v in _PERFIS_VALIDOS else None


def _val_tipo_calendario(v):
    """Valida tipo de calendário. Fallback para 'normal'."""
    v = (v or "").strip().lower()
    return v if v in _TIPOS_CALENDARIO else "normal"


def _val_refeicao(v):
    """Valida opção de refeição. Fallback para string vazia."""
    v = (v or "").strip()
    return v if v in _REFEICAO_OPCOES else ""


def _val_text(v, max_len=_MAX_TEXT):
    """Limita texto a max_len caracteres."""
    return (v or "").strip()[:max_len]


def _val_int_id(v):
    """Valida ID numérico. Devolve int ou None."""
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _val_date_range(d1, d2, max_dias=_MAX_DATE_RANGE):
    """Valida range de datas. Devolve (ok, msg_erro)."""
    if d2 < d1:
        return False, "Data final anterior à inicial."
    if (d2 - d1).days > max_dias:
        return False, f"Intervalo máximo permitido: {max_dias} dias."
    return True, ""


def _val_cap(v, max_val=9999):
    """Valida capacidade (inteiro 0-max_val). Devolve int ou None."""
    try:
        c = int(v)
        return c if 0 <= c <= max_val else None
    except (TypeError, ValueError):
        return None


def _bar_html(val, cap):
    if cap is None or cap <= 0:
        return f'<div class="occ-label">{val} (sem limite)</div>'
    pct = min(100, int(round(100 * val / cap)))
    color = "#1e8449" if pct < 80 else ("#d68910" if pct < 95 else "#c0392b")
    return (
        f'<div class="occ-bar"><span style="width:{pct}%;background:{color}"></span></div>'
        f'<div class="occ-label">{val} / {cap} ({pct}%)</div>'
    )


def _prazo_label(d):
    ok, _ = sr.refeicao_editavel(d)
    if ok:
        return ""
    if sr.PRAZO_LIMITE_HORAS is not None:
        prazo_dt = datetime(d.year, d.month, d.day) - timedelta(
            hours=sr.PRAZO_LIMITE_HORAS
        )
        h = (prazo_dt - datetime.now()).total_seconds() / 3600
        if h <= 0:
            return '<span class="prazo-lock">🔒 Prazo expirado</span>'
        if h <= 24:
            return f'<span class="prazo-warn">⚠️ Prazo em {int(h)}h</span>'
    return '<span class="prazo-lock">🔒 Prazo expirado</span>'


NOMES_DIAS = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
ABREV_DIAS = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]


def _back_btn(href, label="Voltar"):
    return f'<a class="back-btn" href="{href}">← {label}</a>'


# ═══════════════════════════════════════════════════════════════════════════
# AUTH / DECORADORES
# ═══════════════════════════════════════════════════════════════════════════


def login_required(f):
    @wraps(f)
    def d(*a, **kw):
        if "user" not in session:
            return redirect(url_for("login"))
        if session.get("must_change_password") and f.__name__ != "aluno_password":
            flash("Deves alterar a tua password antes de continuar.", "warn")
            return redirect(url_for("aluno_password"))
        return f(*a, **kw)

    return d


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def d(*a, **kw):
            if "user" not in session:
                return redirect(url_for("login"))
            if session.get("must_change_password") and f.__name__ != "aluno_password":
                flash("Deves alterar a tua password antes de continuar.", "warn")
                return redirect(url_for("aluno_password"))
            if session["user"]["perfil"] not in roles:
                flash("Acesso não autorizado.", "error")
                return redirect(url_for("dashboard"))
            return f(*a, **kw)

        return d

    return decorator


def current_user():
    return session.get("user", {})


_WAL_CHECKPOINT_INTERVAL = 300  # checkpoint WAL a cada 5 min
_last_wal_checkpoint = 0.0


@app.before_request
def before():
    global _last_wal_checkpoint
    g._t0 = time.perf_counter()

    # Refrescar sessão permanente em cada request (reset timeout inatividade)
    if "user" in session:
        session.modified = True

    # WAL checkpoint periódico
    now = time.time()
    if now - _last_wal_checkpoint > _WAL_CHECKPOINT_INTERVAL:
        _last_wal_checkpoint = now
        sr.wal_checkpoint()

    if request.method == "POST":
        t = session.get("_csrf_token", "")
        ft = request.form.get("csrf_token", "")
        if not t or not ft or not secrets.compare_digest(t, ft):
            # Sessão expirada sem CSRF — redirecionar para login
            if "user" not in session and request.endpoint not in {None}:
                flash(
                    "A sessão expirou. Inicia sessão novamente e repete a operação.",
                    "warn",
                )
                return redirect(url_for("login"))
            # CSRF inválido (inclui login POST) — rejeitar sempre
            abort(400)


@app.after_request
def after(r):
    r.headers.setdefault("X-Content-Type-Options", "nosniff")
    r.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    r.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    r.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline'",
    )
    # Logging de performance — avisar para requests lentos
    t0 = getattr(g, "_t0", None)
    if t0 is not None:
        dt_ms = (time.perf_counter() - t0) * 1000
        if dt_ms > 500:
            app.logger.warning(
                "Slow request: %s %s %.0fms", request.method, request.path, dt_ms
            )
    return r


@app.errorhandler(400)
def err400(e):
    return render(
        "<div class='container'><div class='page-header'><div class='page-title'>⚠️ Pedido inválido</div></div>"
        "<div class='card'><p>Token CSRF inválido ou pedido malformado. Se a sessão expirou, volta a iniciar sessão.</p><br>"
        "<a class='btn btn-primary' href='/'>Início</a></div></div>",
        400,
    )


@app.errorhandler(404)
def err404(e):
    return render(
        "<div class='container'><div class='page-header'><div class='page-title'>🔎 Não encontrado</div></div>"
        "<div class='card'><p>Página não encontrada.</p><br>"
        "<a class='btn btn-primary' href='/'>Início</a></div></div>",
        404,
    )


@app.errorhandler(500)
def err500(e):
    app.logger.exception("Erro 500")
    return render(
        "<div class='container'><div class='page-header'><div class='page-title'>💥 Erro interno</div></div>"
        "<div class='card'><p>Erro inesperado. Consulta os logs do servidor para detalhes.</p><br>"
        "<a class='btn btn-primary' href='/'>Início</a></div></div>",
        500,
    )


# ═══════════════════════════════════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/", methods=["GET", "POST"])
@app.route("/login", methods=["GET", "POST"])
def login():
    if "user" in session:
        return redirect(url_for("dashboard"))
    error = None
    if request.method == "POST":
        nii = request.form.get("nii", "").strip()[:32]
        pw = request.form.get("pw", request.form.get("password", "")).strip()[:256]
        # Rate limiting por IP (proteção contra ataques distribuídos)
        ip = _client_ip()
        ip_falhas = sr.recent_failures_by_ip(ip, 15)
        if ip_falhas >= 20:
            error = "Demasiadas tentativas falhadas deste endereço. Aguarda 15 minutos."
            app.logger.warning("IP rate-limited: IP=%s falhas=%d", ip, ip_falhas)
            _audit(nii or "unknown", "ip_rate_limited", f"IP={ip} falhas={ip_falhas}")
        # Autenticação via BD (contas de sistema sincronizadas para a BD em desenvolvimento)
        perfis = {}
        u = None
        db_u = None
        if error:
            pass  # IP bloqueado — não tentar autenticação
        elif False and nii in perfis:
            error = "Login legado desativado."
        elif (
            not _is_production
            and not sr.existe_admin()
            and nii == sr.FALLBACK_ADMIN["nii"]
            and pw == sr.FALLBACK_ADMIN["pw"]
        ):
            u = {
                "id": 0,
                "nii": nii,
                "ni": "",
                "nome": sr.FALLBACK_ADMIN["nome"],
                "ano": "",
                "perfil": "admin",
            }
            sr.reg_login(nii, 1, ip=_client_ip())
            app.logger.warning(f"Login via FALLBACK_ADMIN: NII={nii} IP={_client_ip()}")
        else:
            # Busca directa à BD por NII
            db_u = sr.user_by_nii(nii)
            if db_u:
                locked = db_u.get("locked_until")
                if locked:
                    try:
                        lock_dt = datetime.fromisoformat(locked)
                        if lock_dt > datetime.now():
                            mins = max(
                                1, int((lock_dt - datetime.now()).total_seconds() / 60)
                            )
                            error = f"Conta bloqueada por demasiadas tentativas falhadas. Tenta novamente em {mins} min."
                            app.logger.warning(
                                f"Login bloqueado: NII={nii} IP={_client_ip()}"
                            )
                            db_u = None
                    except ValueError:
                        pass
                if db_u:
                    ph = db_u.get("Palavra_chave", "") or ""
                    ok = _check_password(ph, pw)
                    if ok:
                        _perfil = (
                            db_u.get("perfil")
                            if hasattr(db_u, "get")
                            else db_u["perfil"]
                        ) or "aluno"
                        u = {
                            "id": db_u["id"],
                            "nii": db_u["NII"],
                            "ni": db_u["NI"],
                            "nome": db_u["Nome_completo"],
                            "ano": str(db_u["ano"] or ""),
                            "perfil": _perfil,
                        }
                        sr.reg_login(nii, 1, ip=_client_ip())
                        app.logger.info(
                            f"Login OK: NII={nii} perfil={u['perfil']} IP={_client_ip()}"
                        )
                        # Migração transparente: se ainda é plain-text, converter para hash
                        if not ph.startswith(("pbkdf2:", "scrypt:", "argon2:")):
                            _migrate_password_hash(db_u["id"], pw)
                    else:
                        sr.reg_login(nii, 0, ip=_client_ip())
                        falhas = sr.recent_failures(nii, 10)
                        if falhas >= 5:
                            sr.block_user(nii, 15)
                            error = "Conta bloqueada por 15 minutos após 5 tentativas falhadas."
                            app.logger.warning(
                                f"Conta bloqueada: NII={nii} IP={_client_ip()}"
                            )
                        else:
                            restam = max(0, 5 - falhas)
                            error = f"Password incorreta. ({restam} tentativa(s) restante(s) antes de bloqueio)"
            else:
                sr.reg_login(nii, 0, ip=_client_ip())
                error = "NII não encontrado."
        if u:
            session["user"] = u
            session.permanent = True  # ativa timeout de inatividade
            _audit(nii, "login", f"perfil={u['perfil']} IP={_client_ip()}")
            # Forçar alteração de password se necessário
            if db_u and db_u.get("must_change_password"):
                session["must_change_password"] = True
                flash(
                    "Por segurança, deves alterar a tua password antes de continuar.",
                    "warn",
                )
                return redirect(url_for("aluno_password"))
            return redirect(url_for("dashboard"))

    content = f"""
    <div class="login-wrap">
      <div class="login-box">
        <div class="login-header" style="flex-direction:column;text-align:center">
          <img src="/static/logo_escola_naval.jpg" style="height:100px;width:auto;margin-bottom:.5rem" alt="Escola Naval">
          <div class="login-title">Escola Naval</div>
        </div>
        {'<div class="alert alert-error">' + esc(error) + "</div>" if error else ""}
        <form method="post">
          {csrf_input()}
          <div class="form-group">
            <label>NII</label>
            <input type="text" name="nii" maxlength="32" autofocus autocomplete="username" required placeholder="O teu NII (ex: 24123)">
          </div>
          <div class="form-group">
            <label>Password</label>
            <input type="password" name="pw" maxlength="256" autocomplete="current-password" required placeholder="••••••••">
          </div>
          <button type="submit" class="btn btn-primary" style="width:100%;justify-content:center;padding:.72rem;font-size:.95rem;margin-top:.2rem">
            Entrar
          </button>
        </form>
      </div>
    </div>"""
    return render(content)


@app.route("/logout", methods=["POST"])
def logout():
    # Verifica CSRF para logout via POST
    t = session.get("_csrf_token", "")
    ft = request.form.get("csrf_token", "")
    if not t or not secrets.compare_digest(t, ft):
        abort(403)
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    p = current_user().get("perfil", "aluno")
    if p == "admin":
        return redirect(url_for("admin_home"))
    if p in ("cozinha", "oficialdia", "cmd"):
        return redirect(url_for("painel_dia"))
    return redirect(url_for("aluno_home"))


# ═══════════════════════════════════════════════════════════════════════════
# ALUNO
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/aluno/licenca-fds", methods=["POST"])
@login_required
def aluno_licenca_fds():
    """Marcar ou cancelar licença de fim de semana."""
    u = current_user()
    uid = sr.user_id_by_nii(u["nii"])
    if not uid:
        flash("Conta de sistema — funcionalidade não disponível.", "error")
        return redirect(url_for("aluno_home"))

    sexta_str = request.form.get("sexta", "")
    acao = request.form.get("acao_fds", "marcar")  # "marcar" ou "cancelar"
    sexta = _parse_date_strict(sexta_str)
    if not sexta or sexta.weekday() != 4:  # 4 = sexta-feira
        flash("Data inválida — apenas sextas-feiras.", "error")
        return redirect(url_for("aluno_home"))

    # Verificar se a sexta ainda é editável
    ok_edit, msg = _dia_editavel_aluno(sexta)
    if not ok_edit:
        flash(f"Não é possível editar: {msg}", "warn")
        return redirect(url_for("aluno_home"))

    # Verificar ausência
    if _tem_ausencia_ativa(uid, sexta):
        flash("Tens uma ausência registada para este período.", "warn")
        return redirect(url_for("aluno_home"))

    # Verificar detenção
    if _tem_detencao_ativa(uid, sexta):
        flash("Estás detido — não podes marcar licença de fim de semana.", "error")
        return redirect(url_for("aluno_home"))

    if acao == "cancelar":
        ok, err = _cancelar_licenca_fds(uid, sexta, u["nii"])
        flash(
            "Licença de fim de semana cancelada." if ok else (err or "Erro."),
            "ok" if ok else "error",
        )
    else:
        # Verificar regras de licença para a sexta
        with sr.db() as conn:
            aluno_row = conn.execute(
                "SELECT ano, NI FROM utilizadores WHERE id=?", (uid,)
            ).fetchone()
        ano_aluno = int(aluno_row["ano"]) if aluno_row else 1
        ni_aluno = aluno_row["NI"] if aluno_row else ""

        pode, motivo = _pode_marcar_licenca(uid, sexta, ano_aluno, ni_aluno)
        if not pode:
            flash(motivo, "warn")
            return redirect(url_for("aluno_home"))

        ok, err = _marcar_licenca_fds(uid, sexta, u["nii"])
        flash(
            "✅ Licença de fim de semana marcada! Jantar de sexta e refeições de sábado/domingo cancelados."
            if ok
            else (err or "Erro."),
            "ok" if ok else "error",
        )

    return redirect(url_for("aluno_home"))


@app.route("/aluno")
@login_required
def aluno_home():
    u = current_user()
    uid = sr.user_id_by_nii(u["nii"])
    hoje = date.today()
    menu = sr.get_menu_do_dia(hoje)

    # Banner ausência ativa hoje
    ausente_hoje = uid and _tem_ausencia_ativa(uid, hoje)
    ausente_html = ""
    if ausente_hoje:
        ausente_html = '<div class="ausente-banner">⚓ Tens uma <strong>ausência registada</strong> para hoje. As tuas refeições não serão contabilizadas.</div>'

    menu_html = ""
    if menu:

        def mv(k):
            return esc(menu.get(k) or "—")

        menu_html = f"""
        <div class="card">
          <div class="card-title">🍽️ Ementa de hoje — {hoje.strftime("%d/%m/%Y")}</div>
          <div class="grid grid-4">
            <div><strong>Peq. Almoço</strong><br><span class="text-muted">{mv("pequeno_almoco")}</span></div>
            <div><strong>Lanche</strong><br><span class="text-muted">{mv("lanche")}</span></div>
            <div><strong>Almoço</strong><br>N: {mv("almoco_normal")}<br>V: {mv("almoco_veg")}<br>D: {mv("almoco_dieta")}</div>
            <div><strong>Jantar</strong><br>N: {mv("jantar_normal")}<br>V: {mv("jantar_veg")}<br>D: {mv("jantar_dieta")}</div>
          </div>
        </div>"""

    def chip(val, label, tp=None):
        if val:
            return f'<span class="meal-chip chip-{"type" if tp else "ok"}">{tp or label} ✓</span>'
        return f'<span class="meal-chip chip-no">{label} ✗</span>'

    # Batch-load: carregar todos os dados de uma vez (elimina N+1)
    d_ate = hoje + timedelta(days=DIAS_ANTECEDENCIA)
    cal_map = sr.dias_operacionais_batch(hoje, d_ate)
    if uid:
        ref_map, ref_defaults = sr.refeicoes_batch(uid, hoje, d_ate)
        aus_set = sr.ausencias_batch(uid, hoje, d_ate)
        det_set = sr.detencoes_batch(uid, hoje, d_ate)
        lic_map = sr.licencas_batch(uid, hoje, d_ate)
    else:
        ref_map, ref_defaults = {}, {}
        aus_set = set()
        det_set = set()
        lic_map = {}

    dias_html = ""
    for i in range(DIAS_ANTECEDENCIA + 1):
        d = hoje + timedelta(days=i)
        d_iso = d.isoformat()
        tipo = cal_map.get(d_iso, "fim_semana" if d.weekday() >= 5 else "normal")
        r = ref_map.get(d_iso, ref_defaults) if uid else {}
        ok_edit, _ = _dia_editavel_aluno(d)
        prazo = _prazo_label(d)
        ausente_d = d_iso in aus_set
        detido_d = d_iso in det_set
        is_weekend = d.weekday() >= 5
        is_off = tipo in ("feriado", "exercicio")

        if is_off:
            ic = {"feriado": "🔴", "exercicio": "🟡"}.get(tipo, "⚪")
            lb = {"feriado": "Feriado", "exercicio": "Exercício"}.get(tipo, tipo)
            dias_html += f"""
            <div class="week-card day-off">
              <div class="week-dow">{ABREV_DIAS[d.weekday()]}</div>
              <div class="week-date">{d.strftime("%d/%m")}</div>
              <span class="text-muted small">{ic} {lb}</span>
            </div>"""
            continue

        aus_chip = (
            '<span class="meal-chip chip-type" style="background:#fef3cd;color:#856404;margin-bottom:.3rem;display:block">⚓ Ausente</span>'
            if ausente_d
            else ""
        )
        det_chip = (
            '<span class="meal-chip chip-type" style="background:#fdecea;color:#7a1c1c;margin-bottom:.3rem;display:block">🚫 Detido</span>'
            if detido_d
            else ""
        )
        # Licença do batch
        lic_chip = ""
        lic_tipo = lic_map.get(d_iso)
        if uid and not detido_d and lic_tipo:
            lic_lbl = "Antes jantar" if lic_tipo == "antes_jantar" else "Após jantar"
            lic_chip = f'<span class="meal-chip chip-type" style="background:#d4efdf;color:#1e8449;margin-bottom:.3rem;display:block">🚪 {lic_lbl}</span>'
        alm_t = r.get("almoco")
        jan_t = r.get("jantar_tipo")
        meals = f"""<div class="week-meals">
            {chip(r.get("pequeno_almoco"), "PA")}
            {chip(r.get("lanche"), "Lan")}
            {chip(alm_t, "Alm", alm_t[:3] if alm_t else None)}
            {chip(jan_t, "Jan", jan_t[:3] if jan_t else None)}
          </div>{prazo}"""
        btn = (
            f'<a class="btn btn-primary btn-sm" style="margin-top:.38rem" href="{url_for("aluno_editar", d=d.isoformat())}">✏️ Editar</a>'
            if ok_edit and not ausente_d
            else ""
        )

        # ── Botão licença FDS (sextas-feiras) ────────────────────────
        fds_btn_html = ""
        is_friday = d.weekday() == 4
        if is_friday and uid and not is_off:
            tem_licenca_fds = lic_map.get(d_iso) == "antes_jantar"
            nao_detido = not detido_d
            nao_ausente = not ausente_d

            if ok_edit and nao_detido and nao_ausente:
                if tem_licenca_fds:
                    fds_btn_html = f"""
                    <form method="post" action="{url_for("aluno_licenca_fds")}" style="margin-top:.4rem">
                      {csrf_input()}
                      <input type="hidden" name="sexta" value="{d_iso}">
                      <input type="hidden" name="acao_fds" value="cancelar">
                      <button class="btn btn-danger btn-sm" style="width:100%;font-size:.7rem"
                        onclick="return confirm('Cancelar licença de fim de semana?')">
                        🔄 Cancelar licença FDS
                      </button>
                    </form>"""
                else:
                    fds_btn_html = f"""
                    <form method="post" action="{url_for("aluno_licenca_fds")}" style="margin-top:.4rem">
                      {csrf_input()}
                      <input type="hidden" name="sexta" value="{d_iso}">
                      <input type="hidden" name="acao_fds" value="marcar">
                      <button class="btn btn-gold btn-sm" style="width:100%;font-size:.7rem"
                        onclick="return confirm('Marcar licença de fim de semana?\\nIsto vai:\\n• Retirar o jantar de sexta\\n• Apagar refeições de sábado e domingo')">
                        🏖️ Licença FDS
                      </button>
                    </form>"""

        card_cls = "weekend-active" if is_weekend else ""
        dow_cls = "weekend" if is_weekend else ""

        dias_html += f"""
        <div class="week-card {card_cls}">
          <div class="week-dow {dow_cls}">{ABREV_DIAS[d.weekday()]}</div>
          <div class="week-date">{d.strftime("%d/%m/%Y")}</div>
          {aus_chip}{det_chip}{lic_chip}{meals}{fds_btn_html}{btn}
        </div>"""

    stats_html = ""
    if uid:
        d0 = (hoje - timedelta(days=30)).isoformat()
        with sr.db() as conn:
            rows = conn.execute(
                "SELECT pequeno_almoco,lanche,almoco,jantar_tipo FROM refeicoes WHERE utilizador_id=? AND data>=?",
                (uid, d0),
            ).fetchall()
        if rows:
            stats_html = f"""
            <div class="card">
              <div class="card-title">📊 Últimos 30 dias</div>
              <div class="grid grid-4">
                <div class="stat-box"><div class="stat-num">{sum(1 for r in rows if r["pequeno_almoco"])}</div><div class="stat-lbl">Pequenos Almoços</div></div>
                <div class="stat-box"><div class="stat-num">{sum(1 for r in rows if r["lanche"])}</div><div class="stat-lbl">Lanches</div></div>
                <div class="stat-box"><div class="stat-num">{sum(1 for r in rows if r["almoco"])}</div><div class="stat-lbl">Almoços</div></div>
                <div class="stat-box"><div class="stat-num">{sum(1 for r in rows if r["jantar_tipo"])}</div><div class="stat-lbl">Jantares</div></div>
              </div>
            </div>"""

    content = f"""
    <div class="container">
      <div class="page-header"><div class="page-title">Olá, {esc(u["nome"])} 👋</div></div>
      {ausente_html}{menu_html}
      <div class="card">
        <div class="card-title">📆 Próximos {DIAS_ANTECEDENCIA} dias

        </div>
        <div class="week-grid">{dias_html}</div>
      </div>
      {stats_html}
      <div class="gap-btn">
        <a class="btn btn-ghost" href="{url_for("aluno_historico")}">🕘 Histórico (30 dias)</a>
        <a class="btn btn-gold" href="{url_for("aluno_ausencias")}">🚫 Gerir ausências</a>
        <a class="btn btn-ghost" href="{url_for("aluno_password")}">🔑 Alterar password</a>
        <a class="btn btn-ghost" href="{url_for("calendario_publico")}">📅 Calendário</a>
        <a class="btn btn-primary" href="{url_for("aluno_perfil")}">👤 O meu perfil</a>
      </div>
    </div>"""
    return render(content)


@app.route("/aluno/editar/<d>", methods=["GET", "POST"])
@login_required
def aluno_editar(d):
    u = current_user()
    uid = sr.user_id_by_nii(u["nii"])
    dt = _parse_date(d)

    if not uid:
        flash("Conta de sistema — não é possível editar refeições.", "error")
        return redirect(url_for("aluno_home"))

    # Bloquear edição se tem ausência ativa
    if _tem_ausencia_ativa(uid, dt):
        flash(
            "Tens uma ausência registada para este dia. Remove a ausência primeiro.",
            "warn",
        )
        return redirect(url_for("aluno_home"))

    ok_edit, msg = _dia_editavel_aluno(dt)
    if not ok_edit:
        flash(f"Não é possível editar: {msg}", "warn")
        return redirect(url_for("aluno_home"))

    r = sr.refeicao_get(uid, dt)
    occ = _get_ocupacao_dia(dt)
    is_weekend = dt.weekday() >= 5
    detido = _tem_detencao_ativa(uid, dt)

    # Dados do aluno para regras de licença
    with sr.db() as conn:
        aluno_row = conn.execute(
            "SELECT ano, NI FROM utilizadores WHERE id=?", (uid,)
        ).fetchone()
    ano_aluno = int(aluno_row["ano"]) if aluno_row else 1
    ni_aluno = aluno_row["NI"] if aluno_row else ""

    # Licença existente para este dia
    with sr.db() as conn:
        lic_row = conn.execute(
            "SELECT tipo FROM licencas WHERE utilizador_id=? AND data=?",
            (uid, dt.isoformat()),
        ).fetchone()
    licenca_atual = lic_row["tipo"] if lic_row else ""

    pode_lic, motivo_lic = _pode_marcar_licenca(uid, dt, ano_aluno, ni_aluno)

    if request.method == "POST":
        pa = 1 if request.form.get("pa") in ("1", "on") else 0
        lanche = 1 if request.form.get("lanche") in ("1", "on") else 0
        alm = _val_refeicao(request.form.get("almoco"))
        jan = _val_refeicao(request.form.get("jantar"))
        sai = 0 if detido else (1 if request.form.get("sai") else 0)

        # Processar licença (antes_jantar / apos_jantar / vazio)
        licenca_tipo = request.form.get("licenca", "")
        with sr.db() as conn:
            if licenca_tipo in ("antes_jantar", "apos_jantar"):
                pode, motivo = _pode_marcar_licenca(uid, dt, ano_aluno, ni_aluno)
                if pode:
                    conn.execute(
                        """INSERT INTO licencas(utilizador_id, data, tipo)
                        VALUES(?,?,?)
                        ON CONFLICT(utilizador_id, data) DO UPDATE SET tipo=excluded.tipo""",
                        (uid, dt.isoformat(), licenca_tipo),
                    )
                    if licenca_tipo == "antes_jantar":
                        jan = ""
                        sai = 1
                    else:
                        sai = 1
                else:
                    flash(motivo, "warn")
            else:
                conn.execute(
                    "DELETE FROM licencas WHERE utilizador_id=? AND data=?",
                    (uid, dt.isoformat()),
                )
            conn.commit()

        if _refeicao_set(uid, dt, pa, lanche, alm, jan, sai, alterado_por=u["nii"]):
            flash("Refeições atualizadas!", "ok")
        else:
            flash("Erro ao guardar.", "error")
        return redirect(url_for("aluno_home"))

    def occ_row(nome):
        val, cap = occ.get(nome, (0, -1))
        return f'<div style="margin-bottom:.65rem"><strong style="font-size:.84rem">{nome}</strong>{_bar_html(val, cap)}</div>'

    wknd_note = (
        '<div class="alert alert-info" style="margin-bottom:.8rem">Fim de semana — refeições opcionais.</div>'
        if is_weekend
        else ""
    )

    detido_note = (
        '<div class="alert alert-warn" style="margin-bottom:.8rem">'
        "🚫 Estás detido neste dia. Não podes sair da unidade."
        "</div>"
        if detido
        else ""
    )

    # Valores atuais
    pa_on = 1 if r.get("pequeno_almoco") else 0
    lan_on = 1 if r.get("lanche") else 0
    alm_val = r.get("almoco") or ""
    jan_val = r.get("jantar_tipo") or ""
    jan_blocked = licenca_atual == "antes_jantar"

    # Secção de licença — oculta se detido
    licenca_html = ""
    if not detido:
        regras = _regras_licenca(ano_aluno, ni_aluno)
        usadas_semana = _licencas_semana_usadas(uid, dt)
        max_uteis = regras["max_dias_uteis"]
        lic_disabled = "" if pode_lic else " disabled"
        lic_warn = (
            f'<div class="alert alert-warn" style="margin-top:.5rem">{esc(motivo_lic)}</div>'
            if not pode_lic and motivo_lic
            else ""
        )

        sel_antes = " checked" if licenca_atual == "antes_jantar" else ""
        sel_apos = " checked" if licenca_atual == "apos_jantar" else ""
        sel_nenhuma = " checked" if not licenca_atual else ""

        if dt.weekday() < 4:
            quota_info = f'<span class="text-muted small">Seg-Qui usadas: <strong>{usadas_semana}/{max_uteis}</strong></span>'
        else:
            quota_info = (
                '<span class="text-muted small">Fim de semana — sem limite.</span>'
            )

        licenca_html = f"""
      <div class="card" style="border-top:3px solid #2e86c1">
        <div class="card-title">🚪 Licença de saída</div>
        {quota_info}
        <div class="sw-group" style="margin-top:.6rem">
          <label class="sw-row{"  sw-on" if not licenca_atual else ""}" data-lic>
            <input type="radio" name="licenca" value=""{sel_nenhuma}{lic_disabled}>
            <span class="sw-label">Sem licença</span>
          </label>
          <label class="sw-row{"  sw-on" if licenca_atual == "antes_jantar" else ""}" data-lic>
            <input type="radio" name="licenca" value="antes_jantar"{sel_antes}{lic_disabled}>
            <span class="sw-icon">🌅</span>
            <span class="sw-label">Antes do jantar</span>
            <span class="sw-hint">(não janta)</span>
          </label>
          <label class="sw-row{"  sw-on" if licenca_atual == "apos_jantar" else ""}" data-lic>
            <input type="radio" name="licenca" value="apos_jantar"{sel_apos}{lic_disabled}>
            <span class="sw-icon">🌙</span>
            <span class="sw-label">Após o jantar</span>
            <span class="sw-hint">(janta na unidade)</span>
          </label>
        </div>
        {lic_warn}
      </div>"""

    # Ementa do dia
    menu = sr.get_menu_do_dia(dt)
    ementa_html = ""
    if menu:

        def _mv(k):
            return esc(menu.get(k) or "—")

        ementa_html = f"""
      <div class="card" style="border-top:3px solid #f39c12">
        <div class="card-title">📋 Ementa — {dt.strftime("%d/%m/%Y")}</div>
        <div class="grid grid-4" style="font-size:.85rem">
          <div><strong>Peq. Almoço</strong><br><span class="text-muted">{_mv("pequeno_almoco")}</span></div>
          <div><strong>Lanche</strong><br><span class="text-muted">{_mv("lanche")}</span></div>
          <div><strong>Almoço</strong><br>N: {_mv("almoco_normal")}<br>V: {_mv("almoco_veg")}<br>D: {_mv("almoco_dieta")}</div>
          <div><strong>Jantar</strong><br>N: {_mv("jantar_normal")}<br>V: {_mv("jantar_veg")}<br>D: {_mv("jantar_dieta")}</div>
        </div>
      </div>"""

    content = f"""
    <style>
      .sw-group{{display:flex;flex-direction:column;gap:.45rem}}
      .sw-row{{display:flex;align-items:center;gap:.55rem;cursor:pointer;padding:.7rem .85rem;
        border:2px solid var(--border);border-radius:12px;transition:all .2s;
        user-select:none;-webkit-tap-highlight-color:transparent;background:#fff}}
      .sw-row:active{{transform:scale(.97)}}
      .sw-row.sw-on{{background:#eafaf1;border-color:#27ae60}}
      .sw-row input[type=hidden],.sw-row input[type=radio]{{display:none}}
      .sw-icon{{font-size:1.25rem;flex-shrink:0}}
      .sw-label{{flex:1;font-weight:600;font-size:.9rem}}
      .sw-hint{{font-size:.75rem;color:var(--muted);font-weight:400}}
      .sw-mark{{width:28px;height:28px;border-radius:8px;border:2px solid var(--border);
        display:flex;align-items:center;justify-content:center;font-size:.85rem;font-weight:800;
        color:transparent;transition:all .2s;background:#fff;flex-shrink:0}}
      .sw-row.sw-on .sw-mark{{background:#27ae60;border-color:#27ae60;color:#fff}}
      .sw-pills{{display:flex;gap:.35rem;flex-wrap:wrap}}
      .sw-pill{{padding:.5rem .75rem;border:2px solid var(--border);border-radius:10px;cursor:pointer;
        font-weight:600;font-size:.82rem;transition:all .2s;user-select:none;
        -webkit-tap-highlight-color:transparent;text-align:center;min-width:60px;background:#fff}}
      .sw-pill:active{{transform:scale(.95)}}
      .sw-pill.sw-sel{{background:#eafaf1;border-color:#27ae60;color:#1a5c38}}
      .sw-pill-group{{display:flex;flex-direction:column;gap:.35rem}}
      .sw-pill-row{{display:flex;align-items:center;gap:.55rem;padding:.5rem .85rem;
        border:2px solid var(--border);border-radius:12px;background:#fff}}
      .sw-pill-row.sw-on{{border-color:#27ae60;background:#f0faf4}}
      .sw-pill-label{{font-weight:600;font-size:.9rem;min-width:auto;white-space:nowrap}}
      .sw-pill-icon{{font-size:1.25rem;flex-shrink:0}}
      .sw-pill-opts{{display:flex;gap:.3rem;flex:1;justify-content:flex-end;flex-wrap:wrap}}
    </style>
    <div class="container">
      <div class="page-header">
        {_back_btn(url_for("aluno_home"))}
        <div class="page-title">🍽️ {NOMES_DIAS[dt.weekday()]}, {dt.strftime("%d/%m/%Y")}</div>
      </div>
      {wknd_note}{detido_note}{ementa_html}
      <div class="card">
        <div class="card-title">📊 Ocupação</div>
        {occ_row("Pequeno Almoço")}{occ_row("Lanche")}{occ_row("Almoço")}{occ_row("Jantar")}
      </div>
      <div class="card">
        <div class="card-title">✏️ Marcar refeições</div>
        <form method="post" id="mealForm">
          {csrf_input()}
          <input type="hidden" name="pa" id="h_pa" value="{pa_on}">
          <input type="hidden" name="lanche" id="h_lanche" value="{lan_on}">
          <input type="hidden" name="almoco" id="h_almoco" value="{esc(alm_val)}">
          <input type="hidden" name="jantar" id="h_jantar" value="{esc(jan_val)}">
          <div class="sw-group">
            <!-- PA toggle -->
            <div class="sw-row{"  sw-on" if pa_on else ""}" data-meal="pa" onclick="toggleMeal(this)">
              <span class="sw-icon">☕</span>
              <span class="sw-label">Pequeno Almoço</span>
              <span class="sw-mark">{"✓" if pa_on else ""}</span>
            </div>
            <!-- Lanche toggle -->
            <div class="sw-row{"  sw-on" if lan_on else ""}" data-meal="lanche" onclick="toggleMeal(this)">
              <span class="sw-icon">🥐</span>
              <span class="sw-label">Lanche</span>
              <span class="sw-mark">{"✓" if lan_on else ""}</span>
            </div>
            <!-- Almoço pill selector -->
            <div class="sw-pill-row{"  sw-on" if alm_val else ""}" id="alm_row">
              <span class="sw-pill-icon">🍽️</span>
              <span class="sw-pill-label">Almoço</span>
              <div class="sw-pill-opts">
                <div class="sw-pill{" sw-sel" if alm_val == "Normal" else ""}" onclick="setPill('almoco','Normal',this)">Normal</div>
                <div class="sw-pill{" sw-sel" if alm_val == "Vegetariano" else ""}" onclick="setPill('almoco','Vegetariano',this)">Veg</div>
                <div class="sw-pill{" sw-sel" if alm_val == "Dieta" else ""}" onclick="setPill('almoco','Dieta',this)">Dieta</div>
              </div>
            </div>
            <!-- Jantar pill selector -->
            <div class="sw-pill-row{"  sw-on" if jan_val and not jan_blocked else ""}" id="jan_row"{"  style=opacity:.4;pointer-events:none" if jan_blocked else ""}>
              <span class="sw-pill-icon">🌙</span>
              <span class="sw-pill-label">Jantar</span>
              <div class="sw-pill-opts">
                <div class="sw-pill{" sw-sel" if jan_val == "Normal" else ""}" onclick="setPill('jantar','Normal',this)">Normal</div>
                <div class="sw-pill{" sw-sel" if jan_val == "Vegetariano" else ""}" onclick="setPill('jantar','Vegetariano',this)">Veg</div>
                <div class="sw-pill{" sw-sel" if jan_val == "Dieta" else ""}" onclick="setPill('jantar','Dieta',this)">Dieta</div>
              </div>
            </div>
          </div>
          {licenca_html}
          <hr>
          <div class="gap-btn">
            <button class="btn btn-ok" style="flex:1;justify-content:center;padding:.7rem">💾 Guardar</button>
            <a class="btn btn-ghost" href="{url_for("aluno_home")}">Cancelar</a>
          </div>
        </form>
      </div>
    </div>
    <script>
    function toggleMeal(el){{
      var key=el.getAttribute('data-meal');
      var h=document.getElementById('h_'+key);
      var on=h.value==='1'||h.value==='on';
      h.value=on?'0':'1';
      el.classList.toggle('sw-on',!on);
      el.querySelector('.sw-mark').textContent=on?'':'✓';
    }}
    function setPill(meal,val,pill){{
      var h=document.getElementById('h_'+meal);
      var row=pill.closest('.sw-pill-row');
      var pills=row.querySelectorAll('.sw-pill');
      if(h.value===val){{
        h.value='';
        pills.forEach(function(p){{p.classList.remove('sw-sel')}});
        row.classList.remove('sw-on');
      }} else {{
        h.value=val;
        pills.forEach(function(p){{p.classList.remove('sw-sel')}});
        pill.classList.add('sw-sel');
        row.classList.add('sw-on');
      }}
    }}
    // Licença radio: highlight active + block jantar se antes_jantar
    document.querySelectorAll('[data-lic] input[type=radio]').forEach(function(r){{
      r.addEventListener('change',function(){{
        document.querySelectorAll('[data-lic]').forEach(function(l){{l.classList.remove('sw-on')}});
        r.closest('[data-lic]').classList.add('sw-on');
        syncJantar();
      }});
    }});
    function syncJantar(){{
      var antes=document.querySelector('input[name=licenca][value=antes_jantar]');
      var jr=document.getElementById('jan_row');
      if(!jr)return;
      if(antes && antes.checked){{
        jr.style.opacity='.4';jr.style.pointerEvents='none';
        document.getElementById('h_jantar').value='';
        jr.querySelectorAll('.sw-pill').forEach(function(p){{p.classList.remove('sw-sel')}});
        jr.classList.remove('sw-on');
      }} else {{
        jr.style.opacity='1';jr.style.pointerEvents='auto';
      }}
    }}
    </script>"""
    return render(content)


# ─── Aluno: Gerir ausências próprias ─────────────────────────────────────


@app.route("/aluno/ausencias", methods=["GET", "POST"])
@login_required
def aluno_ausencias():
    u = current_user()
    uid = sr.user_id_by_nii(u["nii"])
    if not uid:
        flash("Conta de sistema — funcionalidade não disponível.", "error")
        return redirect(url_for("aluno_home"))

    if request.method == "POST":
        acao = request.form.get("acao", "")
        if acao == "criar":
            de = request.form.get("de", "")
            ate = request.form.get("ate", "")
            motivo = _val_text(request.form.get("motivo", ""))[:500]
            ok, err = _registar_ausencia(uid, de, ate, motivo, u["nii"])
            flash(
                "Ausência registada com sucesso!" if ok else (err or "Erro."),
                "ok" if ok else "error",
            )
        elif acao == "editar":
            aid = _val_int_id(request.form.get("id", ""))
            de = request.form.get("de", "")
            ate = request.form.get("ate", "")
            motivo = _val_text(request.form.get("motivo", ""))[:500]
            if aid is None:
                flash("ID inválido.", "error")
            else:
                ok, err = _editar_ausencia(aid, uid, de, ate, motivo)
                flash(
                    "Ausência atualizada!" if ok else (err or "Erro."),
                    "ok" if ok else "error",
                )
        elif acao == "remover":
            aid = _val_int_id(request.form.get("id", ""))
            if aid is not None:
                with sr.db() as conn:
                    conn.execute(
                        "DELETE FROM ausencias WHERE id=? AND utilizador_id=?",
                        (aid, uid),
                    )
                    conn.commit()
                flash("Ausência removida.", "ok")
        return redirect(url_for("aluno_ausencias"))

    with sr.db() as conn:
        rows = [
            dict(r)
            for r in conn.execute(
                "SELECT id,ausente_de,ausente_ate,motivo FROM ausencias WHERE utilizador_id=? ORDER BY ausente_de DESC",
                (uid,),
            ).fetchall()
        ]

    hoje = date.today().isoformat()
    edit_id = request.args.get("edit", "")
    edit_row = next((r for r in rows if str(r["id"]) == edit_id), None)

    if edit_row:
        form_title = "✏️ Editar ausência"
        form_action = "editar"
        form_de = edit_row["ausente_de"]
        form_ate = edit_row["ausente_ate"]
        form_motivo = edit_row["motivo"] or ""
        form_id_inp = f'<input type="hidden" name="id" value="{edit_row["id"]}">'
        cancel_btn = (
            f'<a class="btn btn-ghost" href="{url_for("aluno_ausencias")}">Cancelar</a>'
        )
    else:
        form_title = "➕ Nova ausência"
        form_action = "criar"
        form_de = form_ate = form_motivo = ""
        form_id_inp = ""
        cancel_btn = ""

    rows_html = ""
    for r in rows:
        is_atual = r["ausente_de"] <= hoje <= r["ausente_ate"]
        is_futura = r["ausente_de"] > hoje
        estado = (
            '<span class="badge badge-warn">Atual</span>'
            if is_atual
            else (
                '<span class="badge badge-info">Futura</span>'
                if is_futura
                else '<span class="badge badge-muted">Passada</span>'
            )
        )
        pode = is_atual or is_futura
        edit_btn = (
            f'<a class="btn btn-ghost btn-sm" href="{url_for("aluno_ausencias")}?edit={r["id"]}">✏️</a>'
            if pode
            else ""
        )
        rem_form = (
            (
                f'<form method="post" style="display:inline">{csrf_input()}'
                f'<input type="hidden" name="acao" value="remover">'
                f'<input type="hidden" name="id" value="{r["id"]}">'
                f'<button class="btn btn-danger btn-sm" onclick="return confirm(\'Remover ausência?\')">🗑</button></form>'
            )
            if pode
            else ""
        )
        rows_html += f"""<tr>
          <td>{r["ausente_de"]}</td><td>{r["ausente_ate"]}</td>
          <td>{esc(r["motivo"] or "—")}</td><td>{estado}</td>
          <td><div class="gap-btn">{edit_btn}{rem_form}</div></td>
        </tr>"""

    content = f"""
    <div class="container">
      <div class="page-header">
        {_back_btn(url_for("aluno_home"))}
        <div class="page-title">🚫 As minhas ausências</div>
      </div>
      <div class="alert alert-info">
        📌 Com uma ausência ativa as tuas refeições não são contabilizadas e não podes editar refeições para esse período.
      </div>
      <div class="card" style="max-width:520px">
        <div class="card-title">{form_title}</div>
        <form method="post">
          {csrf_input()}
          <input type="hidden" name="acao" value="{form_action}">
          {form_id_inp}
          <div class="grid grid-2">
            <div class="form-group">
              <label>De</label>
              <input type="date" name="de" value="{form_de}" required min="{date.today().isoformat()}">
            </div>
            <div class="form-group">
              <label>Até</label>
              <input type="date" name="ate" value="{form_ate}" required min="{date.today().isoformat()}">
            </div>
          </div>
          <div class="form-group">
            <label>Motivo (opcional)</label>
            <input type="text" name="motivo" maxlength="500" value="{esc(form_motivo)}" placeholder="Ex: deslocação, exercício, visita...">
          </div>
          <div class="gap-btn">
            <button class="btn btn-ok">{"Atualizar" if edit_row else "Registar ausência"}</button>
            {cancel_btn}
          </div>
        </form>
      </div>
      <div class="card">
        <div class="card-title">Histórico de ausências</div>
        <div class="table-wrap">
          <table>
            <thead><tr><th>De</th><th>Até</th><th>Motivo</th><th>Estado</th><th>Ações</th></tr></thead>
            <tbody>{rows_html or '<tr><td colspan="5" class="text-muted" style="padding:1.5rem;text-align:center">Sem ausências registadas.</td></tr>'}</tbody>
          </table>
        </div>
      </div>
    </div>"""
    return render(content)


@app.route("/aluno/historico")
@login_required
def aluno_historico():
    u = current_user()
    uid = sr.user_id_by_nii(u["nii"])
    hoje = date.today()
    rows = []
    if uid:
        with sr.db() as conn:
            rows = conn.execute(
                """SELECT data,pequeno_almoco,lanche,almoco,jantar_tipo,jantar_sai_unidade
              FROM refeicoes WHERE utilizador_id=? AND data>=? ORDER BY data DESC""",
                (uid, (hoje - timedelta(days=30)).isoformat()),
            ).fetchall()

    def yn(v):
        return "✅" if v else "❌"

    rows_html = "".join(
        f"<tr><td>{r['data']}</td><td>{yn(r['pequeno_almoco'])}</td><td>{yn(r['lanche'])}</td>"
        f"<td>{r['almoco'] or '—'}</td><td>{r['jantar_tipo'] or '—'}</td>"
        f"<td>{'✅' if r['jantar_sai_unidade'] else '—'}</td></tr>"
        for r in rows
    )

    export_btns = ""
    if rows:
        export_btns = f"""
      <div class="gap-btn" style="margin-top:.8rem">
        <a class="btn btn-ghost btn-sm" href="{url_for("exportar_historico_aluno", fmt="csv")}">📄 Exportar CSV</a>
        <a class="btn btn-ghost btn-sm" href="{url_for("exportar_historico_aluno", fmt="xlsx")}">📊 Exportar Excel</a>
      </div>"""

    content = f"""
    <div class="container">
      <div class="page-header">{_back_btn(url_for("aluno_home"))}<div class="page-title">🕘 Histórico — 30 dias</div></div>
      <div class="card">
        <div class="table-wrap">
          <table>
            <thead><tr><th>Data</th><th>PA</th><th>Lanche</th><th>Almoço</th><th>Jantar</th><th>Sai</th></tr></thead>
            <tbody>{rows_html or '<tr><td colspan="6" class="text-muted center" style="padding:1.5rem">Sem registos.</td></tr>'}</tbody>
          </table>
        </div>
        {export_btns}
      </div>
    </div>"""
    return render(content)


@app.route("/aluno/exportar-historico")
@login_required
def exportar_historico_aluno():
    import io
    import csv as _csv

    u = current_user()
    uid = sr.user_id_by_nii(u["nii"])
    fmt = (request.args.get("fmt", "csv") or "csv").strip().lower()
    if fmt not in {"csv", "xlsx"}:
        abort(400)

    hoje = date.today()
    rows = []
    if uid:
        with sr.db() as conn:
            rows = conn.execute(
                """SELECT data,pequeno_almoco,lanche,almoco,jantar_tipo,jantar_sai_unidade
              FROM refeicoes WHERE utilizador_id=? AND data>=? ORDER BY data DESC""",
                (uid, (hoje - timedelta(days=30)).isoformat()),
            ).fetchall()

    headers = ["Data", "PA", "Lanche", "Almoço", "Jantar", "Sai Unidade"]

    def make_row(r):
        return [
            r["data"],
            "Sim" if r["pequeno_almoco"] else "Não",
            "Sim" if r["lanche"] else "Não",
            r["almoco"] or "—",
            r["jantar_tipo"] or "—",
            "Sim" if r["jantar_sai_unidade"] else "Não",
        ]

    nome_ficheiro = f"historico_{u['nii']}_{hoje.isoformat()}"

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Histórico {u['nome'][:20]}"

            header_fill = PatternFill("solid", fgColor="0A2D4E")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")
                c.border = border

            alt_fill = PatternFill("solid", fgColor="EBF5FB")
            for i, r in enumerate(rows, 2):
                row_data = make_row(r)
                fill = alt_fill if i % 2 == 0 else PatternFill()
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=i, column=col, value=val)
                    c.fill = fill
                    c.border = border
                    c.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col) + 4
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 22)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(
                buf.read(),
                headers={
                    "Content-Disposition": f"attachment; filename={nome_ficheiro}.xlsx",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
            )
        except ImportError:
            fmt = "csv"

    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=";")
    writer.writerow(headers)
    for r in rows:
        writer.writerow(make_row(r))
    csv_bytes = ("\ufeff" + buf.getvalue()).encode("utf-8")
    return Response(
        csv_bytes,
        headers={
            "Content-Disposition": f"attachment; filename={nome_ficheiro}.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


@app.route("/exportar/mensal")
@role_required("cozinha", "oficialdia", "cmd", "admin")
def exportar_mensal():
    import io
    import csv as _csv

    mes = request.args.get("mes", "")
    fmt = (request.args.get("fmt", "csv") or "csv").strip().lower()
    if fmt not in {"csv", "xlsx"}:
        abort(400)

    # Default: mês atual
    hoje = date.today()
    if mes:
        try:
            ano_m, mes_m = mes.split("-")
            d0 = date(int(ano_m), int(mes_m), 1)
        except (ValueError, IndexError):
            abort(400)
    else:
        d0 = date(hoje.year, hoje.month, 1)

    # Último dia do mês
    if d0.month == 12:
        d1 = date(d0.year + 1, 1, 1) - timedelta(days=1)
    else:
        d1 = date(d0.year, d0.month + 1, 1) - timedelta(days=1)

    MESES_PT = {
        1: "Janeiro",
        2: "Fevereiro",
        3: "Março",
        4: "Abril",
        5: "Maio",
        6: "Junho",
        7: "Julho",
        8: "Agosto",
        9: "Setembro",
        10: "Outubro",
        11: "Novembro",
        12: "Dezembro",
    }
    nome_mes = MESES_PT.get(d0.month, str(d0.month))

    dias_data = []
    totais = {
        k: 0
        for k in [
            "pa",
            "lan",
            "alm_norm",
            "alm_veg",
            "alm_dieta",
            "jan_norm",
            "jan_veg",
            "jan_dieta",
            "jan_sai",
        ]
    }
    _men_map, _men_empty = sr.get_totais_periodo(d0.isoformat(), d1.isoformat())
    _men_cal = sr.dias_operacionais_batch(d0, d1)
    di = d0
    while di <= d1:
        t = _men_map.get(di.isoformat(), _men_empty)
        tipo = _men_cal.get(
            di.isoformat(), "fim_semana" if di.weekday() >= 5 else "normal"
        )
        alm = t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]
        jan = t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]
        dias_data.append((di, tipo, t, alm, jan))
        for k in totais:
            totais[k] += t[k]
        di += timedelta(days=1)

    HEADERS = [
        "Data",
        "Dia da Semana",
        "Tipo Dia",
        "PA",
        "Lanche",
        "Alm. Normal",
        "Alm. Veg.",
        "Alm. Dieta",
        "Total Almoços",
        "Jan. Normal",
        "Jan. Veg.",
        "Jan. Dieta",
        "Total Jantares",
        "Sai Unidade",
    ]

    def make_row(di, tipo, t, alm, jan):
        return [
            di.isoformat(),
            NOMES_DIAS[di.weekday()],
            tipo,
            t["pa"],
            t["lan"],
            t["alm_norm"],
            t["alm_veg"],
            t["alm_dieta"],
            alm,
            t["jan_norm"],
            t["jan_veg"],
            t["jan_dieta"],
            jan,
            t["jan_sai"],
        ]

    nome_ficheiro = f"relatorio_mensal_{d0.strftime('%Y-%m')}"

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{nome_mes} {d0.year}"

            header_fill = PatternFill("solid", fgColor="0A2D4E")
            header_font = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col, h in enumerate(HEADERS, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")
                c.border = border

            tipo_fills = {
                "feriado": PatternFill("solid", fgColor="FFD6D6"),
                "exercicio": PatternFill("solid", fgColor="FFFACD"),
                "fim_semana": PatternFill("solid", fgColor="DDEEFF"),
            }

            for i, (di, tipo, t, alm, jan) in enumerate(dias_data, 2):
                row_data = make_row(di, tipo, t, alm, jan)
                fill = tipo_fills.get(tipo, PatternFill())
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=i, column=col, value=val)
                    c.fill = fill
                    c.border = border
                    c.alignment = Alignment(horizontal="center")

            # Total row
            total_row_idx = len(dias_data) + 2
            total_alm = totais["alm_norm"] + totais["alm_veg"] + totais["alm_dieta"]
            total_jan = totais["jan_norm"] + totais["jan_veg"] + totais["jan_dieta"]
            total_data = [
                "TOTAL",
                "",
                "",
                totais["pa"],
                totais["lan"],
                totais["alm_norm"],
                totais["alm_veg"],
                totais["alm_dieta"],
                total_alm,
                totais["jan_norm"],
                totais["jan_veg"],
                totais["jan_dieta"],
                total_jan,
                totais["jan_sai"],
            ]
            total_fill = PatternFill("solid", fgColor="D5F5E3")
            total_font = Font(bold=True)
            for col, val in enumerate(total_data, 1):
                c = ws.cell(row=total_row_idx, column=col, value=val)
                c.fill = total_fill
                c.font = total_font
                c.border = border
                c.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col) + 4
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 22)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(
                buf.read(),
                headers={
                    "Content-Disposition": f"attachment; filename={nome_ficheiro}.xlsx",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
            )
        except ImportError:
            fmt = "csv"

    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=";")
    writer.writerow(HEADERS)
    for di, tipo, t, alm, jan in dias_data:
        writer.writerow(make_row(di, tipo, t, alm, jan))
    total_alm = totais["alm_norm"] + totais["alm_veg"] + totais["alm_dieta"]
    total_jan = totais["jan_norm"] + totais["jan_veg"] + totais["jan_dieta"]
    writer.writerow(
        [
            "TOTAL",
            "",
            "",
            totais["pa"],
            totais["lan"],
            totais["alm_norm"],
            totais["alm_veg"],
            totais["alm_dieta"],
            total_alm,
            totais["jan_norm"],
            totais["jan_veg"],
            totais["jan_dieta"],
            total_jan,
            totais["jan_sai"],
        ]
    )
    csv_bytes = ("\ufeff" + buf.getvalue()).encode("utf-8")
    return Response(
        csv_bytes,
        headers={
            "Content-Disposition": f"attachment; filename={nome_ficheiro}.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


@app.route("/aluno/password", methods=["GET", "POST"])
@login_required
def aluno_password():
    u = current_user()
    if request.method == "POST":
        old = request.form.get("old", "")
        new = request.form.get("new", "")
        conf = request.form.get("conf", "")
        if new != conf:
            flash("As passwords não coincidem.", "error")
        else:
            ok, err = _alterar_password(u["nii"], old, new)
            flash(
                "Password alterada!" if ok else (err or "Erro."),
                "ok" if ok else "error",
            )
            if ok:
                session.pop("must_change_password", None)
                return redirect(url_for("dashboard"))

    is_forced = session.get("must_change_password")
    title = "🔐 Definir nova password" if is_forced else "🔑 Alterar password"
    old_hint = "A tua password atual (NII se é o primeiro login)" if is_forced else ""
    forced_note = (
        '<div class="alert alert-warn" style="margin-bottom:1rem">⚠️ É o teu primeiro login. Define uma password pessoal para continuar.</div>'
        if is_forced
        else ""
    )
    cancel_btn = (
        ""
        if is_forced
        else f'<a class="btn btn-ghost" href="{url_for("aluno_home")}">Cancelar</a>'
    )
    back_btn = "" if is_forced else _back_btn(url_for("aluno_home"))

    content = f"""
    <div class="container">
      <div class="page-header">{back_btn}<div class="page-title">{title}</div></div>
      {forced_note}
      <div class="card" style="max-width:440px">
        <form method="post">
          {csrf_input()}
          <div class="form-group"><label>Password atual</label><input type="password" name="old" maxlength="256" required placeholder="{old_hint}"></div>
          <div class="form-group"><label>Nova password (mín. 8 caracteres, letras e números)</label><input type="password" name="new" maxlength="256" required minlength="8"></div>
          <div class="form-group"><label>Confirmar nova password</label><input type="password" name="conf" maxlength="256" required></div>
          <div class="gap-btn"><button class="btn btn-ok">💾 Guardar</button>{cancel_btn}</div>
        </form>
      </div>
    </div>"""
    return render(content)


# ─── Aluno: Perfil próprio ────────────────────────────────────────────────


@app.route("/aluno/perfil", methods=["GET", "POST"])
@login_required
def aluno_perfil():
    u = current_user()
    uid = sr.user_id_by_nii(u["nii"])
    if not uid:
        flash("Conta de sistema — funcionalidade não disponível.", "error")
        return redirect(url_for("aluno_home"))

    with sr.db() as conn:
        row = dict(
            conn.execute(
                "SELECT NII, NI, Nome_completo, ano, email, telemovel FROM utilizadores WHERE id=?",
                (uid,),
            ).fetchone()
        )

    if request.method == "POST":
        email_n = _val_email(request.form.get("email", ""))
        if email_n is False:
            flash("Email inválido.", "error")
            return redirect(url_for("aluno_perfil"))
        telef_n = _val_phone(request.form.get("telemovel", ""))
        if telef_n is False:
            flash("Telemóvel inválido.", "error")
            return redirect(url_for("aluno_perfil"))
        try:
            with sr.db() as conn:
                conn.execute(
                    "UPDATE utilizadores SET email=?, telemovel=? WHERE id=?",
                    (email_n, telef_n, uid),
                )
                conn.commit()
            flash("Perfil atualizado com sucesso!", "ok")
            return redirect(url_for("aluno_perfil"))
        except Exception as ex:
            flash(f"Erro: {ex}", "error")

    hoje = date.today()
    with sr.db() as conn:
        total_ref = conn.execute(
            "SELECT COUNT(*) c FROM refeicoes WHERE utilizador_id=?", (uid,)
        ).fetchone()["c"]
        ausencias_ativas = conn.execute(
            """SELECT COUNT(*) c FROM ausencias WHERE utilizador_id=?
               AND ausente_de<=? AND ausente_ate>=?""",
            (uid, hoje.isoformat(), hoje.isoformat()),
        ).fetchone()["c"]

    content = f"""
    <div class="container">
      <div class="page-header">
        {_back_btn(url_for("aluno_home"))}
        <div class="page-title">👤 O meu perfil</div>
      </div>
      <div class="grid grid-2">
        <div class="card">
          <div class="card-title">ℹ️ Informação pessoal</div>
          <div style="display:flex;flex-direction:column;gap:.6rem;font-size:.9rem">
            <div><span class="text-muted">Nome completo:</span><br><strong>{esc(row["Nome_completo"])}</strong></div>
            <div><span class="text-muted">NII:</span><br><strong>{esc(row["NII"])}</strong></div>
            <div><span class="text-muted">NI:</span><br><strong>{esc(row["NI"] or "—")}</strong></div>
            <div><span class="text-muted">Ano:</span><br><strong>{row["ano"]}º Ano</strong></div>
          </div>
          <hr style="margin:1rem 0">
          <div class="grid grid-2">
            <div class="stat-box"><div class="stat-num">{total_ref}</div><div class="stat-lbl">Refeições registadas</div></div>
            <div class="stat-box"><div class="stat-num" style="color:{"var(--warn)" if ausencias_ativas else "var(--ok)"}">{ausencias_ativas}</div><div class="stat-lbl">Ausências ativas</div></div>
          </div>
        </div>
        <div class="card">
          <div class="card-title">✉️ Contactos</div>
          <form method="post">
            {csrf_input()}
            <div class="form-group">
              <label>📧 Email</label>
              <input type="email" name="email" value="{esc(row.get("email") or "")}" placeholder="o-teu-email@exemplo.pt">
            </div>
            <div class="form-group">
              <label>📱 Telemóvel</label>
              <input type="tel" name="telemovel" value="{esc(row.get("telemovel") or "")}" placeholder="+351XXXXXXXXX">
            </div>
            <div class="gap-btn">
              <button class="btn btn-ok">💾 Guardar contactos</button>
              <a class="btn btn-ghost" href="{url_for("aluno_home")}">Cancelar</a>
            </div>
          </form>
        </div>
      </div>
      <div class="card">
        <div class="card-title">⚡ Ações rápidas</div>
        <div class="gap-btn">
          <a class="btn btn-ghost" href="{url_for("aluno_ausencias")}">🚫 Gerir ausências</a>
          <a class="btn btn-ghost" href="{url_for("aluno_historico")}">🕘 Histórico de refeições</a>
          <a class="btn btn-ghost" href="{url_for("aluno_password")}">🔑 Alterar password</a>
        </div>
      </div>
    </div>"""
    return render(content)


# ═══════════════════════════════════════════════════════════════════════════
# PAINEL OPERACIONAL
# ═══════════════════════════════════════════════════════════════════════════


def _alertas_painel(d_str: str, perfil: str) -> list:
    """Gera alertas operacionais para o painel do dia (sem tabela extra)."""
    alertas: list = []
    if perfil not in ("oficialdia", "cmd", "admin"):
        return alertas

    hoje = date.today().isoformat()
    amanha = (date.today() + timedelta(days=1)).isoformat()

    with sr.db() as conn:
        # 1. Detenções que expiram hoje
        det_exp = conn.execute(
            """SELECT COUNT(*) c FROM detencoes d
               JOIN utilizadores u ON u.id=d.utilizador_id
               WHERE d.detido_ate=? AND u.perfil='aluno'""",
            (hoje,),
        ).fetchone()["c"]
        if det_exp:
            alertas.append(
                {
                    "icon": "⛔",
                    "msg": f"{det_exp} detenção(ões) expira(m) hoje.",
                    "cat": "warn",
                }
            )

        # 2. Licenças sem registo de saída
        lic_pend = conn.execute(
            """SELECT COUNT(*) c FROM licencas
               WHERE data=? AND hora_saida IS NULL""",
            (hoje,),
        ).fetchone()["c"]
        if lic_pend:
            alertas.append(
                {
                    "icon": "🚪",
                    "msg": f"{lic_pend} licença(s) sem registo de saída.",
                    "cat": "warn",
                }
            )

        # 3. Alunos ativos sem refeições para amanhã (dias úteis)
        amanha_dt = date.today() + timedelta(days=1)
        if amanha_dt.weekday() < 5:
            sem_ref = conn.execute(
                """SELECT COUNT(*) c FROM utilizadores u
                   WHERE u.perfil='aluno' AND u.is_active=1
                   AND NOT EXISTS (
                       SELECT 1 FROM refeicoes r
                       WHERE r.utilizador_id=u.id AND r.data=?
                   )
                   AND NOT EXISTS (
                       SELECT 1 FROM ausencias a
                       WHERE a.utilizador_id=u.id
                       AND a.ausente_de<=? AND a.ausente_ate>=?
                   )""",
                (amanha, amanha, amanha),
            ).fetchone()["c"]
            if sem_ref:
                alertas.append(
                    {
                        "icon": "📋",
                        "msg": f"{sem_ref} aluno(s) sem refeições marcadas para amanhã.",
                        "cat": "info",
                    }
                )

        # 4. Ausências registadas hoje
        novas_aus = conn.execute(
            """SELECT COUNT(*) c FROM ausencias
               WHERE date(criado_em)=?""",
            (hoje,),
        ).fetchone()["c"]
        if novas_aus:
            alertas.append(
                {
                    "icon": "🚫",
                    "msg": f"{novas_aus} ausência(s) registada(s) hoje.",
                    "cat": "info",
                }
            )

    return alertas


@app.route("/painel", methods=["GET", "POST"])
@role_required("cozinha", "oficialdia", "cmd", "admin")
def painel_dia():
    u = current_user()
    perfil = u.get("perfil")
    d_str = request.args.get("d", date.today().isoformat())
    dt = _parse_date(d_str)

    if request.method == "POST":
        acao = request.form.get("acao", "")
        if acao == "backup":
            try:
                sr.ensure_daily_backup()
                flash("Backup criado.", "ok")
            except Exception as e:
                flash(f"Falha: {e}", "error")
        return redirect(url_for("painel_dia", d=dt.isoformat()))

    ano_int = int(u["ano"]) if perfil == "cmd" and u.get("ano") else None
    t = sr.get_totais_dia(dt.isoformat(), ano_int)
    occ = _get_ocupacao_dia(dt)

    def occ_card(nome, icon):
        val, cap = occ.get(nome, (0, -1))
        bar = _bar_html(val, cap) if cap > 0 else ""
        return f'<div class="stat-box"><div class="stat-num">{val}</div><div class="stat-lbl">{icon} {nome}</div>{bar}</div>'

    detail = f"""
    <div class="grid grid-3" style="margin-top:.9rem">
      <div class="stat-box"><div class="stat-num">{t["alm_norm"]}</div><div class="stat-lbl">Almoço Normal</div></div>
      <div class="stat-box"><div class="stat-num">{t["alm_veg"]}</div><div class="stat-lbl">Almoço Vegetariano</div></div>
      <div class="stat-box"><div class="stat-num">{t["alm_dieta"]}</div><div class="stat-lbl">Almoço Dieta</div></div>
      <div class="stat-box"><div class="stat-num">{t["jan_norm"]}</div><div class="stat-lbl">Jantar Normal</div></div>
      <div class="stat-box"><div class="stat-num">{t["jan_veg"]}</div><div class="stat-lbl">Jantar Vegetariano</div></div>
      <div class="stat-box"><div class="stat-num">{t["jan_dieta"]}</div><div class="stat-lbl">Jantar Dieta</div></div>
    </div>"""

    # ── Alertas operacionais ──────────────────────────────────────────────
    alertas = _alertas_painel(d_str, perfil)
    alertas_html = ""
    if alertas:
        items = "".join(
            f'<div class="alert alert-{a["cat"]}" style="margin-bottom:.4rem">'
            f"{a['icon']} {esc(a['msg'])}</div>"
            for a in alertas
        )
        alertas_html = f'<div style="margin-bottom:1rem">{items}</div>'

    # ── Previsão de amanhã (cozinha / admin) ─────────────────
    previsao_html = ""
    if perfil in ("cozinha", "admin"):
        amanha = dt + timedelta(days=1)
        t_am = sr.get_totais_dia(amanha.isoformat(), ano_int)

        def _delta(h, a):
            d = a - h
            if d > 0:
                return f'<span style="color:#1e8449;font-size:.72rem"> ↑{d}</span>'
            if d < 0:
                return f'<span style="color:#c0392b;font-size:.72rem"> ↓{abs(d)}</span>'
            return '<span style="color:#6c757d;font-size:.72rem"> =</span>'

        alm_h = t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]
        jan_h = t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]
        alm_a = t_am["alm_norm"] + t_am["alm_veg"] + t_am["alm_dieta"]
        jan_a = t_am["jan_norm"] + t_am["jan_veg"] + t_am["jan_dieta"]

        previsao_html = f"""
      <div class="card" style="border-top:3px solid #2e86c1">
        <div class="card-title">🔮 Previsão Amanhã — {NOMES_DIAS[amanha.weekday()]} {amanha.strftime("%d/%m")}</div>
        <div class="grid grid-4">
          <div class="stat-box"><div class="stat-num">{t_am["pa"]}{_delta(t["pa"], t_am["pa"])}</div><div class="stat-lbl">☕ PA</div></div>
          <div class="stat-box"><div class="stat-num">{t_am["lan"]}{_delta(t["lan"], t_am["lan"])}</div><div class="stat-lbl">🥐 Lanche</div></div>
          <div class="stat-box"><div class="stat-num">{alm_a}{_delta(alm_h, alm_a)}</div><div class="stat-lbl">🍽️ Almoços</div></div>
          <div class="stat-box"><div class="stat-num">{jan_a}{_delta(jan_h, jan_a)}</div><div class="stat-lbl">🌙 Jantares</div></div>
        </div>
        <div class="grid grid-3" style="margin-top:.5rem">
          <div class="stat-box"><div class="stat-num">{t_am["alm_norm"]}</div><div class="stat-lbl">Alm Normal</div></div>
          <div class="stat-box"><div class="stat-num">{t_am["alm_veg"]}</div><div class="stat-lbl">Alm Veg</div></div>
          <div class="stat-box"><div class="stat-num">{t_am["alm_dieta"]}</div><div class="stat-lbl">Alm Dieta</div></div>
        </div>
      </div>"""

    prev_d = (dt - timedelta(days=1)).isoformat()
    next_d = (dt + timedelta(days=1)).isoformat()
    nav_data = f"""
    <div class="flex-between" style="margin-bottom:1.1rem">
      <div class="flex">
        <a class="btn btn-ghost btn-sm" href="{url_for("painel_dia", d=prev_d)}">← Anterior</a>
        <strong>{NOMES_DIAS[dt.weekday()]}, {dt.strftime("%d/%m/%Y")}</strong>
        <a class="btn btn-ghost btn-sm" href="{url_for("painel_dia", d=next_d)}">Próximo →</a>
      </div>
      <form method="get" style="display:flex;gap:.3rem">
        <input type="date" name="d" value="{d_str}" style="width:auto">
        <button class="btn btn-primary btn-sm">Ir</button>
      </form>
    </div>"""

    # Ações rápidas por perfil
    acoes = []
    if perfil in ("cozinha", "oficialdia", "admin"):
        acoes.append(
            f'<a class="btn btn-ghost" href="{url_for("dashboard_semanal")}">📊 Dashboard</a>'
        )
        acoes.append(
            f'<a class="btn btn-ghost" href="{url_for("admin_menus")}">🍽️ Menus &amp; Capacidade</a>'
        )
        acoes.append(
            f'<a class="btn btn-ghost" href="{url_for("calendario_publico")}">📅 Calendário</a>'
        )
        acoes.append(
            f'<a class="btn btn-ghost" href="{url_for("relatorio_semanal")}">📈 Relatório Semanal</a>'
        )

    if perfil in ("oficialdia", "admin"):
        anos = _get_anos_disponiveis()
        for ano in anos:
            acoes.append(
                f'<a class="btn btn-ghost" href="{url_for("lista_alunos_ano", ano=ano, d=d_str)}">👥 {ano}º Ano</a>'
            )
        acoes.append(
            f'<a class="btn btn-primary" href="{url_for("controlo_presencas", d=dt.isoformat())}">🎯 Controlo Presenças</a>'
        )
        acoes.append(
            f'<a class="btn btn-warn" href="{url_for("excecoes_dia", d=dt.isoformat())}">📝 Exceções</a>'
        )
        acoes.append(
            f'<a class="btn btn-ghost" href="{url_for("ausencias")}">🚫 Ausências</a>'
        )
        acoes.append(
            f'<a class="btn btn-gold" href="{url_for("licencas_entradas_saidas", d=d_str)}">🚪 Licenças / Entradas</a>'
        )

    if perfil == "cmd" and u.get("ano"):
        acoes.append(
            f'<a class="btn btn-ghost" href="{url_for("lista_alunos_ano", ano=u["ano"], d=d_str)}">👥 Lista do {u["ano"]}º Ano</a>'
        )
        acoes.append(
            f'<a class="btn btn-ghost" href="{url_for("imprimir_ano", ano=u["ano"], d=d_str)}" target="_blank">🖨 Imprimir mapa</a>'
        )
        acoes.append(
            f'<a class="btn btn-gold" href="{url_for("ausencias_cmd")}">🚫 Ausências do {u["ano"]}º Ano</a>'
        )
        acoes.append(
            f'<a class="btn btn-warn" href="{url_for("detencoes_cmd")}">⛔ Detenções</a>'
        )
        acoes.append(
            f'<a class="btn btn-ghost" href="{url_for("calendario_publico")}">📅 Calendário</a>'
        )

    backup_btn = ""
    if perfil in ("oficialdia", "admin"):
        backup_btn = f'<form method="post" style="display:inline">{csrf_input()}<input type="hidden" name="acao" value="backup"><button class="btn btn-ghost">💾 Backup BD</button></form>'
    if perfil == "admin":
        backup_btn += f' <a class="btn btn-ghost" href="{url_for("admin_backup_download")}">📥 Download BD</a>'

    # Painel de detidos (oficialdia / admin)
    detidos_html = ""
    if perfil in ("oficialdia", "admin"):
        with sr.db() as conn:
            detidos = [
                dict(r)
                for r in conn.execute(
                    """SELECT uu.NI, uu.Nome_completo, uu.ano, d.detido_de, d.detido_ate, d.motivo
                    FROM detencoes d JOIN utilizadores uu ON uu.id=d.utilizador_id
                    WHERE uu.perfil='aluno' AND d.detido_de<=? AND d.detido_ate>=?
                    ORDER BY uu.ano, uu.NI""",
                    (d_str, d_str),
                ).fetchall()
            ]
        if detidos:
            det_rows = "".join(
                f"<tr><td>{esc(r['NI'])}</td><td>{esc(r['Nome_completo'])}</td>"
                f"<td>{r['ano']}º</td><td>{r['detido_de']} a {r['detido_ate']}</td>"
                f"<td>{esc(r['motivo'] or '—')}</td></tr>"
                for r in detidos
            )
            detidos_html = f"""
      <div class="card" style="border-top:3px solid #e74c3c">
        <div class="card-title">⛔ Detidos hoje ({len(detidos)})</div>
        <div class="table-wrap"><table>
          <thead><tr><th>NI</th><th>Nome</th><th>Ano</th><th>Período</th><th>Motivo</th></tr></thead>
          <tbody>{det_rows}</tbody>
        </table></div>
      </div>"""

    # Licenças do dia (oficialdia / admin)
    licencas_html = ""
    if perfil in ("oficialdia", "admin"):
        with sr.db() as conn:
            lics = [
                dict(r)
                for r in conn.execute(
                    """SELECT uu.NI, uu.Nome_completo, uu.ano, l.tipo, l.hora_saida, l.hora_entrada
                    FROM licencas l JOIN utilizadores uu ON uu.id=l.utilizador_id
                    WHERE l.data=? ORDER BY uu.ano, uu.NI""",
                    (d_str,),
                ).fetchall()
            ]
        if lics:

            def _lic_tipo_badge(tp):
                if tp == "antes_jantar":
                    return '<span class="badge badge-info" style="font-size:.6rem">🌅 Antes</span>'
                return '<span class="badge badge-muted" style="font-size:.6rem">🌙 Após</span>'

            lic_rows = "".join(
                f"<tr><td>{esc(r['NI'])}</td><td>{esc(r['Nome_completo'])}</td>"
                f"<td>{r['ano']}º</td><td>{_lic_tipo_badge(r['tipo'])}</td>"
                f"<td>{'✅ ' + r['hora_saida'] if r['hora_saida'] else '—'}</td>"
                f"<td>{'✅ ' + r['hora_entrada'] if r['hora_entrada'] else '—'}</td></tr>"
                for r in lics
            )
            licencas_html = f"""
      <div class="card" style="border-top:3px solid #2e86c1">
        <div class="card-title">🚪 Licenças hoje ({len(lics)})</div>
        <div class="table-wrap"><table>
          <thead><tr><th>NI</th><th>Nome</th><th>Ano</th><th>Tipo</th><th>Saída</th><th>Entrada</th></tr></thead>
          <tbody>{lic_rows}</tbody>
        </table></div>
        <a class="btn btn-gold btn-sm" href="{url_for("licencas_entradas_saidas", d=d_str)}" style="margin-top:.5rem">Gerir entradas/saídas</a>
      </div>"""

    back = _back_btn(url_for("admin_home")) if perfil == "admin" else ""
    label_ano = f" — {ano_int}º Ano" if ano_int else ""

    content = f"""
    <div class="container">
      <div class="page-header">
        {back}
        <div class="page-title">📋 Painel Operacional{label_ano}</div>
        {backup_btn}
      </div>
      {nav_data}
      {alertas_html}
      <div class="card">
        <div class="card-title">Ocupação geral</div>
        <div class="grid grid-4">
          {occ_card("Pequeno Almoço", "☕")}
          {occ_card("Lanche", "🥐")}
          {occ_card("Almoço", "🍽️")}
          {occ_card("Jantar", "🌙")}
        </div>
        {detail}
        {'<div style="margin-top:.65rem;font-size:.81rem;color:var(--muted)">🚪 Saem após jantar: <strong>' + str(t["jan_sai"]) + "</strong></div>" if perfil != "cozinha" else ""}
      </div>
      {previsao_html}
      {detidos_html}
      {licencas_html}
      <div class="card">
        <div class="card-title">⬇ Exportar</div>
        <div class="gap-btn">
          <a class="btn btn-primary" href="{url_for("exportar_dia", d=dt.isoformat(), fmt="csv")}">CSV</a>
          <a class="btn btn-primary" href="{url_for("exportar_dia", d=dt.isoformat(), fmt="xlsx")}">Excel</a>
        </div>
      </div>
      {'<div class="card"><div class="card-title">⚡ Ações rápidas</div><div class="gap-btn">' + chr(10).join(acoes) + "</div></div>" if acoes else ""}
    </div>"""
    return render(content)


# ═══════════════════════════════════════════════════════════════════════════
# LISTA DE ALUNOS POR ANO (Oficial de Dia / CMD / Admin)
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/alunos/<int:ano>", methods=["GET", "POST"])
@role_required("oficialdia", "cmd", "admin")
def lista_alunos_ano(ano):
    u = current_user()
    perfil = u.get("perfil")

    # CMD só pode ver o seu ano
    if perfil == "cmd" and str(ano) != str(u.get("ano", "")):
        flash("Acesso restrito ao teu ano.", "error")
        return redirect(url_for("painel_dia"))

    d_str = request.args.get("d", date.today().isoformat())
    dt = _parse_date(d_str)

    # POST: marcar/desmarcar ausência via lista
    if request.method == "POST":
        acao = request.form.get("acao", "")
        uid_t = request.form.get("uid", "")
        if acao == "marcar_ausente" and uid_t:
            _registar_ausencia(
                int(uid_t),
                dt.isoformat(),
                dt.isoformat(),
                f"Marcado por {u['nome']} ({perfil})",
                u["nii"],
            )
        elif acao == "marcar_presente" and uid_t:
            with sr.db() as conn:
                conn.execute(
                    """DELETE FROM ausencias WHERE utilizador_id=?
                                AND ausente_de=? AND ausente_ate=?""",
                    (uid_t, dt.isoformat(), dt.isoformat()),
                )
                conn.commit()
        return redirect(url_for("lista_alunos_ano", ano=ano, d=d_str))

    with sr.db() as conn:
        alunos = [
            dict(r)
            for r in conn.execute(
                """
            SELECT u.id, u.NII, u.NI, u.Nome_completo,
                   r.pequeno_almoco, r.lanche, r.almoco, r.jantar_tipo, r.jantar_sai_unidade,
                   EXISTS(SELECT 1 FROM ausencias a WHERE a.utilizador_id=u.id
                          AND a.ausente_de <= ? AND a.ausente_ate >= ?) AS ausente,
                   (SELECT l.tipo FROM licencas l WHERE l.utilizador_id=u.id AND l.data=?) AS licenca_tipo
            FROM utilizadores u
            LEFT JOIN refeicoes r ON r.utilizador_id=u.id AND r.data=?
            WHERE u.ano=?
            ORDER BY u.NI
        """,
                (
                    dt.isoformat(),
                    dt.isoformat(),
                    dt.isoformat(),
                    dt.isoformat(),
                    ano,
                ),
            ).fetchall()
        ]

    t = sr.get_totais_dia(dt.isoformat(), ano)
    total_alunos = len(alunos)
    com_ref = sum(
        1
        for a in alunos
        if any([a["almoco"], a["jantar_tipo"], a["pequeno_almoco"], a["lanche"]])
    )
    ausentes = sum(1 for a in alunos if a["ausente"])

    prev_d = (dt - timedelta(days=1)).isoformat()
    next_d = (dt + timedelta(days=1)).isoformat()

    # Tabs de ano
    anos = _get_anos_disponiveis()
    tabs = ""
    if perfil in ("oficialdia", "admin"):
        tabs = (
            '<div class="year-tabs">'
            + "".join(
                f'<a class="year-tab {"active" if a == ano else ""}" href="{url_for("lista_alunos_ano", ano=a, d=d_str)}">{_ano_label(a)}</a>'
                for a in anos
            )
            + "</div>"
        )

    def chip_ref(val, label, tp=None):
        if val:
            return f'<span class="meal-chip chip-{"type" if tp else "ok"}">{tp or label} ✓</span>'
        return f'<span class="meal-chip chip-no">{label} ✗</span>'

    # Validação de prazo para esta data
    ok_prazo, _ = sr.refeicao_editavel(dt)

    rows_html = ""
    for a in alunos:
        sem = not any([a["pequeno_almoco"], a["lanche"], a["almoco"], a["jantar_tipo"]])
        row_bg = (
            "background:#fdecea"
            if a["ausente"]
            else ("background:#fff3cd" if sem else "background:#d5f5e3")
        )
        ausente_b = (
            '<span class="badge badge-warn" style="font-size:.65rem">Ausente</span>'
            if a["ausente"]
            else ""
        )
        sai_b = (
            '<span class="badge badge-muted" style="font-size:.65rem">🚪</span>'
            if a["jantar_sai_unidade"]
            else ""
        )
        lic_b = ""
        if a.get("licenca_tipo") == "antes_jantar":
            lic_b = '<span class="badge badge-info" style="font-size:.6rem">🌅 Lic. antes</span>'
        elif a.get("licenca_tipo") == "apos_jantar":
            lic_b = '<span class="badge badge-muted" style="font-size:.6rem">🌙 Lic. após</span>'
        exc_btn = ""  # Edição de refeições disponível no módulo de Exceções/Controlo de Presenças

        # Botão perfil do aluno — OD só pode VER, cmd e admin podem EDITAR
        edit_aluno_btn = ""
        if perfil == "oficialdia":
            edit_aluno_btn = f'<a class="btn btn-ghost btn-sm" href="{url_for("ver_perfil_aluno", nii=a["NII"], ano=ano, d=d_str)}" title="Ver perfil do aluno">👁</a>'
        elif perfil in ("admin", "cmd"):
            edit_aluno_btn = f'<a class="btn btn-ghost btn-sm" href="{url_for("cmd_editar_aluno", nii=a["NII"], ano=ano, d=d_str)}" title="Editar dados do aluno">👤</a>'

        # Botão de presença/ausência — removido da lista (usar módulo Controlo de Presenças)
        presenca_btn = ""

        rows_html += f"""
        <tr style="{row_bg}">
          <td class="small text-muted">{esc(a["NI"])}</td>
          <td><strong>{esc(a["Nome_completo"])}</strong> {ausente_b} {lic_b}</td>
          <td>{chip_ref(a["pequeno_almoco"], "PA")}</td>
          <td>{chip_ref(a["lanche"], "Lan")}</td>
          <td>{chip_ref(a["almoco"], "Almoço", a["almoco"][:3] if a["almoco"] else None)}</td>
          <td>{chip_ref(a["jantar_tipo"], "Jantar", a["jantar_tipo"][:3] if a["jantar_tipo"] else None)} {sai_b}</td>
          <td><div class="gap-btn">{presenca_btn}{exc_btn}{edit_aluno_btn}</div></td>
        </tr>"""

    wknd_badge = ""
    prazo_info_banner = ""
    if ok_prazo:
        prazo_info_banner = '<div class="alert alert-ok" style="margin-bottom:.7rem">✅ Os alunos ainda podem editar as próprias refeições (prazo não expirou).</div>'
    else:
        prazo_info_banner = '<div class="alert alert-info" style="margin-bottom:.7rem">🔒 Prazo expirado — os alunos já não podem alterar. Usa o botão <strong>✏️</strong> para fazer exceções.</div>'

    imprimir_btn = f'<a class="btn btn-ghost" href="{url_for("imprimir_ano", ano=ano, d=d_str)}" target="_blank">🖨 Imprimir mapa</a>'

    content = f"""
    <div class="container">
      <div class="page-header">
        {_back_btn(url_for("painel_dia", d=d_str), "Painel")}
        <div class="page-title">👥 {_ano_label(ano)} — {NOMES_DIAS[dt.weekday()]}, {dt.strftime("%d/%m/%Y")}{wknd_badge}</div>
        {imprimir_btn}
      </div>
      {tabs}
      {prazo_info_banner}
      <div class="grid grid-4" style="margin-bottom:1.1rem">
        <div class="stat-box"><div class="stat-num">{total_alunos}</div><div class="stat-lbl">Total alunos</div></div>
        <div class="stat-box"><div class="stat-num" style="color:var(--ok)">{com_ref}</div><div class="stat-lbl">Com refeições</div></div>
        <div class="stat-box"><div class="stat-num" style="color:var(--danger)">{total_alunos - com_ref}</div><div class="stat-lbl">Sem refeições</div></div>
        <div class="stat-box"><div class="stat-num" style="color:var(--warn)">{ausentes}</div><div class="stat-lbl">Ausentes</div></div>
      </div>

      <div class="card" style="padding:.9rem 1.2rem;margin-bottom:.8rem">
        <div class="flex-between">
          <div class="flex">
            <a class="btn btn-ghost btn-sm" href="{url_for("lista_alunos_ano", ano=ano, d=prev_d)}">← Anterior</a>
            <strong>{dt.strftime("%d/%m/%Y")}</strong>
            <a class="btn btn-ghost btn-sm" href="{url_for("lista_alunos_ano", ano=ano, d=next_d)}">Próximo →</a>
          </div>
          <form method="get" style="display:flex;gap:.3rem">
            <input type="date" name="d" value="{d_str}" style="width:auto">
            <button class="btn btn-primary btn-sm">Ir</button>
          </form>
        </div>
      </div>

      <div class="card">
        <div class="card-title">Lista de presenças
          {'<span class="badge badge-info" style="margin-left:.5rem;font-weight:400;font-size:.7rem">Usa o módulo <a href="' + url_for("controlo_presencas", d=d_str) + '">Controlo de Presenças</a> para marcar entradas/saídas</span>' if perfil in ("oficialdia", "admin") else ""}
        </div>
        <div class="table-wrap">
          <table>
            <thead>
              <tr><th>NI</th><th>Nome</th><th>PA</th><th>Lanche</th><th>Almoço</th><th>Jantar</th><th>Presença / Exc.</th></tr>
            </thead>
            <tbody>{rows_html or '<tr><td colspan="7" class="text-muted center" style="padding:1.5rem">Sem dados.</td></tr>'}</tbody>
          </table>
        </div>
        <div style="margin-top:.7rem;font-size:.78rem;color:var(--muted);display:flex;gap:.8rem;flex-wrap:wrap">
          <span style="display:inline-flex;align-items:center;gap:.3rem"><span style="width:.75rem;height:.75rem;background:#d5f5e3;border:1px solid #a9dfbf;border-radius:3px;display:inline-block"></span>Presente com refeições</span>
          <span style="display:inline-flex;align-items:center;gap:.3rem"><span style="width:.75rem;height:.75rem;background:#fff3cd;border:1px solid #ffc107;border-radius:3px;display:inline-block"></span>Sem refeições marcadas</span>
          <span style="display:inline-flex;align-items:center;gap:.3rem"><span style="width:.75rem;height:.75rem;background:#fdecea;border:1px solid #f1948a;border-radius:3px;display:inline-block"></span>Ausente</span>
        </div>
      </div>

      <div class="card">
        <div class="card-title">📊 Totais do {ano}º Ano</div>
        <div class="grid grid-4">
          <div class="stat-box"><div class="stat-num">{t["pa"]}</div><div class="stat-lbl">Pequenos Almoços</div></div>
          <div class="stat-box"><div class="stat-num">{t["lan"]}</div><div class="stat-lbl">Lanches</div></div>
          <div class="stat-box"><div class="stat-num">{t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]}</div><div class="stat-lbl">Almoços</div></div>
          <div class="stat-box"><div class="stat-num">{t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]}</div><div class="stat-lbl">Jantares</div></div>
        </div>
        <div class="gap-btn" style="margin-top:.8rem">
          <a class="btn btn-primary" href="{url_for("exportar_dia", d=d_str, fmt="csv")}">⬇ CSV</a>
          <a class="btn btn-primary" href="{url_for("exportar_dia", d=d_str, fmt="xlsx")}">⬇ Excel</a>
          <a class="btn btn-ghost" href="{url_for("imprimir_ano", ano=ano, d=d_str)}" target="_blank">🖨 Imprimir</a>
        </div>
      </div>
    </div>"""
    return render(content)


# ═══════════════════════════════════════════════════════════════════════════
# RELATÓRIO SEMANAL (cozinha + oficialdia + admin)
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/relatorio")
@role_required("cozinha", "oficialdia", "admin")
def relatorio_semanal():
    u = current_user()
    perfil = u.get("perfil")
    hoje = date.today()
    segunda = hoje - timedelta(days=hoje.weekday())
    d0_str = request.args.get("d0", segunda.isoformat())
    d0 = _parse_date(d0_str)
    d1 = d0 + timedelta(days=6)
    ICONE = {
        "normal": "",
        "fim_semana": "🌊",
        "feriado": "🔴",
        "exercicio": "🟡",
        "outro": "⚪",
    }

    # Batch: totais e calendário para a semana toda
    _rel_map, _rel_empty = sr.get_totais_periodo(d0.isoformat(), d1.isoformat())
    _rel_cal = sr.dias_operacionais_batch(d0, d1)
    res = []
    for i in range(7):
        di = d0 + timedelta(days=i)
        t = _rel_map.get(di.isoformat(), _rel_empty)
        tipo = _rel_cal.get(
            di.isoformat(), "fim_semana" if di.weekday() >= 5 else "normal"
        )
        res.append({"data": di, "t": t, "tipo": tipo})

    totais = {
        k: 0
        for k in [
            "pa",
            "lan",
            "alm_norm",
            "alm_veg",
            "alm_dieta",
            "jan_norm",
            "jan_veg",
            "jan_dieta",
            "jan_sai",
        ]
    }
    rows_html = ""
    for r in res:
        is_off = r["tipo"] in ("feriado", "exercicio")
        is_wknd = r["data"].weekday() >= 5
        st = (
            "color:var(--muted);background:#f9fafb"
            if is_off
            else ("background:#fffdf5" if is_wknd else "")
        )
        ic = ICONE.get(r["tipo"], "")
        t = r["t"]
        sai_td = (
            "" if perfil == "cozinha" else f'<td class="center">{t["jan_sai"]}</td>'
        )
        rows_html += f"""
        <tr style="{st}">
          <td><strong>{ABREV_DIAS[r["data"].weekday()]}</strong> {r["data"].strftime("%d/%m")} {ic}</td>
          <td class="center">{t["pa"]}</td><td class="center">{t["lan"]}</td>
          <td class="center">{t["alm_norm"]}</td><td class="center">{t["alm_veg"]}</td><td class="center">{t["alm_dieta"]}</td>
          <td class="center">{t["jan_norm"]}</td><td class="center">{t["jan_veg"]}</td><td class="center">{t["jan_dieta"]}</td>
          {sai_td}
        </tr>"""
        for k in totais:
            totais[k] += t[k]

    sai_th = "" if perfil == "cozinha" else "<th>Sai</th>"
    sai_total = (
        "" if perfil == "cozinha" else f'<td class="center">{totais["jan_sai"]}</td>'
    )
    rows_html += f"""
    <tr style="font-weight:800;background:#f0f4f8;border-top:2px solid var(--border)">
      <td>TOTAL</td>
      <td class="center">{totais["pa"]}</td><td class="center">{totais["lan"]}</td>
      <td class="center">{totais["alm_norm"]}</td><td class="center">{totais["alm_veg"]}</td><td class="center">{totais["alm_dieta"]}</td>
      <td class="center">{totais["jan_norm"]}</td><td class="center">{totais["jan_veg"]}</td><td class="center">{totais["jan_dieta"]}</td>
      {sai_total}
    </tr>"""

    prev_w = (d0 - timedelta(days=7)).isoformat()
    next_w = (d0 + timedelta(days=7)).isoformat()
    back_url = url_for("admin_home") if perfil == "admin" else url_for("painel_dia")

    content = f"""
    <div class="container">
      <div class="page-header">
        {_back_btn(back_url)}
        <div class="page-title">📊 Relatório Semanal</div>
      </div>
      <div class="card" style="padding:.9rem 1.2rem;margin-bottom:.8rem">
        <div class="flex-between">
          <div class="flex">
            <a class="btn btn-ghost btn-sm" href="{url_for("relatorio_semanal", d0=prev_w)}">← Semana anterior</a>
            <strong>{d0.strftime("%d/%m/%Y")} — {d1.strftime("%d/%m/%Y")}</strong>
            <a class="btn btn-ghost btn-sm" href="{url_for("relatorio_semanal", d0=next_w)}">Semana seguinte →</a>
          </div>
          <form method="get" style="display:flex;gap:.3rem">
            <input type="date" name="d0" value="{d0_str}" style="width:auto">
            <button class="btn btn-primary btn-sm">Ir</button>
          </form>
        </div>
      </div>
      <div class="card">
        <div class="table-wrap">
          <table>
            <thead><tr><th>Dia</th><th>PA</th><th>Lanche</th><th>Alm N</th><th>Alm V</th><th>Alm D</th><th>Jan N</th><th>Jan V</th><th>Jan D</th>{sai_th}</tr></thead>
            <tbody>{rows_html}</tbody>
          </table>
        </div>
        <div class="gap-btn" style="margin-top:.9rem">
          <a class="btn btn-primary" href="{url_for("exportar_relatorio", d0=d0_str, fmt="csv")}">⬇ CSV</a>
          <a class="btn btn-primary" href="{url_for("exportar_relatorio", d0=d0_str, fmt="xlsx")}">⬇ Excel</a>
        </div>
      </div>
      <div class="grid grid-4">
        <div class="stat-box"><div class="stat-num">{totais["pa"]}</div><div class="stat-lbl">Total PA</div></div>
        <div class="stat-box"><div class="stat-num">{totais["lan"]}</div><div class="stat-lbl">Total Lanches</div></div>
        <div class="stat-box"><div class="stat-num">{totais["alm_norm"] + totais["alm_veg"] + totais["alm_dieta"]}</div><div class="stat-lbl">Total Almoços</div></div>
        <div class="stat-box"><div class="stat-num">{totais["jan_norm"] + totais["jan_veg"] + totais["jan_dieta"]}</div><div class="stat-lbl">Total Jantares</div></div>
      </div>
    </div>"""
    return render(content)


# ═══════════════════════════════════════════════════════════════════════════
# EXCEÇÕES
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/excecoes/<d>", methods=["GET", "POST"])
@role_required("oficialdia", "admin")
def excecoes_dia(d):
    u = current_user()
    dt = _parse_date(d)

    if request.method == "POST":
        nii = request.form.get("nii", "").strip()
        db_u = sr.user_by_nii(nii)
        if not db_u:
            flash("Utilizador não encontrado.", "error")
            return redirect(url_for("excecoes_dia", d=dt.isoformat()))
        pa = 1 if request.form.get("pa") else 0
        lanche = 1 if request.form.get("lanche") else 0
        alm = _val_refeicao(request.form.get("almoco"))
        jan = _val_refeicao(request.form.get("jantar"))
        sai = 1 if request.form.get("sai") else 0
        if _refeicao_set(
            db_u["id"], dt, pa, lanche, alm, jan, sai, alterado_por=u["nii"]
        ):
            flash(f"Exceção guardada para {db_u['Nome_completo']}.", "ok")
        else:
            flash("Erro ao guardar.", "error")
        return redirect(
            url_for("excecoes_dia", d=dt.isoformat(), nii=request.form.get("nii", ""))
        )

    nii_q = request.args.get("nii", "").strip()
    u_info = sr.user_by_nii(nii_q) if nii_q else None
    r = sr.refeicao_get(u_info["id"], dt) if u_info and u_info.get("id") else {}

    def tipos_opt(sel):
        return "".join(
            f'<option value="{t}" {"selected" if sel == t else ""}>{t}</option>'
            for t in ["Normal", "Vegetariano", "Dieta"]
        )

    def chk_label(name, checked, icon, label):
        s = "background:#eafaf1;border-color:#a9dfbf" if checked else ""
        return f'<label style="display:flex;align-items:center;gap:.6rem;cursor:pointer;padding:.6rem;border:1.5px solid var(--border);border-radius:9px;{s}"><input type="checkbox" name="{name}" {"checked" if checked else ""}> {icon} {label}</label>'

    form_html = ""
    if u_info:
        uid_info = u_info.get("id")
        # Ausência ativa
        ausente_hoje = uid_info and _tem_ausencia_ativa(uid_info, dt)
        # Prazo — pode o aluno ainda alterar por si?
        ok_prazo, _ = sr.refeicao_editavel(dt)
        # Histórico recente de ausências
        aus_hist = []
        if uid_info:
            with sr.db() as conn:
                aus_hist = [
                    dict(r)
                    for r in conn.execute(
                        """
                    SELECT ausente_de, ausente_ate, motivo FROM ausencias
                    WHERE utilizador_id=? ORDER BY ausente_de DESC LIMIT 5
                """,
                        (uid_info,),
                    ).fetchall()
                ]

        ausente_alert = ""
        if ausente_hoje:
            ausente_alert = '<div class="alert alert-warn">⚠️ <strong>Utilizador com ausência activa hoje</strong> — esta exceção pode não ter efeito prático.</div>'

        prazo_info = ""
        if ok_prazo:
            prazo_info = '<div class="alert alert-ok" style="margin-bottom:.6rem">✅ O aluno ainda pode alterar refeições por si próprio (prazo não expirou). Esta exceção só é necessária se o aluno não conseguir aceder ao sistema.</div>'
        else:
            prazo_info = '<div class="alert alert-info" style="margin-bottom:.6rem">🔒 Prazo expirado — o aluno já não pode alterar. Esta exceção é necessária para efetuar qualquer alteração.</div>'

        aus_hist_html = ""
        if aus_hist:
            aus_hist_html = '<div style="margin-top:.75rem"><div class="card-title" style="font-size:.8rem;margin-bottom:.4rem">📋 Ausências recentes</div>'
            for ah in aus_hist:
                aus_hist_html += f'<div style="font-size:.78rem;padding:.22rem 0;border-bottom:1px solid var(--border);color:var(--text)">{ah["ausente_de"]} → {ah["ausente_ate"]} <span class="text-muted">{esc(ah["motivo"] or "—")}</span></div>'
            aus_hist_html += "</div>"

        form_html = f"""
        <div class="card">
          <div class="card-title">✏️ {esc(u_info.get("Nome_completo", ""))} — NI {esc(u_info.get("NI", ""))} | {esc(u_info.get("ano", ""))}º Ano</div>
          {ausente_alert}{prazo_info}
          <form method="post">
            {csrf_input()}
            <input type="hidden" name="nii" maxlength="32" value="{esc(nii_q)}">
            <div class="grid grid-2">
              {chk_label("pa", r.get("pequeno_almoco"), "☕", "Pequeno Almoço")}
              {chk_label("lanche", r.get("lanche"), "🥐", "Lanche")}
              <div class="form-group" style="margin:0">
                <label>🍽️ Almoço</label>
                <select name="almoco"><option value="">— Sem almoço —</option>{tipos_opt(r.get("almoco"))}</select>
              </div>
              <div class="form-group" style="margin:0">
                <label>🌙 Jantar</label>
                <select name="jantar"><option value="">— Sem jantar —</option>{tipos_opt(r.get("jantar_tipo"))}</select>
              </div>
            </div>
            <div style="margin-top:.8rem">
              {chk_label("sai", r.get("jantar_sai_unidade"), "🚪", "Sai da unidade após jantar")}
            </div>
            <hr>
            <button class="btn btn-ok">💾 Guardar exceção</button>
          </form>
          {aus_hist_html}
        </div>"""

    content = f"""
    <div class="container">
      <div class="page-header">
        {_back_btn(url_for("painel_dia", d=dt.isoformat()), "Painel")}
        <div class="page-title">📝 Exceções — {NOMES_DIAS[dt.weekday()]}, {dt.strftime("%d/%m/%Y")}</div>
      </div>
      <div class="card">
        <div class="card-title">Pesquisar utilizador</div>
        <form method="get" style="display:flex;gap:.5rem">
          <input type="hidden" name="d" value="{d}">
          <input type="text" name="nii" maxlength="32" placeholder="NII do utilizador" value="{esc(nii_q)}" style="flex:1">
          <button class="btn btn-primary">Pesquisar</button>
        </form>
      </div>
      {form_html or '<div class="card"><div class="text-muted">Introduz um NII para editar exceções.</div></div>'}
    </div>"""
    return render(content)


# ═══════════════════════════════════════════════════════════════════════════
# AUSÊNCIAS
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/ausencias", methods=["GET", "POST"])
@role_required("oficialdia", "admin")
def ausencias():
    u = current_user()
    if request.method == "POST":
        acao = request.form.get("acao", "")
        if acao == "remover":
            _remover_ausencia(request.form.get("id"))
            flash("Ausência removida.", "ok")
            return redirect(url_for("ausencias"))
        nii = request.form.get("nii", "").strip()
        db_u = sr.user_by_nii(nii)
        if not db_u:
            flash("Utilizador não encontrado.", "error")
        else:
            ok, err = _registar_ausencia(
                db_u["id"],
                request.form.get("de", ""),
                request.form.get("ate", ""),
                request.form.get("motivo", "")[:500],
                u["nii"],
            )
            flash(
                f"Ausência registada para {db_u['Nome_completo']}."
                if ok
                else (err or "Falha."),
                "ok" if ok else "error",
            )
        return redirect(url_for("ausencias"))

    with sr.db() as conn:
        rows = [
            dict(r)
            for r in conn.execute("""
            SELECT a.id, u.NII, u.Nome_completo, u.NI, u.ano,
                   a.ausente_de, a.ausente_ate, a.motivo
            FROM ausencias a JOIN utilizadores u ON u.id=a.utilizador_id
            ORDER BY a.ausente_de DESC""").fetchall()
        ]

    hoje = date.today().isoformat()
    rows_html = "".join(
        f"""
      <tr>
        <td><strong>{esc(r["Nome_completo"])}</strong><br><span class="text-muted small">{esc(r["NII"])} · {r["ano"]}º ano</span></td>
        <td>{r["ausente_de"]}</td><td>{r["ausente_ate"]}</td>
        <td>{esc(r["motivo"] or "—")}</td>
        <td>{'<span class="badge badge-warn">Atual</span>' if r["ausente_de"] <= hoje <= r["ausente_ate"] else '<span class="badge badge-muted">Inativa</span>'}</td>
        <td><form method="post" style="display:inline">{csrf_input()}<input type="hidden" name="acao" value="remover"><input type="hidden" name="id" value="{r["id"]}"><button class="btn btn-danger btn-sm">🗑</button></form></td>
      </tr>"""
        for r in rows
    )

    content = f"""
    <div class="container">
      <div class="page-header">{_back_btn(url_for("painel_dia"))}<div class="page-title">🚫 Ausências</div></div>
      <div class="card">
        <div class="card-title">Registar ausência</div>
        <form method="post">
          {csrf_input()}
          <div class="grid grid-2">
            <div class="form-group"><label>NII do utilizador</label><input type="text" name="nii" maxlength="32" required placeholder="NII"></div>
            <div class="form-group"><label>Motivo (opcional)</label><input type="text" name="motivo" maxlength="500" placeholder="Ex: deslocação, prova..."></div>
            <div class="form-group"><label>De</label><input type="date" name="de" required></div>
            <div class="form-group"><label>Até</label><input type="date" name="ate" required></div>
          </div>
          <button class="btn btn-ok">Registar</button>
        </form>
      </div>
      <div class="card">
        <div class="card-title">Lista de ausências</div>
        <div class="table-wrap">
          <table>
            <thead><tr><th>Utilizador</th><th>De</th><th>Até</th><th>Motivo</th><th>Estado</th><th></th></tr></thead>
            <tbody>{rows_html or '<tr><td colspan="6" class="text-muted center" style="padding:1.5rem">Sem ausências.</td></tr>'}</tbody>
          </table>
        </div>
      </div>
    </div>"""
    return render(content)


# ═══════════════════════════════════════════════════════════════════════════
# CMD — Editar dados de aluno do seu ano
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/cmd/editar-aluno/<nii>", methods=["GET", "POST"])
@role_required("cmd", "oficialdia", "admin")
def cmd_editar_aluno(nii):
    u = current_user()
    perfil = u.get("perfil")
    ano_cmd = int(u.get("ano", 0)) if u.get("ano") else 0
    ano_ret = request.args.get("ano", str(ano_cmd) if ano_cmd else "1")
    d_ret = request.args.get("d", date.today().isoformat())

    # Buscar o aluno
    with sr.db() as conn:
        aluno = dict(
            conn.execute(
                "SELECT id,NII,NI,Nome_completo,ano,email,telemovel FROM utilizadores WHERE NII=?",
                (nii,),
            ).fetchone()
            or {}
        )

    if not aluno:
        flash("Aluno não encontrado.", "error")
        back_ano = aluno.get("ano", ano_cmd or 1) if aluno else (ano_cmd or 1)
        return redirect(url_for("lista_alunos_ano", ano=back_ano, d=d_ret))

    # CMD só pode editar alunos do seu ano
    if perfil == "cmd" and int(aluno.get("ano", 0)) != ano_cmd:
        flash("Só podes editar alunos do teu ano.", "error")
        return redirect(url_for("lista_alunos_ano", ano=ano_cmd, d=d_ret))

    if request.method == "POST":
        nome_n = _val_nome(request.form.get("nome", ""))
        ni_n = _val_ni(request.form.get("ni", ""))
        email_n = _val_email(request.form.get("email", ""))
        telef_n = _val_phone(request.form.get("telemovel", ""))
        if not nome_n:
            flash("O nome não pode estar vazio.", "error")
        elif ni_n is None:
            flash("NI inválido (alfanumérico, máx. 20 caracteres).", "error")
        elif email_n is False:
            flash("Email inválido.", "error")
        elif telef_n is False:
            flash("Telemóvel inválido.", "error")
        else:
            try:
                with sr.db() as conn:
                    conn.execute(
                        "UPDATE utilizadores SET Nome_completo=?,NI=?,email=?,telemovel=? WHERE NII=?",
                        (nome_n, ni_n or None, email_n, telef_n, nii),
                    )
                    conn.commit()
                flash(f"Dados de {nome_n} actualizados.", "ok")
                return redirect(
                    url_for(
                        "lista_alunos_ano", ano=ano_ret or aluno.get("ano", 1), d=d_ret
                    )
                )
            except Exception as ex:
                flash(f"Erro: {ex}", "error")

    back_url = url_for("lista_alunos_ano", ano=ano_ret, d=d_ret)
    content = f"""
    <div class="container">
      <div class="page-header">
        {_back_btn(back_url, f"{ano_ret}º Ano")}
        <div class="page-title">👤 Editar aluno — {esc(aluno.get("Nome_completo", ""))}</div>
      </div>
      <div class="card" style="max-width:560px">
        <div class="card-title">ℹ️ Dados do aluno
          <span class="badge badge-info" style="margin-left:.4rem">{aluno["ano"]}º Ano</span>
        </div>
        <form method="post">
          {csrf_input()}
          <div class="grid grid-2">
            <div class="form-group">
              <label>Nome completo</label>
              <input type="text" name="nome" value="{esc(aluno.get("Nome_completo", ""))}" required>
            </div>
            <div class="form-group">
              <label>NI <span class="text-muted small">(número interno)</span></label>
              <input type="text" name="ni" value="{esc(aluno.get("NI") or "")}">
            </div>
            <div class="form-group">
              <label>📧 Email</label>
              <input type="email" name="email" value="{esc(aluno.get("email") or "")}" placeholder="email@exemplo.pt">
            </div>
            <div class="form-group">
              <label>📱 Telemóvel</label>
              <input type="tel" name="telemovel" value="{esc(aluno.get("telemovel") or "")}" placeholder="+351XXXXXXXXX">
            </div>
          </div>
          <div class="alert alert-info" style="font-size:.81rem;margin-bottom:.8rem">
            📌 NII: <strong>{esc(aluno["NII"])}</strong> — Este campo não pode ser alterado aqui.
            Para alterar o NII contacta o administrador.
          </div>
          <div class="gap-btn">
            <button class="btn btn-ok">💾 Guardar alterações</button>
            <a class="btn btn-ghost" href="{back_url}">Cancelar</a>
          </div>
        </form>
        <hr style="margin:1rem 0">
        <form method="post" action="{url_for("cmd_reset_password", nii=nii)}"
              onsubmit="return confirm('Tens a certeza que queres resetar a password de {esc(aluno.get("Nome_completo", ""))}?')">
          {csrf_input()}
          <input type="hidden" name="ano" value="{ano_ret}">
          <input type="hidden" name="d" value="{d_ret}">
          <button class="btn btn-danger btn-sm">🔑 Resetar password</button>
          <span class="text-muted small" style="margin-left:.5rem">Gera uma password temporária (o aluno terá de mudar no próximo login)</span>
        </form>
      </div>
    </div>"""
    return render(content)


@app.route("/cmd/reset-password/<nii>", methods=["POST"])
@role_required("cmd", "oficialdia", "admin")
def cmd_reset_password(nii):
    u = current_user()
    perfil = u.get("perfil")
    ano_cmd = int(u.get("ano", 0)) if u.get("ano") else 0
    ano_ret = request.form.get("ano", str(ano_cmd) if ano_cmd else "1")
    d_ret = request.form.get("d", date.today().isoformat())

    with sr.db() as conn:
        aluno = conn.execute(
            "SELECT NII, Nome_completo, ano, perfil FROM utilizadores WHERE NII=?",
            (nii,),
        ).fetchone()

    if not aluno:
        flash("Aluno não encontrado.", "error")
        return redirect(url_for("lista_alunos_ano", ano=ano_cmd or 1, d=d_ret))

    aluno = dict(aluno)

    # Só pode resetar alunos (não admins/cmd/cozinha/oficialdia)
    if aluno.get("perfil") != "aluno":
        flash("Só é possível resetar passwords de alunos.", "error")
        return redirect(url_for("lista_alunos_ano", ano=ano_ret, d=d_ret))

    # CMD só pode resetar alunos do seu ano
    if perfil == "cmd" and int(aluno.get("ano", 0)) != ano_cmd:
        flash("Só podes resetar passwords de alunos do teu ano.", "error")
        return redirect(url_for("lista_alunos_ano", ano=ano_cmd, d=d_ret))

    ok, result = _reset_pw(nii)
    if ok:
        _audit(
            u["nii"],
            "cmd_reset_password",
            f"NII={nii} por {u['nome']} ({perfil})",
        )
        flash(
            f"Password de {aluno['Nome_completo']} resetada. Temporária: {result}",
            "ok",
        )
    else:
        flash(f"Erro: {result}", "error")

    return redirect(url_for("cmd_editar_aluno", nii=nii, ano=ano_ret, d=d_ret))


# ═══════════════════════════════════════════════════════════════════════════
# VER PERFIL DE ALUNO — Oficial de Dia (apenas leitura)
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/alunos/perfil/<nii>")
@role_required("oficialdia", "admin", "cmd")
def ver_perfil_aluno(nii):
    u = current_user()
    perfil = u.get("perfil")
    ano_ret = request.args.get("ano", "")
    d_ret = request.args.get("d", date.today().isoformat())

    with sr.db() as conn:
        aluno = conn.execute(
            "SELECT id,NII,NI,Nome_completo,ano,email,telemovel FROM utilizadores WHERE NII=?",
            (nii,),
        ).fetchone()

    if not aluno:
        flash("Aluno não encontrado.", "error")
        return redirect(url_for("painel_dia"))
    aluno = dict(aluno)

    # CMD só pode ver alunos do seu ano
    if perfil == "cmd" and str(aluno.get("ano", 0)) != str(u.get("ano", "")):
        flash("Acesso restrito ao teu ano.", "error")
        return redirect(url_for("painel_dia"))

    # Admin é redirecionado para edição
    if perfil == "admin":
        return redirect(
            url_for("cmd_editar_aluno", nii=nii, ano=ano_ret or aluno["ano"], d=d_ret)
        )

    hoje = date.today()
    uid = aluno["id"]
    with sr.db() as conn:
        total_ref = conn.execute(
            "SELECT COUNT(*) c FROM refeicoes WHERE utilizador_id=?", (uid,)
        ).fetchone()["c"]
        ausencias_ativas = conn.execute(
            """SELECT COUNT(*) c FROM ausencias WHERE utilizador_id=?
               AND ausente_de<=? AND ausente_ate>=?""",
            (uid, hoje.isoformat(), hoje.isoformat()),
        ).fetchone()["c"]
        aus_recentes = [
            dict(r)
            for r in conn.execute(
                """SELECT ausente_de, ausente_ate, motivo FROM ausencias
               WHERE utilizador_id=? ORDER BY ausente_de DESC LIMIT 5""",
                (uid,),
            ).fetchall()
        ]
        # Refeições de hoje
        ref_hoje = conn.execute(
            "SELECT * FROM refeicoes WHERE utilizador_id=? AND data=?",
            (uid, hoje.isoformat()),
        ).fetchone()

    ref_hoje = dict(ref_hoje) if ref_hoje else {}

    def yn(v, t=None):
        return (
            f'<span class="badge badge-ok">{t or "✅"}</span>'
            if v
            else '<span class="badge badge-muted">—</span>'
        )

    aus_html = ""
    for a in aus_recentes:
        aus_html += f'<div style="font-size:.82rem;padding:.25rem 0;border-bottom:1px solid var(--border)">{a["ausente_de"]} → {a["ausente_ate"]} <span class="text-muted small">{esc(a["motivo"] or "—")}</span></div>'

    back_url = url_for("lista_alunos_ano", ano=ano_ret or aluno["ano"], d=d_ret)
    content = f"""
    <div class="container">
      <div class="page-header">
        {_back_btn(back_url, f"{ano_ret or aluno['ano']}º Ano")}
        <div class="page-title">👁 Perfil — {esc(aluno.get("Nome_completo", ""))}</div>
        <span class="badge badge-info">Só leitura</span>
      </div>
      <div class="grid grid-2">
        <div class="card">
          <div class="card-title">ℹ️ Informação pessoal</div>
          <div style="display:flex;flex-direction:column;gap:.7rem;font-size:.9rem">
            <div><span class="text-muted">Nome completo:</span><br><strong>{esc(aluno["Nome_completo"])}</strong></div>
            <div><span class="text-muted">NII:</span><br><strong>{esc(aluno["NII"])}</strong></div>
            <div><span class="text-muted">NI:</span><br><strong>{esc(aluno.get("NI") or "—")}</strong></div>
            <div><span class="text-muted">Ano:</span><br><strong>{aluno["ano"]}º Ano</strong></div>
            <div><span class="text-muted">📧 Email:</span><br><strong>{esc(aluno.get("email") or "—")}</strong></div>
            <div><span class="text-muted">📱 Telemóvel:</span><br><strong>{esc(aluno.get("telemovel") or "—")}</strong></div>
          </div>
          <hr style="margin:1rem 0">
          <div class="grid grid-2">
            <div class="stat-box"><div class="stat-num">{total_ref}</div><div class="stat-lbl">Refeições registadas</div></div>
            <div class="stat-box"><div class="stat-num" style="color:{"var(--warn)" if ausencias_ativas else "var(--ok)"}">{ausencias_ativas}</div><div class="stat-lbl">Ausências ativas</div></div>
          </div>
        </div>
        <div class="card">
          <div class="card-title">🍽️ Refeições de hoje — {hoje.strftime("%d/%m/%Y")}</div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:.6rem;margin-bottom:.8rem">
            <div style="padding:.6rem;background:#f8f9fa;border-radius:8px;text-align:center">
              <div class="text-muted small">☕ Pequeno Almoço</div>
              <div style="margin-top:.3rem">{yn(ref_hoje.get("pequeno_almoco"))}</div>
            </div>
            <div style="padding:.6rem;background:#f8f9fa;border-radius:8px;text-align:center">
              <div class="text-muted small">🥐 Lanche</div>
              <div style="margin-top:.3rem">{yn(ref_hoje.get("lanche"))}</div>
            </div>
            <div style="padding:.6rem;background:#f8f9fa;border-radius:8px;text-align:center">
              <div class="text-muted small">🍽️ Almoço</div>
              <div style="margin-top:.3rem"><strong>{ref_hoje.get("almoco") or "—"}</strong></div>
            </div>
            <div style="padding:.6rem;background:#f8f9fa;border-radius:8px;text-align:center">
              <div class="text-muted small">🌙 Jantar</div>
              <div style="margin-top:.3rem"><strong>{ref_hoje.get("jantar_tipo") or "—"}</strong></div>
            </div>
          </div>
          {'<div class="alert alert-warn" style="font-size:.82rem">⚠️ Aluno com ausência ativa hoje</div>' if ausencias_ativas else ""}
          <div class="card-title" style="margin-top:.8rem">📋 Ausências recentes</div>
          {aus_html or '<div class="text-muted small">Sem ausências registadas.</div>'}
        </div>
      </div>
      <div class="alert alert-info" style="font-size:.82rem">
        🔒 Estás no modo de visualização. Para editar dados do aluno, contacta o Comandante de Companhia ou o Administrador.
      </div>
    </div>"""
    return render(content)


# ═══════════════════════════════════════════════════════════════════════════
# AUSÊNCIAS — CMD (acesso restrito ao seu ano)
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/cmd/ausencias", methods=["GET", "POST"])
@role_required("cmd", "admin")
def ausencias_cmd():
    u = current_user()
    perfil = u.get("perfil")
    ano_cmd = int(u.get("ano", 0)) if perfil == "cmd" else 0

    if request.method == "POST":
        acao = request.form.get("acao", "")
        if acao == "remover":
            aid = _val_int_id(request.form.get("id"))
            if aid is None:
                flash("ID inválido.", "error")
                return redirect(url_for("ausencias_cmd"))
            # Validar que a ausência pertence ao ano do cmd
            with sr.db() as conn:
                aus = conn.execute(
                    """SELECT a.id FROM ausencias a
                    JOIN utilizadores u ON u.id=a.utilizador_id
                    WHERE a.id=? AND (u.ano=? OR ?=0)""",
                    (aid, ano_cmd, perfil == "admin"),
                ).fetchone()
            if aus:
                _remover_ausencia(aid)
                flash("Ausência removida.", "ok")
            else:
                flash("Não autorizado.", "error")
            return redirect(url_for("ausencias_cmd"))
        nii = request.form.get("nii", "").strip()
        db_u = sr.user_by_nii(nii)
        if not db_u:
            flash("Utilizador não encontrado.", "error")
        elif perfil == "cmd" and int(db_u.get("ano", 0)) != ano_cmd:
            flash(
                f"Só podes registar ausências para alunos do {ano_cmd}º ano.", "error"
            )
        else:
            ok, err = _registar_ausencia(
                db_u["id"],
                request.form.get("de", ""),
                request.form.get("ate", ""),
                _val_text(request.form.get("motivo", ""))[:500],
                u["nii"],
            )
            flash(
                f"Ausência registada para {db_u['Nome_completo']}."
                if ok
                else (err or "Falha."),
                "ok" if ok else "error",
            )
        return redirect(url_for("ausencias_cmd"))

    filtro_ano = f"AND u.ano={ano_cmd}" if perfil == "cmd" else ""
    with sr.db() as conn:
        rows = [
            dict(r)
            for r in conn.execute(f"""
            SELECT a.id, u.NII, u.Nome_completo, u.NI, u.ano,
                   a.ausente_de, a.ausente_ate, a.motivo
            FROM ausencias a JOIN utilizadores u ON u.id=a.utilizador_id
            WHERE u.perfil='aluno' {filtro_ano}
            ORDER BY a.ausente_de DESC""").fetchall()
        ]

    # Alunos do ano para pesquisa rápida
    with sr.db() as conn:
        alunos_ano = (
            [
                dict(r)
                for r in conn.execute(
                    "SELECT NII, NI, Nome_completo FROM utilizadores WHERE perfil='aluno' AND ano=? ORDER BY NI",
                    (ano_cmd,),
                ).fetchall()
            ]
            if perfil == "cmd"
            else []
        )

    hoje = date.today().isoformat()
    rows_html = "".join(
        f"""
      <tr>
        <td><strong>{esc(r["Nome_completo"])}</strong><br><span class="text-muted small">{esc(r["NII"])} · {r["ano"]}º ano</span></td>
        <td>{r["ausente_de"]}</td><td>{r["ausente_ate"]}</td>
        <td>{esc(r["motivo"] or "—")}</td>
        <td>{'<span class="badge badge-warn">Atual</span>' if r["ausente_de"] <= hoje <= r["ausente_ate"] else '<span class="badge badge-muted">Inativa</span>'}</td>
        <td><form method="post" style="display:inline">{csrf_input()}<input type="hidden" name="acao" value="remover"><input type="hidden" name="id" value="{r["id"]}"><button class="btn btn-danger btn-sm">🗑</button></form></td>
      </tr>"""
        for r in rows
    )

    alunos_options = "".join(
        f'<option value="{esc(a["NII"])}">{esc(a["NI"])} — {esc(a["Nome_completo"])}</option>'
        for a in alunos_ano
    )
    alunos_datalist = (
        f'<datalist id="alunos_list">{alunos_options}</datalist>' if alunos_ano else ""
    )

    titulo = (
        f"🚫 Ausências — {ano_cmd}º Ano"
        if perfil == "cmd"
        else "🚫 Ausências (todos os anos)"
    )
    back_url = url_for("painel_dia") if perfil == "cmd" else url_for("ausencias")

    content = f"""
    <div class="container">
      <div class="page-header">{_back_btn(back_url)}<div class="page-title">{titulo}</div></div>
      <div class="card">
        <div class="card-title">Registar ausência</div>
        {alunos_datalist}
        <form method="post">
          {csrf_input()}
          <div class="grid grid-2">
            <div class="form-group">
              <label>NII do aluno</label>
              <input type="text" name="nii" maxlength="32" required placeholder="NII" list="alunos_list">
              {'<div class="text-muted small" style="margin-top:.25rem">💡 Escreve para ver sugestões de alunos do teu ano</div>' if alunos_ano else ""}
            </div>
            <div class="form-group"><label>Motivo (opcional)</label><input type="text" name="motivo" maxlength="500" placeholder="Ex: deslocação, exercício..."></div>
            <div class="form-group"><label>De</label><input type="date" name="de" required value="{hoje}"></div>
            <div class="form-group"><label>Até</label><input type="date" name="ate" required value="{hoje}"></div>
          </div>
          <button class="btn btn-ok">✅ Registar ausência</button>
        </form>
      </div>
      <div class="card">
        <div class="card-title">Ausências registadas</div>
        <div class="table-wrap">
          <table>
            <thead><tr><th>Aluno</th><th>De</th><th>Até</th><th>Motivo</th><th>Estado</th><th>Ações</th></tr></thead>
            <tbody>{rows_html or '<tr><td colspan="6" class="text-muted center" style="padding:1.5rem">Sem ausências.</td></tr>'}</tbody>
          </table>
        </div>
      </div>
    </div>"""
    return render(content)


# ═══════════════════════════════════════════════════════════════════════════
# DETENÇÕES
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/cmd/detencoes", methods=["GET", "POST"])
@role_required("cmd", "admin")
def detencoes_cmd():
    u = current_user()
    perfil = u.get("perfil")
    ano_cmd = int(u.get("ano", 0)) if perfil == "cmd" else 0

    if request.method == "POST":
        acao = request.form.get("acao", "")

        if acao == "remover":
            did = _val_int_id(request.form.get("id", ""))
            if did is None:
                flash("ID inválido.", "error")
                return redirect(url_for("detencoes_cmd"))
            with sr.db() as conn:
                ok = conn.execute(
                    """SELECT d.id FROM detencoes d
                    JOIN utilizadores uu ON uu.id=d.utilizador_id
                    WHERE d.id=? AND (uu.ano=? OR ?=1)""",
                    (did, ano_cmd, 1 if perfil == "admin" else 0),
                ).fetchone()
                if ok:
                    conn.execute("DELETE FROM detencoes WHERE id=?", (did,))
                    conn.commit()
                    flash("Detenção removida.", "ok")
                else:
                    flash("Não autorizado.", "error")
            return redirect(url_for("detencoes_cmd"))

        # criar
        nii = request.form.get("nii", "").strip()
        de = request.form.get("de", "").strip()
        ate = request.form.get("ate", "").strip()
        motivo = _val_text(request.form.get("motivo", ""))[:500]

        db_u = sr.user_by_nii(nii)
        if not db_u:
            flash("Utilizador não encontrado.", "error")
            return redirect(url_for("detencoes_cmd"))

        if perfil == "cmd" and int(db_u.get("ano", 0)) != ano_cmd:
            flash(
                f"Só podes registar detenções para alunos do {ano_cmd}º ano.", "error"
            )
            return redirect(url_for("detencoes_cmd"))

        try:
            d1 = _parse_date(de)
            d2 = _parse_date(ate)
            if d2 < d1:
                flash(
                    "A data 'Até' tem de ser igual ou posterior à data 'De'.", "error"
                )
                return redirect(url_for("detencoes_cmd"))
        except Exception:
            flash("Datas inválidas.", "error")
            return redirect(url_for("detencoes_cmd"))

        with sr.db() as conn:
            conn.execute(
                """INSERT INTO detencoes(utilizador_id, detido_de, detido_ate, motivo, criado_por)
                            VALUES(?,?,?,?,?)""",
                (db_u["id"], d1.isoformat(), d2.isoformat(), motivo or None, u["nii"]),
            )
            conn.commit()

        # Auto-marcar todas as refeições para os dias de detenção (se não estiverem marcadas)
        _auto_marcar_refeicoes_detido(db_u["id"], d1, d2, u["nii"])

        # Cancelar licenças existentes durante o período de detenção
        with sr.db() as conn:
            conn.execute(
                "DELETE FROM licencas WHERE utilizador_id=? AND data>=? AND data<=?",
                (db_u["id"], d1.isoformat(), d2.isoformat()),
            )
            conn.commit()

        flash(
            f"Detenção registada para {db_u['Nome_completo']}. Refeições auto-marcadas.",
            "ok",
        )
        return redirect(url_for("detencoes_cmd"))

    filtro_ano = f"AND uu.ano={int(ano_cmd)}" if perfil == "cmd" else ""
    with sr.db() as conn:
        rows = [
            dict(r)
            for r in conn.execute(
                f"""
            SELECT d.id, uu.NII, uu.Nome_completo, uu.NI, uu.ano,
                   d.detido_de, d.detido_ate, d.motivo
            FROM detencoes d
            JOIN utilizadores uu ON uu.id=d.utilizador_id
            WHERE uu.perfil='aluno' {filtro_ano}
            ORDER BY d.detido_de DESC
        """
            ).fetchall()
        ]

    with sr.db() as conn:
        if perfil == "cmd":
            alunos_ano = [
                dict(r)
                for r in conn.execute(
                    "SELECT NII, NI, Nome_completo FROM utilizadores WHERE perfil='aluno' AND ano=? ORDER BY NI",
                    (ano_cmd,),
                ).fetchall()
            ]
        elif perfil == "admin":
            alunos_ano = [
                dict(r)
                for r in conn.execute(
                    "SELECT NII, NI, Nome_completo FROM utilizadores WHERE perfil='aluno' ORDER BY ano, NI"
                ).fetchall()
            ]
        else:
            alunos_ano = []

    hoje = date.today().isoformat()
    rows_html = "".join(
        f"""
      <tr>
        <td><strong>{esc(r["Nome_completo"])}</strong><br><span class="text-muted small">{esc(r["NII"])} · {r["ano"]}º ano</span></td>
        <td>{r["detido_de"]}</td><td>{r["detido_ate"]}</td>
        <td>{esc(r["motivo"] or "—")}</td>
        <td>{'<span class="badge badge-warn">Atual</span>' if r["detido_de"] <= hoje <= r["detido_ate"] else '<span class="badge badge-muted">Inativa</span>'}</td>
        <td><form method="post" style="display:inline">{csrf_input()}<input type="hidden" name="acao" value="remover"><input type="hidden" name="id" value="{r["id"]}"><button class="btn btn-danger btn-sm">🗑</button></form></td>
      </tr>"""
        for r in rows
    )

    alunos_options = "".join(
        f'<option value="{esc(a["NII"])}">{esc(a["NI"])} — {esc(a["Nome_completo"])}</option>'
        for a in alunos_ano
    )
    alunos_datalist = (
        f'<datalist id="alunos_list">{alunos_options}</datalist>' if alunos_ano else ""
    )

    titulo = (
        f"⛔ Detenções — {ano_cmd}º Ano"
        if perfil == "cmd"
        else "⛔ Detenções (todos os anos)"
    )
    back_url = url_for("painel_dia")

    content = f"""
    <div class="container">
      <div class="page-header">{_back_btn(back_url)}<div class="page-title">{titulo}</div></div>
      <div class="card">
        <div class="card-title">Registar detenção</div>
        {alunos_datalist}
        <form method="post">
          {csrf_input()}
          <div class="grid grid-2">
            <div class="form-group">
              <label>NII do aluno</label>
              <input type="text" name="nii" maxlength="32" required placeholder="NII" list="alunos_list">
            </div>
            <div class="form-group"><label>Motivo (opcional)</label><input type="text" name="motivo" maxlength="500" placeholder="Ex: detido por..."></div>
            <div class="form-group"><label>De</label><input type="date" name="de" required value="{hoje}"></div>
            <div class="form-group"><label>Até</label><input type="date" name="ate" required value="{hoje}"></div>
          </div>
          <button class="btn btn-ok">⛔ Registar detenção</button>
        </form>
      </div>
      <div class="card">
        <div class="card-title">Detenções registadas</div>
        <div class="table-wrap">
          <table>
            <thead><tr><th>Aluno</th><th>De</th><th>Até</th><th>Motivo</th><th>Estado</th><th>Ações</th></tr></thead>
            <tbody>{rows_html or '<tr><td colspan="6" class="text-muted center" style="padding:1.5rem">Sem detenções.</td></tr>'}</tbody>
          </table>
        </div>
      </div>
    </div>"""
    return render(content)


# ═══════════════════════════════════════════════════════════════════════════
# LICENÇAS — ENTRADAS / SAÍDAS (Oficial de Dia)
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/oficialdia/licencas-es", methods=["GET", "POST"])
@role_required("oficialdia", "admin")
def licencas_entradas_saidas():
    hoje = date.today()
    d_str = request.args.get("d", hoje.isoformat())
    dt = _parse_date(d_str)

    if request.method == "POST":
        acao = request.form.get("acao", "")
        lic_id = request.form.get("lic_id", "")
        agora = datetime.now().strftime("%H:%M")

        with sr.db() as conn:
            if acao == "saida" and lic_id:
                conn.execute(
                    "UPDATE licencas SET hora_saida=? WHERE id=? AND hora_saida IS NULL",
                    (agora, lic_id),
                )
                conn.commit()
                flash("✅ Saída registada.", "ok")

            elif acao == "entrada" and lic_id:
                conn.execute(
                    "UPDATE licencas SET hora_entrada=? WHERE id=? AND hora_entrada IS NULL",
                    (agora, lic_id),
                )
                conn.commit()
                flash("✅ Entrada registada.", "ok")

            elif acao == "limpar_saida" and lic_id:
                conn.execute(
                    "UPDATE licencas SET hora_saida=NULL WHERE id=?", (lic_id,)
                )
                conn.commit()

            elif acao == "limpar_entrada" and lic_id:
                conn.execute(
                    "UPDATE licencas SET hora_entrada=NULL WHERE id=?", (lic_id,)
                )
                conn.commit()

        return redirect(url_for("licencas_entradas_saidas", d=d_str))

    # ── Contadores ────────────────────────────────────────────────────────
    with sr.db() as conn:
        # Total de licenças marcadas para hoje
        total = conn.execute(
            """SELECT COUNT(*) c FROM licencas l
               JOIN utilizadores uu ON uu.id=l.utilizador_id
               WHERE l.data=? AND uu.perfil='aluno'""",
            (d_str,),
        ).fetchone()["c"]

        # Saíram hoje (têm hora_saida)
        saidas = conn.execute(
            """SELECT COUNT(*) c FROM licencas l
               JOIN utilizadores uu ON uu.id=l.utilizador_id
               WHERE l.data=? AND uu.perfil='aluno' AND l.hora_saida IS NOT NULL""",
            (d_str,),
        ).fetchone()["c"]

        # Regressaram (têm hora_entrada)
        entradas = conn.execute(
            """SELECT COUNT(*) c FROM licencas l
               JOIN utilizadores uu ON uu.id=l.utilizador_id
               WHERE l.data=? AND uu.perfil='aluno' AND l.hora_entrada IS NOT NULL""",
            (d_str,),
        ).fetchone()["c"]

        # Fora da unidade = saíram (em qualquer data) mas ainda não regressaram
        fora = conn.execute(
            """SELECT COUNT(*) c FROM licencas l
               JOIN utilizadores uu ON uu.id=l.utilizador_id
               WHERE uu.perfil='aluno'
                 AND l.hora_saida IS NOT NULL
                 AND l.hora_entrada IS NULL""",
        ).fetchone()["c"]

        # ── Lista principal: licenças do dia selecionado ──────────────────
        rows_hoje = [
            dict(r)
            for r in conn.execute(
                """SELECT l.id, uu.NI, uu.Nome_completo, uu.ano,
                          l.data, l.tipo, l.hora_saida, l.hora_entrada
                   FROM licencas l
                   JOIN utilizadores uu ON uu.id=l.utilizador_id
                   WHERE l.data=? AND uu.perfil='aluno'
                   ORDER BY uu.ano, uu.NI""",
                (d_str,),
            ).fetchall()
        ]

        # ── Alunos ainda fora de dias anteriores ─────────────────────────
        rows_fora = [
            dict(r)
            for r in conn.execute(
                """SELECT l.id, uu.NI, uu.Nome_completo, uu.ano,
                          l.data, l.tipo, l.hora_saida, l.hora_entrada
                   FROM licencas l
                   JOIN utilizadores uu ON uu.id=l.utilizador_id
                   WHERE uu.perfil='aluno'
                     AND l.hora_saida IS NOT NULL
                     AND l.hora_entrada IS NULL
                     AND l.data != ?
                   ORDER BY l.data ASC, uu.ano, uu.NI""",
                (d_str,),
            ).fetchall()
        ]

    prev_d = (dt - timedelta(days=1)).isoformat()
    next_d = (dt + timedelta(days=1)).isoformat()

    def _tipo_badge(tp):
        if tp == "antes_jantar":
            return '<span class="badge badge-info" style="font-size:.65rem">🌅 Antes jantar</span>'
        return '<span class="badge badge-muted" style="font-size:.65rem">🌙 Após jantar</span>'

    def _build_row(r, mostrar_data=False):
        saiu = r["hora_saida"]
        entrou = r["hora_entrada"]

        if saiu and not entrou:
            estado = '<span class="badge badge-warn">Fora</span>'
        elif saiu and entrou:
            estado = '<span class="badge badge-ok">Regressou</span>'
        else:
            estado = '<span class="badge badge-muted">Pendente</span>'

        # Saída: botão se ainda não saiu, hora se já saiu
        if not saiu:
            col_saida = (
                f'<form method="post" style="display:inline">{csrf_input()}'
                f'<input type="hidden" name="acao" value="saida">'
                f'<input type="hidden" name="lic_id" value="{r["id"]}">'
                f'<button class="btn btn-warn btn-sm">🚶 Registar saída</button></form>'
            )
        else:
            col_saida = f'<span class="text-muted small">{saiu}</span>'

        # Entrada: botão só se saiu mas ainda não entrou
        if saiu and not entrou:
            col_entrada = (
                f'<form method="post" style="display:inline">{csrf_input()}'
                f'<input type="hidden" name="acao" value="entrada">'
                f'<input type="hidden" name="lic_id" value="{r["id"]}">'
                f'<button class="btn btn-ok btn-sm">✅ Registar entrada</button></form>'
            )
        elif entrou:
            col_entrada = f'<span class="text-muted small">{entrou}</span>'
        else:
            col_entrada = "—"

        data_td = (
            f'<td class="small text-muted">{r["data"]}</td>' if mostrar_data else ""
        )

        return (
            f"<tr>"
            f"<td>{esc(r['NI'])}</td>"
            f"<td><strong>{esc(r['Nome_completo'])}</strong></td>"
            f"<td>{r['ano']}º</td>"
            f"{data_td}"
            f"<td>{_tipo_badge(r['tipo'])}</td>"
            f"<td>{estado}</td>"
            f"<td>{col_saida}</td>"
            f"<td>{col_entrada}</td>"
            f"</tr>"
        )

    rows_html = "".join(_build_row(r) for r in rows_hoje)

    # Secção de alunos ainda fora de dias anteriores
    fora_html = ""
    if rows_fora:
        fora_rows_html = "".join(_build_row(r, mostrar_data=True) for r in rows_fora)
        fora_html = f"""
        <div class="card" style="border-left:4px solid var(--danger)">
          <div class="card-title" style="color:var(--danger)">
            ⚠️ Ainda fora da unidade — dias anteriores ({len(rows_fora)})
          </div>
          <div class="alert alert-warn" style="font-size:.82rem">
            Estes alunos saíram em dias anteriores e ainda não regressaram.
            Regista a entrada aqui quando voltarem.
          </div>
          <div class="table-wrap">
            <table>
              <thead>
                <tr><th>NI</th><th>Nome</th><th>Ano</th><th>Data saída</th><th>Tipo</th><th>Estado</th><th>Saída</th><th>Entrada</th></tr>
              </thead>
              <tbody>{fora_rows_html}</tbody>
            </table>
          </div>
        </div>"""

    content = f"""
    <div class="container">
      <div class="page-header">
        {_back_btn(url_for("painel_dia", d=d_str))}
        <div class="page-title">🚪 Licenças / Entradas &amp; Saídas</div>
      </div>
      <div class="flex-between" style="margin-bottom:1rem">
        <div class="flex">
          <a class="btn btn-ghost btn-sm" href="{url_for("licencas_entradas_saidas", d=prev_d)}">← Anterior</a>
          <strong>{NOMES_DIAS[dt.weekday()]}, {dt.strftime("%d/%m/%Y")}</strong>
          <a class="btn btn-ghost btn-sm" href="{url_for("licencas_entradas_saidas", d=next_d)}">Próximo →</a>
        </div>
      </div>

      <div class="grid grid-4" style="margin-bottom:1rem">
        <div class="stat-box">
          <div class="stat-num">{total}</div>
          <div class="stat-lbl">Licenças hoje</div>
        </div>
        <div class="stat-box">
          <div class="stat-num">{saidas}</div>
          <div class="stat-lbl">Saíram hoje</div>
        </div>
        <div class="stat-box">
          <div class="stat-num">{entradas}</div>
          <div class="stat-lbl">Regressaram hoje</div>
        </div>
        <div class="stat-box" style="{"background:#fef3cd" if fora > 0 else ""}">
          <div class="stat-num" style="{"color:var(--danger)" if fora > 0 else ""}">{fora}</div>
          <div class="stat-lbl">Fora da unidade</div>
        </div>
      </div>

      {fora_html}

      <div class="card">
        <div class="card-title">Licenças de {dt.strftime("%d/%m/%Y")}</div>
        <div class="table-wrap">
          <table>
            <thead>
              <tr><th>NI</th><th>Nome</th><th>Ano</th><th>Tipo</th><th>Estado</th><th>Saída</th><th>Entrada</th></tr>
            </thead>
            <tbody>
              {rows_html or '<tr><td colspan="7" class="text-muted center" style="padding:1.5rem">Sem licenças para este dia.</td></tr>'}
            </tbody>
          </table>
        </div>
      </div>
    </div>"""
    return render(content)


# ═══════════════════════════════════════════════════════════════════════════
# ADMIN
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/admin")
@role_required("admin")
def admin_home():
    hoje = date.today()
    t = sr.get_totais_dia(hoje.isoformat())
    with sr.db() as conn:
        n_users = conn.execute("SELECT COUNT(*) c FROM utilizadores").fetchone()["c"]

    action_cards = [
        (url_for("painel_dia"), "📋", "Painel do dia", "Ocupação e totais"),
        (
            url_for("admin_utilizadores"),
            "👥",
            f"Utilizadores ({n_users})",
            "Gerir contas",
        ),
        (url_for("admin_menus"), "🍽️", "Menus & Capacidade", "Ementas e limites"),
        (
            url_for("dashboard_semanal"),
            "📊",
            "Dashboard Semanal",
            "Gráficos e relatório",
        ),
        (url_for("relatorio_semanal"), "📈", "Relatório Semanal", "Exportar dados"),
        (url_for("admin_log"), "📜", "Log de Refeições", "Alterações de refeições"),
        (
            url_for("admin_audit"),
            "🔐",
            "Auditoria de Ações",
            "Logins e alterações admin",
        ),
        (url_for("admin_calendario"), "⚙️", "Gerir Calendário", "Dias operacionais"),
        (url_for("ausencias"), "🚫", "Ausências", "Gerir ausências"),
        (url_for("detencoes_cmd"), "⛔", "Detenções", "Registar detenções"),
        (
            url_for("licencas_entradas_saidas"),
            "🚪",
            "Licenças / Entradas",
            "Controlo de saídas",
        ),
        (url_for("calendario_publico"), "📅", "Calendário", "Ver calendário"),
        (
            url_for("admin_companhias"),
            "⚓",
            "Gestão de Companhias",
            "Turmas, promoções e cursos",
        ),
        (
            url_for("controlo_presencas"),
            "🎯",
            "Controlo Presenças",
            "Pesquisa rápida por NI",
        ),
        (url_for("admin_importar_csv"), "📥", "Importar CSV", "Criar alunos em massa"),
        (
            url_for("admin_backup_download"),
            "💾",
            "Download BD",
            "Descarregar base de dados",
        ),
    ]

    anos = _get_anos_disponiveis()
    ano_cards = "".join(
        f'<a class="action-card" href="{url_for("lista_alunos_ano", ano=a, d=hoje.isoformat())}">'
        f'<div class="icon">👥</div><div class="label">{_ano_label(a)}</div><div class="desc">Lista de presenças</div></a>'
        for a in anos
    )

    cards_html = "".join(
        f'<a class="action-card" href="{href}"><div class="icon">{icon}</div>'
        f'<div class="label">{label}</div><div class="desc">{desc}</div></a>'
        for href, icon, label, desc in action_cards
    )

    total_alm = t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]
    total_jan = t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]

    content = f"""
    <div class="container">
      <div class="page-header"><div class="page-title">⚓ Administração — Escola Naval</div></div>
      <div class="card">
        <div class="card-title">📊 Hoje — {hoje.strftime("%d/%m/%Y")}</div>
        <div class="grid grid-4">
          <div class="stat-box"><div class="stat-num">{t["pa"]}</div><div class="stat-lbl">Pequenos Almoços</div></div>
          <div class="stat-box"><div class="stat-num">{t["lan"]}</div><div class="stat-lbl">Lanches</div></div>
          <div class="stat-box"><div class="stat-num">{total_alm}</div><div class="stat-lbl">Almoços</div></div>
          <div class="stat-box"><div class="stat-num">{total_jan}</div><div class="stat-lbl">Jantares</div></div>
        </div>
      </div>
      <div class="card">
        <div class="card-title">⚡ Módulos</div>
        <div class="grid grid-4">{cards_html}</div>
      </div>
      <div class="card">
        <div class="card-title">👥 Lista por ano</div>
        <div class="grid grid-4">{ano_cards}</div>
      </div>
    </div>"""
    return render(content)


@app.route("/admin/utilizadores", methods=["GET", "POST"])
@role_required("admin")
def admin_utilizadores():
    if request.method == "POST":
        acao = request.form.get("acao", "")
        if acao == "criar":
            ok, err = _criar_utilizador(
                request.form.get("nii", "").strip(),
                request.form.get("ni", "").strip(),
                request.form.get("nome", "").strip(),
                request.form.get("ano", "").strip(),
                request.form.get("perfil", "aluno"),
                request.form.get("pw", "").strip(),
            )
            flash(
                "Utilizador criado." if ok else (err or "Erro."),
                "ok" if ok else "error",
            )
        elif acao == "editar_user":
            nii_e = _val_nii(request.form.get("nii", ""))
            nome_e = _val_nome(request.form.get("nome", ""))
            ni_e = _val_ni(request.form.get("ni", ""))
            ano_e = _val_ano(request.form.get("ano", ""))
            perfil_e = _val_perfil(request.form.get("perfil", "aluno"))
            email_e = _val_email(request.form.get("email", ""))
            tel_e = _val_phone(request.form.get("telemovel", ""))
            pw_e = request.form.get("pw", "").strip()[:256]
            if not nii_e:
                flash("NII inválido.", "error")
            elif not nome_e:
                flash("Nome inválido ou vazio.", "error")
            elif ni_e is None:
                flash("NI inválido.", "error")
            elif ano_e is None:
                flash("Ano inválido (0-8).", "error")
            elif not perfil_e:
                flash("Perfil inválido.", "error")
            elif email_e is False:
                flash("Email inválido.", "error")
            elif tel_e is False:
                flash("Telemóvel inválido.", "error")
            else:
                try:
                    with sr.db() as conn:
                        conn.execute(
                            "UPDATE utilizadores SET Nome_completo=?,NI=?,ano=?,perfil=?,email=?,telemovel=? WHERE NII=?",
                            (nome_e, ni_e, ano_e, perfil_e, email_e, tel_e, nii_e),
                        )
                        conn.commit()
                    if pw_e:
                        with sr.db() as conn:
                            conn.execute(
                                "UPDATE utilizadores SET Palavra_chave=?,must_change_password=1 WHERE NII=?",
                                (generate_password_hash(pw_e), nii_e),
                            )
                            conn.commit()
                    _audit(
                        current_user().get("nii", "admin"),
                        "editar_utilizador",
                        f"NII={nii_e}",
                    )
                    flash("Utilizador atualizado.", "ok")
                except Exception as ex:
                    flash(f"Erro: {ex}", "error")
        elif acao == "editar_contactos":
            nii_e = _val_nii(request.form.get("nii", ""))
            email_e = _val_email(request.form.get("email", ""))
            tel_e = _val_phone(request.form.get("telemovel", ""))
            if not nii_e:
                flash("NII inválido.", "error")
            elif email_e is False:
                flash("Email inválido.", "error")
            elif tel_e is False:
                flash("Telemóvel inválido.", "error")
            else:
                try:
                    with sr.db() as conn:
                        conn.execute(
                            "UPDATE utilizadores SET email=?, telemovel=? WHERE NII=?",
                            (email_e, tel_e, nii_e),
                        )
                        conn.commit()
                    flash("Contactos atualizados.", "ok")
                except Exception as ex:
                    flash(f"Erro: {ex}", "error")
        elif acao == "reset_pw":
            nii = request.form.get("nii", "")
            ok, nova_pw = _reset_pw(nii)
            flash(
                f"Password resetada. Temporária: {nova_pw}" if ok else nova_pw,
                "ok" if ok else "error",
            )
        elif acao == "desbloquear":
            _unblock_user(request.form.get("nii", ""))
            flash("Desbloqueado.", "ok")
        elif acao == "eliminar":
            nii = request.form.get("nii", "")
            eliminado = _eliminar_utilizador(nii)
            if eliminado:
                _audit(
                    current_user().get("nii", "admin"),
                    "eliminar_utilizador",
                    f"NII={nii}",
                )
            flash(f"'{nii}' eliminado." if eliminado else "NII não encontrado.", "ok")
        return redirect(url_for("admin_utilizadores"))

    q = request.args.get("q", "").strip()
    ano_f = request.args.get("ano", "all")
    edit_nii = request.args.get("edit_contactos", "")
    with sr.db() as conn:
        sql = "SELECT id,NII,NI,Nome_completo,ano,perfil,locked_until,email,telemovel FROM utilizadores WHERE 1=1"
        args = []
        if q:
            sql += " AND Nome_completo LIKE ?"
            args.append(f"%{q}%")
        if ano_f != "all":
            sql += " AND ano=?"
            args.append(ano_f)
        sql += " ORDER BY ano, NI"
        rows = [dict(r) for r in conn.execute(sql, args).fetchall()]

    edit_user_nii = request.args.get("edit_user", "")
    edit_user_row = next((r for r in rows if r["NII"] == edit_user_nii), None)
    edit_row = next((r for r in rows if r["NII"] == edit_nii), None)

    def action_btns(r):
        ne = esc(r["NII"])
        b = f'<a class="btn btn-gold btn-sm" href="?edit_user={ne}" title="Editar utilizador">✏️ Editar</a>'
        b += f'<a class="btn btn-ghost btn-sm" href="?edit_contactos={ne}" title="Editar email/telemóvel">✉️</a>'
        if r.get("locked_until"):
            b += f'<form method="post" style="display:inline">{csrf_input()}<input type="hidden" name="acao" value="desbloquear"><input type="hidden" name="nii" maxlength="32" value="{ne}"><button class="btn btn-ghost btn-sm">🔓</button></form>'
        b += f'<form method="post" style="display:inline" onsubmit="return confirm(\'Eliminar {ne}?\');">{csrf_input()}<input type="hidden" name="acao" value="eliminar"><input type="hidden" name="nii" maxlength="32" value="{ne}"><button class="btn btn-danger btn-sm">🗑</button></form>'
        return b

    rows_html = "".join(
        f"""
      <tr{'style="background:#f0f7ff"' if r["NII"] == edit_user_nii or r["NII"] == edit_nii else ""}>
        <td class="small text-muted">{esc(r["NII"])}</td><td>{esc(r["NI"])}</td>
        <td><strong>{esc(r["Nome_completo"])}</strong></td>
        <td class="center">{esc(r["ano"])}</td>
        <td><span class="badge badge-info">{esc(r["perfil"])}</span></td>
        <td class="small text-muted">{esc(r.get("email") or "—")}</td>
        <td class="small text-muted">{esc(r.get("telemovel") or "—")}</td>
        <td>{'<span class="badge badge-warn">Bloqueado</span>' if r.get("locked_until") else '<span class="badge badge-ok">Ativo</span>'}</td>
        <td>{action_btns(r)}</td>
      </tr>"""
        for r in rows
    )

    edit_user_form = ""
    if edit_user_row:
        er = edit_user_row
        perfil_opts = "".join(
            f'<option value="{p}" {"selected" if er["perfil"] == p else ""}>{p}</option>'
            for p in ["aluno", "oficialdia", "cozinha", "cmd", "admin"]
        )
        edit_user_form = f'''
        <div class="card" style="border:1.5px solid var(--primary);max-width:640px">
          <div class="card-title">✏️ Editar Utilizador — {esc(er["Nome_completo"])}</div>
          <form method="post">
            {csrf_input()}
            <input type="hidden" name="acao" value="editar_user">
            <input type="hidden" name="nii" maxlength="32" value="{esc(er["NII"])}">
            <div class="grid grid-3">
              <div class="form-group"><label>Nome completo</label><input type="text" name="nome" value="{esc(er["Nome_completo"])}" required></div>
              <div class="form-group"><label>NI</label><input type="text" name="ni" value="{esc(er["NI"] or "")}"></div>
              <div class="form-group"><label>Ano</label>
                <select name="ano">
                  <option value="0">0 — Concluído/Inativo</option>
                  {"".join(f'<option value="{a}" {"selected" if str(er["ano"]) == str(a) else ""}>{_ano_label(a)}</option>' for a, _ in ANOS_OPCOES)}
                </select>
              </div>
              <div class="form-group"><label>Perfil</label><select name="perfil">{perfil_opts}</select></div>
              <div class="form-group"><label>Email</label><input type="email" name="email" value="{esc(er.get("email") or "")}"></div>
              <div class="form-group"><label>Telemóvel</label><input type="tel" name="telemovel" value="{esc(er.get("telemovel") or "")}"></div>
            </div>
            <div class="form-group"><label>Nova password (deixa em branco para não alterar)</label><input type="text" name="pw" placeholder="Nova password opcional..."></div>
            <div class="gap-btn">
              <button class="btn btn-ok">💾 Guardar alterações</button>
              <a class="btn btn-ghost" href="{url_for("admin_utilizadores")}">Cancelar</a>
            </div>
          </form>
        </div>'''

    edit_contactos_form = ""
    if edit_row:
        edit_contactos_form = f"""
        <div class="card" style="border:1.5px solid var(--gold);max-width:520px">
          <div class="card-title">✉️ Contactos — {esc(edit_row["Nome_completo"])}</div>
          <form method="post">
            {csrf_input()}
            <input type="hidden" name="acao" value="editar_contactos">
            <input type="hidden" name="nii" maxlength="32" value="{esc(edit_row["NII"])}">
            <div class="grid grid-2">
              <div class="form-group"><label>Email</label>
                <input type="email" name="email" value="{esc(edit_row.get("email") or "")}" placeholder="nome@exemplo.pt">
              </div>
              <div class="form-group"><label>Telemóvel</label>
                <input type="tel" name="telemovel" value="{esc(edit_row.get("telemovel") or "")}" placeholder="+351XXXXXXXXX">
              </div>
            </div>
            <div class="gap-btn">
              <button class="btn btn-ok">💾 Guardar contactos</button>
              <a class="btn btn-ghost" href="{url_for("admin_utilizadores")}">Cancelar</a>
            </div>
          </form>
        </div>"""

    content = f"""
    <div class="container">
      <div class="page-header">{_back_btn(url_for("admin_home"))}<div class="page-title">👥 Utilizadores ({len(rows)})</div>
        <a class="btn btn-primary btn-sm" href="{url_for("admin_importar_csv")}">📥 Importar CSV</a>
      </div>
      {edit_user_form}
      {edit_contactos_form}
      <div class="card">
        <form method="get" style="display:flex;gap:.5rem;flex-wrap:wrap">
          <input type="text" name="q" placeholder="Pesquisar por nome..." value="{esc(q)}" style="flex:1;min-width:200px">
          <select name="ano" style="width:auto">
            <option value="all" {"selected" if ano_f == "all" else ""}>Todos os anos</option>
            {"".join(f"<option value='{a}' {'selected' if ano_f == str(a) else ''}>{_ano_label(a)}</option>" for a, _ in ANOS_OPCOES)}
          </select>
          <button class="btn btn-primary">Filtrar</button>
        </form>
      </div>
      <div class="card">
        <div class="card-title">🆕 Criar utilizador</div>
        <form method="post">
          {csrf_input()}
          <input type="hidden" name="acao" value="criar">
          <div class="grid grid-3">
            <div class="form-group"><label>NII</label><input type="text" name="nii" maxlength="32" required></div>
            <div class="form-group"><label>NI</label><input type="text" name="ni" required></div>
            <div class="form-group"><label>Nome completo</label><input type="text" name="nome" required></div>
            <div class="form-group"><label>Ano</label>
              <select name="ano" required>
                {"".join(f"<option value='{a}'>{_ano_label(a)}</option>" for a, _ in ANOS_OPCOES)}
              </select>
            </div>
            <div class="form-group"><label>Perfil</label>
              <select name="perfil">{"".join(f"<option value='{p}'>{p}</option>" for p in ["aluno", "oficialdia", "cozinha", "cmd", "admin"])}</select>
            </div>
            <div class="form-group"><label>Password inicial</label><input type="text" name="pw" required></div>
          </div>
          <button class="btn btn-ok">Criar</button>
        </form>
      </div>
      <div class="card">
        <div class="card-title">Lista
<span style="font-size:.74rem;font-weight:400;color:var(--muted);margin-left:.5rem">Clica em ✉️ para editar email/telemóvel</span>
        </div>
        <div class="table-wrap">
          <table>
            <thead><tr><th>NII</th><th>NI</th><th>Nome</th><th>Ano</th><th>Perfil</th><th>Email</th><th>Telemóvel</th><th>Estado</th><th>Ações</th></tr></thead>
            <tbody>{rows_html or '<tr><td colspan="9" class="text-muted center" style="padding:1.5rem">Sem utilizadores.</td></tr>'}</tbody>
          </table>
        </div>
      </div>
    </div>"""
    return render(content)


@app.route("/admin/importar-csv", methods=["GET", "POST"])
@role_required("admin")
def admin_importar_csv():
    """Importação de alunos em massa via CSV.

    Formato esperado (com ou sem cabeçalho):
        NII, NI, Nome_completo, ano
    Colunas opcionais na mesma linha: perfil, password
    Se perfil omitido → 'aluno'
    Se password omitida → NII do aluno (deve alterar no 1.º login)
    """
    import csv
    import io

    resultado = None

    if request.method == "POST":
        acao = request.form.get("acao", "")

        if acao == "preview":
            f = request.files.get("csvfile")
            if not f or not f.filename:
                flash("Nenhum ficheiro selecionado.", "error")
                return redirect(url_for("admin_importar_csv"))

            raw = f.read().decode("utf-8-sig", errors="replace")
            linhas = list(csv.reader(io.StringIO(raw)))

            # Detectar cabeçalho (primeira célula da 1.ª linha)
            if (
                linhas
                and linhas[0]
                and linhas[0][0].strip().upper() in ("NII", "#", "ID", "NUM")
            ):
                linhas = linhas[1:]

            preview_rows = []
            erros = []
            with sr.db() as conn:
                existentes = {
                    r["NII"]
                    for r in conn.execute("SELECT NII FROM utilizadores").fetchall()
                }

            for i, row in enumerate(linhas, 1):
                row = [c.strip() for c in row]
                if not any(row):
                    continue
                if len(row) < 4:
                    erros.append(
                        f"Linha {i}: colunas insuficientes ({len(row)} — esperadas: NII, NI, Nome, Ano)."
                    )
                    continue
                nii, ni, nome, ano_raw = row[0], row[1], row[2], row[3]
                perfil = row[4] if len(row) > 4 and row[4] else "aluno"
                pw = row[5] if len(row) > 5 and row[5] else nii

                if not nii or not ni or not nome:
                    erros.append(f"Linha {i}: NII, NI e Nome são obrigatórios.")
                    continue
                try:
                    ano = int(ano_raw)
                    if ano not in [a for a, _ in ANOS_OPCOES]:
                        erros.append(
                            f"Linha {i} ({nii}): ano inválido '{ano_raw}'. Usa 1–8."
                        )
                        continue
                except ValueError:
                    erros.append(f"Linha {i} ({nii}): ano não é número ('{ano_raw}').")
                    continue

                duplicado = nii in existentes
                preview_rows.append(
                    {
                        "linha": i,
                        "nii": nii,
                        "ni": ni,
                        "nome": nome,
                        "ano": ano,
                        "perfil": perfil,
                        "pw": pw,
                        "duplicado": duplicado,
                    }
                )

            resultado = {"preview": preview_rows, "erros": erros, "raw": raw}

        elif acao == "confirmar":
            raw = request.form.get("raw_csv", "")
            linhas = list(csv.reader(io.StringIO(raw)))
            if (
                linhas
                and linhas[0]
                and linhas[0][0].strip().upper() in ("NII", "#", "ID", "NUM")
            ):
                linhas = linhas[1:]

            criados = 0
            ignorados = 0
            erros_conf = []
            with sr.db() as conn:
                existentes = {
                    r["NII"]
                    for r in conn.execute("SELECT NII FROM utilizadores").fetchall()
                }

            for i, row in enumerate(linhas, 1):
                row = [c.strip() for c in row]
                if not any(row) or len(row) < 4:
                    continue
                nii, ni, nome, ano_raw = row[0], row[1], row[2], row[3]
                perfil = row[4] if len(row) > 4 and row[4] else "aluno"
                pw = row[5] if len(row) > 5 and row[5] else nii

                if nii in existentes:
                    ignorados += 1
                    continue

                try:
                    ano = int(ano_raw)
                except ValueError:
                    erros_conf.append(f"Linha {i} ({nii}): ano inválido.")
                    continue

                ok, err = _criar_utilizador(nii, ni, nome, str(ano), perfil, pw)
                if ok:
                    criados += 1
                    existentes.add(nii)
                else:
                    erros_conf.append(f"Linha {i} ({nii}): {err}")

            _audit(
                current_user().get("nii", "admin"),
                "importar_csv",
                f"criados={criados} ignorados={ignorados} erros={len(erros_conf)}",
            )
            msgs = [f"✅ {criados} aluno(s) criado(s)."]
            if ignorados:
                msgs.append(f"⚠️ {ignorados} ignorado(s) (NII já existe).")
            if erros_conf:
                msgs.append(
                    f"❌ {len(erros_conf)} erro(s): " + "; ".join(erros_conf[:5])
                )
            flash(" ".join(msgs), "ok" if not erros_conf else "warn")
            return redirect(url_for("admin_utilizadores"))

    # ── Render ───────────────────────────────────────────────────────────────
    preview_html = ""
    erros_html = ""
    hidden_raw = ""

    if resultado:
        rows_prev = resultado["preview"]
        erros_list = resultado["erros"]

        if erros_list:
            erros_html = (
                '<div class="alert alert-warn">⚠️ <strong>Avisos de parsing:</strong><ul style="margin:.4rem 0 0 1.2rem">'
                + "".join(f"<li>{esc(e)}</li>" for e in erros_list)
                + "</ul></div>"
            )

        novos = [r for r in rows_prev if not r["duplicado"]]
        dupls = [r for r in rows_prev if r["duplicado"]]
        raw_csv_escaped = esc(resultado["raw"])
        hidden_raw = f'<input type="hidden" name="raw_csv" value="{raw_csv_escaped}">'

        def _ano_badge(a):
            return f'<span class="badge badge-info">{_ano_label(a)}</span>'

        trs = "".join(
            f"""
          <tr style="{"background:#f0fff4" if not r["duplicado"] else "background:#fff9e6;opacity:.7"}">
            <td class="small text-muted">{r["linha"]}</td>
            <td><strong>{esc(r["nii"])}</strong></td>
            <td>{esc(r["ni"])}</td>
            <td>{esc(r["nome"])}</td>
            <td>{_ano_badge(r["ano"])}</td>
            <td><span class="badge badge-{"info" if r["perfil"] == "aluno" else "warn"}">{esc(r["perfil"])}</span></td>
            <td class="small text-muted">{esc(r["pw"]) if not r["duplicado"] else "—"}</td>
            <td>{'<span class="badge badge-warn">⚠️ Já existe</span>' if r["duplicado"] else '<span class="badge badge-ok">✅ Novo</span>'}</td>
          </tr>"""
            for r in rows_prev
        )

        sumario = (
            f'<div class="alert alert-info" style="margin-bottom:.5rem">'
            f"📊 <strong>{len(novos)} a criar</strong>"
            f"{f', {len(dupls)} ignorados (já existem)' if dupls else ''}"
            f", {len(erros_list)} avisos de formato.</div>"
        )

        confirmar_btn = (
            f'''
        <form method="post" style="margin-top:.9rem">
          {csrf_input()}
          <input type="hidden" name="acao" value="confirmar">
          {hidden_raw}
          <button class="btn btn-ok" {"disabled" if not novos else ""}>
            ✅ Confirmar e importar {len(novos)} aluno(s)
          </button>
          <a class="btn btn-ghost" href="{url_for("admin_importar_csv")}" style="margin-left:.5rem">↩️ Cancelar</a>
        </form>'''
            if novos
            else '<div class="alert alert-warn">Nenhum aluno novo para importar.</div>'
        )

        preview_html = f"""
        <div class="card">
          <div class="card-title">👁️ Pré-visualização ({len(rows_prev)} linha(s))</div>
          {sumario}
          {erros_html}
          <div class="table-wrap">
            <table>
              <thead><tr><th>#</th><th>NII</th><th>NI</th><th>Nome</th><th>Ano</th><th>Perfil</th><th>Password inicial</th><th>Estado</th></tr></thead>
              <tbody>{trs}</tbody>
            </table>
          </div>
          {confirmar_btn}
        </div>"""

    content = f"""
    <div class="container">
      <div class="page-header">
        {_back_btn(url_for("admin_utilizadores"))}
        <div class="page-title">📥 Importar Alunos via CSV</div>
      </div>

      <div class="card" style="max-width:680px">
        <div class="card-title">📋 Instruções</div>
        <p style="font-size:.85rem;color:var(--muted);line-height:1.6">
          Carrega um ficheiro <strong>.csv</strong> com uma linha por aluno. Colunas aceites:<br>
          <code style="background:#f0f4f8;padding:.1rem .4rem;border-radius:4px;font-size:.83rem">NII, NI, Nome_completo, Ano [, Perfil] [, Password]</code><br><br>
          • <strong>Perfil</strong> omitido → <code>aluno</code><br>
          • <strong>Password</strong> omitida → igual ao NII (deve alterar no 1.º login)<br>
          • <strong>Ano</strong>: 1–6 para anos curriculares, 7 para CFBO, 8 para CFCO<br>
          • Linhas com NII já existente são ignoradas (sem sobrescrever)<br>
          • A 1.ª linha é ignorada se começar por <code>NII</code>, <code>#</code>, <code>ID</code> ou <code>NUM</code>
        </p>
        <div class="alert alert-info" style="margin-top:.8rem;font-size:.82rem">
          💡 <strong>Exemplo de CSV:</strong><br>
          <pre style="margin:.4rem 0 0;font-size:.78rem;background:#f0f4f8;padding:.5rem;border-radius:6px;overflow-x:auto">NII,NI,Nome_completo,Ano,Perfil,Password
20240001,A001,João Silva,1,aluno,senha123
20240002,A002,Maria Costa,1
20240003,A003,Pedro Santos,2</pre>
        </div>
      </div>

      <div class="card" style="max-width:680px">
        <div class="card-title">📤 Carregar ficheiro</div>
        <form method="post" enctype="multipart/form-data">
          {csrf_input()}
          <input type="hidden" name="acao" value="preview">
          <div class="form-group">
            <label>Ficheiro CSV</label>
            <input type="file" name="csvfile" accept=".csv,.txt" required style="padding:.42rem .6rem">
          </div>
          <button class="btn btn-primary">🔍 Pré-visualizar</button>
        </form>
      </div>

      {preview_html}
    </div>"""
    return render(content)


@app.route("/admin/menus", methods=["GET", "POST"])
@role_required("cozinha", "admin", "oficialdia")
def admin_menus():
    d_str = request.args.get("d", date.today().isoformat())
    dt = _parse_date(d_str)

    if request.method == "POST":
        d_save = request.form.get("data", dt.isoformat())
        campos = [
            "pequeno_almoco",
            "lanche",
            "almoco_normal",
            "almoco_veg",
            "almoco_dieta",
            "jantar_normal",
            "jantar_veg",
            "jantar_dieta",
        ]
        vals = [request.form.get(c, "").strip() or None for c in campos]
        with sr.db() as conn:
            conn.execute(
                """INSERT OR REPLACE INTO menus_diarios
                (data,pequeno_almoco,lanche,almoco_normal,almoco_veg,almoco_dieta,jantar_normal,jantar_veg,jantar_dieta)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (d_save, *vals),
            )
            for ref in ["Pequeno Almoço", "Lanche", "Almoço", "Jantar"]:
                cap_key = "cap_" + ref.lower().replace(" ", "_").replace(
                    "ç", "c"
                ).replace("ã", "a")
                cap_val = request.form.get(cap_key, "").strip()
                if cap_val:
                    try:
                        cap_int = _val_cap(cap_val)
                        if cap_int is None:
                            continue
                        if cap_int < 0:
                            conn.execute(
                                "DELETE FROM capacidade_refeicao WHERE data=? AND refeicao=?",
                                (d_save, ref),
                            )
                        else:
                            conn.execute(
                                "INSERT OR REPLACE INTO capacidade_refeicao(data,refeicao,max_total) VALUES (?,?,?)",
                                (d_save, ref, cap_int),
                            )
                    except ValueError:
                        pass
            conn.commit()
        flash("Menu e capacidades guardados.", "ok")
        return redirect(url_for("admin_menus", d=d_save))

    with sr.db() as conn:
        menu = conn.execute(
            "SELECT * FROM menus_diarios WHERE data=?", (dt.isoformat(),)
        ).fetchone()
        caps = {
            r["refeicao"]: r["max_total"]
            for r in conn.execute(
                "SELECT refeicao,max_total FROM capacidade_refeicao WHERE data=?",
                (dt.isoformat(),),
            )
        }

    def mv(k):
        return esc(menu[k] if menu and menu[k] else "")

    def cv(ref):
        return caps.get(ref, "")

    back_url = (
        url_for("painel_dia")
        if current_user().get("perfil") in ("cozinha", "oficialdia")
        else url_for("admin_home")
    )

    content = f"""
    <div class="container">
      <div class="page-header">{_back_btn(back_url)}<div class="page-title">🍽️ Menus & Capacidade</div></div>
      <div class="card" style="max-width:640px">
        <form method="post">
          {csrf_input()}
          <div class="form-group"><label>Data</label><input type="date" name="data" value="{dt.isoformat()}" required></div>
          <div class="card-title" style="margin:.7rem 0 .55rem">Ementa</div>
          <div class="grid grid-2">
            <div class="form-group"><label>☕ Pequeno Almoço</label><input type="text" name="pequeno_almoco" value="{mv("pequeno_almoco")}"></div>
            <div class="form-group"><label>🥐 Lanche</label><input type="text" name="lanche" value="{mv("lanche")}"></div>
            <div class="form-group"><label>🍽️ Almoço Normal</label><input type="text" name="almoco_normal" value="{mv("almoco_normal")}"></div>
            <div class="form-group"><label>🥗 Almoço Vegetariano</label><input type="text" name="almoco_veg" value="{mv("almoco_veg")}"></div>
            <div class="form-group"><label>🥙 Almoço Dieta</label><input type="text" name="almoco_dieta" value="{mv("almoco_dieta")}"></div>
            <div class="form-group"><label>🌙 Jantar Normal</label><input type="text" name="jantar_normal" value="{mv("jantar_normal")}"></div>
            <div class="form-group"><label>🌿 Jantar Vegetariano</label><input type="text" name="jantar_veg" value="{mv("jantar_veg")}"></div>
            <div class="form-group"><label>🥗 Jantar Dieta</label><input type="text" name="jantar_dieta" value="{mv("jantar_dieta")}"></div>
          </div>
          <div class="card-title" style="margin:.7rem 0 .55rem">Capacidades <span class="text-muted small">(-1 ou vazio = sem limite)</span></div>
          <div class="grid grid-2">
            <div class="form-group"><label>PA</label><input type="number" name="cap_pequeno_almoco" value="{cv("Pequeno Almoço")}"></div>
            <div class="form-group"><label>Lanche</label><input type="number" name="cap_lanche" value="{cv("Lanche")}"></div>
            <div class="form-group"><label>Almoço</label><input type="number" name="cap_almoco" value="{cv("Almoço")}"></div>
            <div class="form-group"><label>Jantar</label><input type="number" name="cap_jantar" value="{cv("Jantar")}"></div>
          </div>
          <hr>
          <div class="gap-btn"><button class="btn btn-ok">💾 Guardar</button><a class="btn btn-ghost" href="{back_url}">Cancelar</a></div>
        </form>
      </div>
    </div>"""
    return render(content)


@app.route("/admin/log")
@role_required("admin")
def admin_log():
    # ── Filtros ──────────────────────────────────────────────────────────────
    q_nome = request.args.get("q_nome", "").strip()
    q_por = request.args.get("q_por", "").strip()
    q_campo = request.args.get("q_campo", "").strip()
    q_d0 = request.args.get("d0", "").strip()
    q_d1 = request.args.get("d1", "").strip()
    q_limit_str = request.args.get("limite", "500")
    try:
        q_limit = min(int(q_limit_str), 5000)
    except Exception:
        q_limit = 500

    sql = """SELECT l.id, l.alterado_em, u.NII, u.Nome_completo, u.ano,
                    l.data_refeicao, l.campo, l.valor_antes, l.valor_depois, l.alterado_por
             FROM refeicoes_log l LEFT JOIN utilizadores u ON u.id=l.utilizador_id
             WHERE 1=1"""
    args = []

    if q_nome:
        sql += " AND u.Nome_completo LIKE ?"
        args.append(f"%{q_nome}%")
    if q_por:
        sql += " AND l.alterado_por LIKE ?"
        args.append(f"%{q_por}%")
    if q_campo:
        sql += " AND l.campo=?"
        args.append(q_campo)
    if q_d0:
        sql += " AND l.data_refeicao >= ?"
        args.append(q_d0)
    if q_d1:
        sql += " AND l.data_refeicao <= ?"
        args.append(q_d1)

    sql += " ORDER BY l.alterado_em DESC LIMIT ?"
    args.append(q_limit)

    with sr.db() as conn:
        rows = conn.execute(sql, args).fetchall()
        total_logs = conn.execute("SELECT COUNT(*) c FROM refeicoes_log").fetchone()[
            "c"
        ]
        campos_disponiveis = [
            r[0]
            for r in conn.execute(
                "SELECT DISTINCT campo FROM refeicoes_log ORDER BY campo"
            ).fetchall()
        ]

    # Paginação info
    mostrando = len(rows)

    campos_opts = '<option value="">Todos os campos</option>' + "".join(
        f'<option value="{c}" {"selected" if q_campo == c else ""}>{c}</option>'
        for c in campos_disponiveis
    )

    limites_opts = "".join(
        f'<option value="{n}" {"selected" if str(q_limit) == str(n) else ""}>{n} linhas</option>'
        for n in [100, 200, 500, 1000, 2000, 5000]
    )

    rows_html = "".join(
        f"""
      <tr>
        <td class="small" style="white-space:nowrap">{(r["alterado_em"] or "")[:16]}</td>
        <td>
          <span style="font-weight:600">{esc(r["Nome_completo"] or r["NII"] or "—")}</span>
          {'<br><span class="text-muted small">' + esc(r["NII"]) + (f" · {r['ano']}º ano" if r["ano"] else "") + "</span>" if r["Nome_completo"] else ""}
        </td>
        <td style="white-space:nowrap">{r["data_refeicao"]}</td>
        <td><span class="badge badge-info">{esc(r["campo"])}</span></td>
        <td class="small text-muted">{esc(r["valor_antes"] or "—")}</td>
        <td class="small" style="color:var(--ok);font-weight:600">{esc(r["valor_depois"] or "—")}</td>
        <td class="small text-muted">{esc(r["alterado_por"] or "—")}</td>
      </tr>"""
        for r in rows
    )

    filtros_ativos = any([q_nome, q_por, q_campo, q_d0, q_d1])
    limpar_btn = (
        f'<a class="btn btn-ghost btn-sm" href="{url_for("admin_log")}">✕ Limpar filtros</a>'
        if filtros_ativos
        else ""
    )

    content = f"""
    <div class="container">
      <div class="page-header">{_back_btn(url_for("admin_home"))}<div class="page-title">📜 Log de Alterações</div></div>
      <div class="card">
        <div class="card-title">🔍 Filtros
          <span class="badge badge-muted" style="margin-left:.5rem;font-size:.72rem">{total_logs} registos totais</span>
          {f'<span class="badge badge-warn" style="margin-left:.3rem;font-size:.72rem">A mostrar {mostrando}</span>' if filtros_ativos else ""}
        </div>
        <form method="get" style="display:flex;flex-wrap:wrap;gap:.5rem;align-items:flex-end">
          <div class="form-group" style="margin:0;min-width:180px;flex:1">
            <label style="font-size:.77rem">👤 Utilizador (nome)</label>
            <input type="text" name="q_nome" value="{esc(q_nome)}" placeholder="Nome do aluno..." style="font-size:.82rem">
          </div>
          <div class="form-group" style="margin:0;min-width:140px">
            <label style="font-size:.77rem">✏️ Alterado por (NII)</label>
            <input type="text" name="q_por" value="{esc(q_por)}" placeholder="NII..." style="font-size:.82rem">
          </div>
          <div class="form-group" style="margin:0;min-width:140px">
            <label style="font-size:.77rem">🏷 Campo</label>
            <select name="q_campo" style="font-size:.82rem">{campos_opts}</select>
          </div>
          <div class="form-group" style="margin:0">
            <label style="font-size:.77rem">📅 Data ref. de</label>
            <input type="date" name="d0" value="{esc(q_d0)}" style="width:auto;font-size:.82rem">
          </div>
          <div class="form-group" style="margin:0">
            <label style="font-size:.77rem">📅 até</label>
            <input type="date" name="d1" value="{esc(q_d1)}" style="width:auto;font-size:.82rem">
          </div>
          <div class="form-group" style="margin:0">
            <label style="font-size:.77rem">📊 Máx. linhas</label>
            <select name="limite" style="width:auto;font-size:.82rem">{limites_opts}</select>
          </div>
          <button class="btn btn-primary btn-sm" style="align-self:flex-end">🔍 Filtrar</button>
          {limpar_btn}
        </form>
      </div>
      <div class="card">
        <div class="card-title">Resultados
          <span class="badge badge-info" style="margin-left:.5rem;font-size:.72rem;font-weight:400">
            {mostrando} {"(filtrado)" if filtros_ativos else "mais recentes"}
          </span>
        </div>
        <div class="table-wrap">
          <table>
            <thead>
              <tr>
                <th>Quando</th>
                <th>Utilizador</th>
                <th>Data Ref.</th>
                <th>Campo</th>
                <th>Antes</th>
                <th>Depois</th>
                <th>Por (NII)</th>
              </tr>
            </thead>
            <tbody>{rows_html or '<tr><td colspan="7" class="text-muted center" style="padding:2rem">Sem registos com estes filtros.</td></tr>'}</tbody>
          </table>
        </div>
        {f'<div style="margin-top:.6rem;font-size:.8rem;color:var(--muted)">💡 A mostrar os primeiros {q_limit} resultados. Usa os filtros para refinar.</div>' if mostrando == q_limit else ""}
      </div>
    </div>"""
    return render(content)


@app.route("/admin/auditoria")
@role_required("admin")
def admin_audit():
    """Registo de ações administrativas (logins, criação/edição de utilizadores, etc.)."""
    limite = min(_val_int_id(request.args.get("limite", "500")) or 500, 5000)
    q_actor = request.args.get("actor", "").strip()
    q_action = request.args.get("action", "").strip()

    try:
        with sr.db() as conn:
            sql = "SELECT id,ts,actor,action,detail FROM admin_audit_log WHERE 1=1"
            args: list = []
            if q_actor:
                sql += " AND actor LIKE ?"
                args.append(f"%{q_actor}%")
            if q_action:
                sql += " AND action LIKE ?"
                args.append(f"%{q_action}%")
            sql += " ORDER BY id DESC LIMIT ?"
            args.append(limite)
            rows = [dict(r) for r in conn.execute(sql, args).fetchall()]
            total = conn.execute("SELECT COUNT(*) c FROM admin_audit_log").fetchone()[
                "c"
            ]
    except Exception as exc:
        app.logger.error(f"admin_audit: {exc}")
        rows, total = [], 0

    ACTION_ICONS = {
        "login": "🔑",
        "criar_utilizador": "➕",
        "editar_utilizador": "✏️",
        "reset_password": "🔄",
        "eliminar_utilizador": "🗑️",
    }

    rows_html = "".join(
        f"""
      <tr>
        <td class="small text-muted" style="white-space:nowrap">{esc(r["ts"] or "")[:16]}</td>
        <td><strong>{esc(r["actor"])}</strong></td>
        <td>{ACTION_ICONS.get(r["action"], "📌")} {esc(r["action"])}</td>
        <td class="small text-muted">{esc(r.get("detail") or "—")}</td>
      </tr>"""
        for r in rows
    )

    limites_opts = "".join(
        f'<option value="{n}" {"selected" if str(limite) == str(n) else ""}>{n}</option>'
        for n in [100, 200, 500, 1000, 2000, 5000]
    )

    content = f"""
    <div class="container">
      <div class="page-header">{_back_btn(url_for("admin_home"))}<div class="page-title">🔐 Auditoria de Ações</div></div>
      <div class="card">
        <div class="card-title">🔍 Filtros
          <span class="badge badge-muted" style="margin-left:.5rem;font-size:.72rem">{total} entradas</span>
        </div>
        <form method="get" style="display:flex;gap:.5rem;flex-wrap:wrap;align-items:flex-end">
          <div class="form-group" style="margin:0;min-width:160px;flex:1">
            <label style="font-size:.77rem">👤 Actor (NII)</label>
            <input type="text" name="actor" value="{esc(q_actor)}" placeholder="NII...">
          </div>
          <div class="form-group" style="margin:0;min-width:160px;flex:1">
            <label style="font-size:.77rem">📌 Ação</label>
            <input type="text" name="action" value="{esc(q_action)}" placeholder="ex: login, criar_utilizador...">
          </div>
          <div class="form-group" style="margin:0">
            <label style="font-size:.77rem">Máx.</label>
            <select name="limite" style="width:auto">{limites_opts}</select>
          </div>
          <button class="btn btn-primary btn-sm">🔍 Filtrar</button>
          <a class="btn btn-ghost btn-sm" href="{url_for("admin_audit")}">✕ Limpar</a>
        </form>
      </div>
      <div class="card">
        <div class="table-wrap">
          <table>
            <thead><tr><th>Quando</th><th>Actor</th><th>Ação</th><th>Detalhe</th></tr></thead>
            <tbody>{rows_html or '<tr><td colspan="4" class="text-muted center" style="padding:2rem">Sem registos.</td></tr>'}</tbody>
          </table>
        </div>
      </div>
    </div>"""
    return render(content)


@app.route("/admin/calendario", methods=["GET", "POST"])
@role_required("admin", "cmd")
def admin_calendario():
    u = current_user()
    if request.method == "POST":
        acao = request.form.get("acao", "")
        if acao == "adicionar":
            try:
                dia_de = request.form.get("dia_de", "").strip()
                dia_ate = request.form.get("dia_ate", "").strip() or dia_de
                tipo = _val_tipo_calendario(request.form.get("tipo", "normal"))
                nota = _val_text(request.form.get("nota", ""), 200) or None
                if not dia_de:
                    flash("Data de início obrigatória.", "error")
                else:
                    d_de = datetime.strptime(dia_de, "%Y-%m-%d").date()
                    d_ate = datetime.strptime(dia_ate, "%Y-%m-%d").date()
                    range_ok, range_msg = _val_date_range(d_de, d_ate)
                    if not range_ok:
                        flash(range_msg, "error")
                    else:
                        count = 0
                        with sr.db() as conn:
                            cur = d_de
                            while cur <= d_ate:
                                conn.execute(
                                    "INSERT OR REPLACE INTO calendario_operacional(data,tipo,nota) VALUES (?,?,?)",
                                    (cur.isoformat(), tipo, nota),
                                )
                                cur += timedelta(days=1)
                                count += 1
                            conn.commit()
                        flash(
                            f"{count} dia(s) adicionado(s) ao calendário ({dia_de} → {dia_ate}).",
                            "ok",
                        )
            except ValueError as e:
                flash(f"Data inválida: {e}", "error")
            except Exception as e:
                flash(str(e), "error")
        elif acao == "remover":
            with sr.db() as conn:
                conn.execute(
                    "DELETE FROM calendario_operacional WHERE data=?",
                    (request.form.get("dia", ""),),
                )
                conn.commit()
            flash("Removido.", "ok")
        return redirect(url_for("admin_calendario"))

    hoje = date.today()
    with sr.db() as conn:
        entradas = conn.execute(
            "SELECT data,tipo,nota FROM calendario_operacional WHERE data >= ? ORDER BY data LIMIT 90",
            (hoje.isoformat(),),
        ).fetchall()

    TIPOS = ["normal", "fim_semana", "feriado", "exercicio", "outro"]
    ICONES = {
        "normal": "✅",
        "fim_semana": "🔵",
        "feriado": "🔴",
        "exercicio": "🟡",
        "outro": "⚪",
    }

    rows_html = "".join(
        f"""
      <tr><td>{r["data"]}</td><td>{ICONES.get(r["tipo"], "⚪")} {esc(r["tipo"])}</td><td>{esc(r["nota"] or "—")}</td>
      <td><form method="post" style="display:inline">{csrf_input()}<input type="hidden" name="acao" value="remover"><input type="hidden" name="dia" value="{r["data"]}"><button class="btn btn-danger btn-sm">🗑</button></form></td></tr>"""
        for r in entradas
    )

    back_url = (
        url_for("admin_home") if u.get("perfil") == "admin" else url_for("painel_dia")
    )
    content = f"""
    <div class="container">
      <div class="page-header">{_back_btn(back_url)}<div class="page-title">📅 Calendário Operacional</div></div>
      <div class="card">
        <div class="card-title">Adicionar / atualizar período</div>
        <div class="alert alert-info" style="margin-bottom:.8rem">
          💡 Para um único dia, preenche apenas a <strong>Data de início</strong> (ou coloca a mesma data nos dois campos).
          Para um período, preenche ambas as datas — todos os dias do intervalo serão atualizados.
        </div>
        <form method="post">
          {csrf_input()}
          <input type="hidden" name="acao" value="adicionar">
          <div class="grid grid-2" style="max-width:520px">
            <div class="form-group"><label>📅 Data de início</label><input type="date" name="dia_de" required value="{hoje.isoformat()}"></div>
            <div class="form-group"><label>📅 Data de fim <span class="text-muted small">(inclusive)</span></label><input type="date" name="dia_ate" value="{hoje.isoformat()}"></div>
          </div>
          <div class="grid grid-2" style="max-width:520px">
            <div class="form-group"><label>Tipo</label>
              <select name="tipo">{"".join(f"<option value='{t}'>{ICONES.get(t, '')} {t}</option>" for t in TIPOS)}</select>
            </div>
            <div class="form-group"><label>Nota</label><input type="text" name="nota" placeholder="ex: Natal, Exercício..."></div>
          </div>
          <button class="btn btn-ok">💾 Guardar</button>
        </form>
      </div>
      <div class="card">
        <div class="card-title">Próximas entradas (até 90 dias)</div>
        <div class="table-wrap">
          <table>
            <thead><tr><th>Data</th><th>Tipo</th><th>Nota</th><th></th></tr></thead>
            <tbody>{rows_html or '<tr><td colspan="4" class="text-muted center" style="padding:1.5rem">Sem entradas.</td></tr>'}</tbody>
          </table>
        </div>
      </div>
    </div>"""
    return render(content)


# ═══════════════════════════════════════════════════════════════════════════
# CALENDÁRIO PÚBLICO — Visível por todos os utilizadores
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/calendario")
@login_required
def calendario_publico():
    import calendar as _cal

    u = current_user()
    hoje = date.today()
    mes_str = request.args.get("mes", hoje.strftime("%Y-%m"))
    try:
        ano_m, mes_m = int(mes_str[:4]), int(mes_str[5:7])
    except Exception:
        ano_m, mes_m = hoje.year, hoje.month

    ICONES = {
        "normal": "✅",
        "fim_semana": "🔵",
        "feriado": "🔴",
        "exercicio": "🟡",
        "outro": "⚪",
    }
    LABELS = {
        "normal": "Normal",
        "fim_semana": "Fim de semana",
        "feriado": "Feriado",
        "exercicio": "Exercício",
        "outro": "Outro",
    }
    CORES = {
        "normal": "#eafaf1",
        "fim_semana": "#ebf5fb",
        "feriado": "#fdecea",
        "exercicio": "#fef9e7",
        "outro": "#f8f9fa",
    }
    CORES_TEXT = {
        "normal": "#1e8449",
        "fim_semana": "#1a5276",
        "feriado": "#922b21",
        "exercicio": "#9a7d0a",
        "outro": "#6c757d",
    }

    ultimo_dia = _cal.monthrange(ano_m, mes_m)[1]
    d_inicio = date(ano_m, mes_m, 1)
    d_fim = date(ano_m, mes_m, ultimo_dia)
    with sr.db() as conn:
        entradas = {
            r["data"]: dict(r)
            for r in conn.execute(
                "SELECT data,tipo,nota FROM calendario_operacional WHERE data>=? AND data<=?",
                (d_inicio.isoformat(), d_fim.isoformat()),
            ).fetchall()
        }

    cal_grid = _cal.monthcalendar(ano_m, mes_m)
    DIAS_CAB = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]

    grid_html = ""
    for semana in cal_grid:
        grid_html += "<tr>"
        for dia_n in semana:
            if dia_n == 0:
                grid_html += '<td style="background:#f9fafb;border:1px solid var(--border);border-radius:6px"></td>'
                continue
            d_obj = date(ano_m, mes_m, dia_n)
            entrada = entradas.get(d_obj.isoformat())
            tipo = (
                entrada["tipo"]
                if entrada
                else ("fim_semana" if d_obj.weekday() >= 5 else "normal")
            )
            nota = entrada["nota"] if entrada else ""
            is_hoje = d_obj == hoje
            bg = CORES.get(tipo, "#fff")
            tc = CORES_TEXT.get(tipo, "#1a2533")
            ic = ICONES.get(tipo, "✅")
            border_style = (
                "border:2.5px solid var(--primary)"
                if is_hoje
                else "border:1px solid var(--border)"
            )
            hoje_label = (
                '<div style="font-size:.58rem;color:var(--primary);font-weight:900;text-align:center">HOJE</div>'
                if is_hoje
                else ""
            )
            nota_html = (
                '<div style="font-size:.62rem;color:'
                + tc
                + ';margin-top:.12rem">'
                + esc(nota)
                + "</div>"
                if nota
                else ""
            )
            grid_html += (
                '<td style="background:'
                + bg
                + ";"
                + border_style
                + ';border-radius:7px;padding:.38rem;vertical-align:top">'
                + hoje_label
                + '<div style="font-weight:800;font-size:.82rem;color:'
                + tc
                + '">'
                + str(dia_n)
                + '</div><div style="font-size:.6rem">'
                + ic
                + "</div>"
                + nota_html
                + "</td>"
            )
        grid_html += "</tr>"

    if mes_m == 1:
        prev_mes = f"{ano_m - 1}-12"
    else:
        prev_mes = f"{ano_m}-{mes_m - 1:02d}"
    if mes_m == 12:
        next_mes = f"{ano_m + 1}-01"
    else:
        next_mes = f"{ano_m}-{mes_m + 1:02d}"

    MESES_PT = [
        "Janeiro",
        "Fevereiro",
        "Março",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]
    mes_titulo = f"{MESES_PT[mes_m - 1]} {ano_m}"
    perfil = u.get("perfil")
    back_url = (
        url_for("admin_home")
        if perfil == "admin"
        else (url_for("aluno_home") if perfil == "aluno" else url_for("painel_dia"))
    )

    legenda_html = "".join(
        '<span style="display:inline-flex;align-items:center;gap:.3rem;font-size:.78rem">'
        '<span style="width:.75rem;height:.75rem;background:'
        + CORES[t]
        + ";border:1px solid "
        + CORES_TEXT[t]
        + ';border-radius:3px;display:inline-block"></span>'
        + LABELS[t]
        + "</span>"
        for t in ["normal", "fim_semana", "feriado", "exercicio"]
    )

    header_cells = "".join(
        '<th style="text-align:center;padding:.3rem;font-size:.78rem;color:var(--primary);font-weight:700">'
        + d
        + "</th>"
        for d in DIAS_CAB
    )

    admin_link = (
        '<a class="btn btn-primary btn-sm" href="'
        + url_for("admin_calendario")
        + '">⚙️ Gerir calendário</a>'
        if perfil in ("admin", "cmd")
        else '<div class="alert alert-info" style="margin-top:.6rem;font-size:.82rem">📌 O calendário é gerido pelo administrador.</div>'
    )

    c = (
        '<div class="container">'
        '<div class="page-header">'
        + _back_btn(back_url)
        + '<div class="page-title">📅 Calendário Operacional</div></div>'
        '<div class="card">'
        '<div class="flex-between" style="margin-bottom:.9rem">'
        '<a class="btn btn-ghost btn-sm" href="'
        + url_for("calendario_publico", mes=prev_mes)
        + '">← Mês anterior</a>'
        '<strong style="font-size:1.05rem">' + mes_titulo + "</strong>"
        '<a class="btn btn-ghost btn-sm" href="'
        + url_for("calendario_publico", mes=next_mes)
        + '">Mês seguinte →</a>'
        "</div>"
        '<div class="table-wrap"><table style="width:100%;border-collapse:separate;border-spacing:3px">'
        "<thead><tr>" + header_cells + "</tr></thead>"
        "<tbody>" + grid_html + "</tbody></table></div>"
        '<div style="margin-top:.8rem;display:flex;gap:.75rem;flex-wrap:wrap">'
        + legenda_html
        + "</div>"
        + admin_link
        + "</div></div>"
    )
    return render(c)


# ═══════════════════════════════════════════════════════════════════════════
# IMPRESSÃO — Mapa de refeições por ano
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/imprimir/<int:ano>")
@role_required("oficialdia", "cozinha", "cmd", "admin")
def imprimir_ano(ano):
    u = current_user()
    perfil = u.get("perfil")
    if perfil == "cmd" and str(ano) != str(u.get("ano", "")):
        abort(403)

    d_str = request.args.get("d", date.today().isoformat())
    dt = _parse_date(d_str)

    with sr.db() as conn:
        alunos = [
            dict(r)
            for r in conn.execute(
                """
            SELECT u.NI, u.Nome_completo,
                   r.pequeno_almoco, r.lanche, r.almoco, r.jantar_tipo, r.jantar_sai_unidade,
                   EXISTS(SELECT 1 FROM ausencias a WHERE a.utilizador_id=u.id
                          AND a.ausente_de<=? AND a.ausente_ate>=?) AS ausente
            FROM utilizadores u
            LEFT JOIN refeicoes r ON r.utilizador_id=u.id AND r.data=?
            WHERE u.ano=? ORDER BY u.NI
        """,
                (dt.isoformat(), dt.isoformat(), dt.isoformat(), ano),
            ).fetchall()
        ]

    def sim_nao(v):
        return "✓" if v else "–"

    rows = "".join(
        f"""
        <tr{'style="background:#fff9ec"' if a["ausente"] else ""}>
          <td>{esc(a["NI"])}</td>
          <td style="text-align:left">{esc(a["Nome_completo"])}{"  🏖" if a["ausente"] else ""}</td>
          <td>{sim_nao(a["pequeno_almoco"])}</td>
          <td>{sim_nao(a["lanche"])}</td>
          <td>{(a["almoco"] or "–")[:3]}</td>
          <td>{(a["jantar_tipo"] or "–")[:3]}</td>
          <td>{"✓" if a["jantar_sai_unidade"] else "–"}</td>
        </tr>"""
        for a in alunos
    )

    t = sr.get_totais_dia(dt.isoformat(), ano)
    gerado_em = datetime.now().strftime("%d/%m/%Y %H:%M")

    html = f"""<!doctype html>
<html lang="pt"><head><meta charset="utf-8">
<title>Mapa {ano}º Ano — {dt.strftime("%d/%m/%Y")}</title>
<style>
  @page{{size:A4;margin:1.5cm}}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:Arial,sans-serif;font-size:10pt;color:#111}}
  .header{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:.8cm;border-bottom:2px solid #0a2d4e;padding-bottom:.4cm}}
  .header-left h1{{font-size:14pt;color:#0a2d4e;font-weight:900}}
  .header-left p{{font-size:9pt;color:#555;margin-top:.15cm}}
  .header-right{{text-align:right;font-size:8pt;color:#555}}
  table{{width:100%;border-collapse:collapse;font-size:9pt}}
  th{{background:#0a2d4e;color:#fff;padding:.25cm .3cm;text-align:center;font-weight:700}}
  td{{padding:.22cm .3cm;border:1px solid #ccc;text-align:center;vertical-align:middle}}
  tr:nth-child(even){{background:#f5f8fc}}
  tr[style]{{background:#fff9ec!important}}
  .totais{{margin-top:.6cm;font-size:9pt;border:1px solid #ccc;border-radius:6px;padding:.4cm}}
  .totais-title{{font-weight:800;font-size:10pt;color:#0a2d4e;margin-bottom:.25cm}}
  .totais-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:.3cm}}
  .totais-item{{text-align:center;background:#f0f4f8;border-radius:4px;padding:.2cm}}
  .totais-num{{font-size:14pt;font-weight:900;color:#0a2d4e}}
  .totais-lbl{{font-size:8pt;color:#555}}
  .footer{{margin-top:.6cm;font-size:7.5pt;color:#888;text-align:right}}
  .legenda{{margin-top:.4cm;font-size:8pt;color:#555}}
  @media print{{button{{display:none!important}}}}
</style>
</head><body>
<div class="header">
  <div class="header-left">
    <h1>⚓ Escola Naval — Mapa de Refeições</h1>
    <p><strong>{ano}º Ano</strong> &nbsp;|&nbsp; {NOMES_DIAS[dt.weekday()]}, {dt.strftime("%d/%m/%Y")}</p>
  </div>
  <div class="header-right">
    Gerado em: {gerado_em}<br>
    Por: {esc(u["nome"])}
    <br><br>
    <button onclick="window.print()" style="background:#0a2d4e;color:#fff;border:none;padding:.3cm .6cm;border-radius:5px;cursor:pointer;font-size:9pt">🖨 Imprimir</button>
  </div>
</div>

<table>
  <thead><tr>
    <th style="width:1.2cm">NI</th>
    <th style="width:6cm;text-align:left">Nome</th>
    <th>PA</th><th>Lanche</th><th>Almoço</th><th>Jantar</th><th>Sai</th>
  </tr></thead>
  <tbody>{rows}</tbody>
</table>

<div class="totais">
  <div class="totais-title">📊 Totais — {ano}º Ano</div>
  <div class="totais-grid">
    <div class="totais-item"><div class="totais-num">{t["pa"]}</div><div class="totais-lbl">Peq. Almoços</div></div>
    <div class="totais-item"><div class="totais-num">{t["lan"]}</div><div class="totais-lbl">Lanches</div></div>
    <div class="totais-item"><div class="totais-num">{t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]}</div><div class="totais-lbl">Almoços</div></div>
    <div class="totais-item"><div class="totais-num">{t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]}</div><div class="totais-lbl">Jantares</div></div>
    <div class="totais-item"><div class="totais-num">{t["alm_norm"]}</div><div class="totais-lbl">Alm. Normal</div></div>
    <div class="totais-item"><div class="totais-num">{t["alm_veg"]}</div><div class="totais-lbl">Alm. Veg.</div></div>
    <div class="totais-item"><div class="totais-num">{t["alm_dieta"]}</div><div class="totais-lbl">Alm. Dieta</div></div>
    <div class="totais-item"><div class="totais-num">{t["jan_sai"]}</div><div class="totais-lbl">Saem após jantar</div></div>
  </div>
</div>
<div class="legenda">PA=Pequeno Almoço &nbsp;|&nbsp; Nor=Normal &nbsp;|&nbsp; Veg=Vegetariano &nbsp;|&nbsp; Die=Dieta &nbsp;|&nbsp; 🏖=Ausente</div>
<div class="footer">Escola Naval &nbsp;|&nbsp; Documento de uso interno</div>
</body></html>"""

    return Response(html, mimetype="text/html")


# ═══════════════════════════════════════════════════════════════════════════
# DASHBOARD VISUAL SEMANAL
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/dashboard-semanal")
@role_required("cozinha", "oficialdia", "admin")
def dashboard_semanal():
    u = current_user()
    perfil = u.get("perfil")
    hoje = date.today()
    segunda = hoje - timedelta(days=hoje.weekday())
    d0_str = request.args.get("d0", segunda.isoformat())
    d0 = _parse_date(d0_str)
    d1 = d0 + timedelta(days=6)
    prev_w = (d0 - timedelta(days=7)).isoformat()
    next_w = (d0 + timedelta(days=7)).isoformat()

    # Batch: carregar totais e calendário para toda a semana numa query
    totais_map, _t_empty = sr.get_totais_periodo(d0.isoformat(), d1.isoformat())
    cal_map_wk = sr.dias_operacionais_batch(d0, d1)
    dias = []
    for i in range(7):
        di = d0 + timedelta(days=i)
        t = totais_map.get(di.isoformat(), _t_empty)
        tipo = cal_map_wk.get(
            di.isoformat(), "fim_semana" if di.weekday() >= 5 else "normal"
        )
        dias.append({"data": di, "t": t, "tipo": tipo, "is_wknd": di.weekday() >= 5})

    max_alm = (
        max(
            (
                d["t"]["alm_norm"] + d["t"]["alm_veg"] + d["t"]["alm_dieta"]
                for d in dias
            ),
            default=1,
        )
        or 1
    )
    max_jan = (
        max(
            (
                d["t"]["jan_norm"] + d["t"]["jan_veg"] + d["t"]["jan_dieta"]
                for d in dias
            ),
            default=1,
        )
        or 1
    )
    max_pa = max((d["t"]["pa"] for d in dias), default=1) or 1

    def bar(val, maximo, cor, label):
        pct = int(round(100 * val / maximo)) if maximo else 0
        return (
            f'<div style="display:flex;align-items:flex-end;gap:.2rem;height:80px">'
            f'<div style="width:100%;display:flex;flex-direction:column;align-items:center;justify-content:flex-end;height:100%">'
            f'<span style="font-size:.7rem;font-weight:700;color:#1a2533;margin-bottom:.15rem">{val}</span>'
            f'<div style="width:100%;background:{cor};border-radius:5px 5px 0 0;height:{max(4, pct)}%"></div>'
            f"</div></div>"
        )

    # Chart almoços por dia
    alm_chart = ""
    jan_chart = ""
    pa_chart = ""
    table_rows = ""

    for d in dias:
        t = d["t"]
        di = d["data"]
        alm = t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]
        jan = t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]
        tipo = d["tipo"]
        is_wk = d["is_wknd"]
        off = tipo in ("feriado", "exercicio")
        col_bg = "#f9fafb" if off else ("#fffdf5" if is_wk else "#fff")
        dow_col = "#c9a227" if is_wk else "#0a2d4e"

        # Stacked bar almoço
        alm_tot = alm or 0
        pn = int(round(80 * (t["alm_norm"] / max_alm))) if max_alm else 0
        pv = int(round(80 * (t["alm_veg"] / max_alm))) if max_alm else 0
        pd = int(round(80 * (t["alm_dieta"] / max_alm))) if max_alm else 0
        alm_chart += f"""
        <div style="display:flex;flex-direction:column;align-items:center;flex:1;background:{col_bg};padding:.4rem .2rem;border-radius:6px">
          <div style="width:100%;height:80px;display:flex;flex-direction:column;justify-content:flex-end;align-items:center">
            <span style="font-size:.68rem;font-weight:800;color:#1a2533;margin-bottom:.1rem">{alm_tot or "–"}</span>
            <div style="width:70%;display:flex;flex-direction:column;border-radius:4px 4px 0 0;overflow:hidden">
              {'<div style="height:' + str(pd) + 'px;background:#d68910"></div>' if pd else ""}
              {'<div style="height:' + str(pv) + 'px;background:#2471a3"></div>' if pv else ""}
              {'<div style="height:' + str(pn) + 'px;background:#1e8449"></div>' if pn else ""}
            </div>
          </div>
          <div style="font-size:.68rem;font-weight:800;color:{dow_col};margin-top:.2rem">{ABREV_DIAS[di.weekday()]}</div>
          <div style="font-size:.62rem;color:#6c757d">{di.strftime("%d/%m")}</div>
        </div>"""

        # Bar jantar
        pj = int(round(80 * (jan / max_jan))) if max_jan else 0
        jan_chart += f"""
        <div style="display:flex;flex-direction:column;align-items:center;flex:1;background:{col_bg};padding:.4rem .2rem;border-radius:6px">
          <div style="width:100%;height:80px;display:flex;flex-direction:column;justify-content:flex-end;align-items:center">
            <span style="font-size:.68rem;font-weight:800;color:#1a2533;margin-bottom:.1rem">{jan or "–"}</span>
            <div style="width:70%;height:{max(0, pj)}px;background:#1a5276;border-radius:4px 4px 0 0"></div>
          </div>
          <div style="font-size:.68rem;font-weight:800;color:{dow_col};margin-top:.2rem">{ABREV_DIAS[di.weekday()]}</div>
        </div>"""

        # Bar PA
        pp = int(round(80 * (t["pa"] / max_pa))) if max_pa else 0
        pa_chart += f"""
        <div style="display:flex;flex-direction:column;align-items:center;flex:1;background:{col_bg};padding:.4rem .2rem;border-radius:6px">
          <div style="width:100%;height:80px;display:flex;flex-direction:column;justify-content:flex-end;align-items:center">
            <span style="font-size:.68rem;font-weight:800;color:#1a2533;margin-bottom:.1rem">{t["pa"] or "–"}</span>
            <div style="width:70%;height:{max(0, pp)}px;background:#c9a227;border-radius:4px 4px 0 0"></div>
          </div>
          <div style="font-size:.68rem;font-weight:800;color:{dow_col};margin-top:.2rem">{ABREV_DIAS[di.weekday()]}</div>
        </div>"""

        sai_td = (
            "" if perfil == "cozinha" else f'<td class="center">{t["jan_sai"]}</td>'
        )
        table_rows += f"""<tr style="background:{col_bg}">
          <td><strong style="color:{dow_col}">{ABREV_DIAS[di.weekday()]}</strong> {di.strftime("%d/%m")}</td>
          <td class="center">{t["pa"]}</td><td class="center">{t["lan"]}</td>
          <td class="center">{t["alm_norm"]}</td><td class="center">{t["alm_veg"]}</td><td class="center">{t["alm_dieta"]}</td>
          <td class="center">{t["jan_norm"]}</td><td class="center">{t["jan_veg"]}</td><td class="center">{t["jan_dieta"]}</td>
          {sai_td}
        </tr>"""

    sai_th = "" if perfil == "cozinha" else '<th class="center">Sai</th>'
    _keys = [
        "pa",
        "lan",
        "alm_norm",
        "alm_veg",
        "alm_dieta",
        "jan_norm",
        "jan_veg",
        "jan_dieta",
        "jan_sai",
    ]
    totais_semana = {k: sum(d["t"][k] for d in dias) for k in _keys}

    # Totais da semana anterior para comparação (1 query batch)
    prev_d0 = d0 - timedelta(days=7)
    prev_d1 = d0 - timedelta(days=1)
    prev_map, _ = sr.get_totais_periodo(prev_d0.isoformat(), prev_d1.isoformat())
    totais_prev = {k: 0 for k in _keys}
    for t_p in prev_map.values():
        for k in _keys:
            totais_prev[k] += t_p[k]

    def _wk_delta(curr, prev):
        d = curr - prev
        if d > 0:
            return f'<span style="color:#1e8449">↑{d}</span>'
        if d < 0:
            return f'<span style="color:#c0392b">↓{abs(d)}</span>'
        return '<span style="color:#6c757d">=</span>'

    _sai_total = (
        ""
        if perfil == "cozinha"
        else f'<td class="center"><strong>{totais_semana["jan_sai"]}</strong></td>'
    )
    _sai_prev = (
        ""
        if perfil == "cozinha"
        else f'<td class="center">{totais_prev["jan_sai"]}</td>'
    )
    _sai_var = (
        ""
        if perfil == "cozinha"
        else f'<td class="center">{_wk_delta(totais_semana["jan_sai"], totais_prev["jan_sai"])}</td>'
    )
    comparison_rows = f"""
        <tr style="background:#f0f4f8;font-weight:700;border-top:2px solid #0a2d4e">
          <td>Total semana</td>
          <td class="center">{totais_semana["pa"]}</td><td class="center">{totais_semana["lan"]}</td>
          <td class="center">{totais_semana["alm_norm"]}</td><td class="center">{totais_semana["alm_veg"]}</td><td class="center">{totais_semana["alm_dieta"]}</td>
          <td class="center">{totais_semana["jan_norm"]}</td><td class="center">{totais_semana["jan_veg"]}</td><td class="center">{totais_semana["jan_dieta"]}</td>
          {_sai_total}
        </tr>
        <tr style="background:#fef9e7;font-size:.82rem">
          <td>Semana anterior</td>
          <td class="center">{totais_prev["pa"]}</td><td class="center">{totais_prev["lan"]}</td>
          <td class="center">{totais_prev["alm_norm"]}</td><td class="center">{totais_prev["alm_veg"]}</td><td class="center">{totais_prev["alm_dieta"]}</td>
          <td class="center">{totais_prev["jan_norm"]}</td><td class="center">{totais_prev["jan_veg"]}</td><td class="center">{totais_prev["jan_dieta"]}</td>
          {_sai_prev}
        </tr>
        <tr style="background:#fff;font-size:.82rem">
          <td>Variação</td>
          <td class="center">{_wk_delta(totais_semana["pa"], totais_prev["pa"])}</td>
          <td class="center">{_wk_delta(totais_semana["lan"], totais_prev["lan"])}</td>
          <td class="center">{_wk_delta(totais_semana["alm_norm"], totais_prev["alm_norm"])}</td>
          <td class="center">{_wk_delta(totais_semana["alm_veg"], totais_prev["alm_veg"])}</td>
          <td class="center">{_wk_delta(totais_semana["alm_dieta"], totais_prev["alm_dieta"])}</td>
          <td class="center">{_wk_delta(totais_semana["jan_norm"], totais_prev["jan_norm"])}</td>
          <td class="center">{_wk_delta(totais_semana["jan_veg"], totais_prev["jan_veg"])}</td>
          <td class="center">{_wk_delta(totais_semana["jan_dieta"], totais_prev["jan_dieta"])}</td>
          {_sai_var}
        </tr>"""

    back_url = url_for("admin_home") if perfil == "admin" else url_for("painel_dia")

    legenda_alm = "".join(
        f'<span style="display:inline-flex;align-items:center;gap:.3rem;font-size:.72rem">'
        f'<span style="width:.65rem;height:.65rem;background:{c};border-radius:2px;display:inline-block"></span>{lb}</span>'
        for lb, c in [("Normal", "#1e8449"), ("Veg.", "#2471a3"), ("Dieta", "#d68910")]
    )

    content = f"""
    <div class="container">
      <div class="page-header">
        {_back_btn(back_url)}
        <div class="page-title">📊 Dashboard Semanal</div>
      </div>
      <div class="card" style="padding:.85rem 1.1rem;margin-bottom:.75rem">
        <div class="flex-between">
          <div class="flex">
            <a class="btn btn-ghost btn-sm" href="{url_for("dashboard_semanal", d0=prev_w)}">← Semana anterior</a>
            <strong>{d0.strftime("%d/%m/%Y")} — {d1.strftime("%d/%m/%Y")}</strong>
            <a class="btn btn-ghost btn-sm" href="{url_for("dashboard_semanal", d0=next_w)}">Semana seguinte →</a>
          </div>
          <form method="get" style="display:flex;gap:.3rem">
            <input type="date" name="d0" value="{d0_str}" style="width:auto">
            <button class="btn btn-primary btn-sm">Ir</button>
          </form>
        </div>
      </div>

      <div class="grid grid-4" style="margin-bottom:.85rem">
        <div class="stat-box"><div class="stat-num">{totais_semana["pa"]}</div><div class="stat-lbl">PA semana</div></div>
        <div class="stat-box"><div class="stat-num">{totais_semana["alm_norm"] + totais_semana["alm_veg"] + totais_semana["alm_dieta"]}</div><div class="stat-lbl">Almoços semana</div></div>
        <div class="stat-box"><div class="stat-num">{totais_semana["jan_norm"] + totais_semana["jan_veg"] + totais_semana["jan_dieta"]}</div><div class="stat-lbl">Jantares semana</div></div>
        <div class="stat-box"><div class="stat-num">{totais_semana["lan"]}</div><div class="stat-lbl">Lanches semana</div></div>
      </div>

      <div class="card">
        <div class="card-title">🍽️ Almoços por dia
          <span style="margin-left:.6rem;display:inline-flex;gap:.6rem">{legenda_alm}</span>
        </div>
        <div style="display:flex;gap:.3rem;align-items:flex-end;padding:.3rem 0">
          {alm_chart}
        </div>
        <div style="border-top:2px solid #e9ecef;margin-top:.3rem"></div>
      </div>

      <div class="grid grid-2">
        <div class="card">
          <div class="card-title">🌙 Jantares por dia</div>
          <div style="display:flex;gap:.3rem;align-items:flex-end;padding:.3rem 0">{jan_chart}</div>
        </div>
        <div class="card">
          <div class="card-title">☕ Pequenos Almoços por dia</div>
          <div style="display:flex;gap:.3rem;align-items:flex-end;padding:.3rem 0">{pa_chart}</div>
        </div>
      </div>

      <div class="card">
        <div class="card-title">📋 Tabela detalhada</div>
        <div class="table-wrap">
          <table>
            <thead><tr><th>Dia</th><th>PA</th><th>Lan</th><th>Alm N</th><th>Alm V</th><th>Alm D</th><th>Jan N</th><th>Jan V</th><th>Jan D</th>{sai_th}</tr></thead>
            <tbody>{table_rows}{comparison_rows}</tbody>
          </table>
        </div>
        <div class="gap-btn" style="margin-top:.8rem">
          <a class="btn btn-primary" href="{url_for("exportar_relatorio", d0=d0_str, fmt="csv")}">⬇ CSV</a>
          <a class="btn btn-primary" href="{url_for("exportar_relatorio", d0=d0_str, fmt="xlsx")}">⬇ Excel</a>
        </div>
      </div>
      <div class="card">
        <div class="card-title">📅 Relatório Mensal</div>
        <div class="gap-btn">
          <a class="btn btn-gold" href="{url_for("exportar_mensal", mes=d0.strftime("%Y-%m"), fmt="xlsx")}">📊 Excel mês {d0.strftime("%m/%Y")}</a>
          <a class="btn btn-ghost" href="{url_for("exportar_mensal", mes=d0.strftime("%Y-%m"), fmt="csv")}">📄 CSV mês {d0.strftime("%m/%Y")}</a>
        </div>
      </div>
    </div>"""
    return render(content)


# ═══════════════════════════════════════════════════════════════════════════
# EXPORTAÇÕES
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/exportar/dia")
@role_required("cozinha", "oficialdia", "cmd", "admin")
def exportar_dia():
    import io
    import csv as _csv

    d_str = request.args.get("d", date.today().isoformat())
    fmt = (request.args.get("fmt", "csv") or "csv").strip().lower()
    if fmt not in {"csv", "xlsx"}:
        abort(400)
    dt = _parse_date_strict(d_str)
    if dt is None:
        abort(400)
    t = sr.get_totais_dia(dt.isoformat())

    # Tentar xlsx via openpyxl; cair para CSV se não disponível
    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Totais {dt.strftime('%d-%m-%Y')}"

            # Cabeçalho
            header_fill = PatternFill("solid", fgColor="0A2D4E")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            headers = [
                "Data",
                "Dia",
                "PA",
                "Lanche",
                "Alm. Normal",
                "Alm. Veg.",
                "Alm. Dieta",
                "Jan. Normal",
                "Jan. Veg.",
                "Jan. Dieta",
                "Jan. Sai Unidade",
                "Total Almoços",
                "Total Jantares",
            ]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")
                c.border = border

            total_alm = t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]
            total_jan = t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]
            data_row = [
                dt.isoformat(),
                NOMES_DIAS[dt.weekday()],
                t["pa"],
                t["lan"],
                t["alm_norm"],
                t["alm_veg"],
                t["alm_dieta"],
                t["jan_norm"],
                t["jan_veg"],
                t["jan_dieta"],
                t["jan_sai"],
                total_alm,
                total_jan,
            ]
            alt_fill = PatternFill("solid", fgColor="EBF5FB")
            for col, val in enumerate(data_row, 1):
                c = ws.cell(row=2, column=col, value=val)
                c.fill = alt_fill
                c.border = border
                c.alignment = Alignment(horizontal="center")

            # Auto-largura
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col) + 4
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 22)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(
                buf.read(),
                headers={
                    "Content-Disposition": f"attachment; filename=totais_{dt.isoformat()}.xlsx",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
            )
        except ImportError:
            flash("openpyxl não instalado — a exportar CSV.", "warn")
            fmt = "csv"
        except Exception as ex:
            flash(f"Erro ao gerar Excel: {ex} — a exportar CSV.", "warn")
            fmt = "csv"

    # CSV
    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=";")
    writer.writerow(
        [
            "Data",
            "Dia",
            "PA",
            "Lanche",
            "Alm. Normal",
            "Alm. Veg.",
            "Alm. Dieta",
            "Jan. Normal",
            "Jan. Veg.",
            "Jan. Dieta",
            "Jan. Sai Unidade",
            "Total Almoços",
            "Total Jantares",
        ]
    )
    total_alm = t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]
    total_jan = t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]
    writer.writerow(
        [
            dt.isoformat(),
            NOMES_DIAS[dt.weekday()],
            t["pa"],
            t["lan"],
            t["alm_norm"],
            t["alm_veg"],
            t["alm_dieta"],
            t["jan_norm"],
            t["jan_veg"],
            t["jan_dieta"],
            t["jan_sai"],
            total_alm,
            total_jan,
        ]
    )
    csv_bytes = ("\ufeff" + buf.getvalue()).encode("utf-8")
    return Response(
        csv_bytes,
        headers={
            "Content-Disposition": f"attachment; filename=totais_{dt.isoformat()}.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


@app.route("/exportar/relatorio")
@role_required("cozinha", "oficialdia", "admin")
def exportar_relatorio():
    import io
    import csv as _csv

    d0_str = request.args.get("d0", date.today().isoformat())
    fmt = (request.args.get("fmt", "csv") or "csv").strip().lower()
    if fmt not in {"csv", "xlsx"}:
        abort(400)
    d0 = _parse_date_strict(d0_str)
    if d0 is None:
        abort(400)
    d1 = d0 + timedelta(days=6)

    dias_data = []
    totais = {
        k: 0
        for k in [
            "pa",
            "lan",
            "alm_norm",
            "alm_veg",
            "alm_dieta",
            "jan_norm",
            "jan_veg",
            "jan_dieta",
            "jan_sai",
        ]
    }
    _exp_map, _exp_empty = sr.get_totais_periodo(d0.isoformat(), d1.isoformat())
    _exp_cal = sr.dias_operacionais_batch(d0, d1)
    for i in range(7):
        di = d0 + timedelta(days=i)
        t = _exp_map.get(di.isoformat(), _exp_empty)
        tipo = _exp_cal.get(
            di.isoformat(), "fim_semana" if di.weekday() >= 5 else "normal"
        )
        alm = t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]
        jan = t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]
        dias_data.append((di, tipo, t, alm, jan))
        for k in totais:
            totais[k] += t[k]

    HEADERS = [
        "Data",
        "Dia da Semana",
        "Tipo Dia",
        "PA",
        "Lanche",
        "Alm. Normal",
        "Alm. Veg.",
        "Alm. Dieta",
        "Total Almoços",
        "Jan. Normal",
        "Jan. Veg.",
        "Jan. Dieta",
        "Total Jantares",
        "Sai Unidade",
    ]

    def make_row(di, tipo, t, alm, jan):
        return [
            di.isoformat(),
            NOMES_DIAS[di.weekday()],
            tipo,
            t["pa"],
            t["lan"],
            t["alm_norm"],
            t["alm_veg"],
            t["alm_dieta"],
            alm,
            t["jan_norm"],
            t["jan_veg"],
            t["jan_dieta"],
            jan,
            t["jan_sai"],
        ]

    nome = f"relatorio_{d0_str}_a_{d1.isoformat()}"

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Relatório {d0.strftime('%d-%m')} a {d1.strftime('%d-%m-%Y')}"

            header_fill = PatternFill("solid", fgColor="0A2D4E")
            header_font = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col, h in enumerate(HEADERS, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")
                c.border = border

            TIPO_CORES = {
                "feriado": "FFD6D6",
                "exercicio": "FFFACD",
                "fim_semana": "DDEEFF",
                "normal": "FFFFFF",
                "outro": "F0F0F0",
            }
            for ri, (di, tipo, t, alm, jan) in enumerate(dias_data, 2):
                row_fill = PatternFill("solid", fgColor=TIPO_CORES.get(tipo, "FFFFFF"))
                for col, val in enumerate(make_row(di, tipo, t, alm, jan), 1):
                    c = ws.cell(row=ri, column=col, value=val)
                    c.fill = row_fill
                    c.border = border
                    c.alignment = Alignment(horizontal="center" if col > 2 else "left")

            # Linha de totais
            total_alm = totais["alm_norm"] + totais["alm_veg"] + totais["alm_dieta"]
            total_jan = totais["jan_norm"] + totais["jan_veg"] + totais["jan_dieta"]
            total_row = [
                "TOTAL",
                "—",
                "—",
                totais["pa"],
                totais["lan"],
                totais["alm_norm"],
                totais["alm_veg"],
                totais["alm_dieta"],
                total_alm,
                totais["jan_norm"],
                totais["jan_veg"],
                totais["jan_dieta"],
                total_jan,
                totais["jan_sai"],
            ]
            total_fill = PatternFill("solid", fgColor="D5E8F0")
            total_font = Font(bold=True)
            for col, val in enumerate(total_row, 1):
                c = ws.cell(row=9, column=col, value=val)
                c.fill = total_fill
                c.font = total_font
                c.border = border
                c.alignment = Alignment(horizontal="center" if col > 2 else "left")

            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col) + 3
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 20)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(
                buf.read(),
                headers={
                    "Content-Disposition": f"attachment; filename={nome}.xlsx",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
            )
        except ImportError:
            flash("openpyxl não instalado — a exportar CSV.", "warn")
        except Exception as ex:
            flash(f"Erro ao gerar Excel: {ex} — a exportar CSV.", "warn")

    # CSV (com BOM para Excel abrir correctamente)
    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=";")
    writer.writerow(HEADERS)
    for di, tipo, t, alm, jan in dias_data:
        writer.writerow(make_row(di, tipo, t, alm, jan))
    total_alm = totais["alm_norm"] + totais["alm_veg"] + totais["alm_dieta"]
    total_jan = totais["jan_norm"] + totais["jan_veg"] + totais["jan_dieta"]
    writer.writerow(
        [
            "TOTAL",
            "—",
            "—",
            totais["pa"],
            totais["lan"],
            totais["alm_norm"],
            totais["alm_veg"],
            totais["alm_dieta"],
            total_alm,
            totais["jan_norm"],
            totais["jan_veg"],
            totais["jan_dieta"],
            total_jan,
            totais["jan_sai"],
        ]
    )
    csv_bytes = ("\ufeff" + buf.getvalue()).encode("utf-8")
    return Response(
        csv_bytes,
        headers={
            "Content-Disposition": f"attachment; filename={nome}.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


# ═══════════════════════════════════════════════════════════════════════════
# CONTROLO DE PRESENÇAS — Módulo rápido via NI (Oficial de Dia)
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/presencas", methods=["GET", "POST"])
@role_required("oficialdia", "admin", "cmd")
def controlo_presencas():
    u = current_user()
    hoje = date.today()
    d_str = request.args.get("d", hoje.isoformat())
    dt = _parse_date(d_str)

    resultado = None
    ni_q = ""

    if request.method == "POST":
        acao = request.form.get("acao", "")
        ni_q = request.form.get("ni", "").strip()

        if acao == "consultar" and ni_q:
            with sr.db() as conn:
                aluno = conn.execute(
                    "SELECT id,NII,NI,Nome_completo,ano,email,telemovel FROM utilizadores WHERE NI=? AND perfil='aluno'",
                    (ni_q,),
                ).fetchone()
            if aluno:
                aluno = dict(aluno)
                uid = aluno["id"]
                ausente = _tem_ausencia_ativa(uid, dt)
                with sr.db() as conn:
                    ref = conn.execute(
                        "SELECT * FROM refeicoes WHERE utilizador_id=? AND data=?",
                        (uid, dt.isoformat()),
                    ).fetchone()
                    ref = dict(ref) if ref else {}
                    lic = conn.execute(
                        "SELECT tipo, hora_saida, hora_entrada FROM licencas WHERE utilizador_id=? AND data=?",
                        (uid, dt.isoformat()),
                    ).fetchone()
                    lic = dict(lic) if lic else {}
                resultado = {
                    "aluno": aluno,
                    "ausente": ausente,
                    "ref": ref,
                    "ni": ni_q,
                    "licenca": lic,
                }
            else:
                flash(f'NI "{ni_q}" não encontrado.', "error")

        elif acao == "dar_saida" and ni_q:
            with sr.db() as conn:
                aluno = conn.execute(
                    "SELECT id,NII,Nome_completo FROM utilizadores WHERE NI=? AND perfil='aluno'",
                    (ni_q,),
                ).fetchone()
            if aluno:
                aluno = dict(aluno)
                _registar_ausencia(
                    aluno["id"],
                    dt.isoformat(),
                    dt.isoformat(),
                    f"Saída registada por {u['nome']} ({u['perfil']})",
                    u["nii"],
                )
                # Sincronizar hora_saida na licença (se existir)
                agora = datetime.now().strftime("%H:%M")
                with sr.db() as conn:
                    conn.execute(
                        "UPDATE licencas SET hora_saida=? WHERE utilizador_id=? AND data=? AND hora_saida IS NULL",
                        (agora, aluno["id"], dt.isoformat()),
                    )
                    conn.commit()
                flash(
                    f"✅ Saída registada para {aluno['Nome_completo']} (NI {ni_q}).",
                    "ok",
                )
            else:
                flash(f'NI "{ni_q}" não encontrado.', "error")

        elif acao == "dar_entrada" and ni_q:
            with sr.db() as conn:
                aluno = conn.execute(
                    "SELECT id,NII,Nome_completo FROM utilizadores WHERE NI=? AND perfil='aluno'",
                    (ni_q,),
                ).fetchone()
            if aluno:
                aluno = dict(aluno)
                with sr.db() as conn:
                    conn.execute(
                        """DELETE FROM ausencias WHERE utilizador_id=?
                                    AND ausente_de=? AND ausente_ate=?""",
                        (aluno["id"], dt.isoformat(), dt.isoformat()),
                    )
                    conn.commit()
                # Sincronizar hora_entrada na licença (se existir)
                agora = datetime.now().strftime("%H:%M")
                with sr.db() as conn:
                    conn.execute(
                        "UPDATE licencas SET hora_entrada=? WHERE utilizador_id=? AND data=? AND hora_entrada IS NULL",
                        (agora, aluno["id"], dt.isoformat()),
                    )
                    conn.commit()
                flash(
                    f"✅ Entrada registada para {aluno['Nome_completo']} (NI {ni_q}).",
                    "ok",
                )
            else:
                flash(f'NI "{ni_q}" não encontrado.', "error")

        # Após POST sem resultado, redirecionar limpo
        if resultado is None:
            return redirect(url_for("controlo_presencas", d=dt.isoformat()))

    # Resumo de todos os anos para a data
    anos_resumo = []
    for ano in _get_anos_disponiveis():
        with sr.db() as conn:
            total = conn.execute(
                "SELECT COUNT(*) c FROM utilizadores WHERE ano=? AND perfil='aluno'",
                (ano,),
            ).fetchone()["c"]
            ausentes_a = conn.execute(
                """
                SELECT COUNT(*) c FROM utilizadores u
                WHERE u.ano=? AND u.perfil='aluno'
                AND EXISTS(SELECT 1 FROM ausencias a WHERE a.utilizador_id=u.id
                           AND a.ausente_de<=? AND a.ausente_ate>=?)""",
                (ano, dt.isoformat(), dt.isoformat()),
            ).fetchone()["c"]
            com_ref = conn.execute(
                """
                SELECT COUNT(*) c FROM utilizadores u
                WHERE u.ano=? AND u.perfil='aluno'
                AND EXISTS(SELECT 1 FROM refeicoes r WHERE r.utilizador_id=u.id
                           AND r.data=? AND (r.almoco IS NOT NULL OR r.jantar_tipo IS NOT NULL))""",
                (ano, dt.isoformat()),
            ).fetchone()["c"]
        anos_resumo.append(
            {
                "ano": ano,
                "total": total,
                "ausentes": ausentes_a,
                "presentes": total - ausentes_a,
                "com_ref": com_ref,
            }
        )

    resumo_html = ""
    for r in anos_resumo:
        resumo_html += f"""
        <div class="stat-box" style="cursor:pointer" onclick="window.location='{url_for("lista_alunos_ano", ano=r["ano"], d=dt.isoformat())}'">
          <div class="stat-num">{r["presentes"]} <small style="font-size:.6em;color:var(--muted)">/ {r["total"]}</small></div>
          <div class="stat-lbl">{_ano_label(r["ano"])} — Presentes</div>
          <div style="margin-top:.35rem;font-size:.75rem">
            <span style="color:var(--warn)">✖ {r["ausentes"]} ausentes</span> &nbsp;
            <span style="color:var(--ok)">🍽 {r["com_ref"]} c/ refeições</span>
          </div>
        </div>"""

    # Resultado da pesquisa
    resultado_html = ""
    if resultado:
        al = resultado["aluno"]
        ref = resultado["ref"]
        ausente = resultado["ausente"]
        ni_val = resultado["ni"]
        lic_info = resultado.get("licenca", {})

        estado_cor = "#fdecea" if ausente else "#d5f5e3"
        estado_txt = "🔴 AUSENTE" if ausente else "🟢 PRESENTE"

        # Licença badge
        lic_badge = ""
        if lic_info.get("tipo") == "antes_jantar":
            lic_badge = '<div style="margin-top:.4rem"><span class="badge badge-info">🌅 Licença antes do jantar</span>'
            if lic_info.get("hora_saida"):
                lic_badge += f' <span class="text-muted small">Saiu: {lic_info["hora_saida"]}</span>'
            if lic_info.get("hora_entrada"):
                lic_badge += f' <span class="text-muted small">Entrou: {lic_info["hora_entrada"]}</span>'
            lic_badge += "</div>"
        elif lic_info.get("tipo") == "apos_jantar":
            lic_badge = '<div style="margin-top:.4rem"><span class="badge badge-muted">🌙 Licença após o jantar</span>'
            if lic_info.get("hora_saida"):
                lic_badge += f' <span class="text-muted small">Saiu: {lic_info["hora_saida"]}</span>'
            if lic_info.get("hora_entrada"):
                lic_badge += f' <span class="text-muted small">Entrou: {lic_info["hora_entrada"]}</span>'
            lic_badge += "</div>"

        def ref_chip(val, label, tipo=None):
            if val:
                txt = tipo if tipo else "✓"
                return f'<span style="background:#eafaf1;border:1.5px solid #a9dfbf;border-radius:7px;padding:.25rem .55rem;font-size:.8rem;font-weight:700">{label} {txt}</span>'
            return f'<span style="background:#fdecea;border:1.5px solid #f1948a;border-radius:7px;padding:.25rem .55rem;font-size:.8rem;color:var(--muted)">{label} ✗</span>'

        acao_presenca = f"""
        <div style="display:flex;gap:.5rem;flex-wrap:wrap;margin-top:.8rem">
          {
            ""
            if not ausente
            else f'''
          <form method="post">
            {csrf_input()}
            <input type="hidden" name="acao" value="dar_entrada">
            <input type="hidden" name="ni" value="{esc(ni_val)}">
            <button class="btn btn-ok">✅ Dar Entrada (marcar presente)</button>
          </form>'''
        }
          {
            ""
            if ausente
            else f'''
          <form method="post" onsubmit="return confirm('Confirmar saída de {esc(al["Nome_completo"])}?')">
            {csrf_input()}
            <input type="hidden" name="acao" value="dar_saida">
            <input type="hidden" name="ni" value="{esc(ni_val)}">
            <button class="btn btn-danger">🚪 Dar Saída (marcar ausente)</button>
          </form>'''
        }
          <a class="btn btn-ghost" href="{
            url_for("ver_perfil_aluno", nii=al["NII"], ano=al["ano"], d=dt.isoformat())
        }">👁 Ver perfil completo</a>
        </div>"""

        resultado_html = f"""
        <div class="card" style="border-left:4px solid {"var(--danger)" if ausente else "var(--ok)"}">
          <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:.5rem">
            <div>
              <div style="font-size:1.15rem;font-weight:800">{esc(al["Nome_completo"])}</div>
              <div class="text-muted small">NI: <strong>{esc(al["NI"])}</strong> &nbsp;|&nbsp; {al["ano"]}º Ano &nbsp;|&nbsp; NII: {esc(al["NII"])}</div>
              {lic_badge}
            </div>
            <div style="background:{estado_cor};padding:.4rem .9rem;border-radius:20px;font-weight:800;font-size:1rem">{estado_txt}</div>
          </div>
          <hr style="margin:.7rem 0">
          <div class="card-title" style="font-size:.82rem;margin-bottom:.5rem">🍽️ Refeições em {dt.strftime("%d/%m/%Y")}</div>
          <div style="display:flex;gap:.4rem;flex-wrap:wrap">
            {ref_chip(ref.get("pequeno_almoco"), "☕ PA")}
            {ref_chip(ref.get("lanche"), "🥐 Lanche")}
            {ref_chip(ref.get("almoco"), "🍽️ Almoço", ref.get("almoco", "")[:3] if ref.get("almoco") else None)}
            {ref_chip(ref.get("jantar_tipo"), "🌙 Jantar", ref.get("jantar_tipo", "")[:3] if ref.get("jantar_tipo") else None)}
            {'<span style="background:#fef9e7;border:1.5px solid #f9e79f;border-radius:7px;padding:.25rem .55rem;font-size:.8rem">🚪 Sai</span>' if ref.get("jantar_sai_unidade") else ""}
          </div>
          {acao_presenca}
        </div>"""

    prev_d = (dt - timedelta(days=1)).isoformat()
    next_d = (dt + timedelta(days=1)).isoformat()

    content = f"""
    <div class="container">
      <div class="page-header">
        {_back_btn(url_for("painel_dia", d=dt.isoformat()), "Painel")}
        <div class="page-title">🎯 Controlo de Presenças</div>
      </div>

      <!-- Navegação de datas -->
      <div class="card" style="padding:.75rem 1.1rem;margin-bottom:.8rem">
        <div class="flex-between">
          <div class="flex">
            <a class="btn btn-ghost btn-sm" href="{url_for("controlo_presencas", d=prev_d)}">← Anterior</a>
            <strong>{NOMES_DIAS[dt.weekday()]}, {dt.strftime("%d/%m/%Y")}</strong>
            <a class="btn btn-ghost btn-sm" href="{url_for("controlo_presencas", d=next_d)}">Próximo →</a>
          </div>
          <form method="get" style="display:flex;gap:.3rem">
            <input type="date" name="d" value="{dt.isoformat()}" style="width:auto">
            <button class="btn btn-primary btn-sm">Ir</button>
          </form>
        </div>
      </div>

      <!-- Pesquisa rápida por NI -->
      <div class="card" style="border-top:3px solid var(--primary)">
        <div class="card-title">🔍 Pesquisa rápida por NI</div>
        <div class="alert alert-info" style="margin-bottom:.8rem;font-size:.82rem">
          💡 Introduz o NI do aluno (ex: <strong>222</strong>) para consultar o estado de presença e refeições. Podes depois dar entrada ou saída diretamente.
        </div>
        <form method="post" style="display:flex;gap:.5rem;flex-wrap:wrap">
          {csrf_input()}
          <input type="hidden" name="acao" value="consultar">
          <input type="text" name="ni" value="{esc(ni_q)}" placeholder="NI do aluno (ex: 222)"
            style="flex:1;min-width:140px;font-size:1.05rem;font-weight:700;letter-spacing:.05em"
            autofocus autocomplete="off">
          <button class="btn btn-primary" style="font-size:1rem">🔍 Consultar</button>
        </form>
      </div>

      {resultado_html}

      <!-- Botão Licenças / Entradas-Saídas -->
      <div class="card" style="border-top:3px solid var(--gold)">
        <div class="flex-between">
          <div>
            <div class="card-title" style="margin-bottom:.2rem">🚪 Licenças e Entradas/Saídas</div>
            <span class="text-muted small">Ver quem tem licença para hoje e registar entradas/saídas</span>
          </div>
          <a class="btn btn-gold" href="{url_for("licencas_entradas_saidas", d=dt.isoformat())}">Abrir painel</a>
        </div>
      </div>

      <!-- Resumo por ano -->
      <div class="card">
        <div class="card-title">📊 Resumo geral — {dt.strftime("%d/%m/%Y")}</div>
        <div class="grid grid-3">{resumo_html or '<div class="text-muted">Sem dados.</div>'}</div>
        <div style="margin-top:.6rem;font-size:.8rem;color:var(--muted)">Clica num ano para ver a lista completa.</div>
      </div>
    </div>"""
    return render(content)


# ═══════════════════════════════════════════════════════════════════════════
# GESTÃO DE COMPANHIAS — Unificação de Turmas + Promoção de Alunos
# ═══════════════════════════════════════════════════════════════════════════


@app.route("/admin/companhias", methods=["GET", "POST"])
@role_required("admin")
def admin_companhias():
    if request.method == "POST":
        acao = request.form.get("acao", "")

        # ── Criar turma ──────────────────────────────────────────────────
        if acao == "criar_turma":
            nome_turma = _val_text(request.form.get("nome_turma", ""), 100)
            ano_turma = request.form.get("ano_turma", "").strip()
            descricao = _val_text(request.form.get("descricao", ""), 200)
            if not nome_turma or not ano_turma:
                flash("Nome e ano são obrigatórios.", "error")
            else:
                try:
                    ano_int = _val_ano(ano_turma)
                    if ano_int is None:
                        flash("Ano inválido (0-8).", "error")
                        return redirect(url_for("admin_companhias"))
                    with sr.db() as conn:
                        conn.execute(
                            "INSERT INTO turmas (nome, ano, descricao) VALUES (?,?,?)",
                            (nome_turma, ano_int, descricao or None),
                        )
                        conn.commit()
                    flash(
                        f'Turma "{nome_turma}" ({_ano_label(ano_int)}) criada com sucesso!',
                        "ok",
                    )
                except Exception as ex:
                    flash(f"Erro ao criar turma: {ex}", "error")

        # ── Eliminar turma ───────────────────────────────────────────────
        elif acao == "eliminar_turma":
            tid = _val_int_id(request.form.get("tid", ""))
            if tid is None:
                flash("ID de turma inválido.", "error")
                return redirect(url_for("admin_companhias"))
            try:
                with sr.db() as conn:
                    # Desassociar alunos antes de eliminar
                    conn.execute(
                        "UPDATE utilizadores SET turma_id=NULL WHERE turma_id=?",
                        (tid,),
                    )
                    conn.execute("DELETE FROM turmas WHERE id=?", (tid,))
                    conn.commit()
                flash("Turma eliminada.", "ok")
            except Exception as ex:
                flash(f"Erro: {ex}", "error")

        # ── Atribuir aluno a turma ──────────────────────────────────────
        elif acao == "atribuir_turma":
            nii_at = request.form.get("nii_at", "").strip()
            tid_at = request.form.get("turma_id", "").strip()
            if nii_at:
                try:
                    turma_val = int(tid_at) if tid_at else None
                    with sr.db() as conn:
                        conn.execute(
                            "UPDATE utilizadores SET turma_id=? WHERE NII=? AND perfil='aluno'",
                            (turma_val, nii_at),
                        )
                        conn.commit()
                    flash(
                        f"Turma do aluno {nii_at} atualizada.",
                        "ok",
                    )
                except Exception as ex:
                    flash(f"Erro: {ex}", "error")

        # ── Mover aluno de ano ───────────────────────────────────────────
        elif acao == "mover_aluno":
            nii_m = _val_nii(request.form.get("nii_m", ""))
            novo_ano_v = _val_ano(request.form.get("novo_ano", ""))
            if not nii_m:
                flash("NII inválido.", "error")
            elif novo_ano_v is None:
                flash("Ano inválido (0-8).", "error")
            else:
                try:
                    with sr.db() as conn:
                        conn.execute(
                            "UPDATE utilizadores SET ano=? WHERE NII=? AND perfil='aluno'",
                            (novo_ano_v, nii_m),
                        )
                        conn.commit()
                    flash(f"Aluno {nii_m} movido para {_ano_label(novo_ano_v)}.", "ok")
                except Exception as ex:
                    flash(f"Erro: {ex}", "error")

        # ── Promover aluno individual ─────────────────────────────────────
        elif acao == "promover_um":
            uid_p = _val_int_id(request.form.get("uid", ""))
            novo_ni = _val_ni(request.form.get("novo_ni", ""))
            if uid_p is None:
                flash("ID inválido.", "error")
                return redirect(url_for("admin_companhias"))
            with sr.db() as conn:
                al = conn.execute(
                    "SELECT ano,NI FROM utilizadores WHERE id=?", (uid_p,)
                ).fetchone()
            if al:
                ano_a = al["ano"]
                # CFBO(7) e CFCO(8) não têm progressão automática para acima
                if ano_a >= 6:
                    novo_ano_p = 0
                else:
                    novo_ano_p = ano_a + 1
                with sr.db() as conn:
                    conn.execute(
                        "UPDATE utilizadores SET ano=?,NI=? WHERE id=?",
                        (novo_ano_p, novo_ni or al["NI"], uid_p),
                    )
                    conn.commit()
                dest = _ano_label(novo_ano_p) if novo_ano_p else "Concluído"
                flash(f"Aluno promovido para {dest}.", "ok")

        # ── Promoção global de um ano ─────────────────────────────────────
        elif acao == "promover_todos":
            ano_origem = _val_ano(request.form.get("ano_origem", 0))
            if ano_origem is None:
                flash("Ano de origem inválido.", "error")
                return redirect(url_for("admin_companhias"))
            if ano_origem >= 6:
                novo_ano_p = 0
            else:
                novo_ano_p = ano_origem + 1
            with sr.db() as conn:
                conn.execute(
                    "UPDATE utilizadores SET ano=? WHERE perfil='aluno' AND ano=?",
                    (novo_ano_p, ano_origem),
                )
                conn.commit()
            dest = _ano_label(novo_ano_p) if novo_ano_p else "Concluído"
            flash(
                f"Todos os alunos do {_ano_label(ano_origem)} promovidos para {dest}.",
                "ok",
            )

        # ── Promoção global todos os anos ────────────────────────────────
        elif acao == "promover_todos_anos":
            with sr.db() as conn:
                # Promover do maior para o menor para evitar conflitos
                for ano_a in range(6, 0, -1):
                    novo_ano_p = 0 if ano_a >= 6 else ano_a + 1
                    conn.execute(
                        "UPDATE utilizadores SET ano=? WHERE perfil='aluno' AND ano=?",
                        (novo_ano_p, ano_a),
                    )
                conn.commit()
            flash("Promoção global concluída.", "ok")

        return redirect(url_for("admin_companhias"))

    # ── Carregar dados ───────────────────────────────────────────────────
    try:
        with sr.db() as conn:
            turmas = [
                dict(r)
                for r in conn.execute(
                    "SELECT * FROM turmas ORDER BY ano, nome"
                ).fetchall()
            ]
    except Exception:
        turmas = []

    # Contagens por ano (inclui CFBO e CFCO)
    anos_data = {}
    all_anos = list(range(1, 7)) + [7, 8]
    for a in all_anos:
        with sr.db() as conn:
            cnt = conn.execute(
                "SELECT COUNT(*) c FROM utilizadores WHERE ano=? AND perfil='aluno'",
                (a,),
            ).fetchone()["c"]
        anos_data[a] = cnt

    # ── HTML alunos por ano ───────────────────────────────────────────────
    anos_grid = ""
    for a in all_anos:
        n = anos_data.get(a, 0)
        anos_grid += f'<div class="stat-box"><div class="stat-num">{n}</div><div class="stat-lbl">{_ano_label(a)}</div></div>'

    # ── HTML promoção ─────────────────────────────────────────────────────
    def _build_promover_html():
        cards = ""
        promovable = list(range(1, 7)) + [7, 8]
        for a in promovable:
            with sr.db() as conn:
                alunos_a = [
                    dict(r)
                    for r in conn.execute(
                        "SELECT id,NI,Nome_completo,ano FROM utilizadores WHERE perfil='aluno' AND ano=? ORDER BY NI",
                        (a,),
                    ).fetchall()
                ]
            n = len(alunos_a)
            if a >= 6:
                destino = "Concluído"
                cor = "#922b21"
            else:
                destino = _ano_label(a + 1)
                cor = "#1e8449"
            alunos_list = "".join(
                '<div style="display:flex;justify-content:space-between;align-items:center;padding:.3rem 0;border-bottom:1px solid var(--border);font-size:.82rem;gap:.4rem">'
                "<span><strong>"
                + esc(al["NI"])
                + "</strong> — "
                + esc(al["Nome_completo"])
                + "</span>"
                '<form method="post" style="display:inline;display:flex;gap:.3rem;align-items:center">'
                + str(csrf_input())
                + '<input type="hidden" name="acao" value="promover_um">'
                '<input type="hidden" name="uid" value="' + str(al["id"]) + '">'
                '<input type="text" name="novo_ni" placeholder="Novo NI" style="width:110px;padding:.25rem .45rem;font-size:.78rem;border-radius:7px;border:1.5px solid var(--border)">'
                '<button class="btn btn-ghost btn-sm" title="Promover este aluno">↑ Promover</button>'
                "</form></div>"
                for al in alunos_a
            )
            disabled = " disabled" if not alunos_a else ""
            cards += (
                '<div class="card" style="border-top:3px solid ' + cor + '">'
                '<div class="card-title" style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:.4rem">'
                "<span>"
                + _ano_label(a)
                + ' <span class="badge badge-info" style="margin-left:.4rem">'
                + str(n)
                + " alunos</span></span>"
                '<form method="post" style="display:inline" onsubmit="return confirm(\'Promover todos os alunos deste ano?\')">'
                + str(csrf_input())
                + '<input type="hidden" name="acao" value="promover_todos"><input type="hidden" name="ano_origem" value="'
                + str(a)
                + '">'
                '<button class="btn btn-sm" style="background:'
                + cor
                + ';color:#fff"'
                + disabled
                + ">🎖️ Promover todos → "
                + destino
                + "</button></form></div>"
                '<div style="max-height:180px;overflow-y:auto;border-top:1px solid var(--border);padding-top:.4rem">'
                + (
                    alunos_list
                    or '<div class="text-muted small" style="padding:.3rem">Sem alunos.</div>'
                )
                + "</div></div>"
            )
        return cards

    anos_cards_prom = _build_promover_html()

    # ── HTML turmas criadas ───────────────────────────────────────────────
    turmas_html = ""
    for t in turmas:
        turmas_html += f"""
        <tr>
          <td><strong>{esc(t["nome"])}</strong></td>
          <td>{_ano_label(t["ano"])}</td>
          <td>{esc(t.get("descricao") or "—")}</td>
          <td class="small text-muted">{(t.get("criado_em") or "")[:16]}</td>
          <td>
            <form method="post" style="display:inline" onsubmit="return confirm('Eliminar turma?')">
              {csrf_input()}
              <input type="hidden" name="acao" value="eliminar_turma">
              <input type="hidden" name="tid" value="{t["id"]}">
              <button class="btn btn-danger btn-sm">🗑</button>
            </form>
          </td>
        </tr>"""

    # ── Alunos para mover / atribuir turma ──────────────────────────────
    with sr.db() as conn:
        alunos_all = conn.execute(
            "SELECT NII, NI, Nome_completo, ano, turma_id FROM utilizadores WHERE perfil='aluno' ORDER BY ano, NI"
        ).fetchall()
    alunos_opts = "".join(
        f'<option value="{esc(a["NII"])}">[{_ano_label(a["ano"])}] {esc(a["NI"])} — {esc(a["Nome_completo"])}</option>'
        for a in alunos_all
    )
    turma_select_opts = '<option value="">— Sem turma —</option>' + "".join(
        f'<option value="{t["id"]}">{esc(t["nome"])} ({_ano_label(t["ano"])})</option>'
        for t in turmas
    )

    ano_select_opts = "".join(
        f'<option value="{a}">{_ano_label(a)}</option>' for a, _ in ANOS_OPCOES
    )
    ano_select_criar = "".join(
        f'<option value="{a}">{lbl}</option>' for a, lbl in ANOS_OPCOES
    )
    ano_select_mover = (
        ano_select_opts + '<option value="0">Concluído / Inativo</option>'
    )

    # ── Tabs de secção (via hash) ─────────────────────────────────────────
    content = f"""
    <div class="container">
      <div class="page-header">
        {_back_btn(url_for("admin_home"))}
        <div class="page-title">⚓ Gestão de Companhias</div>
      </div>

      <!-- Tabs -->
      <div class="year-tabs" style="margin-bottom:1rem">
        <a class="year-tab" href="#turmas" onclick="showTab('turmas')">📚 Turmas</a>
        <a class="year-tab" href="#atribuir" onclick="showTab('atribuir')">👥 Atribuir Turma</a>
        <a class="year-tab" href="#promocao" onclick="showTab('promocao')">🎖️ Promoção</a>
        <a class="year-tab" href="#mover" onclick="showTab('mover')">🔄 Mover Aluno</a>
      </div>
      <script>
        function showTab(id) {{
          ['turmas','atribuir','promocao','mover'].forEach(function(t) {{
            document.getElementById('tab-'+t).style.display = (t===id) ? '' : 'none';
          }});
          document.querySelectorAll('.year-tab').forEach(function(el) {{
            el.classList.toggle('active', el.getAttribute('href')==='#'+id);
          }});
        }}
        // Mostrar tab por hash ou primeiro
        var hash = window.location.hash.replace('#','') || 'turmas';
        showTab(hash);
      </script>

      <!-- Tab: Turmas -->
      <div id="tab-turmas">
        <div class="card">
          <div class="card-title">📊 Alunos por ano/curso</div>
          <div class="grid grid-4">{anos_grid}</div>
        </div>
        <div class="grid grid-2">
          <div class="card">
            <div class="card-title">➕ Criar nova turma / companhia</div>
            <form method="post">
              {csrf_input()}
              <input type="hidden" name="acao" value="criar_turma">
              <div class="form-group">
                <label>Nome da turma <span class="text-muted small">(ex: Alpha, Bravo...)</span></label>
                <input type="text" name="nome_turma" required placeholder="Ex: Alpha">
              </div>
              <div class="form-group">
                <label>Ano curricular / Curso</label>
                <select name="ano_turma" required>
                  {ano_select_criar}
                </select>
              </div>
              <div class="form-group">
                <label>Descrição <span class="text-muted small">(opcional)</span></label>
                <input type="text" name="descricao" placeholder="Ex: Turma de engenharia naval">
              </div>
              <button class="btn btn-ok">💾 Criar turma</button>
            </form>
          </div>
          <div class="card">
            <div class="card-title">📋 Turmas criadas ({len(turmas)})</div>
            {'<div class="table-wrap"><table><thead><tr><th>Nome</th><th>Ano/Curso</th><th>Descrição</th><th>Criada em</th><th></th></tr></thead><tbody>' + turmas_html + "</tbody></table></div>" if turmas else '<div class="text-muted" style="padding:.8rem">Nenhuma turma criada ainda.</div>'}
          </div>
        </div>
      </div>

      <!-- Tab: Atribuir Turma -->
      <div id="tab-atribuir" style="display:none">
        <div class="card" style="max-width:520px">
          <div class="card-title">👥 Atribuir aluno a uma turma</div>
          <div class="alert alert-info" style="font-size:.81rem;margin-bottom:.8rem">
            💡 Liga um aluno a uma turma/companhia criada no separador Turmas.
          </div>
          <form method="post">
            {csrf_input()}
            <input type="hidden" name="acao" value="atribuir_turma">
            <div class="form-group">
              <label>Aluno (NII)</label>
              <select name="nii_at" required>
                <option value="">— Selecionar aluno —</option>
                {alunos_opts}
              </select>
            </div>
            <div class="form-group">
              <label>Turma</label>
              <select name="turma_id">
                {turma_select_opts}
              </select>
            </div>
            <button class="btn btn-ok">💾 Atribuir turma</button>
          </form>
        </div>
      </div>

      <!-- Tab: Promoção -->
      <div id="tab-promocao" style="display:none">
        <div class="alert alert-warn">⚠️ <strong>Atenção:</strong> A promoção é permanente. Recomenda-se fazer backup antes.</div>
        <div class="card">
          <div class="card-title">🚀 Promoção global — todos os anos em simultâneo</div>
          <p style="font-size:.85rem;color:var(--muted);margin-bottom:.8rem">Promove todos: 1º→2º, 2º→3º, ..., 5º→6º, 6º→Concluído. CFBO e CFCO não são afetados pela promoção global.</p>
          <form method="post" onsubmit="return confirm('Promover TODOS os alunos de todos os anos?')">
            {csrf_input()}<input type="hidden" name="acao" value="promover_todos_anos">
            <button class="btn btn-danger">🎖️ Promoção Global</button>
          </form>
        </div>
        <div class="grid grid-2">{anos_cards_prom}</div>
      </div>

      <!-- Tab: Mover Aluno -->
      <div id="tab-mover" style="display:none">
        <div class="card" style="max-width:520px">
          <div class="card-title">🔄 Mover aluno de ano</div>
          <div class="alert alert-info" style="font-size:.81rem;margin-bottom:.8rem">
            💡 Usa esta função para mover um aluno individualmente para outro ano sem usar a promoção global, incluindo para os cursos CFBO e CFCO.
          </div>
          <form method="post">
            {csrf_input()}
            <input type="hidden" name="acao" value="mover_aluno">
            <div class="form-group">
              <label>Aluno (NII)</label>
              <select name="nii_m" required>
                <option value="">— Selecionar aluno —</option>
                {alunos_opts}
              </select>
            </div>
            <div class="form-group">
              <label>Mover para</label>
              <select name="novo_ano" required>
                {ano_select_mover}
              </select>
            </div>
            <button class="btn btn-warn">🔄 Mover aluno</button>
          </form>
        </div>
      </div>
    </div>"""
    return render(content)


# Rota de compatibilidade — redireciona para o novo módulo
@app.route("/admin/turmas")
@role_required("admin")
def admin_turmas():
    return redirect(url_for("admin_companhias"))


@app.route("/admin/promover", methods=["GET", "POST"])
@role_required("admin")
def admin_promover():
    return redirect(url_for("admin_companhias") + "#promocao")


@app.route("/health")
def health():
    """Health check — verifica BD e devolve JSON + 200 (ou 503 se falhar)."""
    import time as _time

    t0 = _time.monotonic()
    try:
        with sr.db() as conn:
            row = conn.execute("SELECT COUNT(*) as n FROM utilizadores").fetchone()
        latency_ms = round((_time.monotonic() - t0) * 1000, 1)
        resp = {
            "status": "ok",
            "db": "ok",
            "latency_ms": latency_ms,
            "ts": datetime.now().isoformat(),
        }
        if not _is_production:
            resp["utilizadores"] = row["n"]
            resp["db_path"] = sr.BASE_DADOS
        return resp, 200
    except Exception as exc:
        app.logger.error(f"health: BD falhou — {exc}")
        return {
            "status": "error",
            "db": "error",
            "ts": datetime.now().isoformat(),
        }, 503


@app.route("/admin/backup-download")
@login_required
@role_required("admin")
def admin_backup_download():
    """Permite ao admin descarregar o ficheiro da base de dados."""
    from pathlib import Path

    db_path = Path(sr.BASE_DADOS)
    if not db_path.exists():
        flash("Ficheiro da base de dados não encontrado.", "error")
        return redirect(url_for("admin_home"))
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    nome = f"{db_path.stem}_{ts}.db"
    return send_file(
        db_path,
        as_attachment=True,
        download_name=nome,
        mimetype="application/x-sqlite3",
    )


@app.route("/api/backup-cron", methods=["POST"])
def api_backup_cron():
    """Endpoint para cron job externo invocar backup diário.
    Uso: curl -X POST -H "Authorization: Bearer <CRON_API_TOKEN>" http://host/api/backup-cron
    """
    if not _verify_cron_token():
        abort(403)
    try:
        sr.ensure_daily_backup()
        sr.limpar_backups_antigos()
        return {"status": "ok", "ts": datetime.now().isoformat()}
    except Exception as exc:
        app.logger.error(f"api_backup_cron: {exc}")
        return {"status": "error", "msg": str(exc)}, 500


@app.route("/api/autopreencher-cron", methods=["POST"])
def api_autopreencher_cron():
    """Endpoint para cron externo pré-preencher refeições semanalmente.
    Uso: curl -X POST -H "Authorization: Bearer <CRON_API_TOKEN>" http://host/api/autopreencher-cron
    """
    if not _verify_cron_token():
        abort(403)
    try:
        sr.autopreencher_refeicoes_semanais(DIAS_ANTECEDENCIA)
        return {"status": "ok", "ts": datetime.now().isoformat()}
    except Exception as exc:
        app.logger.error(f"api_autopreencher_cron: {exc}")
        return {"status": "error", "msg": str(exc)}, 500


if __name__ == "__main__":
    _init_app_once()
    # Nota: backup e autopreencher NÃO são chamados aqui para evitar duplicação
    # em ambientes multi-worker (gunicorn). Usa /api/backup-cron e /api/autopreencher-cron
    # via cron externo (Railway Cron / crontab).
    cfg.print_startup_banner(sr.BASE_DADOS)
    app.run(debug=cfg.DEBUG, host="0.0.0.0", port=cfg.PORT)
