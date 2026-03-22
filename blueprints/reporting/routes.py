"""Rotas do blueprint reporting."""

import io
import csv as _csv
from datetime import date, timedelta

from flask import (
    render_template,
    Response,
    abort,
    flash,
    request,
    url_for,
)
from core.database import db
from core.meals import dias_operacionais_batch, get_totais_dia, get_totais_periodo
from blueprints.reporting import report_bp
from utils.auth import (
    current_user,
    login_required,
    role_required,
)
from utils.constants import ABREV_DIAS, NOMES_DIAS
from utils.helpers import (
    _parse_date,
    _parse_date_strict,
)


@report_bp.route("/exportar/mensal")
@role_required("cozinha", "oficialdia", "cmd", "admin")
def exportar_mensal():

    mes = request.args.get("mes", "")
    fmt = (request.args.get("fmt", "csv") or "csv").strip().lower()
    if fmt not in {"csv", "xlsx"}:
        abort(400)

    # Default: mês atual
    hoje = date.today()
    if mes:
        try:
            ano_m, mes_m = mes.split("-")
            d0 = date(int(ano_m), int(mes_m), 1)
        except (ValueError, IndexError):
            abort(400)
    else:
        d0 = date(hoje.year, hoje.month, 1)

    # Último dia do mês
    if d0.month == 12:
        d1 = date(d0.year + 1, 1, 1) - timedelta(days=1)
    else:
        d1 = date(d0.year, d0.month + 1, 1) - timedelta(days=1)

    MESES_PT = {
        1: "Janeiro",
        2: "Fevereiro",
        3: "Março",
        4: "Abril",
        5: "Maio",
        6: "Junho",
        7: "Julho",
        8: "Agosto",
        9: "Setembro",
        10: "Outubro",
        11: "Novembro",
        12: "Dezembro",
    }
    nome_mes = MESES_PT.get(d0.month, str(d0.month))

    dias_data = []
    totais = {
        k: 0
        for k in [
            "pa",
            "lan",
            "alm_norm",
            "alm_veg",
            "alm_dieta",
            "alm_estufa",
            "jan_norm",
            "jan_veg",
            "jan_dieta",
            "jan_sai",
            "jan_estufa",
        ]
    }
    _men_map, _men_empty = get_totais_periodo(d0.isoformat(), d1.isoformat())
    _men_cal = dias_operacionais_batch(d0, d1)
    di = d0
    while di <= d1:
        t = _men_map.get(di.isoformat(), _men_empty)
        tipo = _men_cal.get(
            di.isoformat(), "fim_semana" if di.weekday() >= 5 else "normal"
        )
        alm = t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]
        jan = t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]
        dias_data.append((di, tipo, t, alm, jan))
        for k in totais:
            totais[k] += t[k]
        di += timedelta(days=1)

    HEADERS = [
        "Data",
        "Dia da Semana",
        "Tipo Dia",
        "PA",
        "Lanche",
        "Alm. Normal",
        "Alm. Veg.",
        "Alm. Dieta",
        "Alm. Estufa",
        "Total Almoços",
        "Jan. Normal",
        "Jan. Veg.",
        "Jan. Dieta",
        "Total Jantares",
        "Sai Unidade",
        "Jan. Estufa",
    ]

    def make_row(di, tipo, t, alm, jan):
        return [
            di.isoformat(),
            NOMES_DIAS[di.weekday()],
            tipo,
            t["pa"],
            t["lan"],
            t["alm_norm"],
            t["alm_veg"],
            t["alm_dieta"],
            t.get("alm_estufa", 0),
            alm,
            t["jan_norm"],
            t["jan_veg"],
            t["jan_dieta"],
            jan,
            t["jan_sai"],
            t.get("jan_estufa", 0),
        ]

    nome_ficheiro = f"relatorio_mensal_{d0.strftime('%Y-%m')}"

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{nome_mes} {d0.year}"

            header_fill = PatternFill("solid", fgColor="0A2D4E")
            header_font = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col, h in enumerate(HEADERS, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")
                c.border = border

            tipo_fills = {
                "feriado": PatternFill("solid", fgColor="FFD6D6"),
                "exercicio": PatternFill("solid", fgColor="FFFACD"),
                "fim_semana": PatternFill("solid", fgColor="DDEEFF"),
            }

            for i, (di, tipo, t, alm, jan) in enumerate(dias_data, 2):
                row_data = make_row(di, tipo, t, alm, jan)
                fill = tipo_fills.get(tipo, PatternFill())
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=i, column=col, value=val)
                    c.fill = fill
                    c.border = border
                    c.alignment = Alignment(horizontal="center")

            # Total row
            total_row_idx = len(dias_data) + 2
            total_alm = totais["alm_norm"] + totais["alm_veg"] + totais["alm_dieta"]
            total_jan = totais["jan_norm"] + totais["jan_veg"] + totais["jan_dieta"]
            total_data = [
                "TOTAL",
                "",
                "",
                totais["pa"],
                totais["lan"],
                totais["alm_norm"],
                totais["alm_veg"],
                totais["alm_dieta"],
                totais.get("alm_estufa", 0),
                total_alm,
                totais["jan_norm"],
                totais["jan_veg"],
                totais["jan_dieta"],
                total_jan,
                totais["jan_sai"],
                totais.get("jan_estufa", 0),
            ]
            total_fill = PatternFill("solid", fgColor="D5F5E3")
            total_font = Font(bold=True)
            for col, val in enumerate(total_data, 1):
                c = ws.cell(row=total_row_idx, column=col, value=val)
                c.fill = total_fill
                c.font = total_font
                c.border = border
                c.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col) + 4
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 22)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(
                buf.read(),
                headers={
                    "Content-Disposition": f"attachment; filename={nome_ficheiro}.xlsx",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
            )
        except ImportError:
            fmt = "csv"

    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=";")
    writer.writerow(HEADERS)
    for di, tipo, t, alm, jan in dias_data:
        writer.writerow(make_row(di, tipo, t, alm, jan))
    total_alm = totais["alm_norm"] + totais["alm_veg"] + totais["alm_dieta"]
    total_jan = totais["jan_norm"] + totais["jan_veg"] + totais["jan_dieta"]
    writer.writerow(
        [
            "TOTAL",
            "",
            "",
            totais["pa"],
            totais["lan"],
            totais["alm_norm"],
            totais["alm_veg"],
            totais["alm_dieta"],
            totais.get("alm_estufa", 0),
            total_alm,
            totais["jan_norm"],
            totais["jan_veg"],
            totais["jan_dieta"],
            total_jan,
            totais["jan_sai"],
            totais.get("jan_estufa", 0),
        ]
    )
    csv_bytes = ("\ufeff" + buf.getvalue()).encode("utf-8")
    return Response(
        csv_bytes,
        headers={
            "Content-Disposition": f"attachment; filename={nome_ficheiro}.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


# ═══════════════════════════════════════════════════════════════════════════
# PAINEL OPERACIONAL
# ═══════════════════════════════════════════════════════════════════════════


# _alertas_painel — importado de utils/business.py


@report_bp.route("/calendario")
@login_required
def calendario_publico():
    import calendar as _cal

    u = current_user()
    hoje = date.today()
    mes_str = request.args.get("mes", hoje.strftime("%Y-%m"))
    try:
        ano_m, mes_m = int(mes_str[:4]), int(mes_str[5:7])
    except Exception:
        ano_m, mes_m = hoje.year, hoje.month

    ICONES = {
        "normal": "✅",
        "fim_semana": "🔵",
        "feriado": "🔴",
        "exercicio": "🟡",
        "outro": "⚪",
    }
    LABELS = {
        "normal": "Normal",
        "fim_semana": "Fim de semana",
        "feriado": "Feriado",
        "exercicio": "Exercício",
        "outro": "Outro",
    }
    CORES = {
        "normal": "#eafaf1",
        "fim_semana": "#ebf5fb",
        "feriado": "#fdecea",
        "exercicio": "#fef9e7",
        "outro": "#f8f9fa",
    }
    CORES_TEXT = {
        "normal": "#1e8449",
        "fim_semana": "#1a5276",
        "feriado": "#922b21",
        "exercicio": "#9a7d0a",
        "outro": "#6c757d",
    }

    ultimo_dia = _cal.monthrange(ano_m, mes_m)[1]
    d_inicio = date(ano_m, mes_m, 1)
    d_fim = date(ano_m, mes_m, ultimo_dia)
    with db() as conn:
        entradas = {
            r["data"]: dict(r)
            for r in conn.execute(
                "SELECT data,tipo,nota FROM calendario_operacional WHERE data>=? AND data<=?",
                (d_inicio.isoformat(), d_fim.isoformat()),
            ).fetchall()
        }

    cal_grid = _cal.monthcalendar(ano_m, mes_m)
    DIAS_CAB = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]

    # Pré-computar dados de cada dia para o template
    dias_info = {}
    for semana in cal_grid:
        for dia_n in semana:
            if dia_n == 0:
                continue
            d_obj = date(ano_m, mes_m, dia_n)
            entrada = entradas.get(d_obj.isoformat())
            tipo = (
                entrada["tipo"]
                if entrada
                else ("fim_semana" if d_obj.weekday() >= 5 else "normal")
            )
            nota = entrada["nota"] if entrada else ""
            dias_info[dia_n] = {
                "tipo": tipo,
                "nota": nota,
                "is_hoje": d_obj == hoje,
            }

    if mes_m == 1:
        prev_mes = f"{ano_m - 1}-12"
    else:
        prev_mes = f"{ano_m}-{mes_m - 1:02d}"
    if mes_m == 12:
        next_mes = f"{ano_m + 1}-01"
    else:
        next_mes = f"{ano_m}-{mes_m + 1:02d}"

    MESES_PT = [
        "Janeiro",
        "Fevereiro",
        "Março",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]
    mes_titulo = f"{MESES_PT[mes_m - 1]} {ano_m}"
    perfil = u.get("perfil")
    back_url = (
        url_for("admin.admin_home")
        if perfil == "admin"
        else (
            url_for("aluno.aluno_home")
            if perfil == "aluno"
            else url_for("operations.painel_dia")
        )
    )

    return render_template(
        "reporting/calendario.html",
        cal_grid=cal_grid,
        dias_info=dias_info,
        DIAS_CAB=DIAS_CAB,
        ICONES=ICONES,
        LABELS=LABELS,
        CORES=CORES,
        CORES_TEXT=CORES_TEXT,
        prev_mes=prev_mes,
        next_mes=next_mes,
        mes_titulo=mes_titulo,
        perfil=perfil,
        back_url=back_url,
        ano_m=ano_m,
        mes_m=mes_m,
    )


# ═══════════════════════════════════════════════════════════════════════════
# IMPRESSÃO — Mapa de refeições por ano
# ═══════════════════════════════════════════════════════════════════════════


@report_bp.route("/dashboard-semanal")
@role_required("cozinha", "oficialdia", "admin")
def dashboard_semanal():
    u = current_user()
    perfil = u.get("perfil")
    hoje = date.today()
    segunda = hoje - timedelta(days=hoje.weekday())
    d0_str = request.args.get("d0", segunda.isoformat())
    d0 = _parse_date(d0_str)
    d1 = d0 + timedelta(days=6)
    prev_w = (d0 - timedelta(days=7)).isoformat()
    next_w = (d0 + timedelta(days=7)).isoformat()

    # Batch: carregar totais e calendário para toda a semana numa query
    totais_map, _t_empty = get_totais_periodo(d0.isoformat(), d1.isoformat())
    cal_map_wk = dias_operacionais_batch(d0, d1)
    dias = []
    for i in range(7):
        di = d0 + timedelta(days=i)
        t = totais_map.get(di.isoformat(), _t_empty)
        tipo = cal_map_wk.get(
            di.isoformat(), "fim_semana" if di.weekday() >= 5 else "normal"
        )
        dias.append({"data": di, "t": t, "tipo": tipo, "is_wknd": di.weekday() >= 5})

    max_alm = (
        max(
            (
                d["t"]["alm_norm"] + d["t"]["alm_veg"] + d["t"]["alm_dieta"]
                for d in dias
            ),
            default=1,
        )
        or 1
    )
    max_jan = (
        max(
            (
                d["t"]["jan_norm"] + d["t"]["jan_veg"] + d["t"]["jan_dieta"]
                for d in dias
            ),
            default=1,
        )
        or 1
    )
    max_pa = max((d["t"]["pa"] for d in dias), default=1) or 1

    # Pré-computar dados de visualização para cada dia
    for d in dias:
        t = d["t"]
        alm = t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]
        jan = t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]
        off = d["tipo"] in ("feriado", "exercicio")
        d["alm"] = alm
        d["jan"] = jan
        d["col_cls"] = (
            "chart-col-off"
            if off
            else ("chart-col-wknd" if d["is_wknd"] else "chart-col-day")
        )
        d["dow_cls"] = "chart-dow-wknd" if d["is_wknd"] else "chart-dow-day"
        d["dow"] = ABREV_DIAS[d["data"].weekday()]
        d["data_fmt"] = d["data"].strftime("%d/%m")
        # Bar heights (px out of 80)
        d["pn"] = int(round(80 * (t["alm_norm"] / max_alm))) if max_alm else 0
        d["pv"] = int(round(80 * (t["alm_veg"] / max_alm))) if max_alm else 0
        d["pd"] = int(round(80 * (t["alm_dieta"] / max_alm))) if max_alm else 0
        d["pj"] = int(round(80 * (jan / max_jan))) if max_jan else 0
        d["pp"] = int(round(80 * (t["pa"] / max_pa))) if max_pa else 0

    _keys = [
        "pa",
        "lan",
        "alm_norm",
        "alm_veg",
        "alm_dieta",
        "alm_estufa",
        "jan_norm",
        "jan_veg",
        "jan_dieta",
        "jan_sai",
        "jan_estufa",
    ]
    totais_semana = {k: sum(d["t"][k] for d in dias) for k in _keys}
    totais_semana["alm_total"] = (
        totais_semana["alm_norm"]
        + totais_semana["alm_veg"]
        + totais_semana["alm_dieta"]
    )
    totais_semana["jan_total"] = (
        totais_semana["jan_norm"]
        + totais_semana["jan_veg"]
        + totais_semana["jan_dieta"]
    )

    # Totais da semana anterior para comparação
    prev_d0 = d0 - timedelta(days=7)
    prev_d1 = d0 - timedelta(days=1)
    prev_map, _ = get_totais_periodo(prev_d0.isoformat(), prev_d1.isoformat())
    totais_prev = {k: 0 for k in _keys}
    for t_p in prev_map.values():
        for k in _keys:
            totais_prev[k] += t_p[k]

    back_url = (
        url_for("admin.admin_home")
        if perfil == "admin"
        else url_for("operations.painel_dia")
    )

    return render_template(
        "reporting/dashboard_semanal.html",
        dias=dias,
        totais_semana=totais_semana,
        totais_prev=totais_prev,
        perfil=perfil,
        d0=d0,
        d1=d1,
        d0_str=d0_str,
        prev_w=prev_w,
        next_w=next_w,
        back_url=back_url,
    )


# ═══════════════════════════════════════════════════════════════════════════
# EXPORTAÇÕES
# ═══════════════════════════════════════════════════════════════════════════


@report_bp.route("/exportar/dia")
@role_required("cozinha", "oficialdia", "cmd", "admin")
def exportar_dia():

    d_str = request.args.get("d", date.today().isoformat())
    fmt = (request.args.get("fmt", "csv") or "csv").strip().lower()
    if fmt not in {"csv", "xlsx"}:
        abort(400)
    dt = _parse_date_strict(d_str)
    if dt is None:
        abort(400)
    t = get_totais_dia(dt.isoformat())

    # Tentar xlsx via openpyxl; cair para CSV se não disponível
    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Totais {dt.strftime('%d-%m-%Y')}"

            # Cabeçalho
            header_fill = PatternFill("solid", fgColor="0A2D4E")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            headers = [
                "Data",
                "Dia",
                "PA",
                "Lanche",
                "Alm. Normal",
                "Alm. Veg.",
                "Alm. Dieta",
                "Jan. Normal",
                "Jan. Veg.",
                "Jan. Dieta",
                "Jan. Sai Unidade",
                "Total Almoços",
                "Total Jantares",
            ]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")
                c.border = border

            total_alm = t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]
            total_jan = t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]
            data_row = [
                dt.isoformat(),
                NOMES_DIAS[dt.weekday()],
                t["pa"],
                t["lan"],
                t["alm_norm"],
                t["alm_veg"],
                t["alm_dieta"],
                t["jan_norm"],
                t["jan_veg"],
                t["jan_dieta"],
                t["jan_sai"],
                total_alm,
                total_jan,
            ]
            alt_fill = PatternFill("solid", fgColor="EBF5FB")
            for col, val in enumerate(data_row, 1):
                c = ws.cell(row=2, column=col, value=val)
                c.fill = alt_fill
                c.border = border
                c.alignment = Alignment(horizontal="center")

            # Auto-largura
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col) + 4
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 22)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(
                buf.read(),
                headers={
                    "Content-Disposition": f"attachment; filename=totais_{dt.isoformat()}.xlsx",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
            )
        except ImportError:
            flash("openpyxl não instalado — a exportar CSV.", "warn")
            fmt = "csv"
        except Exception as ex:
            flash(f"Erro ao gerar Excel: {ex} — a exportar CSV.", "warn")
            fmt = "csv"

    # CSV
    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=";")
    writer.writerow(
        [
            "Data",
            "Dia",
            "PA",
            "Lanche",
            "Alm. Normal",
            "Alm. Veg.",
            "Alm. Dieta",
            "Jan. Normal",
            "Jan. Veg.",
            "Jan. Dieta",
            "Jan. Sai Unidade",
            "Total Almoços",
            "Total Jantares",
        ]
    )
    total_alm = t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]
    total_jan = t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]
    writer.writerow(
        [
            dt.isoformat(),
            NOMES_DIAS[dt.weekday()],
            t["pa"],
            t["lan"],
            t["alm_norm"],
            t["alm_veg"],
            t["alm_dieta"],
            t["jan_norm"],
            t["jan_veg"],
            t["jan_dieta"],
            t["jan_sai"],
            total_alm,
            total_jan,
        ]
    )
    csv_bytes = ("\ufeff" + buf.getvalue()).encode("utf-8")
    return Response(
        csv_bytes,
        headers={
            "Content-Disposition": f"attachment; filename=totais_{dt.isoformat()}.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


@report_bp.route("/exportar/relatorio")
@role_required("cozinha", "oficialdia", "admin")
def exportar_relatorio():

    d0_str = request.args.get("d0", date.today().isoformat())
    fmt = (request.args.get("fmt", "csv") or "csv").strip().lower()
    if fmt not in {"csv", "xlsx"}:
        abort(400)
    d0 = _parse_date_strict(d0_str)
    if d0 is None:
        abort(400)
    d1 = d0 + timedelta(days=6)

    dias_data = []
    totais = {
        k: 0
        for k in [
            "pa",
            "lan",
            "alm_norm",
            "alm_veg",
            "alm_dieta",
            "alm_estufa",
            "jan_norm",
            "jan_veg",
            "jan_dieta",
            "jan_sai",
            "jan_estufa",
        ]
    }
    _exp_map, _exp_empty = get_totais_periodo(d0.isoformat(), d1.isoformat())
    _exp_cal = dias_operacionais_batch(d0, d1)
    for i in range(7):
        di = d0 + timedelta(days=i)
        t = _exp_map.get(di.isoformat(), _exp_empty)
        tipo = _exp_cal.get(
            di.isoformat(), "fim_semana" if di.weekday() >= 5 else "normal"
        )
        alm = t["alm_norm"] + t["alm_veg"] + t["alm_dieta"]
        jan = t["jan_norm"] + t["jan_veg"] + t["jan_dieta"]
        dias_data.append((di, tipo, t, alm, jan))
        for k in totais:
            totais[k] += t[k]

    HEADERS = [
        "Data",
        "Dia da Semana",
        "Tipo Dia",
        "PA",
        "Lanche",
        "Alm. Normal",
        "Alm. Veg.",
        "Alm. Dieta",
        "Alm. Estufa",
        "Total Almoços",
        "Jan. Normal",
        "Jan. Veg.",
        "Jan. Dieta",
        "Total Jantares",
        "Sai Unidade",
        "Jan. Estufa",
    ]

    def make_row(di, tipo, t, alm, jan):
        return [
            di.isoformat(),
            NOMES_DIAS[di.weekday()],
            tipo,
            t["pa"],
            t["lan"],
            t["alm_norm"],
            t["alm_veg"],
            t["alm_dieta"],
            t.get("alm_estufa", 0),
            alm,
            t["jan_norm"],
            t["jan_veg"],
            t["jan_dieta"],
            jan,
            t["jan_sai"],
            t.get("jan_estufa", 0),
        ]

    nome = f"relatorio_{d0_str}_a_{d1.isoformat()}"

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Relatório {d0.strftime('%d-%m')} a {d1.strftime('%d-%m-%Y')}"

            header_fill = PatternFill("solid", fgColor="0A2D4E")
            header_font = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col, h in enumerate(HEADERS, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")
                c.border = border

            TIPO_CORES = {
                "feriado": "FFD6D6",
                "exercicio": "FFFACD",
                "fim_semana": "DDEEFF",
                "normal": "FFFFFF",
                "outro": "F0F0F0",
            }
            for ri, (di, tipo, t, alm, jan) in enumerate(dias_data, 2):
                row_fill = PatternFill("solid", fgColor=TIPO_CORES.get(tipo, "FFFFFF"))
                for col, val in enumerate(make_row(di, tipo, t, alm, jan), 1):
                    c = ws.cell(row=ri, column=col, value=val)
                    c.fill = row_fill
                    c.border = border
                    c.alignment = Alignment(horizontal="center" if col > 2 else "left")

            # Linha de totais
            total_alm = totais["alm_norm"] + totais["alm_veg"] + totais["alm_dieta"]
            total_jan = totais["jan_norm"] + totais["jan_veg"] + totais["jan_dieta"]
            total_row = [
                "TOTAL",
                "—",
                "—",
                totais["pa"],
                totais["lan"],
                totais["alm_norm"],
                totais["alm_veg"],
                totais["alm_dieta"],
                totais.get("alm_estufa", 0),
                total_alm,
                totais["jan_norm"],
                totais["jan_veg"],
                totais["jan_dieta"],
                total_jan,
                totais["jan_sai"],
                totais.get("jan_estufa", 0),
            ]
            total_fill = PatternFill("solid", fgColor="D5E8F0")
            total_font = Font(bold=True)
            for col, val in enumerate(total_row, 1):
                c = ws.cell(row=9, column=col, value=val)
                c.fill = total_fill
                c.font = total_font
                c.border = border
                c.alignment = Alignment(horizontal="center" if col > 2 else "left")

            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col) + 3
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 20)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(
                buf.read(),
                headers={
                    "Content-Disposition": f"attachment; filename={nome}.xlsx",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
            )
        except ImportError:
            flash("openpyxl não instalado — a exportar CSV.", "warn")
        except Exception as ex:
            flash(f"Erro ao gerar Excel: {ex} — a exportar CSV.", "warn")

    # CSV (com BOM para Excel abrir correctamente)
    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=";")
    writer.writerow(HEADERS)
    for di, tipo, t, alm, jan in dias_data:
        writer.writerow(make_row(di, tipo, t, alm, jan))
    total_alm = totais["alm_norm"] + totais["alm_veg"] + totais["alm_dieta"]
    total_jan = totais["jan_norm"] + totais["jan_veg"] + totais["jan_dieta"]
    writer.writerow(
        [
            "TOTAL",
            "—",
            "—",
            totais["pa"],
            totais["lan"],
            totais["alm_norm"],
            totais["alm_veg"],
            totais["alm_dieta"],
            totais.get("alm_estufa", 0),
            total_alm,
            totais["jan_norm"],
            totais["jan_veg"],
            totais["jan_dieta"],
            total_jan,
            totais["jan_sai"],
            totais.get("jan_estufa", 0),
        ]
    )
    csv_bytes = ("\ufeff" + buf.getvalue()).encode("utf-8")
    return Response(
        csv_bytes,
        headers={
            "Content-Disposition": f"attachment; filename={nome}.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


# ═══════════════════════════════════════════════════════════════════════════
# CONTROLO DE PRESENÇAS — Módulo rápido via NI (Oficial de Dia)
# ═══════════════════════════════════════════════════════════════════════════
