"""Rotas do aluno: home, editar refeições, ausências, histórico, password, perfil."""

import io
import csv as _csv
import logging
from datetime import date, timedelta

from flask import (
    render_template,
    Response,
    abort,
    flash,
    redirect,
    request,
    session,
    url_for,
)

import config as cfg
from core.auth_db import user_id_by_nii
from core.meals import (
    dias_operacionais_batch,
    get_menu_do_dia,
    refeicao_get,
    refeicoes_batch,
)
from core.absences import ausencias_batch, detencoes_batch, licencas_batch
from core.users import (
    delete_ausencia_propria,
    delete_licenca,
    get_aluno_ano_ni,
    get_aluno_historico,
    get_aluno_licenca,
    get_aluno_profile_data,
    get_aluno_stats,
    get_ausencias_aluno,
    update_aluno_contacts,
    upsert_licenca,
)
from blueprints.aluno import aluno_bp
from utils.auth import current_user, login_required
from utils.business import (
    _cancelar_licenca_fds,
    _dia_editavel_aluno,
    _editar_ausencia,
    _get_ocupacao_dia,
    _licencas_semana_usadas,
    _marcar_licenca_fds,
    _pode_marcar_licenca,
    _registar_ausencia,
    _regras_licenca,
    _tem_ausencia_ativa,
    _tem_detencao_ativa,
)
from utils.constants import ABREV_DIAS, NOMES_DIAS
from utils.helpers import (
    _audit,
    _parse_date,
    _parse_date_strict,
    _refeicao_set,
)
from utils.passwords import _alterar_password
from utils.validators import (
    _val_email,
    _val_int_id,
    _val_phone,
    _val_refeicao,
    _val_text,
)

log = logging.getLogger(__name__)


@aluno_bp.route("/aluno/licenca-fds", methods=["POST"])
@login_required
def aluno_licenca_fds():
    """Marcar ou cancelar licença de fim de semana."""
    u = current_user()
    uid = user_id_by_nii(u["nii"])
    if not uid:
        flash("Conta de sistema — funcionalidade não disponível.", "error")
        return redirect(url_for(".aluno_home"))

    if not _check_rate_limit("_meal_ops"):
        flash("Demasiadas alterações. Aguarda um minuto.", "warn")
        return redirect(url_for(".aluno_home"))

    sexta_str = request.form.get("sexta", "")
    acao = request.form.get("acao_fds", "marcar")  # "marcar" ou "cancelar"
    sexta = _parse_date_strict(sexta_str)
    if not sexta or sexta.weekday() != 4:  # 4 = sexta-feira
        flash("Data inválida — apenas sextas-feiras.", "error")
        return redirect(url_for(".aluno_home"))

    # Verificar se a sexta ainda é editável
    ok_edit, msg = _dia_editavel_aluno(sexta)
    if not ok_edit:
        flash(f"Não é possível editar: {msg}", "warn")
        return redirect(url_for(".aluno_home"))

    # Verificar ausência
    if _tem_ausencia_ativa(uid, sexta):
        flash("Tens uma ausência registada para este período.", "warn")
        return redirect(url_for(".aluno_home"))

    # Verificar detenção
    if _tem_detencao_ativa(uid, sexta):
        flash("Estás detido — não podes marcar licença de fim de semana.", "error")
        return redirect(url_for(".aluno_home"))

    if acao == "cancelar":
        ok, err = _cancelar_licenca_fds(uid, sexta, u["nii"])
        flash(
            "Licença de fim de semana cancelada." if ok else (err or "Erro."),
            "ok" if ok else "error",
        )
    else:
        # Verificar regras de licença para a sexta
        ano_aluno, ni_aluno = get_aluno_ano_ni(uid)

        pode, motivo = _pode_marcar_licenca(uid, sexta, ano_aluno, ni_aluno)
        if not pode:
            flash(motivo, "warn")
            return redirect(url_for(".aluno_home"))

        ok, err = _marcar_licenca_fds(uid, sexta, u["nii"])
        flash(
            "✅ Licença de fim de semana marcada! Jantar de sexta e refeições de sábado/domingo cancelados."
            if ok
            else (err or "Erro."),
            "ok" if ok else "error",
        )

    return redirect(url_for(".aluno_home"))


@aluno_bp.route("/aluno")
@login_required
def aluno_home():
    u = current_user()
    uid = user_id_by_nii(u["nii"])
    hoje = date.today()
    menu = get_menu_do_dia(hoje)

    # Banner ausência ativa hoje
    ausente_hoje = uid and _tem_ausencia_ativa(uid, hoje)

    # Batch-load: carregar todos os dados de uma vez (elimina N+1)
    d_ate = hoje + timedelta(days=cfg.DIAS_ANTECEDENCIA)
    cal_map = dias_operacionais_batch(hoje, d_ate)
    if uid:
        ref_map, ref_defaults = refeicoes_batch(uid, hoje, d_ate)
        aus_set = ausencias_batch(uid, hoje, d_ate)
        det_set = detencoes_batch(uid, hoje, d_ate)
        lic_map = licencas_batch(uid, hoje, d_ate)
    else:
        ref_map, ref_defaults = {}, {}
        aus_set = set()
        det_set = set()
        lic_map = {}

    dias = []
    for i in range(cfg.DIAS_ANTECEDENCIA + 1):
        d = hoje + timedelta(days=i)
        d_iso = d.isoformat()
        tipo = cal_map.get(d_iso, "fim_semana" if d.weekday() >= 5 else "normal")
        r = ref_map.get(d_iso, ref_defaults) if uid else {}
        ok_edit, _ = _dia_editavel_aluno(d)
        ausente_d = d_iso in aus_set
        detido_d = d_iso in det_set
        is_weekend = d.weekday() >= 5
        is_off = tipo in ("feriado", "exercicio")
        is_friday = d.weekday() == 4

        lic_tipo = lic_map.get(d_iso)
        lic_label = ""
        if uid and not detido_d and lic_tipo:
            lic_label = "Antes jantar" if lic_tipo == "antes_jantar" else "Após jantar"

        # FDS button logic
        show_fds_btn = is_friday and uid and not is_off
        tem_licenca_fds = lic_map.get(d_iso) == "antes_jantar"
        show_fds_marcar = (
            show_fds_btn
            and ok_edit
            and not detido_d
            and not ausente_d
            and not tem_licenca_fds
        )

        dia = {
            "d_iso": d_iso,
            "date_obj": d,
            "date_str": d.strftime("%d/%m"),
            "date_full": d.strftime("%d/%m/%Y"),
            "abrev": ABREV_DIAS[d.weekday()],
            "r": r,
            "ok_edit": ok_edit,
            "ausente": ausente_d,
            "detido": detido_d,
            "is_weekend": is_weekend,
            "is_off": is_off,
            "icon": {"feriado": "\U0001f534", "exercicio": "\U0001f7e1"}.get(
                tipo, "\u26aa"
            ),
            "label": {"feriado": "Feriado", "exercicio": "Exercício"}.get(tipo, tipo),
            "lic_label": lic_label,
            "show_fds_btn": show_fds_btn and ok_edit and not detido_d and not ausente_d,
            "tem_licenca_fds": tem_licenca_fds,
            "show_fds_marcar": show_fds_marcar,
        }
        dias.append(dia)

    stats = None
    if uid:
        d0 = (hoje - timedelta(days=30)).isoformat()
        stats = get_aluno_stats(uid, d0)

    return render_template(
        "aluno/home.html",
        u=u,
        hoje=hoje,
        menu=menu,
        ausente_hoje=ausente_hoje,
        dias=dias,
        dias_antecedencia=cfg.DIAS_ANTECEDENCIA,
        stats=stats,
    )


def _check_rate_limit(key: str, max_ops: int = 30, window: int = 60) -> bool:
    """Rate limiter por sessão. Retorna True se dentro do limite."""
    import time as _t

    now = _t.time()
    ops = session.get(key, [])
    ops = [t for t in ops if now - t < window]
    if len(ops) >= max_ops:
        return False
    ops.append(now)
    session[key] = ops
    return True


@aluno_bp.route("/aluno/editar/<d>", methods=["GET", "POST"])
@login_required
def aluno_editar(d):
    u = current_user()
    uid = user_id_by_nii(u["nii"])
    dt = _parse_date(d)

    if not uid:
        flash("Conta de sistema — não é possível editar refeições.", "error")
        return redirect(url_for(".aluno_home"))

    # Bloquear edição se tem ausência ativa
    if _tem_ausencia_ativa(uid, dt):
        flash(
            "Tens uma ausência registada para este dia. Remove a ausência primeiro.",
            "warn",
        )
        return redirect(url_for(".aluno_home"))

    ok_edit, msg = _dia_editavel_aluno(dt)
    if not ok_edit:
        flash(f"Não é possível editar: {msg}", "warn")
        return redirect(url_for(".aluno_home"))

    r = refeicao_get(uid, dt)
    occ = _get_ocupacao_dia(dt)
    is_weekend = dt.weekday() >= 5
    detido = _tem_detencao_ativa(uid, dt)

    # Dados do aluno para regras de licença
    ano_aluno, ni_aluno = get_aluno_ano_ni(uid)

    # Licença existente para este dia
    licenca_atual = get_aluno_licenca(uid, dt.isoformat())

    pode_lic, motivo_lic = _pode_marcar_licenca(uid, dt, ano_aluno, ni_aluno)

    if request.method == "POST":
        if not _check_rate_limit("_meal_ops"):
            flash("Demasiadas alterações. Aguarda um minuto.", "warn")
            return redirect(url_for(".aluno_home"))
        pa = 1 if request.form.get("pa") in ("1", "on") else 0
        lanche = 1 if request.form.get("lanche") in ("1", "on") else 0
        alm = _val_refeicao(request.form.get("almoco"))
        jan = _val_refeicao(request.form.get("jantar"))
        sai = 0 if detido else (1 if request.form.get("sai") else 0)
        alm_estufa = 1 if (alm and request.form.get("almoco_estufa") == "1") else 0
        jan_estufa = 1 if (jan and request.form.get("jantar_estufa") == "1") else 0

        # Processar licença (antes_jantar / apos_jantar / vazio)
        licenca_tipo = request.form.get("licenca", "")
        if licenca_tipo in ("antes_jantar", "apos_jantar"):
            pode, motivo = _pode_marcar_licenca(uid, dt, ano_aluno, ni_aluno)
            if pode:
                upsert_licenca(uid, dt.isoformat(), licenca_tipo)
                if licenca_tipo == "antes_jantar":
                    jan = ""
                    sai = 1
                else:
                    sai = 1
            else:
                flash(motivo, "warn")
        else:
            delete_licenca(uid, dt.isoformat())

        if _refeicao_set(
            uid,
            dt,
            pa,
            lanche,
            alm,
            jan,
            sai,
            alterado_por=u["nii"],
            alm_estufa=alm_estufa,
            jan_estufa=jan_estufa,
        ):
            flash("Refeições atualizadas!", "ok")
        else:
            flash("Erro ao guardar.", "error")
        return redirect(url_for(".aluno_home"))

    # Valores atuais
    pa_on = 1 if r.get("pequeno_almoco") else 0
    lan_on = 1 if r.get("lanche") else 0
    alm_val = r.get("almoco") or ""
    jan_val = r.get("jantar_tipo") or ""
    jan_blocked = licenca_atual == "antes_jantar"
    alm_estufa = r.get("almoco_estufa", 0)
    jan_estufa = r.get("jantar_estufa", 0)

    # Ementa do dia
    menu = get_menu_do_dia(dt)

    # Quota info for licença section
    usadas_semana = 0
    max_uteis = 0
    is_weekday = dt.weekday() < 4
    if not detido:
        regras = _regras_licenca(ano_aluno, ni_aluno)
        usadas_semana = _licencas_semana_usadas(uid, dt)
        max_uteis = regras["max_dias_uteis"]

    return render_template(
        "aluno/editar.html",
        dt=dt,
        nome_dia=NOMES_DIAS[dt.weekday()],
        is_weekend=is_weekend,
        detido=detido,
        menu=menu,
        occ=occ,
        pa_on=pa_on,
        lan_on=lan_on,
        alm_val=alm_val,
        jan_val=jan_val,
        jan_blocked=jan_blocked,
        alm_estufa=alm_estufa,
        jan_estufa=jan_estufa,
        licenca_atual=licenca_atual,
        pode_lic=pode_lic,
        motivo_lic=motivo_lic,
        usadas_semana=usadas_semana,
        max_uteis=max_uteis,
        is_weekday=is_weekday,
    )


# ─── Aluno: Gerir ausências próprias ─────────────────────────────────────


@aluno_bp.route("/aluno/ausencias", methods=["GET", "POST"])
@login_required
def aluno_ausencias():
    u = current_user()
    uid = user_id_by_nii(u["nii"])
    if not uid:
        flash("Conta de sistema — funcionalidade não disponível.", "error")
        return redirect(url_for(".aluno_home"))

    if request.method == "POST":
        if not _check_rate_limit("_meal_ops"):
            flash("Demasiadas alterações. Aguarda um minuto.", "warn")
            return redirect(url_for(".aluno_ausencias"))
        acao = request.form.get("acao", "")
        if acao == "criar":
            de = request.form.get("de", "")
            ate = request.form.get("ate", "")
            motivo = _val_text(request.form.get("motivo", ""))[:500]
            ok, err = _registar_ausencia(uid, de, ate, motivo, u["nii"])
            flash(
                "Ausência registada com sucesso!" if ok else (err or "Erro."),
                "ok" if ok else "error",
            )
        elif acao == "editar":
            aid = _val_int_id(request.form.get("id", ""))
            de = request.form.get("de", "")
            ate = request.form.get("ate", "")
            motivo = _val_text(request.form.get("motivo", ""))[:500]
            if aid is None:
                flash("ID inválido.", "error")
            else:
                ok, err = _editar_ausencia(aid, uid, de, ate, motivo)
                flash(
                    "Ausência atualizada!" if ok else (err or "Erro."),
                    "ok" if ok else "error",
                )
        elif acao == "remover":
            aid = _val_int_id(request.form.get("id", ""))
            if aid is not None:
                delete_ausencia_propria(aid, uid)
                flash("Ausência removida.", "ok")
        return redirect(url_for(".aluno_ausencias"))

    rows = get_ausencias_aluno(uid)

    hoje = date.today().isoformat()
    edit_id = request.args.get("edit", "")
    edit_row = next((r for r in rows if str(r["id"]) == edit_id), None)

    if edit_row:
        form_title = "✏️ Editar ausência"
        form_action = "editar"
        form_de = edit_row["ausente_de"]
        form_ate = edit_row["ausente_ate"]
        form_motivo = edit_row["motivo"] or ""
    else:
        form_title = "➕ Nova ausência"
        form_action = "criar"
        form_de = form_ate = form_motivo = ""

    return render_template(
        "aluno/ausencias.html",
        rows=rows,
        hoje=hoje,
        edit_row=edit_row,
        form_title=form_title,
        form_action=form_action,
        form_de=form_de,
        form_ate=form_ate,
        form_motivo=form_motivo,
    )


@aluno_bp.route("/aluno/historico")
@login_required
def aluno_historico():
    u = current_user()
    uid = user_id_by_nii(u["nii"])
    hoje = date.today()
    rows = []
    if uid:
        rows = get_aluno_historico(uid, (hoje - timedelta(days=30)).isoformat())

    return render_template("aluno/historico.html", rows=rows)


@aluno_bp.route("/aluno/exportar-historico")
@login_required
def exportar_historico_aluno():
    u = current_user()
    uid = user_id_by_nii(u["nii"])
    fmt = (request.args.get("fmt", "csv") or "csv").strip().lower()
    if fmt not in {"csv", "xlsx"}:
        abort(400)

    hoje = date.today()
    rows = []
    if uid:
        rows = get_aluno_historico(uid, (hoje - timedelta(days=30)).isoformat())

    headers = [
        "Data",
        "PA",
        "Lanche",
        "Almoço",
        "♨️ Alm",
        "Jantar",
        "Sai Unidade",
        "♨️ Jan",
    ]

    def make_row(r):
        return [
            r["data"],
            "Sim" if r["pequeno_almoco"] else "Não",
            "Sim" if r["lanche"] else "Não",
            r["almoco"] or "—",
            "Sim" if r["almoco_estufa"] else "Não",
            r["jantar_tipo"] or "—",
            "Sim" if r["jantar_sai_unidade"] else "Não",
            "Sim" if r["jantar_estufa"] else "Não",
        ]

    nome_ficheiro = f"historico_{u['nii']}_{hoje.isoformat()}"

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Histórico {u['nome'][:20]}"

            header_fill = PatternFill("solid", fgColor="0A2D4E")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")
                c.border = border

            alt_fill = PatternFill("solid", fgColor="EBF5FB")
            for i, r in enumerate(rows, 2):
                row_data = make_row(r)
                fill = alt_fill if i % 2 == 0 else PatternFill()
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=i, column=col, value=val)
                    c.fill = fill
                    c.border = border
                    c.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col) + 4
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 22)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(
                buf.read(),
                headers={
                    "Content-Disposition": f"attachment; filename={nome_ficheiro}.xlsx",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
            )
        except ImportError:
            fmt = "csv"

    buf = io.StringIO()
    writer = _csv.writer(buf, delimiter=";")
    writer.writerow(headers)
    for r in rows:
        writer.writerow(make_row(r))
    csv_bytes = ("\ufeff" + buf.getvalue()).encode("utf-8")
    return Response(
        csv_bytes,
        headers={
            "Content-Disposition": f"attachment; filename={nome_ficheiro}.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


@aluno_bp.route("/aluno/password", methods=["GET", "POST"])
@login_required
def aluno_password():
    u = current_user()
    if request.method == "POST":
        # Rate limiting: máx 10 tentativas por 5 minutos
        import time as _t

        pw_attempts = session.get("_pw_attempts", [])
        now = _t.time()
        pw_attempts = [t for t in pw_attempts if now - t < 300]
        if len(pw_attempts) >= 10:
            flash("Demasiadas tentativas. Aguarda uns minutos.", "error")
            return redirect(url_for(".aluno_password"))
        pw_attempts.append(now)
        session["_pw_attempts"] = pw_attempts

        old = request.form.get("old", "")
        new = request.form.get("new", "")
        conf = request.form.get("conf", "")
        if new != conf:
            flash("As passwords não coincidem.", "error")
        else:
            ok, err = _alterar_password(u["nii"], old, new)
            flash(
                "Password alterada!" if ok else (err or "Erro."),
                "ok" if ok else "error",
            )
            if ok:
                session.pop("must_change_password", None)
                session.pop("_pw_attempts", None)
                return redirect(url_for("auth.dashboard"))

    is_forced = session.get("must_change_password")
    title = (
        "\U0001f510 Definir nova password"
        if is_forced
        else "\U0001f511 Alterar password"
    )
    old_hint = "A tua password atual (NII se é o primeiro login)" if is_forced else ""

    return render_template(
        "aluno/password.html",
        is_forced=is_forced,
        title=title,
        old_hint=old_hint,
    )


# ─── Aluno: Perfil próprio ────────────────────────────────────────────────


@aluno_bp.route("/aluno/perfil", methods=["GET", "POST"])
@login_required
def aluno_perfil():
    u = current_user()
    uid = user_id_by_nii(u["nii"])
    if not uid:
        flash("Conta de sistema — funcionalidade não disponível.", "error")
        return redirect(url_for(".aluno_home"))

    from core.users import get_user_by_nii_fields

    row = get_user_by_nii_fields(u["nii"], "NII,NI,Nome_completo,ano,email,telemovel")
    if not row:
        flash("Erro ao carregar perfil.", "error")
        return redirect(url_for(".aluno_home"))

    if request.method == "POST":
        email_n = _val_email(request.form.get("email", ""))
        if email_n is False:
            flash("Email inválido.", "error")
            return redirect(url_for(".aluno_perfil"))
        telef_n = _val_phone(request.form.get("telemovel", ""))
        if telef_n is False:
            flash("Telemóvel inválido.", "error")
            return redirect(url_for(".aluno_perfil"))
        try:
            update_aluno_contacts(uid, email_n, telef_n)
            _audit(
                current_user().get("nii", "?"),
                "aluno_perfil_update",
                f"uid={uid}",
            )
            flash("Perfil atualizado com sucesso!", "ok")
            return redirect(url_for(".aluno_perfil"))
        except Exception as ex:
            log.exception("aluno_perfil: erro ao atualizar contactos")
            flash(f"Erro: {ex}", "error")

    hoje = date.today()
    profile = get_aluno_profile_data(uid, hoje.isoformat())

    return render_template(
        "aluno/perfil.html",
        aluno=row,
        total_ref=profile["total_ref"],
        ausencias_ativas=profile["ausencias_ativas"],
    )
