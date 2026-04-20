"""Funções de exportação CSV, XLSX e PDF."""

from __future__ import annotations

import csv
import html
import logging
import os
from datetime import date, datetime

from core.constants import EXPORT_DIR
from core.database import db
from core.meals import (
    _HEADERS_DISTRIBUICAO,
    _HEADERS_TOTAIS,
    _totais_para_csv_row,
    get_totais_dia,
)

log = logging.getLogger(__name__)


def export_csv(rows: list[dict], headers: list[str], name: str) -> str:
    path = os.path.join(EXPORT_DIR, name + ".csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow([r.get(h, "") for h in headers])
    return path


def export_xlsx(rows: list[dict], headers: list[str], name: str) -> str:
    """Exporta para Excel (.xlsx). Requer openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl não instalado. Instala com: pip install openpyxl")
        return export_csv(rows, headers, name)

    path = os.path.join(EXPORT_DIR, name + ".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = name[:31]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))

    for col_idx, h in enumerate(headers, 1):
        max_len = (
            max(len(str(h)), *(len(str(row.get(h, ""))) for row in rows))
            if rows
            else len(str(h))
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    wb.save(path)
    return path


def export_both(rows: list[dict], headers: list[str], name: str) -> tuple[str, str]:
    """Exporta CSV e XLSX e devolve ambos os caminhos."""
    p1 = export_csv(rows, headers, name)
    p2 = export_xlsx(rows, headers, name)
    return p1, p2


# ── PDF export ────────────────────────────────────────────────────────────
#
# `export_pdf` tenta usar reportlab (output .pdf); se reportlab não estiver
# instalado, cai num render HTML auto-contido (.html) — menos ideal mas
# preservável e imprimível pelo browser. Nunca falha silenciosamente: o
# caller recebe sempre um caminho.


def _export_pdf_html_fallback(
    rows: list[dict], headers: list[str], name: str, title: str
) -> str:
    """Fallback HTML — self-contained, printer-friendly."""
    path = os.path.join(EXPORT_DIR, name + ".html")
    esc = html.escape
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    parts = [
        "<!doctype html><html lang='pt'><head><meta charset='utf-8'>",
        f"<title>{esc(title)}</title>",
        "<style>",
        "body{font-family:system-ui,sans-serif;margin:2rem;color:#111}",
        "h1{font-size:1.4rem;margin:0 0 .25rem}",
        ".meta{color:#666;font-size:.85rem;margin-bottom:1rem}",
        "table{border-collapse:collapse;width:100%;font-size:.9rem}",
        "th,td{border:1px solid #ccc;padding:.35rem .5rem;text-align:left}",
        "th{background:#1F4E79;color:#fff}",
        "tr:nth-child(even) td{background:#f7f7f7}",
        "@media print{.meta{color:#000}}",
        "</style></head><body>",
        f"<h1>{esc(title)}</h1>",
        f"<div class='meta'>Gerado em {esc(ts)}</div>",
        "<table><thead><tr>",
        *(f"<th>{esc(str(h))}</th>" for h in headers),
        "</tr></thead><tbody>",
    ]
    for r in rows:
        parts.append("<tr>")
        for h in headers:
            parts.append(f"<td>{esc(str(r.get(h, '')))}</td>")
        parts.append("</tr>")
    parts.extend(["</tbody></table></body></html>"])
    with open(path, "w", encoding="utf-8") as f:
        f.write("".join(parts))
    return path


def export_pdf(
    rows: list[dict],
    headers: list[str],
    name: str,
    title: str | None = None,
) -> str:
    """Exporta um relatório para PDF via reportlab.

    Se reportlab não estiver disponível (deploy sem a dep opcional), cai num
    render HTML auto-contido com o mesmo conteúdo — imprimível pelo browser.
    Retorna sempre o caminho do ficheiro gerado.
    """
    title = title or name

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        log.info("reportlab ausente — export_pdf a cair no fallback HTML.")
        return _export_pdf_html_fallback(rows, headers, name, title)

    path = os.path.join(EXPORT_DIR, name + ".pdf")
    try:
        doc = SimpleDocTemplate(
            path,
            pagesize=landscape(A4),
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
            title=title,
            author="Sistema de Refeições",
        )
        styles = getSampleStyleSheet()
        story = [
            Paragraph(f"<b>{html.escape(title)}</b>", styles["Title"]),
            Paragraph(f"Gerado em {datetime.now():%Y-%m-%d %H:%M}", styles["Normal"]),
            Spacer(1, 6 * mm),
        ]

        if not rows:
            story.append(
                Paragraph("<i>Sem dados para este relatório.</i>", styles["Normal"])
            )
        else:
            data = [list(headers)] + [
                [str(r.get(h, "")) for h in headers] for r in rows
            ]
            table = Table(data, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.whitesmoke, colors.white],
                        ),
                    ]
                )
            )
            story.append(table)

        doc.build(story)
        return path
    except Exception:
        log.exception("export_pdf: reportlab falhou — fallback HTML.")
        return _export_pdf_html_fallback(rows, headers, name, title)


def exportacao_pdf_do_dia(d: date, ano: int | None = None) -> str:
    """Gera um PDF diário com totais + distribuição por aluno.

    Combina as views existentes num único documento de gestão.
    Retorna o caminho do PDF (ou HTML se reportlab ausente).
    """
    di = d.isoformat()
    tag = f"_ano{ano}" if ano else ""
    t = get_totais_dia(di, ano)

    row_sum = _totais_para_csv_row(di, t, {"ano": ano} if ano else {})

    with db() as conn:
        if ano is None:
            det = [
                dict(r)
                for r in conn.execute(
                    """
                    SELECT u.ano, u.NI, u.Nome_completo, r.data,
                           r.pequeno_almoco, r.lanche, r.almoco,
                           r.jantar_tipo, r.jantar_sai_unidade
                    FROM refeicoes r JOIN utilizadores u ON u.id=r.utilizador_id
                    WHERE r.data=?
                    ORDER BY u.ano, u.NI
                    """,
                    (di,),
                )
            ]
        else:
            det = [
                dict(r)
                for r in conn.execute(
                    """
                    SELECT u.NII, u.NI, u.Nome_completo, u.ano, r.data,
                           r.pequeno_almoco, r.lanche, r.almoco,
                           r.jantar_tipo, r.jantar_sai_unidade
                    FROM refeicoes r JOIN utilizadores u ON u.id=r.utilizador_id
                    WHERE r.data=? AND u.ano=?
                    ORDER BY u.NI
                    """,
                    (di, ano),
                )
            ]

    # Primeiro: totais (1 linha); depois: distribuição detalhada
    title = f"Relatório diário — {di}" + (f" (ano {ano})" if ano else "")
    name = f"relatorio{tag}_{di}"

    # Para o PDF combinamos totais numa tabela + a distribuição por baixo.
    # Aqui mantemos simples: exportamos só distribuição (mais útil para cozinha)
    # e incluímos a linha de totais no topo como cabeçalho auxiliar.
    hdrs_det = (
        [
            "NII",
            "NI",
            "Nome_completo",
            "ano",
            "data",
            "pequeno_almoco",
            "lanche",
            "almoco",
            "jantar_tipo",
            "jantar_sai_unidade",
        ]
        if ano
        else _HEADERS_DISTRIBUICAO
    )
    # Linha sumário (serve como topo visual): inserir como "row 0" adicionando
    # um placeholder no cabeçalho. Simples: acrescentamos totais no title.
    totais_str = ", ".join(
        f"{k}={v}" for k, v in row_sum.items() if k not in ("data", "ano")
    )
    return export_pdf(det, hdrs_det, name, title=f"{title} — {totais_str}")


def exportacoes_do_dia(d: date, ano: int | None = None) -> None:
    """Gera CSV + XLSX do dia d."""

    di = d.isoformat()
    tag = f"_ano{ano}" if ano else ""
    t = get_totais_dia(di, ano)

    row_sum = _totais_para_csv_row(di, t, {"ano": ano} if ano else {})
    hdrs = (["data", "ano"] + _HEADERS_TOTAIS[1:]) if ano else _HEADERS_TOTAIS
    export_both([row_sum], hdrs, f"totais{tag}_{di}")

    with db() as conn:
        if ano is None:
            det = [
                dict(r)
                for r in conn.execute(
                    """
                SELECT u.ano, u.NI, u.Nome_completo, r.data,
                       r.pequeno_almoco, r.lanche, r.almoco, r.jantar_tipo, r.jantar_sai_unidade
                FROM refeicoes r JOIN utilizadores u ON u.id=r.utilizador_id
                WHERE r.data=?
                ORDER BY u.ano, u.NI
            """,
                    (di,),
                )
            ]
        else:
            det = [
                dict(r)
                for r in conn.execute(
                    """
                SELECT u.NII, u.NI, u.Nome_completo, u.ano, r.data,
                       r.pequeno_almoco, r.lanche, r.almoco, r.jantar_tipo, r.jantar_sai_unidade
                FROM refeicoes r JOIN utilizadores u ON u.id=r.utilizador_id
                WHERE r.data=? AND u.ano=?
                ORDER BY u.NI
            """,
                    (di, ano),
                )
            ]
    hdrs_det = (
        [
            "NII",
            "NI",
            "Nome_completo",
            "ano",
            "data",
            "pequeno_almoco",
            "lanche",
            "almoco",
            "jantar_tipo",
            "jantar_sai_unidade",
        ]
        if ano
        else _HEADERS_DISTRIBUICAO
    )
    export_both(det, hdrs_det, f"distribuicao{tag}_{di}")

    with db() as conn:
        occ_rows = [
            dict(r)
            for r in conn.execute("SELECT * FROM v_ocupacao_dia WHERE data=?", (di,))
        ]
    export_both(
        occ_rows,
        ["data", "refeicao", "ocupacao", "capacidade"],
        f"ocupacao_vs_capacidade_{di}",
    )
