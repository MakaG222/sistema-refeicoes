#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import sqlite3
import logging
import csv
import shutil
import secrets
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional, List

try:
    from werkzeug.security import check_password_hash as _wz_check_password_hash
except Exception:
    _wz_check_password_hash = None

# ---------------------------------------------------------------------------
# Caminhos e diretórios
# ---------------------------------------------------------------------------
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sistema.db")
BASE_DADOS = os.getenv("DB_PATH", DB_PATH)  # pode ser sobrescrito por env var
BACKUP_DIR = "backups"
EXPORT_DIR = "exportacoes"
Path(BACKUP_DIR).mkdir(exist_ok=True)
Path(EXPORT_DIR).mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Logging — stdout (compatível com Railway/Docker/Heroku; sem ficheiro local)
# ---------------------------------------------------------------------------
logging.basicConfig(
    stream=sys.stdout,
    level=logging.INFO,
    format="%(asctime)s %(levelname)s [sr]: %(message)s",
)

PRAZO_LIMITE_HORAS: Optional[int] = 48

BACKUP_RETENCAO_DIAS: Optional[int] = 30

PERFIS_ADMIN = {
    "admin": {
        "senha": "admin123",  # ⚠️ Alterar em produção
        "nome": "Administrador Geral",
        "perfil": "admin",
        "ano": "",
    },
    "cmd1": {
        "senha": "cmd1123",
        "nome": "Comandante 1º Ano",
        "perfil": "cmd",
        "ano": "1",
    },
    "cmd2": {
        "senha": "cmd2123",
        "nome": "Comandante 2º Ano",
        "perfil": "cmd",
        "ano": "2",
    },
    "cmd3": {
        "senha": "cmd3123",
        "nome": "Comandante 3º Ano",
        "perfil": "cmd",
        "ano": "3",
    },
    "cmd4": {
        "senha": "cmd4123",
        "nome": "Comandante 4º Ano",
        "perfil": "cmd",
        "ano": "4",
    },
    "cmd5": {
        "senha": "cmd5123",
        "nome": "Comandante 5º Ano",
        "perfil": "cmd",
        "ano": "5",
    },
    "cmd6": {
        "senha": "cmd6123",
        "nome": "Comandante 6º Ano",
        "perfil": "cmd",
        "ano": "6",
    },
    "cmd7": {
        "senha": "cmd7123",
        "nome": "Comandante CFBO",
        "perfil": "cmd",
        "ano": "7",
    },
    "cmd8": {
        "senha": "cmd8123",
        "nome": "Comandante CFCO",
        "perfil": "cmd",
        "ano": "8",
    },
    "cozinha": {
        "senha": "cozinha123",
        "nome": "Responsável da Cozinha",
        "perfil": "cozinha",
        "ano": "",
    },
    "oficialdia": {
        "senha": "oficial123",
        "nome": "Oficial de Dia",
        "perfil": "oficialdia",
        "ano": "",
    },
}

# ⚠️  Perfis de TESTE — apenas para demonstrações. Desativados em produção.
_ENV = os.getenv("ENV", "development").lower()
if _ENV != "production":
    PERFIS_TESTE = {
        f"teste{i}": {
            "senha": f"teste{i}",
            "nome": f"Utilizador Teste {i}",
            "perfil": "aluno",
            "ano": "1",
        }
        for i in range(1, 16)
    }
else:
    PERFIS_TESTE = {}

# ⚠️  Admin de emergência — só ativo fora de produção ou quando não existe nenhum admin na BD.
#     Em produção, criar um admin real na BD e não depender deste fallback.
if _ENV != "production":
    FALLBACK_ADMIN = {
        "nii": "admin",
        "pw": "admin123",
        "nome": "Administrador (fallback)",
    }
else:
    FALLBACK_ADMIN = {"nii": "", "pw": secrets.token_urlsafe(32), "nome": ""}

# ===========================================================================
# FUNÇÕES UTILITÁRIAS DE TERMINAL (UI)
# Versão única — substitui o par bar/ui_bar, print_kv/ui_print_kv, etc.
# ===========================================================================


def clear():
    os.system("cls" if os.name == "nt" else "clear")


def line(t="", w=72):
    print("\n" + "=" * w + "\n" + t.center(w) + "\n" + "=" * w)


def _pct(x, m):
    """Percentagem segura; devolve None se m < 0 (sem limite)."""
    try:
        if m is None or m < 0:
            return None
        return (x / m) if m else 0.0
    except Exception:
        return None


def bar(x, m, width=22, fill="█", empty="░"):
    """Barra de ocupação vs capacidade. m=-1 mostra ∞ (sem limite)."""
    p = _pct(x, m)
    if p is None:
        return f"{str(x).rjust(4)}  " + ("-" * width) + "  ∞"
    p = max(0.0, min(1.0, p))
    n = int(round(p * width))
    return (
        f"{str(x).rjust(4)}/{str(m).ljust(4)}  "
        + (fill * n + empty * (width - n))
        + f"  {int(p * 100):>3}%"
    )


def print_kv(rows, left=22):
    """Imprime pares (chave, valor) alinhados à esquerda."""
    for k, v in rows:
        print(f"  {k:<{left}}: {v}")


def print_table(headers, rows):
    """Tabela simples em texto sem dependências externas."""
    if not rows:
        print(" (sem dados)")
        return
    widths = [
        max(len(str(h)), *(len(str(r[i])) for r in rows)) for i, h in enumerate(headers)
    ]

    def fmt(row):
        return " | ".join(str(c).ljust(widths[i]) for i, c in enumerate(row))

    print(" " + fmt(headers))
    print(" " + "-+-".join("-" * w for w in widths))
    for r in rows:
        print(" " + fmt(r))


# Sparklines (tendências)
_SPARKS = "▁▂▃▄▅▆▇█"


def sparkline(vals):
    arr = [v for v in vals if v is not None]
    if not arr:
        return "(sem dados)"
    lo, hi = min(arr), max(arr)
    if hi == lo:
        return _SPARKS[0] * len(vals)
    out = []
    for v in vals:
        if v is None:
            out.append(" ")
            continue
        i = int((v - lo) / (hi - lo) * (len(_SPARKS) - 1))
        out.append(_SPARKS[i])
    return "".join(out)


def masked_input(prompt="Password: ", mask="*") -> str:
    """Input de password com mascaramento de caracteres."""
    if not sys.stdin.isatty() or not sys.stdout.isatty():
        return input(prompt + " (visível): ")
    try:
        if os.name == "nt":
            import msvcrt

            sys.stdout.write(prompt)
            sys.stdout.flush()
            buf = []
            while True:
                ch = msvcrt.getwch()
                if ch in ("\r", "\n"):
                    sys.stdout.write("\n")
                    break
                if ch == "\x03":
                    raise KeyboardInterrupt
                if ch in ("\x00", "\xe0"):
                    msvcrt.getwch()
                    continue
                if ch in ("\b", "\x7f"):
                    if buf:
                        buf.pop()
                        sys.stdout.write("\b \b")
                        sys.stdout.flush()
                    continue
                buf.append(ch)
                sys.stdout.write(mask)
                sys.stdout.flush()
            return "".join(buf)
        else:
            import termios
            import tty

            fd = sys.stdin.fileno()
            old = termios.tcgetattr(fd)
            try:
                tty.setraw(fd)
                sys.stdout.write(prompt)
                sys.stdout.flush()
                buf = []
                while True:
                    ch = sys.stdin.read(1)
                    if ch in ("\r", "\n"):
                        sys.stdout.write("\n")
                        break
                    if ch == "\x03":
                        raise KeyboardInterrupt
                    if ch in ("\x7f", "\b"):
                        if buf:
                            buf.pop()
                            sys.stdout.write("\b \b")
                            sys.stdout.flush()
                        continue
                    buf.append(ch)
                    sys.stdout.write(mask)
                    sys.stdout.flush()
                return "".join(buf)
            finally:
                termios.tcsetattr(fd, termios.TCSADRAIN, old)
    except Exception:
        return input(prompt + " (visível): ")


# ===========================================================================
# BASE DE DADOS
# ===========================================================================


def _new_conn() -> sqlite3.Connection:
    """Cria uma nova conexão SQLite com pragmas de performance."""
    conn = sqlite3.connect(BASE_DADOS)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA busy_timeout=8000")
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA cache_size=-4000")  # 4 MB cache
    return conn


def db() -> sqlite3.Connection:
    """Devolve conexão SQLite reutilizável por request (via Flask g) ou nova."""
    try:
        from flask import g, has_request_context

        if has_request_context():
            conn = getattr(g, "_sr_db", None)
            if conn is None:
                conn = _new_conn()
                g._sr_db = conn
            return conn
    except ImportError:
        pass
    return _new_conn()


def close_request_db(exc=None) -> None:
    """Fecha a conexão da request (chamado pelo teardown do Flask)."""
    try:
        from flask import g

        conn = getattr(g, "_sr_db", None)
        if conn is not None:
            g._sr_db = None
            try:
                conn.close()
            except Exception:
                pass
    except ImportError:
        pass


def wal_checkpoint() -> None:
    """Força checkpoint do WAL para libertar espaço no ficheiro -wal."""
    try:
        conn = _new_conn()
        conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        conn.close()
    except Exception:
        pass


def sqlite_quick_check() -> bool:
    try:
        with db() as conn:
            row = conn.execute("PRAGMA quick_check").fetchone()
            return bool(row and row[0] == "ok")
    except Exception:
        return False


def ensure_daily_backup():
    """Cria backup automático 1x por dia (nome inclui data)."""
    try:
        ts_date = datetime.now().strftime("%Y%m%d")
        stem = Path(BASE_DADOS).stem
        dest = Path(BACKUP_DIR) / f"{stem}_{ts_date}.db"
        if not dest.exists() and Path(BASE_DADOS).exists():
            shutil.copy2(BASE_DADOS, dest)
            print(f"💾 Backup diário criado: {dest}")
    except Exception as e:
        print("⚠️ Falha a criar backup diário:", e)


def limpar_backups_antigos():
    """Remove backups mais antigos que BACKUP_RETENCAO_DIAS dias."""
    if BACKUP_RETENCAO_DIAS is None:
        return
    try:
        limite = datetime.now() - timedelta(days=BACKUP_RETENCAO_DIAS)
        removidos = 0
        for f in Path(BACKUP_DIR).glob("*.db"):
            try:
                if datetime.fromtimestamp(f.stat().st_mtime) < limite:
                    f.unlink()
                    removidos += 1
            except Exception:
                pass
        if removidos:
            print(
                f"🗑️  {removidos} backup(s) antigo(s) removido(s) (retenção: {BACKUP_RETENCAO_DIAS} dias)."
            )
    except Exception as e:
        logging.warning(f"limpar_backups_antigos falhou: {e}")


SCHEMA_SQL = r"""
PRAGMA foreign_keys=ON;

DROP TRIGGER IF EXISTS refeicoes_no_holiday_i;
DROP TRIGGER IF EXISTS refeicoes_no_holiday_u;
DROP TRIGGER IF EXISTS capacidade_check_pa_i;
DROP TRIGGER IF EXISTS capacidade_check_pa_u;
DROP TRIGGER IF EXISTS capacidade_check_lanche_i;
DROP TRIGGER IF EXISTS capacidade_check_lanche_u;
DROP TRIGGER IF EXISTS capacidade_check_almoco_i;
DROP TRIGGER IF EXISTS capacidade_check_almoco_u;
DROP TRIGGER IF EXISTS capacidade_check_jantar_i;
DROP TRIGGER IF EXISTS capacidade_check_jantar_u;

-- -----------------------------------------------------------------------
-- TABELAS PRINCIPAIS (criadas se não existirem)
-- -----------------------------------------------------------------------
CREATE TABLE IF NOT EXISTS utilizadores (
  id                   INTEGER PRIMARY KEY AUTOINCREMENT,
  NII                  TEXT UNIQUE NOT NULL,
  NI                   TEXT UNIQUE NOT NULL,
  Nome_completo        TEXT NOT NULL,
  Palavra_chave        TEXT NOT NULL,
  ano                  INTEGER NOT NULL,
  perfil               TEXT DEFAULT 'aluno',
  locked_until         TEXT,
  must_change_password INTEGER DEFAULT 0,
  password_updated_at  TEXT,
  is_active            INTEGER NOT NULL DEFAULT 1,
  email                TEXT,
  telemovel            TEXT,
  data_criacao         TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS refeicoes (
  id                 INTEGER PRIMARY KEY AUTOINCREMENT,
  utilizador_id      INTEGER NOT NULL
                       REFERENCES utilizadores(id) ON DELETE CASCADE,
  data               TEXT NOT NULL,
  pequeno_almoco     BOOLEAN DEFAULT 0,
  lanche             BOOLEAN DEFAULT 0,
  almoco             TEXT CHECK(almoco IN ('Normal','Vegetariano','Dieta')),
  jantar_tipo        TEXT CHECK(jantar_tipo IN ('Normal','Vegetariano','Dieta')),
  jantar_sai_unidade BOOLEAN DEFAULT 0,
  UNIQUE(utilizador_id, data)
);

CREATE TABLE IF NOT EXISTS login_eventos (
  id        INTEGER PRIMARY KEY AUTOINCREMENT,
  nii       TEXT NOT NULL,
  sucesso   INTEGER NOT NULL,  -- 0/1
  ip        TEXT,
  criado_em TEXT NOT NULL DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_login_eventos_nii_data ON login_eventos(nii, criado_em);
CREATE INDEX IF NOT EXISTS idx_login_eventos_ip_data  ON login_eventos(ip, criado_em);

CREATE TABLE IF NOT EXISTS calendario_operacional (
  data TEXT PRIMARY KEY,
  tipo TEXT NOT NULL CHECK (tipo IN ('normal','fim_semana','feriado','exercicio','outro')),
  nota TEXT
);

-- Ausências prolongadas de utilizadores
-- O auto-preenchimento não cria refeições para utilizadores ausentes neste intervalo.
CREATE TABLE IF NOT EXISTS ausencias (
  id           INTEGER PRIMARY KEY AUTOINCREMENT,
  utilizador_id INTEGER NOT NULL REFERENCES utilizadores(id) ON DELETE CASCADE,
  ausente_de   TEXT NOT NULL,  -- data ISO YYYY-MM-DD
  ausente_ate  TEXT NOT NULL,  -- data ISO YYYY-MM-DD (inclusive)
  motivo       TEXT,
  criado_em    TEXT NOT NULL DEFAULT (datetime('now','localtime')),
  criado_por   TEXT           -- NII de quem registou
);
CREATE INDEX IF NOT EXISTS idx_ausencias_uid  ON ausencias(utilizador_id);
CREATE INDEX IF NOT EXISTS idx_ausencias_datas ON ausencias(ausente_de, ausente_ate);

-- Detenções de cadetes (impede sair da unidade após jantar)
CREATE TABLE IF NOT EXISTS detencoes (
  id            INTEGER PRIMARY KEY AUTOINCREMENT,
  utilizador_id INTEGER NOT NULL REFERENCES utilizadores(id) ON DELETE CASCADE,
  detido_de     TEXT NOT NULL,  -- YYYY-MM-DD
  detido_ate    TEXT NOT NULL,  -- YYYY-MM-DD (inclusive)
  motivo        TEXT,
  criado_em     TEXT NOT NULL DEFAULT (datetime('now','localtime')),
  criado_por    TEXT            -- NII de quem registou (cmd/admin)
);
CREATE INDEX IF NOT EXISTS idx_detencoes_uid   ON detencoes(utilizador_id);
CREATE INDEX IF NOT EXISTS idx_detencoes_datas ON detencoes(detido_de, detido_ate);

-- Licenças de saída
CREATE TABLE IF NOT EXISTS licencas (
  id            INTEGER PRIMARY KEY AUTOINCREMENT,
  utilizador_id INTEGER NOT NULL REFERENCES utilizadores(id) ON DELETE CASCADE,
  data          TEXT NOT NULL,                -- YYYY-MM-DD
  tipo          TEXT NOT NULL CHECK(tipo IN ('antes_jantar','apos_jantar')),
  aprovado_por  TEXT,                         -- NULL = auto-aprovado por regra, NII se exceção
  criado_em     TEXT NOT NULL DEFAULT (datetime('now','localtime')),
  hora_saida    TEXT,                         -- HH:MM — registado pelo oficial de dia
  hora_entrada  TEXT,                         -- HH:MM — registado pelo oficial de dia
  UNIQUE(utilizador_id, data)
);
CREATE INDEX IF NOT EXISTS idx_licencas_uid  ON licencas(utilizador_id);
CREATE INDEX IF NOT EXISTS idx_licencas_data ON licencas(data);

-- Log de auditoria de alterações de refeições
CREATE TABLE IF NOT EXISTS refeicoes_log (
  id            INTEGER PRIMARY KEY AUTOINCREMENT,
  utilizador_id INTEGER NOT NULL,
  data_refeicao TEXT NOT NULL,
  campo         TEXT NOT NULL,   -- ex: 'almoco', 'jantar_tipo', 'lanche'
  valor_antes   TEXT,
  valor_depois  TEXT,
  alterado_por  TEXT NOT NULL,   -- NII de quem fez a alteração
  alterado_em   TEXT NOT NULL DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_rlog_uid  ON refeicoes_log(utilizador_id);
CREATE INDEX IF NOT EXISTS idx_rlog_data ON refeicoes_log(data_refeicao);
CREATE INDEX IF NOT EXISTS idx_rlog_por  ON refeicoes_log(alterado_por);

CREATE TABLE IF NOT EXISTS menus_diarios (
  data           TEXT PRIMARY KEY,
  pequeno_almoco TEXT,
  lanche         TEXT,
  almoco_normal  TEXT,
  almoco_veg     TEXT,
  almoco_dieta   TEXT,
  jantar_normal  TEXT,
  jantar_veg     TEXT,
  jantar_dieta   TEXT
);

CREATE TABLE IF NOT EXISTS capacidade_refeicao (
  data      TEXT NOT NULL,
  refeicao  TEXT NOT NULL CHECK (refeicao IN ('Pequeno Almoço','Lanche','Almoço','Jantar')),
  max_total INTEGER NOT NULL CHECK (max_total >= 0),
  PRIMARY KEY (data, refeicao)
);

CREATE TABLE IF NOT EXISTS capacidade_excessos (
  id         INTEGER PRIMARY KEY AUTOINCREMENT,
  data       TEXT NOT NULL,
  refeicao   TEXT NOT NULL,
  ocupacao   INTEGER NOT NULL,
  capacidade INTEGER NOT NULL,
  criado_em  TEXT NOT NULL DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_capex_data ON capacidade_excessos(data);

CREATE TABLE IF NOT EXISTS admin_audit_log (
  id     INTEGER PRIMARY KEY AUTOINCREMENT,
  ts     TEXT DEFAULT (datetime('now','localtime')),
  actor  TEXT NOT NULL,
  action TEXT NOT NULL,
  detail TEXT
);

CREATE TABLE IF NOT EXISTS turmas (
  id        INTEGER PRIMARY KEY AUTOINCREMENT,
  nome      TEXT NOT NULL UNIQUE,
  ano       INTEGER NOT NULL,
  descricao TEXT,
  criado_em TEXT DEFAULT (datetime('now','localtime'))
);

CREATE INDEX IF NOT EXISTS idx_refeicoes_data ON refeicoes(data);
CREATE INDEX IF NOT EXISTS idx_refeicoes_user ON refeicoes(utilizador_id);
CREATE INDEX IF NOT EXISTS idx_utilizadores_ano ON utilizadores(ano);

CREATE VIRTUAL TABLE IF NOT EXISTS utilizadores_fts USING fts5(
  Nome_completo,
  content='utilizadores',
  content_rowid='id'
);
CREATE TRIGGER IF NOT EXISTS utilizadores_ai_fts
AFTER INSERT ON utilizadores BEGIN
  INSERT INTO utilizadores_fts(rowid, Nome_completo) VALUES (NEW.id, NEW.Nome_completo);
END;
CREATE TRIGGER IF NOT EXISTS utilizadores_ad_fts
AFTER DELETE ON utilizadores BEGIN
  INSERT INTO utilizadores_fts(utilizadores_fts, rowid) VALUES('delete', OLD.id);
END;
CREATE TRIGGER IF NOT EXISTS utilizadores_au_fts
AFTER UPDATE ON utilizadores BEGIN
  INSERT INTO utilizadores_fts(utilizadores_fts, rowid) VALUES('delete', OLD.id);
  INSERT INTO utilizadores_fts(rowid, Nome_completo) VALUES (NEW.id, NEW.Nome_completo);
END;

CREATE TRIGGER IF NOT EXISTS refeicoes_chk_values
BEFORE INSERT ON refeicoes
BEGIN
  SELECT
    CASE
      WHEN NEW.pequeno_almoco NOT IN (0,1) THEN RAISE(ABORT,'pequeno_almoco inválido')
      WHEN NEW.lanche NOT IN (0,1) THEN RAISE(ABORT,'lanche inválido')
      WHEN NEW.jantar_sai_unidade NOT IN (0,1) THEN RAISE(ABORT,'jantar_sai_unidade inválido')
      WHEN NEW.almoco IS NOT NULL AND NEW.almoco NOT IN ('Normal','Vegetariano','Dieta') THEN RAISE(ABORT,'almoco inválido')
      WHEN NEW.jantar_tipo IS NOT NULL AND NEW.jantar_tipo NOT IN ('Normal','Vegetariano','Dieta') THEN RAISE(ABORT,'jantar_tipo inválido')
    END;
END;
CREATE TRIGGER IF NOT EXISTS refeicoes_chk_values_u
BEFORE UPDATE ON refeicoes
BEGIN
  SELECT
    CASE
      WHEN NEW.pequeno_almoco NOT IN (0,1) THEN RAISE(ABORT,'pequeno_almoco inválido')
      WHEN NEW.lanche NOT IN (0,1) THEN RAISE(ABORT,'lanche inválido')
      WHEN NEW.jantar_sai_unidade NOT IN (0,1) THEN RAISE(ABORT,'jantar_sai_unidade inválido')
      WHEN NEW.almoco IS NOT NULL AND NEW.almoco NOT IN ('Normal','Vegetariano','Dieta') THEN RAISE(ABORT,'almoco inválido')
      WHEN NEW.jantar_tipo IS NOT NULL AND NEW.jantar_tipo NOT IN ('Normal','Vegetariano','Dieta') THEN RAISE(ABORT,'jantar_tipo inválido')
    END;
END;

CREATE VIEW IF NOT EXISTS v_ocupacao_dia AS
SELECT
  d.data,
  'Pequeno Almoço' AS refeicao,
  (SELECT COUNT(*) FROM refeicoes r
   JOIN utilizadores u ON u.id=r.utilizador_id AND u.is_active=1
   AND NOT EXISTS (SELECT 1 FROM ausencias a WHERE a.utilizador_id=u.id AND a.ausente_de<=r.data AND a.ausente_ate>=r.data)
   WHERE r.data=d.data AND r.pequeno_almoco=1) AS ocupacao,
  COALESCE((SELECT max_total FROM capacidade_refeicao c WHERE c.data=d.data AND c.refeicao='Pequeno Almoço'), -1) AS capacidade
FROM (SELECT DISTINCT data FROM refeicoes) d
UNION ALL
SELECT
  d.data, 'Lanche',
  (SELECT COUNT(*) FROM refeicoes r
   JOIN utilizadores u ON u.id=r.utilizador_id AND u.is_active=1
   AND NOT EXISTS (SELECT 1 FROM ausencias a WHERE a.utilizador_id=u.id AND a.ausente_de<=r.data AND a.ausente_ate>=r.data)
   WHERE r.data=d.data AND r.lanche=1),
  COALESCE((SELECT max_total FROM capacidade_refeicao c WHERE c.data=d.data AND c.refeicao='Lanche'), -1)
FROM (SELECT DISTINCT data FROM refeicoes) d
UNION ALL
SELECT
  d.data, 'Almoço',
  (SELECT COUNT(*) FROM refeicoes r
   JOIN utilizadores u ON u.id=r.utilizador_id AND u.is_active=1
   AND NOT EXISTS (SELECT 1 FROM ausencias a WHERE a.utilizador_id=u.id AND a.ausente_de<=r.data AND a.ausente_ate>=r.data)
   WHERE r.data=d.data AND r.almoco IS NOT NULL),
  COALESCE((SELECT max_total FROM capacidade_refeicao c WHERE c.data=d.data AND c.refeicao='Almoço'), -1)
FROM (SELECT DISTINCT data FROM refeicoes) d
UNION ALL
SELECT
  d.data, 'Jantar',
  (SELECT COUNT(*) FROM refeicoes r
   JOIN utilizadores u ON u.id=r.utilizador_id AND u.is_active=1
   AND NOT EXISTS (SELECT 1 FROM ausencias a WHERE a.utilizador_id=u.id AND a.ausente_de<=r.data AND a.ausente_ate>=r.data)
   WHERE r.data=d.data AND r.jantar_tipo IS NOT NULL),
  COALESCE((SELECT max_total FROM capacidade_refeicao c WHERE c.data=d.data AND c.refeicao='Jantar'), -1)
FROM (SELECT DISTINCT data FROM refeicoes) d;

CREATE TRIGGER IF NOT EXISTS cap_log_pa
AFTER INSERT ON refeicoes
WHEN NEW.pequeno_almoco=1 AND EXISTS (
  SELECT 1 FROM capacidade_refeicao c WHERE c.data=NEW.data AND c.refeicao='Pequeno Almoço'
) AND (
  (SELECT COUNT(*) FROM refeicoes r WHERE r.data=NEW.data AND r.pequeno_almoco=1) >
  (SELECT max_total FROM capacidade_refeicao c WHERE c.data=NEW.data AND c.refeicao='Pequeno Almoço')
)
BEGIN
  INSERT INTO capacidade_excessos(data,refeicao,ocupacao,capacidade)
  SELECT NEW.data,'Pequeno Almoço',
         (SELECT COUNT(*) FROM refeicoes r WHERE r.data=NEW.data AND r.pequeno_almoco=1),
         (SELECT max_total FROM capacidade_refeicao c WHERE c.data=NEW.data AND c.refeicao='Pequeno Almoço');
END;

CREATE TRIGGER IF NOT EXISTS cap_log_lanche
AFTER INSERT ON refeicoes
WHEN NEW.lanche=1 AND EXISTS (
  SELECT 1 FROM capacidade_refeicao c WHERE c.data=NEW.data AND c.refeicao='Lanche'
) AND (
  (SELECT COUNT(*) FROM refeicoes r WHERE r.data=NEW.data AND r.lanche=1) >
  (SELECT max_total FROM capacidade_refeicao c WHERE c.data=NEW.data AND c.refeicao='Lanche')
)
BEGIN
  INSERT INTO capacidade_excessos(data,refeicao,ocupacao,capacidade)
  SELECT NEW.data,'Lanche',
         (SELECT COUNT(*) FROM refeicoes r WHERE r.data=NEW.data AND r.lanche=1),
         (SELECT max_total FROM capacidade_refeicao c WHERE c.data=NEW.data AND c.refeicao='Lanche');
END;

CREATE TRIGGER IF NOT EXISTS cap_log_almoco
AFTER INSERT ON refeicoes
WHEN NEW.almoco IS NOT NULL AND EXISTS (
  SELECT 1 FROM capacidade_refeicao c WHERE c.data=NEW.data AND c.refeicao='Almoço'
) AND (
  (SELECT COUNT(*) FROM refeicoes r WHERE r.data=NEW.data AND r.almoco IS NOT NULL) >
  (SELECT max_total FROM capacidade_refeicao c WHERE c.data=NEW.data AND c.refeicao='Almoço')
)
BEGIN
  INSERT INTO capacidade_excessos(data,refeicao,ocupacao,capacidade)
  SELECT NEW.data,'Almoço',
         (SELECT COUNT(*) FROM refeicoes r WHERE r.data=NEW.data AND r.almoco IS NOT NULL),
         (SELECT max_total FROM capacidade_refeicao c WHERE c.data=NEW.data AND c.refeicao='Almoço');
END;

CREATE TRIGGER IF NOT EXISTS cap_log_jantar
AFTER INSERT ON refeicoes
WHEN NEW.jantar_tipo IS NOT NULL AND EXISTS (
  SELECT 1 FROM capacidade_refeicao c WHERE c.data=NEW.data AND c.refeicao='Jantar'
) AND (
  (SELECT COUNT(*) FROM refeicoes r WHERE r.data=NEW.data AND r.jantar_tipo IS NOT NULL) >
  (SELECT max_total FROM capacidade_refeicao c WHERE c.data=NEW.data AND c.refeicao='Jantar')
)
BEGIN
  INSERT INTO capacidade_excessos(data,refeicao,ocupacao,capacidade)
  SELECT NEW.data,'Jantar',
         (SELECT COUNT(*) FROM refeicoes r WHERE r.data=NEW.data AND r.jantar_tipo IS NOT NULL),
         (SELECT max_total FROM capacidade_refeicao c WHERE c.data=NEW.data AND c.refeicao='Jantar');
END;
"""


def ensure_schema():
    with db() as conn:
        # Verificar se a FTS5 está íntegra antes de recriar
        fts_ok = False
        try:
            conn.execute("SELECT COUNT(*) FROM utilizadores_fts").fetchone()
            fts_ok = True
        except Exception:
            pass

        if not fts_ok:
            # FTS corrompida ou inexistente — recriar
            try:
                conn.execute("DROP TABLE IF EXISTS utilizadores_fts")
            except sqlite3.Error:
                pass

        conn.executescript(SCHEMA_SQL)

        if not fts_ok:
            try:
                conn.execute(
                    "INSERT INTO utilizadores_fts(utilizadores_fts) VALUES('rebuild')"
                )
            except sqlite3.Error:
                pass
        conn.commit()


# ===========================================================================
# AUTENTICAÇÃO
# ===========================================================================


def verify_password(pw: str, stored: str) -> bool:
    """Verifica password; suporta hashes werkzeug e password em claro (legado)."""
    stored = stored or ""
    if stored.startswith(("pbkdf2:", "scrypt:", "argon2:")) and _wz_check_password_hash:
        try:
            return bool(_wz_check_password_hash(stored, pw))
        except Exception:
            return False
    return pw == stored


def reg_login(nii: str, ok: int, ip: Optional[str] = None):
    """Regista evento de login na BD (com IP opcional)."""
    try:
        ip = (ip or "127.0.0.1")[:64]
        with db() as conn:
            conn.execute(
                "INSERT INTO login_eventos(nii,sucesso,ip) VALUES (?,?,?)",
                (nii, ok, ip),
            )
            conn.commit()
    except sqlite3.Error:
        pass


def recent_failures(nii: str, minutes: int = 10) -> int:
    """Conta tentativas falhadas recentes. Bug da v7.x corrigido."""
    with db() as conn:
        modifier = f"-{minutes} minutes"
        r = conn.execute(
            """SELECT COUNT(*) c FROM login_eventos
               WHERE nii=? AND sucesso=0
               AND criado_em >= datetime('now','localtime',?)""",
            (nii, modifier),
        ).fetchone()
        return r["c"] if r else 0


def recent_failures_by_ip(ip: str, minutes: int = 15) -> int:
    """Conta tentativas falhadas recentes por IP (proteção contra ataques distribuídos)."""
    with db() as conn:
        modifier = f"-{minutes} minutes"
        r = conn.execute(
            """SELECT COUNT(*) c FROM login_eventos
               WHERE ip=? AND sucesso=0
               AND criado_em >= datetime('now','localtime',?)""",
            (ip, modifier),
        ).fetchone()
        return r["c"] if r else 0


def block_user(nii: str, minutes: int = 15):
    with db() as conn:
        conn.execute(
            "UPDATE utilizadores SET locked_until=datetime('now','localtime',?) WHERE NII=?",
            (f"+{minutes} minutes", nii),
        )
        conn.commit()


def existe_admin() -> bool:
    with db() as conn:
        r = conn.execute(
            "SELECT COUNT(*) c FROM utilizadores WHERE perfil='admin'"
        ).fetchone()
        return bool(r and r["c"] > 0)


def user_by_nii(nii: str):
    nii = (nii or "").strip()
    if not nii:
        return None
    with db() as conn:
        r = conn.execute(
            "SELECT * FROM utilizadores WHERE NII = ? COLLATE NOCASE", (nii,)
        ).fetchone()
        return dict(r) if r else None


def user_by_ni(ni: str):
    ni = (ni or "").strip()
    if not ni:
        return None
    with db() as conn:
        r = conn.execute("SELECT * FROM utilizadores WHERE NI = ?", (ni,)).fetchone()
        return r


def user_id_by_nii(nii: str) -> Optional[int]:
    u = user_by_nii(nii)
    return u["id"] if u else None


def login_flow() -> Optional[dict]:
    tries = 0
    while tries < 3:
        clear()
        line("🔐 LOGIN")
        nii = input("NII: ").strip()
        print("(Nota: a password é mascarada com '*'.)")
        pw = masked_input("Password: ")

        # Login via BD (contas de sistema devem existir na BD; sem bypass por PERFIS_ADMIN)
        # Fallback admin de emergência
        if (
            not existe_admin()
            and nii == FALLBACK_ADMIN["nii"]
            and pw == FALLBACK_ADMIN["pw"]
        ):
            reg_login(nii, 1)
            return {
                "id": 0,
                "nii": nii,
                "ni": "",
                "nome": FALLBACK_ADMIN["nome"],
                "ano": "",
                "perfil": "admin",
            }

        # Utilizador na BD
        u = user_by_nii(nii)
        if not u:
            print("❌ NII não encontrado.")
            reg_login(nii, 0)
            tries += 1
            input("ENTER...")
            continue

        # Bloqueio temporário
        if u.get("locked_until"):
            try:
                if datetime.fromisoformat(u["locked_until"]) > datetime.now():
                    print(f"⛔ Conta bloqueada até {u['locked_until']}.")
                    reg_login(nii, 0)
                    input("ENTER...")
                    return None
            except Exception:
                pass

        # Verificar password
        ok = verify_password(pw, u["Palavra_chave"])

        reg_login(nii, 1 if ok else 0)
        if not ok:
            print("❌ Password errada.")
            if recent_failures(nii, 10) >= 5:
                block_user(nii, 15)
                print("⚠️ Múltiplas falhas: conta bloqueada por 15 minutos.")
                return None
            tries += 1
            input("ENTER...")
            continue

        return {
            "id": u["id"],
            "nii": u["NII"],
            "ni": u["NI"],
            "nome": u["Nome_completo"],
            "ano": str(u["ano"]),
            "perfil": u["perfil"] or "aluno",
        }

    print("⛔ Tentativas esgotadas.")
    return None


# ===========================================================================
# PRAZO LIMITE PARA ALTERAÇÃO DE REFEIÇÕES
# ===========================================================================


def refeicao_editavel(d: date) -> tuple[bool, str]:
    """
    Devolve (True, '') se a data d ainda pode ser editada,
    ou (False, motivo) caso contrário.

    Regras:
    - Não se pode editar datas passadas.
    - Se PRAZO_LIMITE_HORAS estiver definido, não se pode alterar uma refeição
      a menos de PRAZO_LIMITE_HORAS horas do início desse dia (meia-noite).
      Com 48h: a refeição de 4ª feira só pode ser alterada até 2ª à meia-noite.
    """
    agora_dt = datetime.now()
    hoje = agora_dt.date()

    if d < hoje:
        return (
            False,
            f"Não é possível alterar refeições de datas passadas ({d.strftime('%d/%m/%Y')}).",
        )

    if PRAZO_LIMITE_HORAS is not None:
        # Prazo = meia-noite do dia d menos PRAZO_LIMITE_HORAS
        prazo_dt = datetime(d.year, d.month, d.day, 0, 0, 0) - timedelta(
            hours=PRAZO_LIMITE_HORAS
        )
        if agora_dt >= prazo_dt:
            prazo_str = prazo_dt.strftime("%d/%m/%Y às %H:%M")
            return False, (
                f"⛔ Prazo excedido para alterar a refeição de {d.strftime('%d/%m/%Y')}.\n"
                f"   O prazo terminou em {prazo_str} ({PRAZO_LIMITE_HORAS}h antes da refeição).\n"
                f"   Para efetuar alterações, fala com o Oficial de Dia."
            )

    return True, ""


# ===========================================================================
# QUERIES CONSOLIDADAS — substitui código repetido em múltiplos sítios
# ===========================================================================


def get_totais_dia(di: str, ano: Optional[int] = None) -> dict:
    """
    Devolve totais de todas as refeições para uma data ISO (di).
    Se ano for fornecido, filtra apenas utilizadores desse ano.
    Exclui utilizadores inativos e ausentes nesse dia.
    Resultado: dict com chaves pa, lan, alm_norm, alm_veg, alm_dieta,
                              jan_norm, jan_veg, jan_dieta, jan_sai
    """
    # Filtro comum: utilizador ativo e sem ausência registada para o dia
    _active = (
        "JOIN utilizadores u ON u.id=r.utilizador_id"
        " AND u.is_active=1"
        " AND NOT EXISTS ("
        "SELECT 1 FROM ausencias a"
        " WHERE a.utilizador_id=u.id AND a.ausente_de<=r.data AND a.ausente_ate>=r.data)"
    )
    _ano_cond = " AND u.ano=?" if ano is not None else ""

    with db() as conn:
        params_base = (di, ano) if ano is not None else (di,)

        pa = (
            conn.execute(
                f"SELECT COUNT(*) c FROM refeicoes r {_active}"
                f" WHERE r.data=? {_ano_cond} AND r.pequeno_almoco=1",
                params_base,
            ).fetchone()["c"]
            or 0
        )
        ln = (
            conn.execute(
                f"SELECT COUNT(*) c FROM refeicoes r {_active}"
                f" WHERE r.data=? {_ano_cond} AND r.lanche=1",
                params_base,
            ).fetchone()["c"]
            or 0
        )
        alm = conn.execute(
            f"""
            SELECT
              SUM(CASE WHEN r.almoco='Normal'      THEN 1 ELSE 0 END) norm,
              SUM(CASE WHEN r.almoco='Vegetariano' THEN 1 ELSE 0 END) veg,
              SUM(CASE WHEN r.almoco='Dieta'       THEN 1 ELSE 0 END) dieta
            FROM refeicoes r {_active}
            WHERE r.data=? {_ano_cond}
        """,
            params_base,
        ).fetchone()
        jan = conn.execute(
            f"""
            SELECT
              SUM(CASE WHEN r.jantar_tipo='Normal'      THEN 1 ELSE 0 END) norm,
              SUM(CASE WHEN r.jantar_tipo='Vegetariano' THEN 1 ELSE 0 END) veg,
              SUM(CASE WHEN r.jantar_tipo='Dieta'       THEN 1 ELSE 0 END) dieta,
              SUM(COALESCE(r.jantar_sai_unidade, 0))                        sai
            FROM refeicoes r {_active}
            WHERE r.data=? {_ano_cond}
        """,
            params_base,
        ).fetchone()

    return {
        "pa": pa,
        "lan": ln,
        "alm_norm": alm["norm"] or 0,
        "alm_veg": alm["veg"] or 0,
        "alm_dieta": alm["dieta"] or 0,
        "jan_norm": jan["norm"] or 0,
        "jan_veg": jan["veg"] or 0,
        "jan_dieta": jan["dieta"] or 0,
        "jan_sai": jan["sai"] or 0,
    }


def get_ocupacao_capacidade(d: date) -> dict:
    """Devolve ocupação e capacidade por refeição (capacidade -1 => sem limite)."""
    t = get_totais_dia(d.isoformat())
    with db() as conn:
        caps = {
            r["refeicao"]: r["max_total"]
            for r in conn.execute(
                "SELECT refeicao,max_total FROM capacidade_refeicao WHERE data=?",
                (d.isoformat(),),
            )
        }
    return {
        "Pequeno Almoço": (t["pa"], caps.get("Pequeno Almoço", -1)),
        "Lanche": (t["lan"], caps.get("Lanche", -1)),
        "Almoço": (
            t["alm_norm"] + t["alm_veg"] + t["alm_dieta"],
            caps.get("Almoço", -1),
        ),
        "Jantar": (
            t["jan_norm"] + t["jan_veg"] + t["jan_dieta"],
            caps.get("Jantar", -1),
        ),
    }


def get_menu_do_dia(d: date) -> dict:
    with db() as conn:
        r = conn.execute(
            "SELECT * FROM menus_diarios WHERE data=?", (d.isoformat(),)
        ).fetchone()
        return dict(r) if r else {}


def _totais_para_csv_row(di: str, t: dict, extra: dict = None) -> dict:
    row = {
        "data": di,
        "PA_total": t["pa"],
        "Lanche_total": t["lan"],
        "Almoco_Normal": t["alm_norm"],
        "Almoco_Vegetariano": t["alm_veg"],
        "Almoco_Dieta": t["alm_dieta"],
        "Jantar_Normal": t["jan_norm"],
        "Jantar_Vegetariano": t["jan_veg"],
        "Jantar_Dieta": t["jan_dieta"],
        "Jantar_Saem_Unidade": t["jan_sai"],
    }
    if extra:
        row.update(extra)
    return row


_HEADERS_TOTAIS = [
    "data",
    "PA_total",
    "Lanche_total",
    "Almoco_Normal",
    "Almoco_Vegetariano",
    "Almoco_Dieta",
    "Jantar_Normal",
    "Jantar_Vegetariano",
    "Jantar_Dieta",
    "Jantar_Saem_Unidade",
]
_HEADERS_DISTRIBUICAO = [
    "ano",
    "NI",
    "Nome_completo",
    "data",
    "pequeno_almoco",
    "lanche",
    "almoco",
    "jantar_tipo",
    "jantar_sai_unidade",
]

# ===========================================================================
# REFEIÇÕES — CRUD
# ===========================================================================


def refeicao_get(uid: int, d: date) -> dict:
    with db() as conn:
        r = conn.execute(
            "SELECT * FROM refeicoes WHERE utilizador_id=? AND data=?",
            (uid, d.isoformat()),
        ).fetchone()
        if r:
            return dict(r)
    return {
        "pequeno_almoco": 0,
        "lanche": 0,
        "almoco": None,
        "jantar_tipo": None,
        "jantar_sai_unidade": 0,
    }


def refeicoes_batch(uid: int, d_de: date, d_ate: date) -> dict:
    """Carrega refeições de um aluno para um intervalo de datas. Devolve {iso_date: dict}."""
    defaults = {
        "pequeno_almoco": 0,
        "lanche": 0,
        "almoco": None,
        "jantar_tipo": None,
        "jantar_sai_unidade": 0,
    }
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM refeicoes WHERE utilizador_id=? AND data>=? AND data<=?",
            (uid, d_de.isoformat(), d_ate.isoformat()),
        ).fetchall()
    result = {}
    for r in rows:
        result[r["data"]] = dict(r)
    return result, defaults


def dias_operacionais_batch(d_de: date, d_ate: date) -> dict:
    """Carrega tipos de dia do calendário operacional. Devolve {iso_date: tipo}."""
    with db() as conn:
        rows = conn.execute(
            "SELECT data, tipo FROM calendario_operacional WHERE data>=? AND data<=?",
            (d_de.isoformat(), d_ate.isoformat()),
        ).fetchall()
    result = {}
    for r in rows:
        result[r["data"]] = r["tipo"]
    return result


def ausencias_batch(uid: int, d_de: date, d_ate: date) -> set:
    """Devolve conjunto de datas (ISO str) com ausência ativa no intervalo."""
    with db() as conn:
        rows = conn.execute(
            """SELECT ausente_de, ausente_ate FROM ausencias
               WHERE utilizador_id=? AND ausente_ate>=? AND ausente_de<=?""",
            (uid, d_de.isoformat(), d_ate.isoformat()),
        ).fetchall()
    dates = set()
    for r in rows:
        a_de = date.fromisoformat(r["ausente_de"])
        a_ate = date.fromisoformat(r["ausente_ate"])
        d = max(a_de, d_de)
        while d <= min(a_ate, d_ate):
            dates.add(d.isoformat())
            d += timedelta(days=1)
    return dates


def detencoes_batch(uid: int, d_de: date, d_ate: date) -> set:
    """Devolve conjunto de datas (ISO str) com detenção ativa no intervalo."""
    with db() as conn:
        rows = conn.execute(
            """SELECT detido_de, detido_ate FROM detencoes
               WHERE utilizador_id=? AND detido_ate>=? AND detido_de<=?""",
            (uid, d_de.isoformat(), d_ate.isoformat()),
        ).fetchall()
    dates = set()
    for r in rows:
        d_de_r = date.fromisoformat(r["detido_de"])
        d_ate_r = date.fromisoformat(r["detido_ate"])
        d = max(d_de_r, d_de)
        while d <= min(d_ate_r, d_ate):
            dates.add(d.isoformat())
            d += timedelta(days=1)
    return dates


def licencas_batch(uid: int, d_de: date, d_ate: date) -> dict:
    """Carrega licenças de um aluno para um intervalo. Devolve {iso_date: tipo}."""
    with db() as conn:
        rows = conn.execute(
            "SELECT data, tipo FROM licencas WHERE utilizador_id=? AND data>=? AND data<=?",
            (uid, d_de.isoformat(), d_ate.isoformat()),
        ).fetchall()
    return {r["data"]: r["tipo"] for r in rows}


def refeicao_save(uid: int, d: date, r: dict, alterado_por: str = "sistema") -> bool:
    """
    Guarda refeição e regista no log de auditoria os campos que mudaram.
    alterado_por: NII de quem fez a alteração (ou 'sistema' para auto-preenchimento).
    """
    try:
        with db() as conn:
            # Ler estado anterior para auditoria
            anterior = conn.execute(
                "SELECT * FROM refeicoes WHERE utilizador_id=? AND data=?",
                (uid, d.isoformat()),
            ).fetchone()
            anterior = dict(anterior) if anterior else {}

            # Se o cadete estiver detido nesse dia, não pode sair da unidade
            dd = d.isoformat()
            det = conn.execute(
                """SELECT 1 FROM detencoes
                WHERE utilizador_id=? AND detido_de<=? AND detido_ate>=?
                LIMIT 1""",
                (uid, dd, dd),
            ).fetchone()
            if det:
                r["jantar_sai_unidade"] = 0

            conn.execute(
                """
                INSERT INTO refeicoes
                  (utilizador_id, data, pequeno_almoco, lanche, almoco, jantar_tipo, jantar_sai_unidade)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(utilizador_id, data) DO UPDATE SET
                    pequeno_almoco=excluded.pequeno_almoco,
                    lanche=excluded.lanche,
                    almoco=excluded.almoco,
                    jantar_tipo=excluded.jantar_tipo,
                    jantar_sai_unidade=excluded.jantar_sai_unidade
            """,
                (
                    uid,
                    d.isoformat(),
                    r.get("pequeno_almoco", 0),
                    r.get("lanche", 0),
                    r.get("almoco"),
                    r.get("jantar_tipo"),
                    r.get("jantar_sai_unidade", 0),
                ),
            )

            # Registar apenas campos que realmente mudaram
            campos = [
                "pequeno_almoco",
                "lanche",
                "almoco",
                "jantar_tipo",
                "jantar_sai_unidade",
            ]
            for campo in campos:
                val_antes = (
                    str(anterior.get(campo))
                    if anterior.get(campo) is not None
                    else None
                )
                val_depois = str(r.get(campo)) if r.get(campo) is not None else None
                if val_antes != val_depois:
                    conn.execute(
                        """
                        INSERT INTO refeicoes_log
                          (utilizador_id, data_refeicao, campo, valor_antes, valor_depois, alterado_por)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """,
                        (
                            uid,
                            d.isoformat(),
                            campo,
                            val_antes,
                            val_depois,
                            alterado_por,
                        ),
                    )

            conn.commit()
            return True
    except sqlite3.IntegrityError as e:
        print(f"❌ Rejeitado pela BD: {e}")
        return False
    except sqlite3.Error as e:
        print(f"Erro ao salvar: {e}")
        return False


def refeicao_exists(uid: int, d: date) -> bool:
    try:
        with db() as conn:
            r = conn.execute(
                "SELECT 1 FROM refeicoes WHERE utilizador_id=? AND data=? LIMIT 1",
                (uid, d.isoformat()),
            ).fetchone()
            return r is not None
    except Exception:
        return False


# ===========================================================================
# EXPORTAÇÃO — CSV e Excel (novo)
# ===========================================================================


def export_csv(rows: List[dict], headers: List[str], name: str) -> str:
    path = os.path.join(EXPORT_DIR, name + ".csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow([r.get(h, "") for h in headers])
    return path


def export_xlsx(rows: List[dict], headers: List[str], name: str) -> str:
    """Exporta para Excel (.xlsx). Requer openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠️ openpyxl não instalado. Instala com: pip install openpyxl")
        return export_csv(rows, headers, name)

    path = os.path.join(EXPORT_DIR, name + ".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = name[:31]  # Excel limita a 31 chars

    # Cabeçalho com estilo
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))

    # Auto-largura das colunas
    for col_idx, h in enumerate(headers, 1):
        max_len = (
            max(len(str(h)), *(len(str(row.get(h, ""))) for row in rows))
            if rows
            else len(str(h))
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    wb.save(path)
    return path


def export_both(rows: List[dict], headers: List[str], name: str) -> tuple[str, str]:
    """Exporta CSV e XLSX e devolve ambos os caminhos."""
    p1 = export_csv(rows, headers, name)
    p2 = export_xlsx(rows, headers, name)
    return p1, p2


def do_backup() -> bool:
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = Path(BACKUP_DIR) / f"{Path(BASE_DADOS).stem}_{ts}.db"
        shutil.copy2(BASE_DADOS, dest)
        print(f"✅ Backup: {dest}")
        return True
    except Exception as e:
        print("❌ Falha no backup:", e)
        return False


# ===========================================================================
# EXPORTAÇÕES DO DIA (único ponto — sem duplicados)
# ===========================================================================


def exportacoes_do_dia(d: date, ano: Optional[int] = None):
    """
    Gera CSV + XLSX do dia d.
    Se ano for fornecido (modo CMD), filtra só esse ano.
    Gera 3 ficheiros: totais, distribuição nominal e ocupação vs capacidade.
    """
    di = d.isoformat()
    tag = f"_ano{ano}" if ano else ""
    t = get_totais_dia(di, ano)

    # 1) Totais por turno/tipo
    row_sum = _totais_para_csv_row(di, t, {"ano": ano} if ano else {})
    hdrs = (["data", "ano"] + _HEADERS_TOTAIS[1:]) if ano else _HEADERS_TOTAIS
    p1c, p1x = export_both([row_sum], hdrs, f"totais{tag}_{di}")
    print(f"✅ Totais: {p1c} | {p1x}")

    # 2) Distribuição nominal
    with db() as conn:
        if ano is None:
            det = [
                dict(r)
                for r in conn.execute(
                    """
                SELECT u.ano, u.NI, u.Nome_completo, r.data,
                       r.pequeno_almoco, r.lanche, r.almoco, r.jantar_tipo, r.jantar_sai_unidade
                FROM refeicoes r JOIN utilizadores u ON u.id=r.utilizador_id
                WHERE r.data=?
                ORDER BY u.ano, u.NI
            """,
                    (di,),
                )
            ]
        else:
            det = [
                dict(r)
                for r in conn.execute(
                    """
                SELECT u.NII, u.NI, u.Nome_completo, u.ano, r.data,
                       r.pequeno_almoco, r.lanche, r.almoco, r.jantar_tipo, r.jantar_sai_unidade
                FROM refeicoes r JOIN utilizadores u ON u.id=r.utilizador_id
                WHERE r.data=? AND u.ano=?
                ORDER BY u.NI
            """,
                    (di, ano),
                )
            ]
    hdrs_det = (
        [
            "NII",
            "NI",
            "Nome_completo",
            "ano",
            "data",
            "pequeno_almoco",
            "lanche",
            "almoco",
            "jantar_tipo",
            "jantar_sai_unidade",
        ]
        if ano
        else _HEADERS_DISTRIBUICAO
    )
    p2c, p2x = export_both(det, hdrs_det, f"distribuicao{tag}_{di}")
    print(f"✅ Distribuição nominal: {p2c} | {p2x}")

    # 3) Ocupação vs capacidade
    with db() as conn:
        occ_rows = [
            dict(r)
            for r in conn.execute("SELECT * FROM v_ocupacao_dia WHERE data=?", (di,))
        ]
    p3c, p3x = export_both(
        occ_rows,
        ["data", "refeicao", "ocupacao", "capacidade"],
        f"ocupacao_vs_capacidade_{di}",
    )
    print(f"✅ Ocupação vs capacidade: {p3c} | {p3x}")


# ===========================================================================
# AUSÊNCIAS
# ===========================================================================


def utilizador_ausente(uid: int, d: date) -> bool:
    """Devolve True se o utilizador tem ausência registada para a data d."""
    with db() as conn:
        r = conn.execute(
            """
            SELECT 1 FROM ausencias
            WHERE utilizador_id=? AND ausente_de <= ? AND ausente_ate >= ?
            LIMIT 1
        """,
            (uid, d.isoformat(), d.isoformat()),
        ).fetchone()
        return r is not None


def admin_gerir_ausencias(actor_nii: str):
    """Menu de gestão de ausências (admin / oficial de dia)."""
    while True:
        clear()
        line("🏖️  GESTÃO DE AUSÊNCIAS")
        print("  [1] Registar ausência")
        print("  [2] Ver ausências de um utilizador")
        print("  [3] Remover ausência")
        print("  [0] Voltar")
        op = input("Opção: ").strip()

        if op == "1":
            nii = input("NII do utilizador: ").strip()
            with db() as conn:
                u = conn.execute(
                    "SELECT id, Nome_completo FROM utilizadores WHERE NII=?", (nii,)
                ).fetchone()
            if not u:
                print("❌ NII não encontrado.")
                input("ENTER...")
                continue
            de = input("Ausente de (YYYY-MM-DD): ").strip()
            ate = input("Ausente até (YYYY-MM-DD): ").strip()
            mot = input("Motivo (ex: licença, exercício): ").strip() or None
            try:
                datetime.strptime(de, "%Y-%m-%d")
                datetime.strptime(ate, "%Y-%m-%d")
            except ValueError:
                print("❌ Data inválida.")
                input("ENTER...")
                continue
            with db() as conn:
                conn.execute(
                    """
                    INSERT INTO ausencias (utilizador_id, ausente_de, ausente_ate, motivo, criado_por)
                    VALUES (?, ?, ?, ?, ?)
                """,
                    (u["id"], de, ate, mot, actor_nii),
                )
                conn.commit()
            print(f"✅ Ausência registada para {u['Nome_completo']}: {de} → {ate}.")
            input("ENTER...")

        elif op == "2":
            nii = input("NII do utilizador: ").strip()
            with db() as conn:
                u = conn.execute(
                    "SELECT id, Nome_completo FROM utilizadores WHERE NII=?", (nii,)
                ).fetchone()
                if not u:
                    print("❌ NII não encontrado.")
                    input("ENTER...")
                    continue
                rows = conn.execute(
                    """
                    SELECT id, ausente_de, ausente_ate, motivo, criado_em, criado_por
                    FROM ausencias WHERE utilizador_id=? ORDER BY ausente_de DESC
                """,
                    (u["id"],),
                ).fetchall()
            clear()
            line(f"🏖️  Ausências — {u['Nome_completo']}")
            if not rows:
                print("  Sem ausências registadas.")
            else:
                print_table(
                    ["ID", "De", "Até", "Motivo", "Registado em", "Por"],
                    [
                        [
                            r["id"],
                            r["ausente_de"],
                            r["ausente_ate"],
                            r["motivo"] or "-",
                            r["criado_em"],
                            r["criado_por"] or "-",
                        ]
                        for r in rows
                    ],
                )
            input("\nENTER...")

        elif op == "3":
            ausencia_id = input("ID da ausência a remover: ").strip()
            if not ausencia_id.isdigit():
                print("❌ ID inválido.")
                input("ENTER...")
                continue
            with db() as conn:
                cur = conn.execute(
                    "DELETE FROM ausencias WHERE id=?", (int(ausencia_id),)
                )
                conn.commit()
            print("✅ Removida." if cur.rowcount else "❌ ID não encontrado.")
            input("ENTER...")

        elif op == "0":
            break


# ===========================================================================
# CALENDÁRIO OPERACIONAL
# ===========================================================================


def dia_operacional(d: date) -> str:
    """
    Devolve o tipo do dia segundo o calendário operacional.
    Se não estiver no calendário, infere 'fim_semana' ou 'normal'.
    """
    with db() as conn:
        r = conn.execute(
            "SELECT tipo FROM calendario_operacional WHERE data=?", (d.isoformat(),)
        ).fetchone()
    if r:
        return r["tipo"]
    return "fim_semana" if d.weekday() >= 5 else "normal"


def dia_tem_refeicoes(d: date) -> bool:
    """Dias normais têm refeições; feriados e exercícios não."""
    return dia_operacional(d) not in ("feriado", "exercicio", "fim_semana")


def admin_gerir_calendario():
    """Menu de gestão do calendário operacional."""
    while True:
        clear()
        line("📅 CALENDÁRIO OPERACIONAL")
        print("  [1] Ver próximos 30 dias")
        print("  [2] Definir dia")
        print("  [3] Remover entrada (volta ao padrão automático)")
        print("  [0] Voltar")
        op = input("Opção: ").strip()

        if op == "1":
            clear()
            line("📅 Próximos 30 dias")
            hoje = date.today()
            with db() as conn:
                entradas = {
                    r["data"]: r
                    for r in conn.execute(
                        """
                    SELECT data, tipo, nota FROM calendario_operacional
                    WHERE data BETWEEN ? AND ?
                    ORDER BY data
                """,
                        (hoje.isoformat(), (hoje + timedelta(days=30)).isoformat()),
                    )
                }
            ICONE = {
                "normal": "✅",
                "fim_semana": "🔵",
                "feriado": "🔴",
                "exercicio": "🟡",
                "outro": "⚪",
            }
            nomes = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
            for i in range(31):
                d = hoje + timedelta(days=i)
                di = d.isoformat()
                if di in entradas:
                    e = entradas[di]
                    ic = ICONE.get(e["tipo"], "⚪")
                    nota = f" — {e['nota']}" if e["nota"] else ""
                    print(f"  {nomes[d.weekday()]} {di}  {ic} {e['tipo']}{nota}")
                elif d.weekday() >= 5:
                    print(f"  {nomes[d.weekday()]} {di}  🔵 fim_semana (auto)")
            input("\nENTER...")

        elif op == "2":
            d_str = input("Data (YYYY-MM-DD): ").strip()
            try:
                datetime.strptime(d_str, "%Y-%m-%d")
            except ValueError:
                print("❌ Data inválida.")
                input("ENTER...")
                continue
            print("Tipos: normal | fim_semana | feriado | exercicio | outro")
            tipo = input("Tipo: ").strip().lower()
            if tipo not in ("normal", "fim_semana", "feriado", "exercicio", "outro"):
                print("❌ Tipo inválido.")
                input("ENTER...")
                continue
            nota = input("Nota (opcional): ").strip() or None
            with db() as conn:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO calendario_operacional (data, tipo, nota)
                    VALUES (?, ?, ?)
                """,
                    (d_str, tipo, nota),
                )
                conn.commit()
            print(f"✅ {d_str} definido como '{tipo}'.")
            input("ENTER...")

        elif op == "3":
            d_str = input("Data a remover (YYYY-MM-DD): ").strip()
            with db() as conn:
                cur = conn.execute(
                    "DELETE FROM calendario_operacional WHERE data=?", (d_str,)
                )
                conn.commit()
            print("✅ Removido." if cur.rowcount else "❌ Data não encontrada.")
            input("ENTER...")

        elif op == "0":
            break


# ===========================================================================
# LOG DE AUDITORIA
# ===========================================================================


def admin_log_alteracoes():
    """Visualiza o log de alterações de refeições."""
    clear()
    line("📋 LOG DE ALTERAÇÕES DE REFEIÇÕES")
    print("Filtrar por (ENTER para ver tudo):")
    nii_filtro = input("  NII do utilizador: ").strip() or None
    data_filtro = input("  Data da refeição (YYYY-MM-DD): ").strip() or None
    por_filtro = input("  Alterado por (NII): ").strip() or None

    where = []
    params = []
    if nii_filtro:
        where.append("u.NII=?")
        params.append(nii_filtro)
    if data_filtro:
        where.append("l.data_refeicao=?")
        params.append(data_filtro)
    if por_filtro:
        where.append("l.alterado_por=?")
        params.append(por_filtro)

    sql = """
        SELECT l.alterado_em, u.NII, u.Nome_completo, l.data_refeicao,
               l.campo, l.valor_antes, l.valor_depois, l.alterado_por
        FROM refeicoes_log l
        JOIN utilizadores u ON u.id = l.utilizador_id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY l.alterado_em DESC LIMIT 200"

    with db() as conn:
        rows = conn.execute(sql, params).fetchall()

    if not rows:
        print("\n  Sem registos para os filtros aplicados.")
    else:
        print(f"\n  {len(rows)} registo(s) encontrado(s):\n")
        print_table(
            ["Quando", "NII", "Nome", "Data Ref.", "Campo", "Antes", "Depois", "Por"],
            [
                [
                    r["alterado_em"][:16],
                    r["NII"],
                    r["Nome_completo"][:20],
                    r["data_refeicao"],
                    r["campo"],
                    r["valor_antes"] or "-",
                    r["valor_depois"] or "-",
                    r["alterado_por"],
                ]
                for r in rows
            ],
        )
    input("\nENTER para voltar...")


# ===========================================================================
# AUTO-PREENCHIMENTO SEMANAL
# ===========================================================================


def _is_weekday_mon_to_fri(d: date) -> bool:
    return 0 <= d.weekday() <= 4


def _is_friday(d: date) -> bool:
    return d.weekday() == 4


def _default_refeicao_para_dia(d: date) -> dict:
    """Marks all meals Mon-Fri, except dinner on Fridays."""
    if not _is_weekday_mon_to_fri(d):
        return {
            "pequeno_almoco": 0,
            "lanche": 0,
            "almoco": None,
            "jantar_tipo": None,
            "jantar_sai_unidade": 0,
        }
    base = {
        "pequeno_almoco": 1,
        "lanche": 1,
        "almoco": "Normal",
        "jantar_tipo": "Normal",
        "jantar_sai_unidade": 0,
    }
    if _is_friday(d):
        base["jantar_tipo"] = None
        base["jantar_sai_unidade"] = 0
    return base


def _carry_forward_from_last_week(uid: int, d: date, base: dict) -> dict:
    prev = refeicao_get(uid, d - timedelta(days=7))
    out = dict(base)
    for k in [
        "pequeno_almoco",
        "lanche",
        "almoco",
        "jantar_tipo",
        "jantar_sai_unidade",
    ]:
        if k in prev and prev[k] is not None:
            out[k] = prev[k]
    if _is_friday(d):
        out["jantar_tipo"] = None
        out["jantar_sai_unidade"] = 0
    return out


def autopreencher_refeicoes_semanais(dias_a_gerar: int = 14):
    """
    Preenche automaticamente refeições para os próximos dias.
    Respeita:
      - Calendário operacional (não preenche feriados/exercícios)
      - Ausências registadas (não preenche utilizadores ausentes)
    """
    try:
        today = date.today()
        with db() as conn:
            users = [dict(r) for r in conn.execute("SELECT id FROM utilizadores")]
        for u in users:
            uid = u["id"]
            for i in range(dias_a_gerar):
                d = today + timedelta(days=i)
                # Respeitar calendário (feriados, exercícios, fim de semana)
                if not dia_tem_refeicoes(d):
                    continue
                # Respeitar ausências
                if utilizador_ausente(uid, d):
                    continue
                if refeicao_exists(uid, d):
                    continue
                base = _default_refeicao_para_dia(d)
                final = _carry_forward_from_last_week(uid, d, base)
                refeicao_save(uid, d, final, alterado_por="sistema")
    except Exception as e:
        logging.warning(f"autopreencher_refeicoes_semanais falhou: {e}")


# ===========================================================================
# TENDÊNCIAS / DASHBOARD
# ===========================================================================


def period_days(base: date, days: int):
    return [(base - timedelta(days=i)) for i in range(days - 1, -1, -1)]


def series_consumo_por_dia(d0: date, d1: date, ano: Optional[int] = None):
    days = (d1 - d0).days + 1
    idx = {(d0 + timedelta(days=i)).isoformat(): i for i in range(days)}
    pa = [0] * days
    ln = [0] * days
    alm = [0] * days
    jan = [0] * days
    exc = [0] * days

    with db() as conn:
        if ano is None:
            q = """
                SELECT r.data d, SUM(r.pequeno_almoco) pa, SUM(r.lanche) lan,
                       SUM(CASE WHEN r.almoco IS NOT NULL THEN 1 ELSE 0 END) alm,
                       SUM(CASE WHEN r.jantar_tipo IS NOT NULL THEN 1 ELSE 0 END) jan
                FROM refeicoes r WHERE r.data BETWEEN ? AND ?
                GROUP BY r.data"""
            args = (d0.isoformat(), d1.isoformat())
        else:
            q = """
                SELECT r.data d, SUM(r.pequeno_almoco) pa, SUM(r.lanche) lan,
                       SUM(CASE WHEN r.almoco IS NOT NULL THEN 1 ELSE 0 END) alm,
                       SUM(CASE WHEN r.jantar_tipo IS NOT NULL THEN 1 ELSE 0 END) jan
                FROM refeicoes r JOIN utilizadores u ON u.id=r.utilizador_id
                WHERE r.data BETWEEN ? AND ? AND u.ano=?
                GROUP BY r.data"""
            args = (d0.isoformat(), d1.isoformat(), ano)

        for r in conn.execute(q, args):
            i = idx.get(r["d"])
            if i is None:
                continue
            pa[i] = r["pa"] or 0
            ln[i] = r["lan"] or 0
            alm[i] = r["alm"] or 0
            jan[i] = r["jan"] or 0

        for r in conn.execute(
            """
            SELECT data, COALESCE(SUM(ocupacao - capacidade),0) AS over
            FROM capacidade_excessos WHERE data BETWEEN ? AND ?
            GROUP BY data
        """,
            (d0.isoformat(), d1.isoformat()),
        ):
            i = idx.get(r["data"])
            if i is not None:
                exc[i] = r["over"] or 0

    days_list = [(d0 + timedelta(days=i)) for i in range(days)]
    return days_list, pa, ln, alm, jan, exc


def dashboard_analitico(ano: str = "all", dias: int = 14):
    clear()
    titulo = f"📈 Dashboard Analítico - Últimos {dias} dias"
    if ano != "all":
        titulo += f" (Ano {ano})"
    line(titulo)

    d1 = date.today()
    d0 = d1 - timedelta(days=dias - 1)
    ano_int = None if ano == "all" else int(ano)
    days_list, pa, ln, alm, jan, exc = series_consumo_por_dia(d0, d1, ano_int)
    labels = f"{d0.strftime('%d/%m')} … {d1.strftime('%d/%m')}"

    def row(nome, serie):
        return (nome, f"{sparkline(serie)}  | Σ={sum(serie)}")

    print_kv(
        [
            row("Pequeno Almoço", pa),
            row("Lanche", ln),
            row("Almoço", alm),
            row("Jantar", jan),
            row("Excedentes (+)", exc),
        ],
        left=18,
    )

    medias = [
        ("Média PA/dia", round(sum(pa) / len(pa), 1) if pa else 0),
        ("Média Lanche/dia", round(sum(ln) / len(ln), 1) if ln else 0),
        ("Média Almoço/dia", round(sum(alm) / len(alm), 1) if alm else 0),
        ("Média Jantar/dia", round(sum(jan) / len(jan), 1) if jan else 0),
        ("Excedente médio", round(sum(exc) / len(exc), 1) if exc else 0),
    ]
    print("\nMédias:")
    print_kv(medias, left=18)
    print(f"\nJanela: {labels}")
    print("\n [1] Exportar CSV + Excel (janela)  [0] Voltar")
    if input("Opção: ").strip() == "1":
        _export_dashboard_janela(d0, d1, ano_int)
        input("ENTER...")


def _export_dashboard_janela(d0: date, d1: date, ano: Optional[int]):
    days_list, pa, ln, alm, jan, exc = series_consumo_por_dia(d0, d1, ano)
    rows = [
        {
            "data": d.isoformat(),
            "PA": pa[i],
            "Lanche": ln[i],
            "Almoco": alm[i],
            "Jantar": jan[i],
            "Excedente_pos": exc[i],
        }
        for i, d in enumerate(days_list)
    ]
    tag = "all" if ano is None else f"ano{ano}"
    name = f"dashboard_{tag}_{d0.isoformat()}_a_{d1.isoformat()}"
    pc, px = export_both(
        rows, ["data", "PA", "Lanche", "Almoco", "Jantar", "Excedente_pos"], name
    )
    print(f"✅ CSV: {pc}\n✅ Excel: {px}")


# ===========================================================================
# PAINÉIS OPERACIONAIS
# ===========================================================================


def _mostrar_barras_occ(occ: dict):
    print("\nOcupação vs Capacidade:")
    print_kv(
        [
            ("Pequeno Almoço", bar(*occ["Pequeno Almoço"])),
            ("Lanche", bar(*occ["Lanche"])),
            ("Almoço", bar(*occ["Almoço"])),
            ("Jantar", bar(*occ["Jantar"])),
        ]
    )


def _mostrar_tabela_tipos(t: dict, com_pa_lanche=True):
    print("\nTotais por tipo:")
    rows = []
    if com_pa_lanche:
        rows += [
            ["Peq.Alm.", t["pa"], "-", "-", "-"],
            ["Lanche", t["lan"], "-", "-", "-"],
        ]
    rows += [
        ["Almoço", t["alm_norm"], t["alm_veg"], t["alm_dieta"], "-"],
        ["Jantar", t["jan_norm"], t["jan_veg"], t["jan_dieta"], f"Sai:{t['jan_sai']}"],
    ]
    print_table(["Turno", "Normal", "Vegetariano", "Dieta", "Extra"], rows)


def painel_operacional_do_dia(d: date):
    clear()
    line(f"🍳 Painel Operacional - {d.strftime('%A %d/%m/%Y')}")
    occ = get_ocupacao_capacidade(d)
    t = get_totais_dia(d.isoformat())
    _mostrar_barras_occ(occ)
    _mostrar_tabela_tipos(t)
    print("\n [1] Exportar CSV + Excel do dia  [0] Voltar")
    if input("Opção: ").strip() == "1":
        exportacoes_do_dia(d)
        input("ENTER...")


def painel_cmd_do_dia(d: date, ano: int):
    clear()
    line(f"🪖 Painel CMD {ano}º - {d.strftime('%A %d/%m/%Y')}")
    # Ocupação filtrada por ano
    di = d.isoformat()
    with db() as conn:
        row = conn.execute(
            """
            SELECT SUM(r.pequeno_almoco) pa, SUM(r.lanche) lan,
                   SUM(CASE WHEN r.almoco IS NOT NULL THEN 1 ELSE 0 END) alm,
                   SUM(CASE WHEN r.jantar_tipo IS NOT NULL THEN 1 ELSE 0 END) jan
            FROM refeicoes r JOIN utilizadores u ON u.id=r.utilizador_id
            WHERE r.data=? AND u.ano=?
        """,
            (di, ano),
        ).fetchone()
        caps = {
            r["refeicao"]: r["max_total"]
            for r in conn.execute(
                "SELECT refeicao,max_total FROM capacidade_refeicao WHERE data=?", (di,)
            )
        }
    occ = {
        "Pequeno Almoço": (row["pa"] or 0, caps.get("Pequeno Almoço", -1)),
        "Lanche": (row["lan"] or 0, caps.get("Lanche", -1)),
        "Almoço": (row["alm"] or 0, caps.get("Almoço", -1)),
        "Jantar": (row["jan"] or 0, caps.get("Jantar", -1)),
    }
    print("\nOcupação do ano vs Capacidade do dia:")
    _mostrar_barras_occ(occ)
    t = get_totais_dia(di, ano)
    _mostrar_tabela_tipos(t, com_pa_lanche=False)
    print("\n [1] Exportar CSV + Excel do dia (ano)  [0] Voltar")
    if input("Opção: ").strip() == "1":
        exportacoes_do_dia(d, ano)
        input("ENTER...")


# ===========================================================================
# ESTATÍSTICAS
# ===========================================================================


def stats_today(ano: str = "all"):
    with db() as conn:
        hoje = date.today().isoformat()
        ano_int = None if ano == "all" else int(ano)
        t = get_totais_dia(hoje, ano_int)
        total = conn.execute(
            "SELECT COUNT(*) c FROM utilizadores"
            + ("" if ano == "all" else " WHERE ano=?"),
            (() if ano == "all" else (int(ano),)),
        ).fetchone()["c"]
    return total, t


def stats_menu(ano: str = "all"):
    total, t = stats_today(ano)
    clear()
    line("📊 ESTATÍSTICAS")
    print(f"👥 Total utilizadores: {total}")
    print("🍽️ Hoje (contagens):")
    print_kv(
        [
            ("Pequeno Almoço", t["pa"]),
            ("Lanche", t["lan"]),
            ("Almoço Normal", t["alm_norm"]),
            ("Almoço Vegetariano", t["alm_veg"]),
            ("Almoço Dieta", t["alm_dieta"]),
            ("Jantar Normal", t["jan_norm"]),
            ("Jantar Vegetariano", t["jan_veg"]),
            ("Jantar Dieta", t["jan_dieta"]),
            ("Saíram após Jantar", t["jan_sai"]),
        ]
    )
    occ = get_ocupacao_capacidade(date.today())
    print("\nOcupação vs Capacidade (hoje):")
    _mostrar_barras_occ(occ)
    input("\nENTER para voltar...")


# ===========================================================================
# PROXIMIDADE / CALENDÁRIO
# ===========================================================================


def proximos_dias(n=7):
    hoje = date.today()
    nomes = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
    return [
        {
            "data": hoje + timedelta(days=i),
            "nome": nomes[(hoje + timedelta(days=i)).weekday()],
            "str": (hoje + timedelta(days=i)).strftime("%d/%m/%Y"),
        }
        for i in range(n)
    ]


def _parse_data_or_today(txt: str) -> date:
    try:
        return datetime.strptime((txt or "").strip(), "%Y-%m-%d").date()
    except Exception:
        return date.today()


def escolher_dia() -> Optional[date]:
    dias = proximos_dias()
    while True:
        clear()
        line("📅 Escolher Dia")
        for i, d in enumerate(dias, start=1):
            print(f"  [{i}] {d['nome']} ({d['str']})")
        print("  [0] Voltar")
        x = input("Opção: ").strip()
        if x == "0":
            return None
        if x.isdigit() and 1 <= int(x) <= len(dias):
            return dias[int(x) - 1]["data"]
        print("❌ Opção inválida.")
        input("ENTER...")


def escolher_tipo(titulo: str) -> str:
    opcoes = {"1": "Normal", "2": "Vegetariano", "3": "Dieta"}
    while True:
        print(f"\nEscolha o tipo de {titulo}:")
        for k, v in opcoes.items():
            print(f"  [{k}] {v}")
        x = input("Opção: ").strip()
        if x in opcoes:
            return opcoes[x]
        print("❌ Opção inválida.")


# ===========================================================================
# MENU ALUNO
# ===========================================================================


def menu_aluno(u: dict):
    while True:
        clear()
        line(f"📌 Menu - {u['nome']} ({u['perfil']})")
        hoje = date.today()
        menu = get_menu_do_dia(hoje)

        if menu:
            print("🍽️ Menu do dia:")

            def a(k):
                return menu.get(k) or "-"

            print(f"  PA: {a('pequeno_almoco')}")
            print(f"  Lanche: {a('lanche')}")
            print(
                f"  Almoço: N:{a('almoco_normal')} | V:{a('almoco_veg')} | D:{a('almoco_dieta')}"
            )
            print(
                f"  Jantar: N:{a('jantar_normal')} | V:{a('jantar_veg')} | D:{a('jantar_dieta')}"
            )

        _resumo_semana_para_utilizador(u)

        print("\n  [1] Alterar refeições")
        print("  [2] Verificar refeições")
        print("  [3] Histórico (30 dias)")
        print("  [4] Alterar password")
        print("  [5] Sair")

        esc = input("Opção: ").strip()
        if esc == "1":
            d = escolher_dia()
            if d:
                ok, motivo = refeicao_editavel(d)
                if not ok:
                    print(f"\n{motivo}")
                    input("\nENTER...")
                else:
                    alterar_refeicao(u, d)
        elif esc == "2":
            ver_refeicoes(u)
        elif esc == "3":
            ver_historico(u)
        elif esc == "4":
            aluno_alterar_password(u)
        elif esc == "5":
            print("👋 Até já!")
            break
        else:
            print("❌ Opção inválida.")
            input("ENTER...")


def alterar_refeicao(u: dict, d: date):
    uid = user_id_by_nii(u["nii"])
    if not uid:
        print("❌ Utilizador não encontrado.")
        input("ENTER...")
        return
    r = refeicao_get(uid, d)
    while True:
        occ = get_ocupacao_capacidade(d)
        clear()
        line(f"🍽️ {d.strftime('%A %d/%m/%Y')}")

        def show(tag, val, cap):
            if cap == -1:
                print(f"   {tag:<15}: {val} (sem limite)")
            else:
                alerta = " ⚠️" if val >= cap else ""
                print(f"   {tag:<15}: {val}/{cap}{alerta}")

        show("Peq. Almoço", *occ["Pequeno Almoço"])
        show("Lanche", *occ["Lanche"])
        show("Almoço", *occ["Almoço"])
        show("Jantar", *occ["Jantar"])

        print("\nSeleção atual:")
        print(f"  [1] Pequeno Almoço: {'✅' if r['pequeno_almoco'] else '❌'}")
        print(f"  [2] Lanche:         {'✅' if r['lanche'] else '❌'}")
        print(f"  [3] Almoço:         {r['almoco'] or '❌'}")
        extra = " (Sai)" if r.get("jantar_sai_unidade") else ""
        print(f"  [4] Jantar:         {r['jantar_tipo'] or '❌'}{extra}")
        print("  ─────────────────────────────────")
        print("  [5] Guardar alterações")
        print("  [6] Cancelar TODAS as refeições deste dia")
        print("  [0] Sair sem guardar")
        op = input("Opção: ").strip()
        try:
            if op == "1":
                r["pequeno_almoco"] = 0 if r["pequeno_almoco"] else 1
            elif op == "2":
                r["lanche"] = 0 if r["lanche"] else 1
            elif op == "3":
                if r["almoco"]:
                    if input("Remover almoço? (s/n): ").strip().lower() == "s":
                        r["almoco"] = None
                else:
                    r["almoco"] = escolher_tipo("Almoço")
            elif op == "4":
                if r["jantar_tipo"]:
                    r["jantar_tipo"] = None
                    r["jantar_sai_unidade"] = 0
                else:
                    r["jantar_tipo"] = escolher_tipo("Jantar")
                    r["jantar_sai_unidade"] = (
                        1
                        if input("Sai após jantar? (1=Sim, 2=Não): ").strip() == "1"
                        else 0
                    )
            elif op == "5":
                ok = refeicao_save(uid, d, r, alterado_por=u["nii"])
                if ok:
                    print("✅ Guardado.")
                    input("ENTER...")
                return
            elif op == "6":
                conf = (
                    input("⚠️  Cancelar TODAS as refeições deste dia? (s/n): ")
                    .strip()
                    .lower()
                )
                if conf == "s":
                    r_vazio = {
                        "pequeno_almoco": 0,
                        "lanche": 0,
                        "almoco": None,
                        "jantar_tipo": None,
                        "jantar_sai_unidade": 0,
                    }
                    ok = refeicao_save(uid, d, r_vazio, alterado_por=u["nii"])
                    if ok:
                        print("✅ Refeições canceladas.")
                        input("ENTER...")
                    return
            elif op == "0":
                return
            else:
                print("❌ Opção inválida.")
                input("ENTER...")
        except sqlite3.IntegrityError as e:
            print("❌ Bloqueado pela BD:", e)
            input("ENTER...")


def ver_refeicoes(u: dict):
    uid = user_id_by_nii(u["nii"])
    if not uid:
        print("❌ Utilizador não encontrado.")
        input("ENTER...")
        return
    clear()
    line("📋 Minhas Refeições")
    for d in proximos_dias():
        r = refeicao_get(uid, d["data"])
        print(
            f"{d['nome']} {d['str']} | PA:{'✅' if r['pequeno_almoco'] else '❌'} "
            f"Lan:{'✅' if r['lanche'] else '❌'} "
            f"Almoço:{r['almoco'] or '❌'} "
            f"Jantar:{r['jantar_tipo'] or '❌'}{' (Sai)' if r['jantar_sai_unidade'] else ''}"
        )
    input("\nENTER para voltar...")


def ver_historico(u: dict, dias=30):
    uid = user_id_by_nii(u["nii"])
    if not uid:
        print("❌ Utilizador não encontrado.")
        input("ENTER...")
        return
    clear()
    line(f"🕘 Histórico ({dias} dias) — {u['nome']}")
    data_min = (date.today() - timedelta(days=dias)).isoformat()
    with db() as conn:
        rows = conn.execute(
            """
            SELECT data, pequeno_almoco, lanche, almoco, jantar_tipo, jantar_sai_unidade
            FROM refeicoes WHERE utilizador_id=? AND data>=?
            ORDER BY data DESC
        """,
            (uid, data_min),
        ).fetchall()

    if not rows:
        print("Sem registos.")
        input("ENTER...")
        return

    # ── Estatísticas pessoais ──────────────────────────────────────────────
    total_pa = sum(1 for r in rows if r["pequeno_almoco"])
    total_lan = sum(1 for r in rows if r["lanche"])
    total_alm = sum(1 for r in rows if r["almoco"])
    total_jan = sum(1 for r in rows if r["jantar_tipo"])
    total_dias = len(rows)

    alm_tipos = {}
    jan_tipos = {}
    for r in rows:
        if r["almoco"]:
            alm_tipos[r["almoco"]] = alm_tipos.get(r["almoco"], 0) + 1
        if r["jantar_tipo"]:
            jan_tipos[r["jantar_tipo"]] = jan_tipos.get(r["jantar_tipo"], 0) + 1

    print(
        f"\n  📊 Estatísticas dos últimos {dias} dias ({total_dias} dias com registo):"
    )
    print(f"  {'─' * 46}")
    print(f"  {'Pequeno Almoço':<22}: {total_pa:>3} refeições")
    print(f"  {'Lanche':<22}: {total_lan:>3} refeições")
    print(f"  {'Almoço (total)':<22}: {total_alm:>3} refeições", end="")
    if alm_tipos:
        detalhe = "  (" + ", ".join(f"{k}:{v}" for k, v in alm_tipos.items()) + ")"
        print(detalhe, end="")
    print()
    print(f"  {'Jantar (total)':<22}: {total_jan:>3} refeições", end="")
    if jan_tipos:
        detalhe = "  (" + ", ".join(f"{k}:{v}" for k, v in jan_tipos.items()) + ")"
        print(detalhe, end="")
    print()
    saidas = sum(1 for r in rows if r["jantar_sai_unidade"])
    if saidas:
        print(f"  {'Saídas após jantar':<22}: {saidas:>3}")
    # Taxa de presença (dias com pelo menos 1 refeição vs dias úteis no período)
    dias_uteis = sum(
        1 for i in range(dias) if (date.today() - timedelta(days=i)).weekday() < 5
    )
    if dias_uteis:
        taxa = round(total_dias / dias_uteis * 100)
        print(
            f"  {'Taxa de presença':<22}: {taxa:>3}%  ({total_dias}/{dias_uteis} dias úteis)"
        )
    print(f"  {'─' * 46}\n")

    # ── Tabela diária ─────────────────────────────────────────────────────
    out = []
    for r in rows:
        dt = datetime.strptime(r["data"], "%Y-%m-%d")
        dia = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"][dt.weekday()]
        out.append(
            [
                dt.strftime("%d/%m/%Y") + f" ({dia})",
                "✅" if r["pequeno_almoco"] else "❌",
                "✅" if r["lanche"] else "❌",
                r["almoco"] or "❌",
                (r["jantar_tipo"] or "❌")
                + (" (Sai)" if r["jantar_sai_unidade"] else ""),
            ]
        )
    print_table(["Data", "PA", "Lan", "Almoço", "Jantar"], out)
    input("\nENTER para voltar...")


def _resumo_semana_para_utilizador(u: dict):
    """
    Resumo semanal compacto no ecrã principal do aluno.
    Mostra seleção atual + aviso se o prazo de alteração expirar em breve (< 24h).
    """
    uid = user_id_by_nii(u["nii"])
    print("\n┌─ 📆 PRÓXIMOS 7 DIAS ──────────────────────────────────────────────┐")
    nomes_abrev = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    for d_info in proximos_dias(7):
        d = d_info["data"]
        dia = nomes_abrev[d.weekday()]

        # Tipo do dia segundo calendário
        tipo = dia_operacional(d)
        if tipo in ("feriado", "exercicio"):
            icone_dia = "🔴" if tipo == "feriado" else "🟡"
            with db() as conn:
                cal = conn.execute(
                    "SELECT nota FROM calendario_operacional WHERE data=?",
                    (d.isoformat(),),
                ).fetchone()
            nota = f" ({cal['nota']})" if cal and cal["nota"] else f" ({tipo})"
            print(f"│  {dia} {d_info['str']}  {icone_dia}{nota}")
            continue
        if tipo == "fim_semana":
            print(f"│  {dia} {d_info['str']}  🔵 Fim de semana")
            continue

        # Dia normal — mostrar seleção do utilizador
        if uid:
            r = refeicao_get(uid, d)
            pa = "✅" if r["pequeno_almoco"] else "❌"
            lan = "✅" if r["lanche"] else "❌"
            alm = r["almoco"] or "❌"
            jan = r["jantar_tipo"] or "❌"
            sel = f"PA:{pa} Lan:{lan} Alm:{alm[:3]} Jan:{jan[:3]}"
        else:
            sel = "(sem dados)"

        # Aviso de prazo a expirar
        aviso = ""
        if PRAZO_LIMITE_HORAS is not None:
            prazo_dt = datetime(d.year, d.month, d.day) - timedelta(
                hours=PRAZO_LIMITE_HORAS
            )
            horas_ate = (prazo_dt - datetime.now()).total_seconds() / 3600
            if 0 < horas_ate <= 24:
                aviso = f"  ⚠️  Prazo em {int(horas_ate)}h!"
            elif horas_ate <= 0:
                aviso = "  🔒 Prazo expirado"

        print(f"│  {dia} {d_info['str']}  {sel}{aviso}")
    print("└────────────────────────────────────────────────────────────────────┘")


# ===========================================================================
# MENU COZINHA
# ===========================================================================


def cozinha_menu(u: dict):
    while True:
        clear()
        line(f"🍽️ COZINHA - {u['nome']}")
        print("  [1] Painel de hoje")
        print("  [2] Painel de outra data")
        print("  [3] Exportar (hoje)")
        print("  [4] Exportar (outra data)")
        print("  [0] Sair")
        x = input("Opção: ").strip()
        if x == "1":
            painel_operacional_do_dia(date.today())
        elif x == "2":
            d = _parse_data_or_today(input("Data (YYYY-MM-DD) [default=hoje]: "))
            painel_operacional_do_dia(d)
        elif x == "3":
            exportacoes_do_dia(date.today())
            input("ENTER...")
        elif x == "4":
            d = _parse_data_or_today(input("Data (YYYY-MM-DD) [default=hoje]: "))
            exportacoes_do_dia(d)
            input("ENTER...")
        elif x == "0":
            break
        else:
            print("❌ Opção inválida.")
            input("ENTER...")


# ===========================================================================
# MENU OFICIAL DE DIA
# ===========================================================================


def _oficialdia_excecoes(d: date, actor: dict = None):
    clear()
    line(f"📝 Exceções - {d.strftime('%A %d/%m/%Y')}")
    nii = input("NII do utilizador: ").strip()
    u = user_by_nii(nii)
    if not u:
        print("❌ NII não encontrado.")
        input("ENTER...")
        return
    uid = u["id"]
    r = refeicao_get(uid, d)
    actor_nii = actor["nii"] if actor else "oficialdia"

    while True:
        clear()
        line(
            f"📝 Exceções - {d.strftime('%A %d/%m/%Y')} | {u['Nome_completo']} (NI {u['NI']})"
        )
        print(f"  [1] Pequeno-almoço: {'✅' if r['pequeno_almoco'] else '❌'}")
        print(f"  [2] Lanche         : {'✅' if r['lanche'] else '❌'}")
        print(f"  [3] Almoço         : {r['almoco'] or '❌'}")
        jt = r["jantar_tipo"] or "❌"
        print(
            f"  [4] Jantar         : {jt}{' (Sai)' if r.get('jantar_sai_unidade') else ''}"
        )
        print("  [5] Guardar e voltar")
        print("  [0] Cancelar")
        op = input("Opção: ").strip()
        if op == "1":
            r["pequeno_almoco"] = 0 if r["pequeno_almoco"] else 1
        elif op == "2":
            r["lanche"] = 0 if r["lanche"] else 1
        elif op == "3":
            if r["almoco"]:
                if input("Remover almoço? (s/n): ").strip().lower() == "s":
                    r["almoco"] = None
            else:
                r["almoco"] = escolher_tipo("Almoço")
        elif op == "4":
            if r["jantar_tipo"]:
                if input("Remover jantar? (s/n): ").strip().lower() == "s":
                    r["jantar_tipo"] = None
                    r["jantar_sai_unidade"] = 0
            else:
                r["jantar_tipo"] = escolher_tipo("Jantar")
                r["jantar_sai_unidade"] = (
                    1
                    if input("Sai após jantar? (1=Sim, 2=Não): ").strip() == "1"
                    else 0
                )
        elif op == "5":
            if refeicao_save(uid, d, r, alterado_por=actor_nii):
                print("✅ Guardado.")
                input("ENTER...")
            return
        elif op == "0":
            return
        else:
            print("❌ Opção inválida.")
            input("ENTER...")


def oficialdia_menu(u: dict):
    while True:
        clear()
        line(f"🪖 OFICIAL DE DIA - {u['nome']}")
        print("  [1] Painel de hoje")
        print("  [2] Painel de outra data")
        print("  [3] Exceções (hoje)")
        print("  [4] Exceções (outra data)")
        print("  [5] Exportar (hoje)")
        print("  [6] Exportar (outra data)")
        print("  [7] Excedentes do dia")
        print("  [8] Gerir ausências")
        print("  [0] Voltar")
        x = input("Opção: ").strip()
        if x == "1":
            painel_operacional_do_dia(date.today())
        elif x == "2":
            d = _parse_data_or_today(input("Data (YYYY-MM-DD) [default=hoje]: "))
            painel_operacional_do_dia(d)
        elif x == "3":
            _oficialdia_excecoes(date.today(), actor=u)
        elif x == "4":
            d = _parse_data_or_today(input("Data (YYYY-MM-DD) [default=hoje]: "))
            _oficialdia_excecoes(d, actor=u)
        elif x == "5":
            exportacoes_do_dia(date.today())
            input("ENTER...")
        elif x == "6":
            d = _parse_data_or_today(input("Data (YYYY-MM-DD) [default=hoje]: "))
            exportacoes_do_dia(d)
            input("ENTER...")
        elif x == "7":
            admin_excedentes()
        elif x == "8":
            admin_gerir_ausencias(u["nii"])
        elif x == "0":
            break
        else:
            print("❌ Opção inválida.")
            input("ENTER...")


# ===========================================================================
# MENU CMD
# ===========================================================================


def cmd_menu(u: dict):
    while True:
        clear()
        line(f"🪖 CMD {u['ano']}º - {u['nome']}")
        print("  [1] Listar utilizadores do ano")
        print("  [2] Ver refeições de um utilizador do ano")
        print("  [3] Estatísticas do ano")
        print("  [4] Painel do dia (barras)")
        print("  [5] Exportar CSV + Excel do dia (ano)")
        print("  [6] Dashboard 14 dias (ano)")
        print("  [0] Voltar")
        op = input("Opção: ").strip()
        if op == "1":
            admin_listar_utilizadores(None, u["ano"])
        elif op == "2":
            nii = input("NII: ").strip()
            with db() as conn:
                r = conn.execute(
                    "SELECT id,Nome_completo,NI,ano FROM utilizadores WHERE NII=?",
                    (nii,),
                ).fetchone()
                if not r:
                    print("❌ Não encontrado.")
                    input("ENTER...")
                    continue
                if str(r["ano"]) != str(u["ano"]):
                    print("❌ Outro ano.")
                    input("ENTER...")
                    continue
                clear()
                line(f"📋 {r['Nome_completo']} (NI {r['NI']})")
                for row in conn.execute(
                    """
                    SELECT data, pequeno_almoco, lanche, almoco, jantar_tipo, jantar_sai_unidade
                    FROM refeicoes WHERE utilizador_id=? ORDER BY data DESC LIMIT 14
                """,
                    (r["id"],),
                ).fetchall():
                    dt = datetime.strptime(row["data"], "%Y-%m-%d")
                    dia = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"][
                        dt.weekday()
                    ]
                    print(
                        f"{dia} {dt.strftime('%d/%m/%Y')} | PA:{'✅' if row['pequeno_almoco'] else '❌'} "
                        f"Lan:{'✅' if row['lanche'] else '❌'} "
                        f"Almoço:{row['almoco'] or '❌'} "
                        f"Jantar:{row['jantar_tipo'] or '❌'}{' (Sai)' if row['jantar_sai_unidade'] else ''}"
                    )
                input("\nENTER...")
        elif op == "3":
            stats_menu(u["ano"])
        elif op == "4":
            d = _parse_data_or_today(input("Data (YYYY-MM-DD) [default=hoje]: "))
            painel_cmd_do_dia(d, int(u["ano"]))
        elif op == "5":
            d = _parse_data_or_today(input("Data (YYYY-MM-DD) [default=hoje]: "))
            exportacoes_do_dia(d, int(u["ano"]))
            input("ENTER...")
        elif op == "6":
            dashboard_analitico(str(u["ano"]), dias=14)
        elif op == "0":
            break
        else:
            print("❌ Opção inválida.")
            input("ENTER...")


# ===========================================================================
# MENU ADMIN
# ===========================================================================


def admin_listar_utilizadores(query=None, ano="all"):
    clear()
    line("👥 LISTA DE UTILIZADORES")
    with db() as conn:
        if query:
            sql = """
                SELECT u.NII, u.NI, u.Nome_completo, u.ano, u.perfil
                FROM utilizadores u JOIN utilizadores_fts f ON f.rowid=u.id
                WHERE utilizadores_fts MATCH ?"""
            args = (query,)
            if ano != "all":
                sql += " AND u.ano=?"
                args += (int(ano),)
            sql += " ORDER BY u.ano, u.NI"
            rows = conn.execute(sql, args).fetchall()
        else:
            if ano == "all":
                rows = conn.execute(
                    "SELECT NII,NI,Nome_completo,ano,perfil FROM utilizadores ORDER BY ano, NI"
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT NII,NI,Nome_completo,ano,perfil FROM utilizadores WHERE ano=? ORDER BY NI",
                    (int(ano),),
                ).fetchall()
    for r in rows:
        print(
            f"NI:{r['NI']:<6}  NII:{r['NII']:<12}  Ano:{r['ano']}  "
            f"Perfil:{r['perfil']:<10}  Nome:{r['Nome_completo']}"
        )
    print(f"\nTotal: {len(rows)}")
    input("ENTER...")


def admin_diagnostico_bd():
    clear()
    line("🩺 DIAGNÓSTICO DA BASE DE DADOS")
    print(f"BD atual: {BASE_DADOS}\n")
    with db() as conn:
        tabs = [
            r["name"]
            for r in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
            )
        ]
        print("Tabelas:", ", ".join(tabs))
        try:
            tot = conn.execute("SELECT COUNT(*) c FROM utilizadores").fetchone()["c"]
            print(f"\nUtilizadores na BD: {tot}")
            amostras = conn.execute("""
                SELECT NII, NI, Nome_completo, ano, perfil
                FROM utilizadores ORDER BY ano, NI LIMIT 10
            """).fetchall()
            print("\nAlguns registos:")
            for r in amostras:
                print(
                    f"  {r['NII']:<12}  NI:{r['NI']:<6}  Ano:{r['ano']}  "
                    f"Perfil:{r['perfil']:<10}  {r['Nome_completo']}"
                )
        except sqlite3.Error as e:
            print("⚠️ Erro a consultar utilizadores:", e)
    input("\nENTER para voltar…")


def admin_set_menu_capacidade():
    clear()
    line("🗓️ Menu & Capacidade (dia)")
    d = input("Data (YYYY-MM-DD) [default=hoje]: ").strip() or date.today().isoformat()
    with db() as conn:
        exist = conn.execute(
            "SELECT * FROM menus_diarios WHERE data=?", (d,)
        ).fetchone()
        if not exist:
            try:
                dt = datetime.strptime(d, "%Y-%m-%d").date()
            except Exception:
                dt = date.today()
            sug = conn.execute(
                """
                SELECT * FROM menus_diarios
                WHERE strftime('%w', data) = strftime('%w', ?)
                ORDER BY data DESC LIMIT 1
            """,
                (dt.isoformat(),),
            ).fetchone()
        else:
            sug = exist

    print("Menus (ENTER para manter sugestão)")

    def ask(label, key):
        sug_val = sug[key] if (sug and key in sug.keys()) else None
        v = input(f"  {label} [{sug_val or ''}]: ").strip()
        return v if v != "" else (sug_val or None)

    pa = ask("Pequeno-almoço", "pequeno_almoco")
    lan = ask("Lanche", "lanche")
    an = ask("Almoço Normal", "almoco_normal")
    av = ask("Almoço Veg", "almoco_veg")
    ad = ask("Almoço Dieta", "almoco_dieta")
    jn = ask("Jantar Normal", "jantar_normal")
    jv = ask("Jantar Veg", "jantar_veg")
    jd = ask("Jantar Dieta", "jantar_dieta")

    caps_sug = {}
    with db() as conn:
        for ref in ("Pequeno Almoço", "Lanche", "Almoço", "Jantar"):
            c = conn.execute(
                "SELECT max_total FROM capacidade_refeicao WHERE data=? AND refeicao=?",
                (d, ref),
            ).fetchone()
            if c:
                caps_sug[ref] = c["max_total"]
            else:
                cs = conn.execute(
                    """
                    SELECT max_total FROM capacidade_refeicao
                    WHERE refeicao=? AND data IN (
                        SELECT data FROM menus_diarios
                        WHERE strftime('%w', data) = strftime('%w', ?)
                    ) ORDER BY data DESC LIMIT 1
                """,
                    (ref, d),
                ).fetchone()
                caps_sug[ref] = cs["max_total"] if cs else None

    caps = {}
    for ref in ("Pequeno Almoço", "Lanche", "Almoço", "Jantar"):
        sug_val = caps_sug.get(ref)
        x = input(
            f"Capacidade {ref} (-1=sem limite) [{'' if sug_val is None else sug_val}]: "
        ).strip()
        if x != "":
            try:
                caps[ref] = int(x)
            except Exception:
                print(f"  ⚠️ ignorado valor inválido em {ref}.")
        elif sug_val is not None:
            caps[ref] = int(sug_val)

    with db() as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO menus_diarios
              (data,pequeno_almoco,lanche,almoco_normal,almoco_veg,almoco_dieta,jantar_normal,jantar_veg,jantar_dieta)
            VALUES (?,?,?,?,?,?,?,?,?)
        """,
            (
                d,
                pa or None,
                lan or None,
                an or None,
                av or None,
                ad or None,
                jn or None,
                jv or None,
                jd or None,
            ),
        )
        for ref, cap in caps.items():
            if cap < 0:
                conn.execute(
                    "DELETE FROM capacidade_refeicao WHERE data=? AND refeicao=?",
                    (d, ref),
                )
            else:
                conn.execute(
                    "INSERT OR REPLACE INTO capacidade_refeicao(data,refeicao,max_total) VALUES (?,?,?)",
                    (d, ref, cap),
                )
        conn.commit()
    print("✅ Atualizado.")
    input("ENTER...")


def admin_exportar():
    clear()
    line("📤 EXPORTAR")
    print(" [1] Utilizadores")
    print(" [2] Refeições (30 dias)")
    print(" [3] Ocupação vs Capacidade (dia)")
    print(" [4] Excedentes (dia)")
    print(" [5] Totais da cozinha (dia)")
    print(" [6] Relatório semanal (exportar)")
    print(" [0] Voltar")
    op = input("Opção: ").strip()
    ts = datetime.now().strftime("%Y%m%d")

    if op == "1":
        with db() as conn:
            rows = [
                dict(r)
                for r in conn.execute(
                    "SELECT NII,NI,Nome_completo,ano,perfil,data_criacao FROM utilizadores ORDER BY ano,NI"
                )
            ]
        pc, px = export_both(
            rows,
            ["NII", "NI", "Nome_completo", "ano", "perfil", "data_criacao"],
            f"utilizadores_{ts}",
        )
        print(f"✅ {pc}\n✅ {px}")
        input("ENTER...")

    elif op == "2":
        d0 = (date.today() - timedelta(days=30)).isoformat()
        with db() as conn:
            rows = [
                dict(r)
                for r in conn.execute(
                    """
                SELECT u.NII,u.NI,u.Nome_completo,r.data,
                       r.pequeno_almoco,r.lanche,r.almoco,r.jantar_tipo,r.jantar_sai_unidade
                FROM refeicoes r JOIN utilizadores u ON u.id=r.utilizador_id
                WHERE r.data>=? ORDER BY r.data DESC, u.ano, u.NI
            """,
                    (d0,),
                )
            ]
        hdrs = [
            "NII",
            "NI",
            "Nome_completo",
            "data",
            "pequeno_almoco",
            "lanche",
            "almoco",
            "jantar_tipo",
            "jantar_sai_unidade",
        ]
        pc, px = export_both(rows, hdrs, f"refeicoes_{ts}")
        print(f"✅ {pc}\n✅ {px}")
        input("ENTER...")

    elif op == "3":
        d = (
            input("Data (YYYY-MM-DD) [default=hoje]: ").strip()
            or date.today().isoformat()
        )
        with db() as conn:
            rows = [
                dict(r)
                for r in conn.execute("SELECT * FROM v_ocupacao_dia WHERE data=?", (d,))
            ]
        pc, px = export_both(
            rows, ["data", "refeicao", "ocupacao", "capacidade"], f"ocupacao_{d}"
        )
        print(f"✅ {pc}\n✅ {px}")
        input("ENTER...")

    elif op == "4":
        d = (
            input("Data (YYYY-MM-DD) [default=hoje]: ").strip()
            or date.today().isoformat()
        )
        with db() as conn:
            rows = [
                dict(r)
                for r in conn.execute(
                    """
                SELECT data, refeicao, ocupacao, capacidade, criado_em
                FROM capacidade_excessos WHERE data=? ORDER BY criado_em ASC
            """,
                    (d,),
                )
            ]
        pc, px = export_both(
            rows,
            ["data", "refeicao", "ocupacao", "capacidade", "criado_em"],
            f"excedentes_{d}",
        )
        print(f"✅ {pc}\n✅ {px}")
        input("ENTER...")

    elif op == "5":
        d = (
            input("Data (YYYY-MM-DD) [default=hoje]: ").strip()
            or date.today().isoformat()
        )
        t = get_totais_dia(d)
        row = _totais_para_csv_row(d, t)
        pc, px = export_both([row], _HEADERS_TOTAIS, f"cozinha_{d}")
        print(f"✅ {pc}\n✅ {px}")
        input("ENTER...")

    elif op == "6":
        admin_relatorio_semanal(exportar_direto=True)


def admin_criar_utilizador():
    clear()
    line("🆕 CRIAR UTILIZADOR")
    nii = input("NII: ").strip()
    ni = input("NI: ").strip()
    nome = input("Nome completo: ").strip()
    perfil = input("Perfil (aluno/admin/cmd/cozinha/oficialdia): ").strip()
    if len(ni) == 3 and ni[0] in "123456":
        ano = int(ni[0])
    elif len(ni) == 4 and ni.startswith("7") and ni[1] in "123":
        ano = int(ni[1])
    else:
        print("❌ NI inválido.")
        input("ENTER...")
        return
    pw1 = masked_input("Password inicial: ")
    with db() as conn:
        try:
            conn.execute(
                """
                INSERT INTO utilizadores
                  (NII,NI,Nome_completo,Palavra_chave,ano,perfil,must_change_password,password_updated_at)
                VALUES (?,?,?,?,?,?,0, datetime('now','localtime'))
            """,
                (nii, ni, nome, pw1, ano, perfil),
            )
            conn.commit()
            print("✅ Criado.")
        except sqlite3.Error as e:
            print("❌", e)
    input("ENTER...")


def admin_reset_password():
    """Permite ao admin fazer reset da password de um utilizador."""
    clear()
    line("🔑 RESET DE PASSWORD")
    nii = input("NII do utilizador: ").strip()
    with db() as conn:
        u = conn.execute(
            "SELECT id, Nome_completo, NI FROM utilizadores WHERE NII=?", (nii,)
        ).fetchone()
    if not u:
        print("❌ NII não encontrado.")
        input("ENTER...")
        return

    print(f"\nUtilizador: {u['Nome_completo']} (NI {u['NI']})")
    print(" [1] Definir nova password manualmente")
    print(" [2] Gerar password temporária aleatória")
    print(" [0] Cancelar")
    op = input("Opção: ").strip()

    if op == "1":
        nova = masked_input("Nova password: ")
        if not nova:
            print("❌ Password não pode ser vazia.")
            input("ENTER...")
            return
        conf = masked_input("Confirmar password: ")
        if nova != conf:
            print("❌ Passwords não coincidem.")
            input("ENTER...")
            return
        nova_hash = nova
        with db() as conn:
            conn.execute(
                """
                UPDATE utilizadores
                SET Palavra_chave=?, must_change_password=1,
                    password_updated_at=datetime('now','localtime')
                WHERE id=?
            """,
                (nova_hash, u["id"]),
            )
            conn.commit()
        print(
            "✅ Password redefinida. O utilizador será obrigado a alterá-la no próximo login."
        )

    elif op == "2":
        import random
        import string

        chars = string.ascii_letters + string.digits
        pw_temp = "".join(random.choices(chars, k=10))
        nova_hash = pw_temp
        with db() as conn:
            conn.execute(
                """
                UPDATE utilizadores
                SET Palavra_chave=?, must_change_password=1,
                    password_updated_at=datetime('now','localtime')
                WHERE id=?
            """,
                (nova_hash, u["id"]),
            )
            conn.commit()
        print("\n  ✅ Password temporária gerada:")
        print("  ┌─────────────────────────────┐")
        print(f"  │  Password: {pw_temp:<17}  │")
        print("  └─────────────────────────────┘")
        print("  Entrega esta password ao utilizador pessoalmente.")
        print("  Será obrigado a alterá-la no próximo login.")
    input("\nENTER...")


def aluno_alterar_password(u: dict):
    """Permite ao aluno alterar a sua própria password."""
    uid = user_id_by_nii(u["nii"])
    if not uid:
        print("❌ Conta de sistema — não é possível alterar password.")
        input("ENTER...")
        return
    clear()
    line("🔑 ALTERAR PASSWORD")
    atual = masked_input("Password atual: ")
    with db() as conn:
        row = conn.execute(
            "SELECT Palavra_chave FROM utilizadores WHERE id=?", (uid,)
        ).fetchone()
    pw_guardada = row["Palavra_chave"] if row else None
    if not verify_password(atual, pw_guardada):
        print("❌ Password atual incorreta.")
        input("ENTER...")
        return
    nova = masked_input("Nova password: ")
    if not nova or len(nova) < 4:
        print("❌ A nova password deve ter pelo menos 4 caracteres.")
        input("ENTER...")
        return
    conf = masked_input("Confirmar nova password: ")
    if nova != conf:
        print("❌ As passwords não coincidem.")
        input("ENTER...")
        return
    with db() as conn:
        conn.execute(
            """
            UPDATE utilizadores
            SET Palavra_chave=?, must_change_password=0,
                password_updated_at=datetime('now','localtime')
            WHERE id=?
        """,
            (nova, uid),
        )
        conn.commit()
    print("✅ Password alterada com sucesso!")
    input("ENTER...")


def admin_editar_utilizador():
    clear()
    line("✏️ EDITAR UTILIZADOR")
    nii = input("NII: ").strip()
    with db() as conn:
        u = conn.execute("SELECT * FROM utilizadores WHERE NII=?", (nii,)).fetchone()
        if not u:
            print("❌ Não encontrado.")
            input("ENTER...")
            return
        print(
            f"Nome: {u['Nome_completo']} | NI: {u['NI']} | Ano: {u['ano']} | Perfil: {u['perfil']}"
        )
        print(" [1] Nome  [2] NI/Ano  [3] Perfil  [4] Desbloquear  [5] Voltar")
        op = input("Opção: ").strip()
        if op == "1":
            novo = input("Novo nome: ")
            conn.execute(
                "UPDATE utilizadores SET Nome_completo=? WHERE id=?", (novo, u["id"])
            )
        elif op == "2":
            novo = input("Novo NI: ").strip()
            if len(novo) == 3 and novo[0] in "123456":
                ano = int(novo[0])
            elif len(novo) == 4 and novo.startswith("7") and novo[1] in "123":
                ano = int(novo[1])
            else:
                print("❌ NI inválido.")
                input("ENTER...")
                return
            conn.execute(
                "UPDATE utilizadores SET NI=?, ano=? WHERE id=?", (novo, ano, u["id"])
            )
        elif op == "3":
            novo = input("Perfil (aluno/admin/cmd/cozinha/oficialdia): ").strip()
            conn.execute("UPDATE utilizadores SET perfil=? WHERE id=?", (novo, u["id"]))
        elif op == "4":
            conn.execute(
                "UPDATE utilizadores SET locked_until=NULL WHERE id=?", (u["id"],)
            )
        conn.commit()
        print("✅ Atualizado.")
        input("ENTER...")


def admin_eliminar_utilizador():
    clear()
    line("🗑️ ELIMINAR UTILIZADOR")
    nii = input("NII do utilizador a eliminar: ").strip()
    conf = input(f"Tem a certeza que quer eliminar '{nii}'? (s/n): ").strip().lower()
    if conf != "s":
        print("Operação cancelada.")
        input("ENTER...")
        return
    with db() as conn:
        try:
            cur = conn.execute("DELETE FROM utilizadores WHERE NII=?", (nii,))
            conn.commit()
            print("✅ Eliminado." if cur.rowcount > 0 else "❌ NII não encontrado.")
        except sqlite3.Error as e:
            print("❌ Erro ao eliminar:", e)
    input("ENTER...")


def admin_eventos_login():
    clear()
    line("📜 EVENTOS DE LOGIN (100 mais recentes)")
    with db() as conn:
        rows = conn.execute(
            "SELECT nii, sucesso, ip, criado_em FROM login_eventos ORDER BY id DESC LIMIT 100"
        ).fetchall()
    if not rows:
        print("Sem registos.")
    else:
        for r in rows:
            status = "✅" if r["sucesso"] else "❌"
            print(f"{r['criado_em']} | {status} | {r['nii']:<12} | {r['ip']}")
    input("\nENTER para voltar...")


def admin_excedentes():
    clear()
    line("⚠️ EXCEDENTES DE CAPACIDADE")
    d = input("Data (YYYY-MM-DD) [default=hoje]: ").strip() or date.today().isoformat()
    with db() as conn:
        rows = conn.execute(
            """
            SELECT data, refeicao, ocupacao, capacidade, criado_em
            FROM capacidade_excessos WHERE data=? ORDER BY criado_em ASC
        """,
            (d,),
        ).fetchall()
    if not rows:
        print(f"Sem registos de excedente em {d}.")
    else:
        for r in rows:
            over = r["ocupacao"] - r["capacidade"]
            print(
                f"{r['criado_em']} | {r['refeicao']:<15} | "
                f"ocupação {r['ocupacao']} / capacidade {r['capacidade']} | +{over}"
            )
        print(f"\nTotal registos: {len(rows)}")
        if input("\nExportar CSV + Excel? (s/n): ").strip().lower() == "s":
            pc, px = export_both(
                [dict(r) for r in rows],
                ["data", "refeicao", "ocupacao", "capacidade", "criado_em"],
                f"excedentes_{d}",
            )
            print(f"✅ {pc}\n✅ {px}")
    input("\nENTER para voltar...")


def admin_relatorio_semanal(exportar_direto=False):
    clear()
    line("🗓️ RELATÓRIO SEMANAL")
    d0_str = input("Data inicial (YYYY-MM-DD) [default=2ª desta semana]: ").strip()
    if d0_str:
        try:
            base = datetime.strptime(d0_str, "%Y-%m-%d").date()
        except Exception:
            base = date.today()
    else:
        today = date.today()
        base = today - timedelta(days=today.weekday())
    d1 = base + timedelta(days=6)
    print(f"Período: {base} a {d1}")

    dias = [(base + timedelta(days=i)).isoformat() for i in range(7)]
    res = []
    with db() as conn:
        for di in dias:
            t = get_totais_dia(di)
            r2 = conn.execute(
                """
                SELECT COALESCE(SUM(ocupacao - capacidade),0) over_sum, COUNT(*) registos
                FROM capacidade_excessos WHERE data=?
            """,
                (di,),
            ).fetchone()
            res.append(
                {
                    "data": di,
                    "PA": t["pa"],
                    "Lanche": t["lan"],
                    "Alm_Normal": t["alm_norm"],
                    "Alm_Veg": t["alm_veg"],
                    "Alm_Dieta": t["alm_dieta"],
                    "Jan_Normal": t["jan_norm"],
                    "Jan_Veg": t["jan_veg"],
                    "Jan_Dieta": t["jan_dieta"],
                    "Jan_Sai": t["jan_sai"],
                    "Excedente_total": r2["over_sum"] or 0,
                    "Excedente_registos": r2["registos"] or 0,
                }
            )

    print("\nDia       | PA  Lan | A:N  A:V  A:D | J:N  J:V  J:D | Sai | Exced(+)/regs")
    tot = {
        k: 0
        for k in [
            "PA",
            "Lanche",
            "Alm_Normal",
            "Alm_Veg",
            "Alm_Dieta",
            "Jan_Normal",
            "Jan_Veg",
            "Jan_Dieta",
            "Jan_Sai",
            "Excedente_total",
            "Excedente_registos",
        ]
    }
    for r in res:
        print(
            f"{r['data']} | {r['PA']:>3} {r['Lanche']:>3} | "
            f"{r['Alm_Normal']:>3} {r['Alm_Veg']:>3} {r['Alm_Dieta']:>3} | "
            f"{r['Jan_Normal']:>3} {r['Jan_Veg']:>3} {r['Jan_Dieta']:>3} | "
            f"{r['Jan_Sai']:>3} | +{r['Excedente_total']:>3}/{r['Excedente_registos']:>2}"
        )
        for k in tot:
            tot[k] += r[k]
    print(
        f"\nTOTAL SEMANAL: PA {tot['PA']} | Lanche {tot['Lanche']} | "
        f"Almoço N/V/D {tot['Alm_Normal']}/{tot['Alm_Veg']}/{tot['Alm_Dieta']} | "
        f"Jantar N/V/D {tot['Jan_Normal']}/{tot['Jan_Veg']}/{tot['Jan_Dieta']} | "
        f"Saem {tot['Jan_Sai']} | Excedentes +{tot['Excedente_total']} em {tot['Excedente_registos']} registos"
    )

    if (
        exportar_direto
        or input("\nExportar CSV + Excel? (s/n): ").strip().lower() == "s"
    ):
        name = f"relatorio_semanal_{base}_a_{d1}"
        hdrs = [
            "data",
            "PA",
            "Lanche",
            "Alm_Normal",
            "Alm_Veg",
            "Alm_Dieta",
            "Jan_Normal",
            "Jan_Veg",
            "Jan_Dieta",
            "Jan_Sai",
            "Excedente_total",
            "Excedente_registos",
        ]
        pc, px = export_both(res, hdrs, name)
        print(f"✅ {pc}\n✅ {px}")
    input("\nENTER para voltar...")


def admin_menu(u: dict):
    while True:
        clear()
        line(f"🔧 ADMIN - {u['nome']}")
        print("  [1]  Listar utilizadores")
        print("  [2]  Pesquisar por nome (FTS)")
        print("  [3]  Gerir utilizadores")
        print("  [4]  Estatísticas")
        print("  [5]  Exportar CSV + Excel")
        print("  [6]  Menu & Capacidade (dia)")
        print("  [7]  Eventos de login")
        print("  [8]  Backup BD")
        print("  [9]  Excedentes (por dia)")
        print("  [10] Relatório semanal")
        print("  [11] Dashboard analítico")
        print("  [12] Diagnóstico da BD")
        print("  [13] Gerir ausências")
        print("  [14] Calendário operacional")
        print("  [15] Log de alterações de refeições")
        print("  [0]  Sair")
        op = input("Opção: ").strip()
        if op == "1":
            admin_listar_utilizadores()
        elif op == "2":
            termo = input("Termo (ex: joao*): ")
            admin_listar_utilizadores(termo)
        elif op == "3":
            while True:
                clear()
                line("👥 GERIR UTILIZADORES")
                print(
                    " [1] Criar  [2] Editar  [3] Eliminar  [4] Reset Password  [5] Voltar"
                )
                x = input("Opção: ").strip()
                if x == "1":
                    admin_criar_utilizador()
                elif x == "2":
                    admin_editar_utilizador()
                elif x == "3":
                    admin_eliminar_utilizador()
                elif x == "4":
                    admin_reset_password()
                elif x == "5":
                    break
        elif op == "4":
            stats_menu("all")
        elif op == "5":
            admin_exportar()
        elif op == "6":
            admin_set_menu_capacidade()
        elif op == "7":
            admin_eventos_login()
        elif op == "8":
            do_backup()
            input("ENTER...")
        elif op == "9":
            admin_excedentes()
        elif op == "10":
            admin_relatorio_semanal()
        elif op == "11":
            y = input("Ano (1-6) ou 'all' [all]: ").strip() or "all"
            if y != "all" and not y.isdigit():
                y = "all"
            dashboard_analitico(y, dias=14)
        elif op == "12":
            admin_diagnostico_bd()
        elif op == "13":
            admin_gerir_ausencias(u["nii"])
        elif op == "14":
            admin_gerir_calendario()
        elif op == "15":
            admin_log_alteracoes()
        elif op == "0":
            break


# ===========================================================================
# PONTO DE ENTRADA
# ===========================================================================


def main():
    ensure_schema()
    if not sqlite_quick_check():
        print(
            "⚠️ PRAGMA quick_check não está OK — considera correr manutenção/recuperação."
        )
    ensure_daily_backup()
    limpar_backups_antigos()

    print(f"📁 Base de dados em uso: {BASE_DADOS}")

    # Auto-preenchimento corre apenas aqui, uma vez, no arranque
    print("🟢 A executar auto-preenchimento semanal...")
    autopreencher_refeicoes_semanais(14)
    print("✅ Auto-preenchimento concluído.")

    u = login_flow()
    if not u:
        return

    if u["perfil"] == "admin":
        admin_menu(u)
    elif u["perfil"] == "cmd":
        cmd_menu(u)
    elif u["perfil"] == "cozinha":
        cozinha_menu(u)
    elif u["perfil"] == "oficialdia":
        oficialdia_menu(u)
    else:
        menu_aluno(u)


if __name__ == "__main__":
    main()
