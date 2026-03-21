"""Tests for core/exports.py — CSV, XLSX, and daily export generation."""

from __future__ import annotations

import csv
import os
from datetime import date

import pytest

from tests.conftest import create_aluno


# ── Fixtures ─────────────────────────────────────────────────────────────


@pytest.fixture(autouse=True)
def _patch_export_dir(monkeypatch, tmp_path):
    """Redirect EXPORT_DIR to a temp directory for every test."""
    d = str(tmp_path)
    monkeypatch.setattr("core.constants.EXPORT_DIR", d)
    monkeypatch.setattr("core.exports.EXPORT_DIR", d)


@pytest.fixture
def sample_rows():
    return [
        {"nome": "Alice", "idade": "20", "curso": "Eng"},
        {"nome": "Bob", "idade": "22", "curso": "Med"},
    ]


@pytest.fixture
def sample_headers():
    return ["nome", "idade", "curso"]


# ── CSV tests ────────────────────────────────────────────────────────────


def test_export_csv_creates_file(tmp_path, sample_rows, sample_headers):
    from core.exports import export_csv

    path = export_csv(sample_rows, sample_headers, "report")

    assert os.path.isfile(path)
    assert path.endswith(".csv")

    with open(path, encoding="utf-8") as f:
        reader = list(csv.reader(f))

    # header + 2 data rows
    assert len(reader) == 3
    assert reader[0] == sample_headers
    assert reader[1] == ["Alice", "20", "Eng"]
    assert reader[2] == ["Bob", "22", "Med"]


def test_export_csv_empty_rows(tmp_path, sample_headers):
    from core.exports import export_csv

    path = export_csv([], sample_headers, "empty")

    assert os.path.isfile(path)
    with open(path, encoding="utf-8") as f:
        reader = list(csv.reader(f))

    # header only, no data rows
    assert len(reader) == 1
    assert reader[0] == sample_headers


# ── XLSX tests ───────────────────────────────────────────────────────────


def test_export_xlsx_creates_file(tmp_path, sample_rows, sample_headers):
    openpyxl = pytest.importorskip("openpyxl")
    from core.exports import export_xlsx

    path = export_xlsx(sample_rows, sample_headers, "report")

    assert os.path.isfile(path)
    assert path.endswith(".xlsx")

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # header row
    assert [ws.cell(1, c).value for c in range(1, 4)] == sample_headers
    # first data row
    assert ws.cell(2, 1).value == "Alice"
    assert ws.cell(2, 2).value == "20"
    # second data row
    assert ws.cell(3, 1).value == "Bob"
    assert ws.cell(3, 3).value == "Med"
    wb.close()


def test_export_xlsx_empty_rows(tmp_path, sample_headers):
    openpyxl = pytest.importorskip("openpyxl")
    from core.exports import export_xlsx

    path = export_xlsx([], sample_headers, "empty")

    assert os.path.isfile(path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # header present
    assert [ws.cell(1, c).value for c in range(1, 4)] == sample_headers
    # no data row
    assert ws.cell(2, 1).value is None
    wb.close()


def test_export_xlsx_column_widths(tmp_path, sample_rows, sample_headers):
    openpyxl = pytest.importorskip("openpyxl")
    from core.exports import export_xlsx

    path = export_xlsx(sample_rows, sample_headers, "widths")

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    for col_idx, h in enumerate(sample_headers, 1):
        from openpyxl.utils import get_column_letter

        letter = get_column_letter(col_idx)
        width = ws.column_dimensions[letter].width
        # width should be at least len(header) + 4 (the padding added by export_xlsx)
        assert width >= len(h) + 4
        # and capped at 50
        assert width <= 50
    wb.close()


# ── export_both ──────────────────────────────────────────────────────────


def test_export_both_returns_two_paths(tmp_path, sample_rows, sample_headers):
    pytest.importorskip("openpyxl")
    from core.exports import export_both

    csv_path, xlsx_path = export_both(sample_rows, sample_headers, "dual")

    assert csv_path.endswith(".csv")
    assert xlsx_path.endswith(".xlsx")
    assert os.path.isfile(csv_path)
    assert os.path.isfile(xlsx_path)


# ── exportacoes_do_dia (DB-dependent) ────────────────────────────────────


def test_exportacoes_do_dia_basic(app, tmp_path):
    pytest.importorskip("openpyxl")
    from core.exports import exportacoes_do_dia
    from core.meals import refeicao_save

    d = date(2025, 6, 15)

    with app.app_context():
        uid = create_aluno("EXP001", "NI001", "Aluno Export", ano="1")
        refeicao_save(
            uid,
            d,
            {
                "pequeno_almoco": "Sim",
                "lanche": "Sim",
                "almoco": "Normal",
                "jantar_tipo": "Normal",
                "jantar_sai_unidade": "Não",
            },
        )

        exportacoes_do_dia(d)

    export_dir = str(tmp_path)
    di = d.isoformat()
    # Should have created totais and distribuicao files (csv + xlsx each)
    assert os.path.isfile(os.path.join(export_dir, f"totais_{di}.csv"))
    assert os.path.isfile(os.path.join(export_dir, f"totais_{di}.xlsx"))
    assert os.path.isfile(os.path.join(export_dir, f"distribuicao_{di}.csv"))
    assert os.path.isfile(os.path.join(export_dir, f"distribuicao_{di}.xlsx"))
    assert os.path.isfile(os.path.join(export_dir, f"ocupacao_vs_capacidade_{di}.csv"))
    assert os.path.isfile(os.path.join(export_dir, f"ocupacao_vs_capacidade_{di}.xlsx"))


def test_exportacoes_do_dia_with_ano(app, tmp_path):
    pytest.importorskip("openpyxl")
    from core.exports import exportacoes_do_dia
    from core.meals import refeicao_save

    d = date(2025, 6, 16)

    with app.app_context():
        uid = create_aluno("EXP002", "NI002", "Aluno Export 2", ano="2")
        refeicao_save(
            uid,
            d,
            {
                "pequeno_almoco": "Sim",
                "lanche": "Não",
                "almoco": "Vegetariano",
                "jantar_tipo": "Normal",
                "jantar_sai_unidade": "Não",
            },
        )

        exportacoes_do_dia(d, ano=2)

    export_dir = str(tmp_path)
    di = d.isoformat()
    # With ano filter, filenames include _anoN tag
    assert os.path.isfile(os.path.join(export_dir, f"totais_ano2_{di}.csv"))
    assert os.path.isfile(os.path.join(export_dir, f"totais_ano2_{di}.xlsx"))
    assert os.path.isfile(os.path.join(export_dir, f"distribuicao_ano2_{di}.csv"))
    assert os.path.isfile(os.path.join(export_dir, f"distribuicao_ano2_{di}.xlsx"))
